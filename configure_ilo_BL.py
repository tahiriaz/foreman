#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import sys
import time
import re
import warnings
import logging
import threading
import concurrent.futures
import csv
from pathlib import Path
from enum import Enum
from typing import Dict, Any, List, Tuple, Optional
from xml.sax.saxutils import escape as xml_escape

# ============================================================================
# WARNING SUPPRESSION (Must be before third-party imports)
# ============================================================================

os.environ["PYTHONWARNINGS"] = "ignore"
warnings.simplefilter("ignore")
warnings.filterwarnings("ignore", category=UserWarning)
warnings.filterwarnings("ignore", category=DeprecationWarning)

with warnings.catch_warnings():
    warnings.simplefilter("ignore")
    try:
        import cryptography
    except Exception:
        pass
    import paramiko

import urllib3
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import pandas as pd
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ============================================================================
# LOGGING SETUP
# ============================================================================

def setup_logging() -> logging.Logger:
    """Configures thread-safe console logging."""
    logger = logging.getLogger("ilo_automation")
    logger.setLevel(logging.DEBUG)
    logger.propagate = False
    if logger.handlers:
        logger.handlers.clear()
    
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter('%(message)s'))
    
    logger.addHandler(ch)
    return logger

logger = setup_logging()

def log(slot: Any, ip: str, message: str):
    """Bridge function to maintain existing log signature while using native logging."""
    try:
        slot_str = f"{int(slot):02}"
    except Exception:
        slot_str = "??"
    logger.info(f"[Slot {slot_str} | {ip}] {message}")

# ============================================================================
# ENUMS & DATA STRUCTURES
# ============================================================================

class TaskStatus(Enum):
    PENDING = "PENDING"
    OK = "OK"
    FAIL = "FAIL"
    SKIP = "SKIP"

# ============================================================================
# CREDENTIALS & GLOBAL SETTINGS
# ============================================================================

OA_USERNAME = "Administrator"
OA_PASSWORD = "Th@les01"

ILO_LOGIN = "thlocaladmin"
ILO_PASSWORD = "Th@les018664"
ILO_ADVANCED_LICENSE_KEY = "35SCR-RYLML-CBK7N-TD3B9-GGBW2"

FENCE_USER_NAME = "Pacemaker Fence"
FENCE_USER_LOGIN = "hpilofence"
FENCE_USER_PASSWORD = "Th@les01"

RIBCL_CONCURRENT_SESSIONS = 4
REDFISH_CONCURRENT_SESSIONS = 16

IPMI_PORT = 623
ILO_DOMAIN = "mak.iss"
ILO_TIMEZONE_SEARCH = "Riyadh"

LDAP_PORT = 636
LDAP_GROUP_NAME = "CN=ILOAdmins,OU=Roles,OU=IT,OU=ISS,DC=mak,DC=iss"
LDAP_GROUP_SID = ""
LDAP_GROUP_PRIVILEGES = "1,2,3,4,5,6"
LDAP_USER_CONTEXTS = [
    "OU=IT,OU=ISS,DC=mak,DC=iss",
    "CN=Users,DC=mak,DC=iss",
    "CN=Builtin,DC=mak,DC=iss",
    "@mak.iss",
]

SCOPE_SETTINGS = {
    "SIL": {"PRIMARY_DNS": "10.130.2.11", "SECONDARY_DNS": "10.130.2.12", "DIRECTORY_SERVER": "INFCOSADU001MP.mak.iss", "NTP_PRIMARY": "10.101.18.1", "NTP_SECONDARY": "10.130.2.11"},
    "MTR": {"PRIMARY_DNS": "10.130.2.11", "SECONDARY_DNS": "10.130.2.12", "DIRECTORY_SERVER": "INFCOSADU001MP.mak.iss", "NTP_PRIMARY": "10.101.18.1", "NTP_SECONDARY": "10.130.2.11"},
    "RTR": {"PRIMARY_DNS": "10.130.4.11", "SECONDARY_DNS": "10.130.4.12", "DIRECTORY_SERVER": "INFCOSADU002MP.mak.iss", "NTP_PRIMARY": "10.102.18.1", "NTP_SECONDARY": "10.130.4.11"},
}

BASE_DIR = Path(__file__).resolve().parent
LOGS_DIR = BASE_DIR / "logs"
timestamp = time.strftime("%Y%m%d-%H%M%S")
report_filename = f"ILO_BL_Report_{timestamp}.csv"
CSV_REPORT_PATH = LOGS_DIR / report_filename

TEMPLATE_DIR = BASE_DIR / "Templates"
EXCEL_FILE = "Resource List-v7.6.xlsx"
SHEET_NAME = "General Resource List"
CERT_FILE = TEMPLATE_DIR / "iss_root_ca.crt"

VALID_EQUIPMENT_TYPES = ["NVR Blade Server", "VCA Blade Server"]
EXCEL_COLUMNS = {"enclosure_physical_name", "equipment_type", "enclosure_slot", "ilo_ip", "scope", "hostname", "ilo_hostname"}
REQUIRED_COLUMNS = ["enclosure_slot", "ilo_ip", "scope", "hostname", "ilo_hostname"]
EXCEL_EMPTY_ROW_STOP = 25

DEBUG = False
DEBUG_ON_FAILURE = True

SSH_PORT = 22
SSH_CONNECT_TIMEOUT = 15
SSH_KEEPALIVE_INTERVAL = 5
COMMAND_TIMEOUT = 120
COMMAND_QUIET_TIME = 20.0
HPONCFG_LINE_DELAY = 0.02

AUTH_RETRIES = 5
AUTH_RETRY_DELAY = 5
LOGIN_PENALTY_WAIT = 32
REDFISH_TIMEOUT = 25
HTTP_SAFE_RETRIES = 2

POST_ERROR_STRING = "UnableToModifyDuringSystemPOST"
POWER_OFF_POLL_INTERVAL = 5
POWER_OFF_TIMEOUT = 90
ILO_SETTLE_AFTER_POWEROFF = 5


# ============================================================================
# PAYLOAD BUILDERS (Decoupled RIBCL XML)
# ============================================================================

def xml_attr(value: Any) -> str:
    """Escape a value for use inside a double-quoted XML attribute."""
    return xml_escape(str(value), {"\"": "&quot;", "'": "&apos;"})


def build_ribcl_user_check() -> str:
    return f"""<RIBCL VERSION="2.0">
<LOGIN USER_LOGIN="{xml_attr(OA_USERNAME)}" PASSWORD="{xml_attr(OA_PASSWORD)}">
<USER_INFO MODE="read">
<GET_USER USER_LOGIN="{xml_attr(ILO_LOGIN)}"/>
</USER_INFO>
</LOGIN>
</RIBCL>"""


def build_ribcl_user_add() -> str:
    return f"""<RIBCL VERSION="2.0">
<LOGIN USER_LOGIN="{xml_attr(OA_USERNAME)}" PASSWORD="{xml_attr(OA_PASSWORD)}">
<USER_INFO MODE="write">
<ADD_USER USER_NAME="{xml_attr(ILO_LOGIN)}" USER_LOGIN="{xml_attr(ILO_LOGIN)}" PASSWORD="{xml_attr(ILO_PASSWORD)}">
<ADMIN_PRIV VALUE="Y"/><REMOTE_CONS_PRIV VALUE="Y"/><RESET_SERVER_PRIV VALUE="Y"/>
<VIRTUAL_MEDIA_PRIV VALUE="Y"/><CONFIG_ILO_PRIV VALUE="Y"/>
</ADD_USER>
</USER_INFO>
</LOGIN>
</RIBCL>"""


def build_ribcl_combined(hostname: str, directory_server: str, cert_data: str) -> str:
    return f"""<RIBCL VERSION="2.0">
<LOGIN USER_LOGIN="{xml_attr(OA_USERNAME)}" PASSWORD="{xml_attr(OA_PASSWORD)}">
<SERVER_INFO MODE="write">
<SERVER_NAME VALUE="{xml_attr(hostname)}"/>
</SERVER_INFO>
<DIR_INFO MODE="write">
<MOD_DIR_CONFIG>
<DIR_AUTHENTICATION_ENABLED VALUE="Yes"/>
<DIR_LOCAL_USER_ACCT VALUE="Yes"/>
<DIR_SERVER_ADDRESS VALUE="{xml_attr(directory_server)}"/>
<DIR_SERVER_PORT VALUE="{LDAP_PORT}"/>
<DIR_USER_CONTEXT_1 VALUE="{xml_attr(LDAP_USER_CONTEXTS[0])}"/>
<DIR_USER_CONTEXT_2 VALUE="{xml_attr(LDAP_USER_CONTEXTS[1])}"/>
<DIR_USER_CONTEXT_3 VALUE="{xml_attr(LDAP_USER_CONTEXTS[2])}"/>
<DIR_USER_CONTEXT_4 VALUE="{xml_attr(LDAP_USER_CONTEXTS[3])}"/>
<DIR_ENABLE_GRP_ACCT VALUE="Yes"/>
<DIR_GENERIC_LDAP_ENABLED VALUE="No"/>
<DIR_GRPACCT1_NAME VALUE="{xml_attr(LDAP_GROUP_NAME)}"/>
<DIR_GRPACCT1_PRIV VALUE="{xml_attr(LDAP_GROUP_PRIVILEGES)}"/>
<DIR_GRPACCT1_SID VALUE="{xml_attr(LDAP_GROUP_SID)}"/>
</MOD_DIR_CONFIG>
<IMPORT_LDAP_CA_CERTIFICATE>
{cert_data.strip()}
</IMPORT_LDAP_CA_CERTIFICATE>
</DIR_INFO>
</LOGIN>
</RIBCL>"""


def build_ribcl_ldap_verify() -> str:
    return f"""<RIBCL VERSION="2.0">
<LOGIN USER_LOGIN="{xml_attr(OA_USERNAME)}" PASSWORD="{xml_attr(OA_PASSWORD)}">
<DIR_INFO MODE="read">
<GET_DIR_CONFIG/>
</DIR_INFO>
</LOGIN>
</RIBCL>"""


# ============================================================================
# SANITIZATION & RESULT HELPERS
# ============================================================================

def sanitize_debug_output(output: str) -> str:
    if not output: return str(output)
    sanitized = str(output)
    for secret in (OA_PASSWORD, ILO_PASSWORD, FENCE_USER_PASSWORD, ILO_ADVANCED_LICENSE_KEY):
        if secret: sanitized = sanitized.replace(secret, "***REDACTED***")
    sanitized = re.sub(
        r"-----BEGIN CERTIFICATE-----.*?-----END CERTIFICATE-----",
        "-----BEGIN CERTIFICATE-----\n*** CERTIFICATE BODY REDACTED ***\n-----END CERTIFICATE-----",
        sanitized, flags=re.DOTALL,
    )
    return sanitized

def debug_block(title: str, output: str, force: bool = False):
    if not DEBUG and not force: return
    logger.debug("\n" + "=" * 80)
    logger.debug(title)
    logger.debug(sanitize_debug_output(output))
    logger.debug("=" * 80 + "\n")

def mark_failure(res_dict: Dict[str, Any], step: str, message: str):
    res_dict[step] = TaskStatus.FAIL.value
    res_dict["status"] = "Failed Tasks"
    res_dict["errors"].append(f"[{step}] {message}")


# ============================================================================
# GENERAL HELPERS
# ============================================================================

def split_ilo_hostname(value: Any) -> Tuple[str, str]:
    value = str(value).strip().rstrip(".")
    if not value: raise ValueError("Empty iLO hostname")
    parts = value.split(".", 1)
    hostname = parts[0]
    domain = parts[1] if len(parts) == 2 else ILO_DOMAIN
    if not re.match(r"^[A-Za-z0-9](?:[A-Za-z0-9-]*[A-Za-z0-9])?$", hostname):
        raise ValueError(f"Invalid iLO hostname label: '{hostname}'")
    if len(hostname) > 63:
        raise ValueError(f"iLO hostname label exceeds 63 characters: '{hostname}'")
    return hostname, domain

def values_equal_case_insensitive(left: Any, right: Any) -> bool:
    return str(left or "").strip().lower() == str(right or "").strip().lower()


# ============================================================================
# EXCEL LOADER
# ============================================================================

def load_resource_excel_fast(excel_path: Path, sheet_name: str, empty_row_stop: int = 100) -> Tuple[pd.DataFrame, float]:
    start = time.monotonic()
    wb = load_workbook(filename=str(excel_path), read_only=True, data_only=True)
    
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Worksheet '{sheet_name}' not found")
        ws = wb[sheet_name]
        header_map = None
        records = []
        empty_count = 0
        data_started = False
        wanted = {c.lower() for c in EXCEL_COLUMNS}

        for row in ws.iter_rows(values_only=True):
            if header_map is None:
                normalized = [str(val).strip().lower() if val is not None else "" for val in row]
                if "enclosure_physical_name" in normalized and "equipment_type" in normalized and "ilo_ip" in normalized:
                    header_map = {name: idx for idx, name in enumerate(normalized) if name in wanted}
                    missing = wanted - set(header_map)
                    if missing: raise RuntimeError("Missing Excel columns: " + ", ".join(sorted(missing)))
                continue

            record = {}
            has_value = False
            for column, index in header_map.items():
                val = row[index] if index < len(row) else None
                if val is None: val = ""
                elif isinstance(val, str): val = val.strip()
                record[column] = val
                if str(val).strip() != "": has_value = True

            if not has_value:
                if data_started:
                    empty_count += 1
                    if empty_count >= empty_row_stop: break
                continue

            data_started = True
            empty_count = 0
            records.append(record)
    finally:
        wb.close()

    if not records: raise RuntimeError("No data rows found in worksheet")
    df = pd.DataFrame(records)
    for column in EXCEL_COLUMNS:
        if column not in df.columns: df[column] = ""

    string_columns = ["enclosure_physical_name", "equipment_type", "ilo_ip", "scope", "hostname", "ilo_hostname"]
    for column in string_columns:
        df[column] = df[column].fillna("").astype(str).str.strip()

    df["scope"] = df["scope"].str.upper()
    return df, time.monotonic() - start


# ============================================================================
# RIBCL PARSING
# ============================================================================

def extract_ribcl_results(output: str) -> str:
    """Return the actual HPONCFG result section, excluding echoed request XML."""
    if not output:
        return ""

    match = re.search(
        r"START\s+RIBCL\s+RESULTS.*?(.*?)(?:END\s+RIBCL\s+RESULTS|$)",
        output,
        flags=re.IGNORECASE | re.DOTALL,
    )
    return match.group(1) if match else output


def normalize_ribcl_status(status: str) -> str:
    value = str(status or "").strip().upper()
    try:
        return f"0X{int(value, 16):04X}"
    except Exception:
        return value


def parse_ribcl_responses(output: str) -> List[Tuple[str, str]]:
    results = []
    result_text = extract_ribcl_results(output)
    response_tags = re.findall(r"<RESPONSE\b(.*?)/>", result_text, flags=re.IGNORECASE | re.DOTALL)
    for tag in response_tags:
        status_match = re.search(r'\bSTATUS\s*=\s*["\']([^"\']+)["\']', tag, flags=re.IGNORECASE)
        message_match = re.search(r'\b(?:MSG|MESSAGE)\s*=\s*["\']([^"\']*)["\']', tag, flags=re.IGNORECASE)
        status = normalize_ribcl_status(status_match.group(1)) if status_match else "UNKNOWN"
        message = message_match.group(1).strip() if message_match else ""
        results.append((status, message))
    return results


def ribcl_errors(output: str, allowed_statuses=("0X0000",)) -> List[str]:
    allowed = {normalize_ribcl_status(s) for s in allowed_statuses}
    responses = parse_ribcl_responses(output)
    if not responses:
        return ["No RIBCL RESPONSE elements returned"]
    return [f"{status}: {message or 'Unknown error'}" for status, message in responses if status not in allowed]


def get_ribcl_value(output: str, tag_name: str) -> Optional[str]:
    result_text = extract_ribcl_results(output)
    pattern = r"<" + re.escape(tag_name) + r"\b[^>]*VALUE\s*=\s*[\"']([^\"']*)[\"']"
    match = re.search(pattern, result_text, flags=re.IGNORECASE | re.DOTALL)
    return match.group(1).strip() if match else None


def ribcl_user_exists(output: str, login_name: str) -> Tuple[bool, Optional[str]]:
    """Determine whether GET_USER returned the requested local account."""
    result_text = extract_ribcl_results(output)

    for tag in re.findall(r"<GET_USER\b[^>]*>", result_text, flags=re.IGNORECASE | re.DOTALL):
        login_match = re.search(
            r'\bUSER_LOGIN\s*=\s*["\']([^"\']+)["\']',
            tag,
            flags=re.IGNORECASE,
        )
        if login_match and login_match.group(1).strip() == login_name.strip():
            return True, None

    if re.search(r"User\s+login\s+name\s+was\s+not\s+found", result_text, flags=re.IGNORECASE):
        return False, "not found"

    errors = ribcl_errors(result_text)
    if errors:
        return False, "; ".join(errors)

    return False, "GET_USER completed successfully but did not return the requested account"


# ============================================================================
# OA CONNECTION (Context Manager)
# ============================================================================

class OAConnection:
    def __init__(self, username, password, port=22):
        self.username = username
        self.password = password
        self.port = port
        self.client = None
        self.channel = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def connect(self, hostname):
        self.close()
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        self.client.connect(
            hostname=hostname, port=self.port, username=self.username, password=self.password,
            timeout=SSH_CONNECT_TIMEOUT, banner_timeout=SSH_CONNECT_TIMEOUT, auth_timeout=SSH_CONNECT_TIMEOUT,
            look_for_keys=False, allow_agent=False,
        )
        transport = self.client.get_transport()
        if transport: transport.set_keepalive(SSH_KEEPALIVE_INTERVAL)
        self.channel = self.client.invoke_shell()
        time.sleep(0.5)
        self._drain_channel()

    def close(self):
        if self.channel:
            try: self.channel.close()
            except Exception: pass
        self.channel = None
        if self.client:
            try: self.client.close()
            except Exception: pass
        self.client = None

    def _drain_channel(self):
        if not self.channel: return ""
        output = ""
        while self.channel.recv_ready():
            data = self.channel.recv(65535)
            if not data: break
            output += data.decode("utf-8", errors="replace")
        return output

    def send_command(self, cmd: str, wait_string: str = ">", timeout: int = 15) -> str:
        if not self.channel: raise RuntimeError("OA SSH channel is not connected")
        self._drain_channel()
        self.channel.sendall(cmd + "\n")
        output = ""
        deadline = time.monotonic() + timeout
        while time.monotonic() < deadline:
            if self.channel.recv_ready():
                data = self.channel.recv(65535)
                if not data: break
                output += data.decode("utf-8", errors="replace")
                if wait_string in output:
                    time.sleep(0.1)
                    while self.channel.recv_ready():
                        more = self.channel.recv(65535)
                        if not more: break
                        output += more.decode("utf-8", errors="replace")
                    break
            else:
                time.sleep(0.05)
        return output

    def is_active(self) -> bool:
        output = self.send_command("SHOW OA STATUS")
        return bool(re.search(r"Role:\s*Active", output, flags=re.IGNORECASE))

    def execute_hponcfg(self, bay_number: int, ribcl: str, end_marker="ILO_RIBCL_EOF", retry_on_post=True) -> str:
        if not self.channel: raise RuntimeError("OA SSH channel is not connected")
        self._drain_channel()
        command = f"HPONCFG {bay_number} << {end_marker}\n{ribcl.rstrip()}\n{end_marker}\n"
        
        for line in command.splitlines():
            self.channel.sendall(line + "\n")
            time.sleep(HPONCFG_LINE_DELAY)

        output = ""
        deadline = time.monotonic() + COMMAND_TIMEOUT
        quiet_since = None
        while time.monotonic() < deadline:
            if self.channel.recv_ready():
                data = self.channel.recv(65535)
                if not data: break
                output += data.decode("utf-8", errors="replace")
                quiet_since = None
                if "END RIBCL RESULTS" in output:
                    time.sleep(0.1)
                    while self.channel.recv_ready():
                        more = self.channel.recv(65535)
                        if not more: break
                        output += more.decode("utf-8", errors="replace")
                    break
            else:
                if quiet_since is None: quiet_since = time.monotonic()
                elif time.monotonic() - quiet_since >= COMMAND_QUIET_TIME: break
                time.sleep(0.05)
                
        if retry_on_post and ("0X00D4" in output.upper() or "POST IN PROGRESS" in output.upper()):
            try: slot_str = f"{int(bay_number):02}"
            except Exception: slot_str = str(bay_number)
            logger.info(f"[Slot {slot_str}] POST lock detected in RIBCL. Forcing power off and retrying...")
            self.send_command(f"POWEROFF SERVER {bay_number} FORCE")
            time.sleep(20)
            return self.execute_hponcfg(bay_number, ribcl, end_marker, retry_on_post=False)

        return output


def find_active_oa(primary_oa_ip: Optional[str], secondary_oa_ip: Optional[str]) -> Optional[str]:
    for ip in (primary_oa_ip, secondary_oa_ip):
        if not ip: continue
        print(f"\nChecking OA {ip}...")
        try:
            with OAConnection(OA_USERNAME, OA_PASSWORD, SSH_PORT) as oa:
                oa.connect(ip)
                if oa.is_active():
                    print(f"SUCCESS: ACTIVE OA is {ip}.")
                    return ip
                print(f"OA {ip} is standby/inactive.")
        except Exception as exc:
            print(f"OA {ip} check failed: {exc}")
    return None


# ============================================================================
# RIBCL BOOTSTRAP USER
# ============================================================================

def ensure_bootstrap_admin(oa: OAConnection, srv: Dict[str, Any], res_dict: Dict[str, Any]) -> bool:
    slot = srv["enclosure_slot"]
    ip = srv["ilo_ip"]
    log(slot, ip, f"RIBCL: Checking '{ILO_LOGIN}'...")

    try:
        output = oa.execute_hponcfg(slot, build_ribcl_user_check(), end_marker="ILO_USER_CHECK_EOF")
    except Exception as exc:
        mark_failure(res_dict, "Auth", f"Bootstrap user check failed: {exc}")
        res_dict["bootstrap_ok"] = False
        return False

    if DEBUG:
        debug_block(f"USER CHECK - BAY {slot}", output)

    user_exists, check_error = ribcl_user_exists(output, ILO_LOGIN)
    if user_exists:
        log(slot, ip, f"  -> '{ILO_LOGIN}' already exists.")
        res_dict["bootstrap_ok"] = True
        return True

    if check_error != "not found":
        if DEBUG_ON_FAILURE:
            debug_block(f"FAILED USER CHECK - BAY {slot}", output, force=True)
        mark_failure(
            res_dict,
            "Auth",
            f"Unable to determine whether '{ILO_LOGIN}' exists: {check_error}",
        )
        res_dict["bootstrap_ok"] = False
        return False

    log(slot, ip, f"  -> '{ILO_LOGIN}' does not exist. Creating account...")
    try:
        add_output = oa.execute_hponcfg(slot, build_ribcl_user_add(), end_marker="ILO_USER_ADD_EOF")
    except Exception as exc:
        mark_failure(res_dict, "Auth", f"Bootstrap account creation failed: {exc}")
        res_dict["bootstrap_ok"] = False
        return False

    errors = ribcl_errors(add_output, allowed_statuses=("0X0000", "0X0007"))
    if errors:
        duplicate_error = any("exist" in error.lower() or "duplicate" in error.lower() for error in errors)
        if duplicate_error:
            try:
                verify_output = oa.execute_hponcfg(slot, build_ribcl_user_check(), end_marker="ILO_USER_RECHECK_EOF")
                exists_after_add, _ = ribcl_user_exists(verify_output, ILO_LOGIN)
                if exists_after_add:
                    log(slot, ip, f"  -> '{ILO_LOGIN}' confirmed present after ADD_USER response.")
                    res_dict["bootstrap_ok"] = True
                    return True
            except Exception:
                pass

        if DEBUG_ON_FAILURE:
            debug_block(f"FAILED USER CREATE - BAY {slot}", add_output, force=True)
        for error in errors:
            res_dict["errors"].append(f"[Bootstrap] {error}")
        res_dict["Auth"] = TaskStatus.FAIL.value
        res_dict["status"] = "Failed Tasks"
        res_dict["bootstrap_ok"] = False
        return False

    # Verify the account after creation instead of trusting ADD_USER alone.
    try:
        verify_output = oa.execute_hponcfg(slot, build_ribcl_user_check(), end_marker="ILO_USER_VERIFY_EOF")
        user_exists, verify_error = ribcl_user_exists(verify_output, ILO_LOGIN)
    except Exception as exc:
        mark_failure(res_dict, "Auth", f"Bootstrap account created but verification failed: {exc}")
        res_dict["bootstrap_ok"] = False
        return False

    if not user_exists:
        mark_failure(
            res_dict,
            "Auth",
            f"ADD_USER completed but '{ILO_LOGIN}' could not be verified: {verify_error}",
        )
        res_dict["bootstrap_ok"] = False
        return False

    log(slot, ip, "  -> Bootstrap administrator ready.")
    res_dict["bootstrap_ok"] = True
    return True


# ============================================================================
# COMBINED RIBCL CONFIGURATION
# ============================================================================

def configure_combined_ribcl(oa: OAConnection, srv: Dict[str, Any], directory_server: str, cert_data: str, res_dict: Dict[str, Any]) -> bool:
    slot = srv["enclosure_slot"]
    ip = srv["ilo_ip"]
    hostname = srv["hostname"]

    log(slot, ip, "RIBCL: Applying Server Name + LDAP + LDAP CA...")
    ribcl = build_ribcl_combined(hostname, directory_server, cert_data)

    try:
        output = oa.execute_hponcfg(slot, ribcl, end_marker="ILO_COMBINED_CONFIG_EOF")
    except Exception as exc:
        for step in ("ID", "LDAP", "Cert"): mark_failure(res_dict, step, str(exc))
        return False

    if DEBUG: debug_block(f"COMBINED CONFIG - BAY {slot}", output)

    errors = ribcl_errors(output)
    if errors:
        if DEBUG_ON_FAILURE: debug_block(f"FAILED COMBINED CONFIG - BAY {slot}", output, force=True)
        error_text = "; ".join(errors)
        for step in ("ID", "LDAP", "Cert"): mark_failure(res_dict, step, error_text)
        log(slot, ip, "  -> Combined RIBCL configuration FAILED.")
        return False

    res_dict["ID"] = TaskStatus.OK.value
    res_dict["LDAP"] = TaskStatus.OK.value
    res_dict["Cert"] = TaskStatus.OK.value
    log(slot, ip, "  -> Combined RIBCL configuration successful.")
    return True

def verify_ldap_ribcl(oa: OAConnection, srv: Dict[str, Any], directory_server: str, res_dict: Dict[str, Any]) -> bool:
    slot = srv["enclosure_slot"]
    ip = srv["ilo_ip"]
    log(slot, ip, "RIBCL: Verifying LDAP settings...")

    try:
        output = oa.execute_hponcfg(slot, build_ribcl_ldap_verify(), end_marker="ILO_LDAP_VERIFY_EOF")
    except Exception as exc:
        mark_failure(res_dict, "LDAP", f"Verification failed: {exc}")
        return False

    errors = ribcl_errors(output)
    if errors:
        if DEBUG_ON_FAILURE: debug_block(f"FAILED LDAP VERIFY - BAY {slot}", output, force=True)
        mark_failure(res_dict, "LDAP", "; ".join(errors))
        return False

    expected = {
        "DIR_SERVER_ADDRESS": directory_server,
        "DIR_SERVER_PORT": str(LDAP_PORT),
        "DIR_USER_CONTEXT_1": LDAP_USER_CONTEXTS[0],
        "DIR_USER_CONTEXT_2": LDAP_USER_CONTEXTS[1],
        "DIR_USER_CONTEXT_3": LDAP_USER_CONTEXTS[2],
        "DIR_USER_CONTEXT_4": LDAP_USER_CONTEXTS[3],
        "DIR_GRPACCT1_NAME": LDAP_GROUP_NAME,
        "DIR_GRPACCT1_PRIV": LDAP_GROUP_PRIVILEGES,
    }

    mismatches = []
    for tag, expected_value in expected.items():
        actual = get_ribcl_value(output, tag)
        if actual is None: mismatches.append(f"{tag}=not returned")
        elif not values_equal_case_insensitive(actual, expected_value): mismatches.append(f"{tag}: expected '{expected_value}', got '{actual}'")

    for tag in ("DIR_AUTHENTICATION_ENABLED", "DIR_ENABLE_GRP_ACCT"):
        actual = get_ribcl_value(output, tag)
        if actual is not None and actual.lower() not in ("y", "yes", "true", "1"): mismatches.append(f"{tag}={actual}")

    generic_ldap = get_ribcl_value(output, "DIR_GENERIC_LDAP_ENABLED")
    if generic_ldap is not None and generic_ldap.lower() not in ("n", "no", "false", "0"):
        mismatches.append(f"DIR_GENERIC_LDAP_ENABLED={generic_ldap}")

    if mismatches:
        if DEBUG_ON_FAILURE: debug_block(f"LDAP VERIFY MISMATCH - BAY {slot}", output, force=True)
        for mismatch in mismatches: res_dict["errors"].append(f"[LDAP Verify] {mismatch}")
        res_dict["LDAP"] = TaskStatus.FAIL.value
        res_dict["status"] = "Failed Tasks"
        return False

    res_dict["LDAP"] = TaskStatus.OK.value
    log(slot, ip, "  -> LDAP configuration verified.")
    return True


def process_ribcl_server(oa: OAConnection, srv: Dict[str, Any], ldap_ca_cert: str, global_results: Dict[str, Any]):
    ip = srv["ilo_ip"]
    slot = srv["enclosure_slot"]
    result = global_results[srv["result_key"]]
    scope_data = SCOPE_SETTINGS[srv["scope"]]
    server_start = time.monotonic()

    try:
        if not ensure_bootstrap_admin(oa, srv, result): return
        if configure_combined_ribcl(oa, srv, scope_data["DIRECTORY_SERVER"], ldap_ca_cert, result):
            verify_ldap_ribcl(oa, srv, scope_data["DIRECTORY_SERVER"], result)
    except Exception as exc:
        result["status"] = "Failed Tasks"
        result["errors"].append(f"[RIBCL Worker] {exc}")
    finally:
        result["ribcl_seconds"] = time.monotonic() - server_start
        log(slot, ip, f"RIBCL phase finished in {result['ribcl_seconds']:.2f}s")


def ribcl_worker(worker_number: int, active_oa_ip: str, servers: List[Dict[str, Any]], ldap_ca_cert: str, global_results: Dict[str, Any], redfish_executor, redfish_futures: dict, redfish_lock, trigger_reboot: bool):
    if not servers: return
    logger.info(f"[RIBCL Worker {worker_number}] Connecting to active OA {active_oa_ip}...")
    
    try:
        with OAConnection(OA_USERNAME, OA_PASSWORD, SSH_PORT) as oa:
            oa.connect(active_oa_ip)
            for srv in servers:
                process_ribcl_server(oa, srv, ldap_ca_cert, global_results)
                
                result = global_results[srv["result_key"]]
                if result.get("bootstrap_ok") is True:
                    future = redfish_executor.submit(process_redfish, srv, global_results, trigger_reboot)
                    with redfish_lock:
                        redfish_futures[future] = srv
                else:
                    result["Auth"] = TaskStatus.FAIL.value
                    result["status"] = "Failed Tasks"

    except Exception as exc:
        for srv in servers:
            result = global_results[srv["result_key"]]
            if result.get("bootstrap_ok") is None:
                result["bootstrap_ok"] = False
                result["Auth"] = TaskStatus.FAIL.value
                result["status"] = "Failed Tasks"
                result["errors"].append(f"[RIBCL Worker] {exc}")


# ============================================================================
# REDFISH PROCESSOR (Context Manager)
# ============================================================================

class ILORedfishProcessor:
    def __init__(self, ip: str, slot: int, login: str, password: str, res_dict: Dict[str, Any]):
        self.ip = ip
        self.slot = slot
        self.base_url = f"https://{ip}"
        self.login = login
        self.password = password
        self.res = res_dict

        self.session = requests.Session()
        self.session.verify = False
        self.session.headers.update({
            "Content-Type": "application/json",
            "Accept": "application/json",
            "OData-Version": "4.0",
        })

        retry_strategy = Retry(
            total=HTTP_SAFE_RETRIES,
            connect=HTTP_SAFE_RETRIES,
            read=HTTP_SAFE_RETRIES,
            status=HTTP_SAFE_RETRIES,
            backoff_factor=0.5,
            status_forcelist=[500, 502, 503, 504],
            allowed_methods=frozenset(["HEAD", "GET", "OPTIONS"]),
            raise_on_status=False,
        )
        adapter = HTTPAdapter(pool_connections=4, pool_maxsize=4, max_retries=retry_strategy)
        self.session.mount("https://", adapter)
        self.session_uri = None

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.close()

    def _get(self, url: str): return self.session.get(url, timeout=REDFISH_TIMEOUT)
    def _post(self, url: str, payload: dict = None): return self.session.post(url, json=payload, timeout=REDFISH_TIMEOUT)
    def _patch(self, url: str, payload: dict): return self.session.patch(url, json=payload, timeout=REDFISH_TIMEOUT)

    def _extract_error_message(self, response: requests.Response) -> str:
        try:
            data = response.json()
            extended = data.get("error", {}).get("@Message.ExtendedInfo", [])
            if extended:
                item = extended[0]
                parts = []
                message_id = item.get("MessageID") or item.get("MessageId")
                if message_id: parts.append(message_id)
                if item.get("Message"): parts.append(item["Message"])
                if item.get("MessageArgs"): parts.append(f"Args={item['MessageArgs']}")
                if parts: return " | ".join(parts)
            messages = data.get("Messages", [])
            if messages:
                item = messages[0]
                message_id = item.get("MessageID") or item.get("MessageId") or ""
                return f"{message_id} Args={item.get('MessageArgs', [])}".strip()
        except Exception: pass
        text = (response.text or "").strip()
        return text[:500] if text else f"HTTP {response.status_code}"

    def _handle_error(self, step: str, error: Exception):
        if isinstance(error, requests.exceptions.HTTPError) and error.response is not None:
            response = error.response
            msg = f"HTTP {response.status_code}: {self._extract_error_message(response)}"
            if DEBUG_ON_FAILURE:
                raw = (response.text or "").strip()
                if raw: msg += " | RAW=" + raw[:500]
        else: msg = str(error)
        mark_failure(self.res, step, msg)
        log(self.slot, self.ip, f"  -> {step} FAILED: {msg}")

    def close(self):
        if self.session_uri:
            try: self.session.delete(f"{self.base_url}{self.session_uri}", timeout=REDFISH_TIMEOUT)
            except Exception: pass
        try: self.session.close()
        except Exception: pass

    def authenticate(self) -> bool:
        session_endpoints = [
            f"{self.base_url}/redfish/v1/SessionService/Sessions",
            f"{self.base_url}/redfish/v1/Sessions",
            f"{self.base_url}/rest/v1/Sessions",
        ]
        last_error = None
        selected_endpoint = None

        for attempt in range(1, AUTH_RETRIES + 1):
            endpoints = [selected_endpoint] if selected_endpoint else session_endpoints
            delayed = False

            for auth_url in endpoints:
                if not auth_url:
                    continue
                try:
                    resp = self._post(auth_url, {"UserName": self.login, "Password": self.password})

                    if resp.status_code in (404, 405) and selected_endpoint is None:
                        continue

                    # Once an endpoint responds as an authentication endpoint, keep using it.
                    if resp.status_code not in (404, 405):
                        selected_endpoint = auth_url

                    if resp.status_code in (400, 401, 403):
                        err_msg = self._extract_error_message(resp)
                        last_error = f"HTTP {resp.status_code}: {err_msg}"

                        if "LoginAttemptDelayed" in err_msg or "LoginAttemptDelayed" in (resp.text or ""):
                            log(
                                self.slot,
                                self.ip,
                                f"iLO authentication temporarily delayed (attempt {attempt}/{AUTH_RETRIES}). "
                                f"Waiting {LOGIN_PENALTY_WAIT}s for the penalty timeout...",
                            )
                            time.sleep(LOGIN_PENALTY_WAIT)
                            delayed = True
                            break

                        # 401/403 proves the endpoint exists. Do not multiply bad-login
                        # attempts across legacy endpoint fallbacks.
                        if resp.status_code in (401, 403):
                            break

                        # A 400 can mean an incompatible legacy/session URI; before an
                        # endpoint is established, allow the next compatibility URI.
                        if selected_endpoint == auth_url and attempt < AUTH_RETRIES:
                            break
                        continue

                    resp.raise_for_status()
                    token = resp.headers.get("X-Auth-Token")
                    if not token:
                        raise RuntimeError("No X-Auth-Token returned")

                    self.session.headers["X-Auth-Token"] = token
                    location = resp.headers.get("Location")
                    if location:
                        if location.startswith("https://"):
                            location = re.sub(r"^https://[^/]+", "", location)
                        self.session_uri = location

                    self.res["Auth"] = TaskStatus.OK.value
                    return True

                except Exception as exc:
                    last_error = str(exc)
                    if selected_endpoint:
                        break

            if delayed:
                continue
            if attempt < AUTH_RETRIES:
                time.sleep(AUTH_RETRY_DELAY)

        mark_failure(self.res, "Auth", last_error or "Authentication failed")
        return False

    def _force_power_off(self) -> bool:
        sys_url = f"{self.base_url}/redfish/v1/Systems/1/"
        try:
            resp = self._get(sys_url)
            resp.raise_for_status()
            if resp.json().get("PowerState") == "Off":
                log(self.slot, self.ip, "Server is already OFF.")
                time.sleep(ILO_SETTLE_AFTER_POWEROFF)
                return True
            log(self.slot, self.ip, "POST is blocking API change. Forcing server power OFF...")
            reset_url = f"{sys_url}Actions/ComputerSystem.Reset"
            resp = self._post(reset_url, {"ResetType": "ForceOff"})
            resp.raise_for_status()
            deadline = time.monotonic() + POWER_OFF_TIMEOUT
            while time.monotonic() < deadline:
                time.sleep(POWER_OFF_POLL_INTERVAL)
                resp = self._get(sys_url)
                resp.raise_for_status()
                if resp.json().get("PowerState") == "Off":
                    log(self.slot, self.ip, "Server is now OFF.")
                    time.sleep(ILO_SETTLE_AFTER_POWEROFF)
                    return True
            log(self.slot, self.ip, f"Server did not reach OFF within {POWER_OFF_TIMEOUT}s.")
            return False
        except Exception as exc:
            log(self.slot, self.ip, f"ForceOff FAILED: {exc}")
            return False

    def _patch_with_post_handling(self, url: str, payload: dict) -> requests.Response:
        resp = self._patch(url, payload)
        if resp.status_code >= 400 and POST_ERROR_STRING in (resp.text or ""):
            log(self.slot, self.ip, "UnableToModifyDuringSystemPOST received.")
            if not self._force_power_off(): return resp
            log(self.slot, self.ip, "Retrying API change after ForceOff...")
            resp = self._patch(url, payload)
        return resp

    def configure_license(self, license_key: str):
        try:
            mgr_url = f"{self.base_url}/redfish/v1/Managers/1/"
            resp = self._get(mgr_url)
            resp.raise_for_status()
            if "Advanced" in resp.json().get("Oem", {}).get("Hp", {}).get("License", {}).get("LicenseString", ""):
                self.res["Lic"] = TaskStatus.SKIP.value
                return
            action_urls = [
                f"{self.base_url}/redfish/v1/Managers/1/LicenseService/Actions/HpeiLOLicense.InstallLicense",
                f"{self.base_url}/redfish/v1/Managers/1/LicenseService/Actions/Oem/Hp/InstallLicense"
            ]
            last_resp = None
            for url in action_urls:
                last_resp = self._post(url, {"LicenseKey": license_key})
                if last_resp.status_code < 400: break
            last_resp.raise_for_status()
            self.res["Lic"] = TaskStatus.OK.value
        except Exception as exc:
            self._handle_error("Lic", exc)

    def configure_fence_user(self, friendly_name: str, login_name: str, password: str):
        try:
            accounts_url = f"{self.base_url}/redfish/v1/AccountService/Accounts/"
            resp = self._get(accounts_url)
            resp.raise_for_status()
            members = resp.json().get("Members", [])

            for member in members:
                if member.get("UserName") == login_name:
                    self.res["Usr"] = TaskStatus.SKIP.value
                    return
            for member in members:
                uri = member.get("@odata.id")
                if not uri: continue
                account_resp = self._get(f"{self.base_url}{uri}")
                if account_resp.status_code >= 400: continue
                if account_resp.json().get("UserName") == login_name:
                    self.res["Usr"] = TaskStatus.SKIP.value
                    return

            payload = {
                "UserName": login_name,
                "Password": password,
                "Oem": {"Hp": {"LoginName": friendly_name, "Privileges": {"RemoteConsolePriv": False, "VirtualMediaPriv": False, "UserConfigPriv": False, "VirtualPowerAndResetPriv": True, "iLOConfigPriv": False}}}
            }
            resp = self._post(accounts_url, payload)
            resp.raise_for_status()
            self.res["Usr"] = TaskStatus.OK.value
        except Exception as exc:
            self._handle_error("Usr", exc)

    def configure_network(self, ilo_hostname: str, primary_dns: str, secondary_dns: str):
        try:
            ilo_short_hostname, ilo_domain = split_ilo_hostname(ilo_hostname)
            collection_url = f"{self.base_url}/redfish/v1/Managers/1/EthernetInterfaces/"
            resp = self._get(collection_url)
            resp.raise_for_status()
            members = resp.json().get("Members", [])
            if not members: raise RuntimeError("No Ethernet interfaces found")
            target_uri = members[0].get("@odata.id")
            if not target_uri: raise RuntimeError("Ethernet interface missing @odata.id")
            target_url = f"{self.base_url}{target_uri}"

            resp = self._get(target_url)
            resp.raise_for_status()
            hp = resp.json().get("Oem", {}).get("Hp", {})
            current_hostname = hp.get("HostName")
            current_domain = hp.get("DomainName")
            dhcp4 = hp.get("DHCPv4", {})
            dhcp6 = hp.get("DHCPv6", {})
            current_dns = hp.get("IPv4", {}).get("DNSServers", [])
            desired_dns = [primary_dns, secondary_dns]

            need_hostname = not values_equal_case_insensitive(current_hostname, ilo_short_hostname)
            need_domain = not values_equal_case_insensitive(current_domain, ilo_domain)
            
            # Explicitly force UseNTPServers to False in DHCP so it doesn't lock our SNTP patch
            need_dhcp4 = dhcp4.get("UseDNSServers") is not False or dhcp4.get("UseDomainName") is not False or dhcp4.get("UseNTPServers") is not False
            need_dhcp6 = dhcp6.get("UseDNSServers") is not False or dhcp6.get("UseDomainName") is not False or dhcp6.get("UseNTPServers") is not False
            need_dns = list(current_dns or []) != desired_dns

            if not any((need_hostname, need_domain, need_dhcp4, need_dhcp6, need_dns)):
                self.res["Net"] = TaskStatus.SKIP.value
                return

            hp_payload = {}
            if need_hostname: hp_payload["HostName"] = ilo_short_hostname
            if need_domain: hp_payload["DomainName"] = ilo_domain
            if need_dhcp4: hp_payload["DHCPv4"] = {"UseDNSServers": False, "UseDomainName": False, "UseNTPServers": False}
            if need_dhcp6: hp_payload["DHCPv6"] = {"UseDNSServers": False, "UseDomainName": False, "UseNTPServers": False}
            if need_dns: hp_payload["IPv4"] = {"DNSServers": desired_dns}

            combined_payload = {"Oem": {"Hp": hp_payload}}
            combined_resp = self._patch_with_post_handling(target_url, combined_payload)
            if combined_resp.status_code < 400:
                self.res["Net"] = TaskStatus.OK.value
                return
            if POST_ERROR_STRING in (combined_resp.text or ""): combined_resp.raise_for_status()

            log(self.slot, self.ip, "Combined network PATCH rejected; using sliced iLO 4 fallback...")
            sliced_payloads = []
            if need_dhcp4: sliced_payloads.append({"Oem": {"Hp": {"DHCPv4": {"UseDNSServers": False, "UseDomainName": False, "UseNTPServers": False}}}})
            if need_dhcp6: sliced_payloads.append({"Oem": {"Hp": {"DHCPv6": {"UseDNSServers": False, "UseDomainName": False, "UseNTPServers": False}}}})
            if need_hostname: sliced_payloads.append({"Oem": {"Hp": {"HostName": ilo_short_hostname}}})
            if need_domain: sliced_payloads.append({"Oem": {"Hp": {"DomainName": ilo_domain}}})
            if need_dns: sliced_payloads.append({"Oem": {"Hp": {"IPv4": {"DNSServers": desired_dns}}}})

            for payload in sliced_payloads:
                self._patch_with_post_handling(target_url, payload).raise_for_status()
            self.res["Net"] = TaskStatus.OK.value

        except Exception as exc:
            self._handle_error("Net", exc)

    def configure_sntp(self, ntp_primary: str, ntp_secondary: str, target_tz_search: str):
        try:
            # 1. Fetch DateTime URL from Manager Link
            mgr_url = f"{self.base_url}/redfish/v1/Managers/1/"
            resp = self._get(mgr_url)
            resp.raise_for_status()
            
            dt_uri = resp.json().get("Oem", {}).get("Hp", {}).get("Links", {}).get("DateTime", {}).get("@odata.id")
            if not dt_uri:
                dt_uri = "/redfish/v1/Managers/1/DateTime/"
                
            dt_url = f"{self.base_url}{dt_uri}"
            dt_resp = self._get(dt_url)
            if dt_resp.status_code == 404:
                self.res["SNTP"] = TaskStatus.SKIP.value
                return
                
            dt_resp.raise_for_status()
            data = dt_resp.json()
            
            # 2. Extract NTP locations (handling iLO 4 Oem structure)
            current_ntp = data.get("StaticNTPServers")
            if current_ntp is None:
                current_ntp = data.get("Oem", {}).get("Hp", {}).get("StaticNTPServers", [])
                
            desired_ntp = []
            if ntp_primary: desired_ntp.append(ntp_primary)
            if ntp_secondary: desired_ntp.append(ntp_secondary)
            
            while len(desired_ntp) < 2:
                desired_ntp.append("")
                
            need_ntp = (list(current_ntp or []) != desired_ntp)

            # 3. Extract Timezone locations (handling iLO 4 Oem structure)
            tz_list = data.get("TimeZoneList")
            if tz_list is None:
                tz_list = data.get("Oem", {}).get("Hp", {}).get("TimeZoneList", [])
                
            target_index = None
            for tz in tz_list:
                name = tz.get("Name", "")
                value = tz.get("Value", "")
                if target_tz_search.lower() in name.lower() or target_tz_search.lower() in value.lower():
                    target_index = tz.get("Index")
                    break
                    
            if target_index is None:
                log(self.slot, self.ip, f"  -> Warning: Timezone matching '{target_tz_search}' not found in iLO TimeZoneList.")
            
            current_tz = data.get("TimeZone", {})
            if not current_tz:
                current_tz = data.get("Oem", {}).get("Hp", {}).get("TimeZone", {})
                
            current_tz_index = current_tz.get("Index")
            need_tz = (target_index is not None and current_tz_index != target_index)

            if not need_ntp and not need_tz:
                self.res["SNTP"] = TaskStatus.SKIP.value
                return
                
            payload = {}
            oem_hp = {}
            
            if need_ntp:
                if "StaticNTPServers" in data:
                    payload["StaticNTPServers"] = desired_ntp
                else:
                    oem_hp["StaticNTPServers"] = desired_ntp
                    
            if need_tz:
                if "TimeZone" in data:
                    payload["TimeZone"] = {"Index": target_index}
                else:
                    oem_hp["TimeZone"] = {"Index": target_index}
                    
            if oem_hp:
                payload["Oem"] = {"Hp": oem_hp}
                
            patch_resp = self._patch_with_post_handling(dt_url, payload)
            
            if patch_resp.status_code >= 400 and "SNTPConfigurationManagedByDHCPAndIsReadOnly" in (patch_resp.text or ""):
                log(self.slot, self.ip, "SNTP is managed by DHCP; unable to override settings via Redfish.")
                self.res["SNTP"] = TaskStatus.SKIP.value
                return
                
            patch_resp.raise_for_status()
            self.res["SNTP"] = TaskStatus.OK.value
            self.res["ilo_reset_needed"] = True
            
        except Exception as exc:
            self._handle_error("SNTP", exc)

    def configure_ipmi(self, target_port: int):
        try:
            mgr_url = f"{self.base_url}/redfish/v1/Managers/1/"
            resp = self._get(mgr_url)
            resp.raise_for_status()
            net_uri = resp.json().get("NetworkService", {}).get("@odata.id")
            if not net_uri: net_uri = "/redfish/v1/Managers/1/NetworkService/"
            
            net_url = f"{self.base_url}{net_uri}"
            resp = self._get(net_url)
            if resp.status_code == 404:
                net_uri = "/redfish/v1/Managers/1/NetworkProtocol/"
                net_url = f"{self.base_url}{net_uri}"
                resp = self._get(net_url)
            resp.raise_for_status()
            
            current = resp.json().get("IPMI", {})
            enabled = current.get("ProtocolEnabled")
            if enabled is None: enabled = current.get("Enabled")

            if enabled is True and current.get("Port") == target_port:
                self.res["IPMI"] = TaskStatus.SKIP.value
                return

            payload = {"IPMI": {"ProtocolEnabled": True, "Port": target_port}}
            patch_resp = self._patch_with_post_handling(net_url, payload)
            if patch_resp.status_code >= 400 and POST_ERROR_STRING not in (patch_resp.text or ""):
                payload = {"IPMI": {"Enabled": True, "Port": target_port}}
                patch_resp = self._patch_with_post_handling(net_url, payload)
            patch_resp.raise_for_status()
            self.res["IPMI"] = TaskStatus.OK.value

        except Exception as exc:
            self._handle_error("IPMI", exc)

    def configure_boot(self):
        try:
            system_url = f"{self.base_url}/redfish/v1/Systems/1/"
            resp = self._get(system_url)
            resp.raise_for_status()
            current_boot = resp.json().get("Boot", {})

            need_pxe = not (current_boot.get("BootSourceOverrideTarget") == "Pxe" and current_boot.get("BootSourceOverrideEnabled") == "Once")
            changed = False

            if need_pxe:
                payload = {"Boot": {"BootSourceOverrideTarget": "Pxe", "BootSourceOverrideEnabled": "Once"}}
                self._patch_with_post_handling(system_url, payload).raise_for_status()
                changed = True

            bios_url = f"{system_url}Bios/Settings/"
            bios_get = self._get(bios_url)
            if bios_get.status_code == 404:
                self.res["Boot"] = TaskStatus.OK.value if changed else TaskStatus.SKIP.value
                return
            if bios_get.status_code >= 400:
                if "ResourceMissingAtURI" in (bios_get.text or ""):
                    self.res["Boot"] = TaskStatus.OK.value if changed else TaskStatus.SKIP.value
                    return
                bios_get.raise_for_status()

            if bios_get.json().get("BootMode") == "Uefi":
                self.res["Boot"] = TaskStatus.OK.value if changed else TaskStatus.SKIP.value
                return

            bios_resp = self._patch_with_post_handling(bios_url, {"BootMode": "Uefi"})
            if bios_resp.status_code >= 400:
                body = bios_resp.text or ""
                if bios_resp.status_code == 404 or "PropertyUnknown" in body or "ResourceMissingAtURI" in body:
                    self.res["Boot"] = TaskStatus.OK.value if changed else TaskStatus.SKIP.value
                    return
                bios_resp.raise_for_status()
            self.res["Boot"] = TaskStatus.OK.value
        except Exception as exc:
            self._handle_error("Boot", exc)

    def reboot_server(self):
        """Sends ResetType command to reboot or power on the host server."""
        try:
            sys_url = f"{self.base_url}/redfish/v1/Systems/1/"
            resp = self._get(sys_url)
            resp.raise_for_status()
            
            power_state = resp.json().get("PowerState", "Unknown")
            reset_type = "ForceRestart" if power_state == "On" else "On"
            
            log(self.slot, self.ip, f"Sending power command '{reset_type}' to trigger network boot...")
            reset_url = f"{sys_url}Actions/ComputerSystem.Reset"
            resp = self._post(reset_url, {"ResetType": reset_type})
            resp.raise_for_status()
            self.res["Reboot"] = TaskStatus.OK.value
            log(self.slot, self.ip, f"  -> Power command '{reset_type}' sent successfully.")
        except Exception as exc:
            self._handle_error("Reboot", exc)

    def reset_ilo(self):
        """Resets the iLO Manager to apply SNTP changes without hanging."""
        try:
            log(self.slot, self.ip, "Sending iLO reset command to apply SNTP changes...")
            
            reset_timeout = 5
            headers = self.session.headers.copy()
            
            endpoints = [
                f"{self.base_url}/redfish/v1/Managers/1/Actions/Manager.Reset",
                f"{self.base_url}/redfish/v1/Managers/1/Actions/Oem/Hp/Manager.Reset",
                f"{self.base_url}/rest/v1/Managers/1/Actions/Manager.Reset"
            ]
            
            payloads = [
                {"ResetType": "ForceRestart"},
                {"ResetType": "GracefulRestart"},
                {"Action": "Manager.Reset"},
                {"Action": "Reset"},
                {}
            ]
            
            success = False
            for url in endpoints:
                for payload in payloads:
                    try:
                        resp = requests.post(url, json=payload, headers=headers, verify=False, timeout=reset_timeout)
                        if resp.status_code in (200, 201, 202, 204):
                            success = True
                            break
                    except (requests.exceptions.ConnectionError, requests.exceptions.ReadTimeout, requests.exceptions.Timeout):
                        # A dropped connection means the iLO is successfully rebooting
                        success = True
                        break
                if success:
                    break
                    
            if success:
                log(self.slot, self.ip, "  -> iLO reset command sent successfully.")
                self.session_uri = None
            else:
                log(self.slot, self.ip, "  -> Warning: All iLO reset payloads failed (HTTP 400). Manual reset may be required.")
            
        except Exception as exc:
            log(self.slot, self.ip, f"  -> Warning: iLO reset failed: {exc}")


def process_redfish(srv: Dict[str, Any], global_results: Dict[str, Dict[str, Any]], trigger_reboot: bool):
    ip = srv["ilo_ip"]
    slot = srv["enclosure_slot"]
    result = global_results[srv["result_key"]]
    scope_data = SCOPE_SETTINGS[srv["scope"]]
    start = time.monotonic()

    log(slot, ip, "Starting Redfish phase...")

    try:
        with ILORedfishProcessor(ip, slot, ILO_LOGIN, ILO_PASSWORD, result) as ilo:
            if not ilo.authenticate(): return
            ilo.configure_license(ILO_ADVANCED_LICENSE_KEY)
            ilo.configure_fence_user(FENCE_USER_NAME, FENCE_USER_LOGIN, FENCE_USER_PASSWORD)
            ilo.configure_network(srv["ilo_hostname"], scope_data["PRIMARY_DNS"], scope_data["SECONDARY_DNS"])
            ilo.configure_sntp(scope_data.get("NTP_PRIMARY", ""), scope_data.get("NTP_SECONDARY", ""), ILO_TIMEZONE_SEARCH)
            ilo.configure_ipmi(IPMI_PORT)
            ilo.configure_boot()

            if trigger_reboot:
                ilo.reboot_server()
            else:
                result["Reboot"] = TaskStatus.SKIP.value
                
            if result.get("ilo_reset_needed"):
                ilo.reset_ilo()

    except Exception as exc:
        result["status"] = "Failed Tasks"
        result["errors"].append(f"[Redfish Worker] {exc}")
    finally:
        result["redfish_seconds"] = time.monotonic() - start
        log(slot, ip, f"Redfish phase finished in {result['redfish_seconds']:.2f}s")


# ============================================================================
# FINAL STATUS
# ============================================================================

def recalculate_overall_status(result: Dict[str, Any]):
    if result["status"] == "Missing Cols":
        return

    tracked = ["Auth", "Lic", "Usr", "Net", "SNTP", "IPMI", "ID", "LDAP", "Cert", "Boot", "Reboot"]
    failed = [step for step in tracked if result.get(step) == TaskStatus.FAIL.value]
    pending = [step for step in tracked if result.get(step) == TaskStatus.PENDING.value]

    if failed or pending:
        result["status"] = "Failed Tasks"
        if pending:
            result["errors"].append("[Final Status] Unfinished tasks: " + ", ".join(pending))
    else:
        result["status"] = "Successful"

# ============================================================================
# MAIN
# ============================================================================

def main():
    total_start = time.monotonic()
    timings = {}

    if RIBCL_CONCURRENT_SESSIONS < 1 or REDFISH_CONCURRENT_SESSIONS < 1 or EXCEL_EMPTY_ROW_STOP < 1:
        logger.error("ERROR: Concurrency and empty row stop settings must be >= 1")
        sys.exit(1)

    excel_path = TEMPLATE_DIR / EXCEL_FILE
    if not excel_path.exists():
        logger.error(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    if not CERT_FILE.exists():
        logger.error(f"ERROR: LDAP certificate not found: {CERT_FILE}")
        sys.exit(1)

    try:
        with open(CERT_FILE, "r", encoding="utf-8") as handle:
            ldap_ca_cert = handle.read().strip()
    except Exception as exc:
        logger.error(f"ERROR reading certificate: {exc}")
        sys.exit(1)

    if "-----BEGIN CERTIFICATE-----" not in ldap_ca_cert or "-----END CERTIFICATE-----" not in ldap_ca_cert:
        logger.error("ERROR: LDAP certificate is not valid PEM text.")
        sys.exit(1)

    logger.info(f"Loading {excel_path.name} (sheet '{SHEET_NAME}')...")
    try:
        df, excel_elapsed = load_resource_excel_fast(excel_path, SHEET_NAME, empty_row_stop=EXCEL_EMPTY_ROW_STOP)
    except Exception as exc:
        logger.error(f"ERROR reading Excel workbook: {exc}")
        sys.exit(1)

    timings["Excel"] = excel_elapsed
    logger.info(f"Excel parsed: {len(df)} meaningful rows in {excel_elapsed:.2f}s")

    target_enclosure = input("\nEnter the Enclosure Name to process: ").strip()
    enc_df = df[df["enclosure_physical_name"] == target_enclosure]
    if enc_df.empty:
        logger.info(f"\nNo resources found for '{target_enclosure}'.")
        sys.exit(0)

    oa_df = enc_df[enc_df["equipment_type"] == "Enclosure OA"]
    primary_oa_ip, secondary_oa_ip = None, None
    for row in oa_df.itertuples(index=False):
        try:
            oa_slot = int(float(row.enclosure_slot))
        except Exception: continue
        ip = str(row.ilo_ip).strip()
        if not ip: continue
        if oa_slot == 1: primary_oa_ip = ip
        elif oa_slot == 2: secondary_oa_ip = ip

    blade_df = enc_df[enc_df["equipment_type"].isin(VALID_EQUIPMENT_TYPES)]
    total_rows = len(blade_df)
    servers_to_process = []
    global_results = {}

    for row_index, row in enumerate(blade_df.itertuples(index=False), start=1):
        try: slot_num = int(float(row.enclosure_slot))
        except Exception: slot_num = 999
        ip = str(row.ilo_ip).strip() or "Unknown"
        values = {"enclosure_slot": row.enclosure_slot, "ilo_ip": row.ilo_ip, "scope": row.scope, "hostname": row.hostname, "ilo_hostname": row.ilo_hostname}
        missing = [column for column, value in values.items() if value is None or str(value).strip() == ""]
        scope = str(row.scope).strip().upper()
        if not missing and scope not in SCOPE_SETTINGS: missing.append(f"Invalid Scope: {scope}")

        result_key = f"{slot_num}:{ip}:{row_index}"
        result = {
            "slot": slot_num, "ip": ip, "status": "Successful",
            "Auth": TaskStatus.PENDING.value, "Lic": TaskStatus.PENDING.value, "Usr": TaskStatus.PENDING.value,
            "Net": TaskStatus.PENDING.value, "SNTP": TaskStatus.PENDING.value, "IPMI": TaskStatus.PENDING.value, 
            "ID": TaskStatus.PENDING.value, "LDAP": TaskStatus.PENDING.value, "Cert": TaskStatus.PENDING.value, 
            "Boot": TaskStatus.PENDING.value, "Reboot": TaskStatus.PENDING.value, 
            "bootstrap_ok": None, "ilo_reset_needed": False, "ribcl_seconds": 0.0, "redfish_seconds": 0.0, "errors": [],
        }
        global_results[result_key] = result

        if missing:
            result["status"] = "Missing Cols"
            result["errors"].append("Missing/Invalid columns: " + ", ".join(missing))
            log(slot_num, ip, "SKIPPING: " + ", ".join(missing))
            continue

        try: ilo_short, ilo_domain = split_ilo_hostname(row.ilo_hostname)
        except Exception as exc:
            result["status"] = "Missing Cols"
            result["errors"].append(f"Invalid ilo_hostname: {exc}")
            log(slot_num, ip, f"SKIPPING: Invalid ilo_hostname: {exc}")
            continue

        servers_to_process.append({
            "result_key": result_key, "enclosure_slot": slot_num, "ilo_ip": ip, "scope": scope,
            "hostname": str(row.hostname).strip(), "ilo_hostname": str(row.ilo_hostname).strip(),
            "ilo_short_hostname": ilo_short, "ilo_domain": ilo_domain,
        })

    total_valid = len(servers_to_process)
    logger.info(f"\nBlade servers found : {total_rows}")
    logger.info(f"Valid for processing: {total_valid}")
    if total_valid == 0: sys.exit(0)

    # ------------------------------------------------------------------------
    # Post-Configuration Reboot Prompt
    # ------------------------------------------------------------------------
    reboot_choice = input("\nDo you want to REBOOT servers after configuration to trigger PXE boot? (yes/no): ").strip().lower()
    trigger_reboot = (reboot_choice == "yes")

    oa_discovery_start = time.monotonic()
    active_oa_ip = find_active_oa(primary_oa_ip, secondary_oa_ip)
    timings["OA Discovery"] = time.monotonic() - oa_discovery_start

    if not active_oa_ip:
        logger.error("\nERROR: Unable to find ACTIVE OA.")
        sys.exit(1)

    logger.info("\n" + "=" * 100)
    logger.info("PHASE 1 & 2: CONCURRENT OA/RIBCL & iLO REDFISH CONFIGURATION")
    logger.info("Redfish configuration will trigger immediately as each server completes its RIBCL phase")
    logger.info("=" * 100)

    config_start = time.monotonic()
    ribcl_worker_count = min(RIBCL_CONCURRENT_SESSIONS, total_valid)
    redfish_worker_count = min(REDFISH_CONCURRENT_SESSIONS, total_valid)
    
    groups = [[] for _ in range(ribcl_worker_count)]
    for index, srv in enumerate(servers_to_process): groups[index % ribcl_worker_count].append(srv)

    redfish_futures = {}
    redfish_lock = threading.Lock()

    with concurrent.futures.ThreadPoolExecutor(max_workers=redfish_worker_count) as redfish_executor:
        with concurrent.futures.ThreadPoolExecutor(max_workers=ribcl_worker_count) as ribcl_executor:
            ribcl_futures = [
                ribcl_executor.submit(
                    ribcl_worker, worker_index, active_oa_ip, group, ldap_ca_cert, global_results, 
                    redfish_executor, redfish_futures, redfish_lock, trigger_reboot
                )
                for worker_index, group in enumerate(groups, start=1)
            ]
            
            for future in concurrent.futures.as_completed(ribcl_futures):
                try: future.result()
                except Exception as exc: logger.error(f"RIBCL worker exception: {exc}")

        timings["RIBCL"] = time.monotonic() - config_start

        for future in concurrent.futures.as_completed(redfish_futures.keys()):
            srv = redfish_futures[future]
            try: future.result()
            except Exception as exc:
                result = global_results[srv["result_key"]]
                result["status"] = "Failed Tasks"
                result["errors"].append(f"[Worker] Unhandled exception: {exc}")
                
        timings["Redfish"] = time.monotonic() - config_start - timings["RIBCL"]

    for result in global_results.values(): recalculate_overall_status(result)

    report_start = time.monotonic()
    report_list = sorted(global_results.values(), key=lambda item: item["slot"])

    logger.info("\n\n" + "=" * 159)
    logger.info(f"FINAL EXECUTION REPORT: Enclosure '{target_enclosure}'")
    logger.info("=" * 159)
    logger.info(f"{'SLOT':<5} | {'iLO IP':<15} | {'OVERALL STATUS':<15} | {'AUTH':<4} | {'LIC':<4} | {'USR':<4} | {'NET':<4} | {'SNTP':<4} | {'IPMI':<4} | {'ID':<4} | {'LDAP':<4} | {'CERT':<4} | {'BOOT':<4} | {'RBT':<4} | {'RIBCL(s)':>8} | {'RF(s)':>8}")
    logger.info("-" * 159)

    success_count = fail_count = skip_count = 0
    all_errors = []

    for result in report_list:
        slot_display = f"{result['slot']:02}" if result["slot"] != 999 else "??"
        status = result["status"]

        if status == "Successful": success_count += 1
        elif status == "Missing Cols": skip_count += 1
        else: fail_count += 1

        logger.info(
            f"{slot_display:<5} | {result['ip']:<15} | {status:<15} | "
            f"{result['Auth']:<4} | {result['Lic']:<4} | {result['Usr']:<4} | "
            f"{result['Net']:<4} | {result['SNTP']:<4} | {result['IPMI']:<4} | {result['ID']:<4} | "
            f"{result['LDAP']:<4} | {result['Cert']:<4} | {result['Boot']:<4} | "
            f"{result['Reboot']:<4} | "
            f"{result['ribcl_seconds']:>8.2f} | {result['redfish_seconds']:>8.2f}"
        )

        if result["errors"]: all_errors.append((slot_display, result["ip"], result["errors"]))

    logger.info("-" * 159)
    logger.info(f"TOTAL FOUND: {total_rows} | SUCCESSFUL: {success_count} | FAILED: {fail_count} | SKIPPED: {skip_count}")
    logger.info("=" * 159)

    if all_errors:
        logger.info("\nFAILURE DETAILS:")
        for slot, ip, errors in all_errors:
            logger.info(f"\n- Slot {slot} ({ip}):")
            for error in errors: logger.info(f"    {error}")

    # ========================================================================
    # CSV EXPORT LOGIC
    # ========================================================================
    try:
        CSV_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(CSV_REPORT_PATH, mode='w', newline='', encoding='utf-8') as csv_file:
            if report_list:
                # Omit tracking variable from CSV
                internal_fields = {"ilo_reset_needed", "bootstrap_ok"}
                fieldnames = [key for key in report_list[0].keys() if key not in internal_fields]
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
                writer.writeheader()
                
                for row in report_list:
                    row_copy = row.copy()
                    row_copy.pop("ilo_reset_needed", None)
                    row_copy.pop("bootstrap_ok", None)
                    if isinstance(row_copy.get('errors'), list):
                        row_copy['errors'] = " | ".join(row_copy['errors'])
                    writer.writerow(row_copy)
                    
        logger.info(f"\nCSV Report successfully saved to: {CSV_REPORT_PATH}")
    except Exception as e:
        logger.error(f"\nFailed to save CSV report: {e}")
    # ========================================================================

    timings["Report"] = time.monotonic() - report_start
    timings["Total"] = time.monotonic() - total_start

    logger.info("\n" + "=" * 65)
    logger.info("PERFORMANCE SUMMARY")
    logger.info("=" * 65)
    logger.info(f"{'Excel load/parse':<30}: {timings.get('Excel', 0):>8.2f} sec")
    logger.info(f"{'OA discovery':<30}: {timings.get('OA Discovery', 0):>8.2f} sec")
    logger.info(f"{'RIBCL Phase (Overlapping)':<30}: {timings.get('RIBCL', 0):>8.2f} sec")
    logger.info(f"{'Redfish Phase (Completion)':<30}: {timings.get('Redfish', 0):>8.2f} sec")
    logger.info(f"{'Final report':<30}: {timings.get('Report', 0):>8.2f} sec")
    logger.info("-" * 65)
    logger.info(f"{'TOTAL EXECUTION TIME':<30}: {timings.get('Total', 0):>8.2f} sec")
    logger.info("=" * 65)

    if fail_count > 0: sys.exit(1)
    sys.exit(0)

if __name__ == "__main__":
    main()