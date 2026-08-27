# BUILD_MARKER: DNS_STANDALONE_CENTRAL_V2_20260828

import concurrent.futures
import csv
import os
import sys
import time

import openpyxl

from functions import dns, vars


# =============================================================================
# EXCEL
# =============================================================================

def load_excel_range(
    file_path,
    sheet_name,
    start_row,
    end_row,
    empty_row_stop,
):
    """
    Read only the requested Excel range using openpyxl read-only streaming.

    START_ROW / END_ROW are the same global row numbers used by the Foreman,
    iLO and RAID automation.
    """
    start_row = int(start_row)
    end_row = int(end_row)
    empty_row_stop = int(empty_row_stop)

    if start_row < 2:
        raise ValueError(
            "START_ROW must be >= 2."
        )

    if end_row < start_row:
        raise ValueError(
            "END_ROW must be >= START_ROW."
        )

    if empty_row_stop < 1:
        raise ValueError(
            "EXCEL_EMPTY_ROW_STOP must be >= 1."
        )

    workbook = openpyxl.load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "Sheet '{}' not found in workbook."
                .format(sheet_name)
            )

        worksheet = workbook[
            sheet_name
        ]

        header_values = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not header_values:
            return []

        headers = []

        for index, cell in enumerate(
            header_values
        ):
            if (
                cell is None
                or str(cell).strip() == ""
            ):
                header = (
                    "Unnamed_{}"
                    .format(index)
                )
            else:
                header = str(
                    cell
                ).strip()

            headers.append(
                header
            )

        records = []
        empty_count = 0

        for excel_row, row in enumerate(
            worksheet.iter_rows(
                min_row=start_row,
                max_row=end_row,
                values_only=True,
            ),
            start=start_row,
        ):
            is_empty = all(
                cell is None
                or str(cell).strip() == ""
                for cell in row
            )

            if is_empty:
                empty_count += 1

                if (
                    empty_count
                    >= empty_row_stop
                ):
                    break

                continue

            empty_count = 0

            values = list(
                row[:len(headers)]
            )

            if len(values) < len(headers):
                values.extend(
                    [None]
                    * (
                        len(headers)
                        - len(values)
                    )
                )

            record = {}

            for header, value in zip(
                headers,
                values,
            ):
                if value is None:
                    value = (
                        vars.EMPTY_VALUE
                    )

                record[
                    header
                ] = value

            record[
                "_excel_row"
            ] = excel_row

            records.append(
                record
            )

        return records

    finally:
        workbook.close()


# =============================================================================
# WORKER
# =============================================================================

def _display_value(
    vm,
    primary_key,
    fallback_key=None,
):
    value = vm.get(
        primary_key
    )

    if (
        value is None
        or str(value).strip().upper()
        in vars.INVALID_VALUES
    ):
        if fallback_key:
            value = vm.get(
                fallback_key
            )

    if (
        value is None
        or str(value).strip().upper()
        in vars.INVALID_VALUES
    ):
        return "Unknown"

    return str(
        value
    ).strip()


def process_dns_row(vm):
    """
    Process one Excel row through the same optimized functions.dns module used
    by process_host.py and process_vm.py.
    """
    start_time = time.time()

    excel_row = vm.get(
        "_excel_row",
        "Unknown",
    )

    logical_name = _display_value(
        vm,
        "logical_name",
        "hostname",
    )

    hostname = _display_value(
        vm,
        "hostname",
        "logical_name",
    )

    equipment_type = _display_value(
        vm,
        "equipment_type",
    )

    print(
        "[Row {} | {}] Starting DNS configuration..."
        .format(
            excel_row,
            logical_name,
        ),
        flush=True,
    )

    try:
        dns_result = dns.create_dns_records(
            vm
        )

        if not isinstance(
            dns_result,
            dict,
        ):
            raise RuntimeError(
                "functions.dns.create_dns_records() "
                "did not return a dictionary."
            )

        status = dns_result.get(
            "status",
            "Failed",
        )

        details = dns_result.get(
            "details",
            "",
        )

        record_count = dns_result.get(
            "records",
            0,
        )

        return_code = dns_result.get(
            "returncode",
            "",
        )

    except Exception as exc:
        status = "Failed"
        details = str(exc)
        record_count = 0
        return_code = ""

    elapsed = (
        time.time()
        - start_time
    )

    print(
        "[Row {} | {}] FINISHED - {}"
        .format(
            excel_row,
            logical_name,
            status,
        ),
        flush=True,
    )

    return {
        "ExcelRow": excel_row,
        "EquipmentType": equipment_type,
        "Hostname": hostname,
        "LogicalName": logical_name,
        "Status": status,
        "DNSRecords": record_count,
        "ReturnCode": return_code,
        "Details": details,
        "TimeSeconds": round(
            elapsed,
            2,
        ),
    }


# =============================================================================
# REPORTING
# =============================================================================

def print_final_report(
    results,
    total_time,
):
    results.sort(
        key=lambda item:
            item["ExcelRow"]
    )

    successful = sum(
        1
        for item in results
        if item["Status"]
        == "Successful"
    )

    skipped = sum(
        1
        for item in results
        if item["Status"]
        == "Skipped"
    )

    failed = sum(
        1
        for item in results
        if item["Status"]
        == "Failed"
    )

    width = 140

    print(
        "\n"
        + "=" * width
    )

    print(
        "FINAL DNS EXECUTION REPORT"
    )

    print(
        "=" * width
    )

    print(
        "{:<7} | {:<22} | {:<20} | {:<11} | "
        "{:<7} | {:<9} | {}".format(
            "ROW",
            "LOGICAL NAME",
            "EQUIPMENT TYPE",
            "STATUS",
            "RECORDS",
            "TIME",
            "DETAILS",
        )
    )

    print(
        "-" * width
    )

    for item in results:
        print(
            "{:<7} | {:<22} | {:<20} | {:<11} | "
            "{:<7} | {:<9.2f} | {}".format(
                item["ExcelRow"],
                str(
                    item["LogicalName"]
                )[:22],
                str(
                    item["EquipmentType"]
                )[:20],
                item["Status"],
                item["DNSRecords"],
                item["TimeSeconds"],
                item["Details"],
            )
        )

    print(
        "=" * width
    )

    print(
        "TOTAL: {} | SUCCESSFUL: {} | "
        "FAILED: {} | SKIPPED: {} | "
        "TIME: {:.2f}s".format(
            len(results),
            successful,
            failed,
            skipped,
            total_time,
        )
    )

    print(
        "=" * width
        + "\n"
    )

    return failed


def write_audit_report(
    results,
):
    if not os.path.isdir(
        vars.LOG_DIR
    ):
        os.makedirs(
            vars.LOG_DIR
        )

    timestamp = time.strftime(
        "%Y%m%d-%H%M%S"
    )

    report_path = os.path.join(
        vars.LOG_DIR,
        "{}_{}.csv".format(
            vars.DNS_REPORT_PREFIX,
            timestamp,
        ),
    )

    fieldnames = [
        "ExcelRow",
        "EquipmentType",
        "Hostname",
        "LogicalName",
        "Status",
        "DNSRecords",
        "ReturnCode",
        "Details",
        "TimeSeconds",
    ]

    with open(
        report_path,
        "w",
        newline="",
    ) as report_file:
        writer = csv.DictWriter(
            report_file,
            fieldnames=fieldnames,
        )

        writer.writeheader()

        for item in results:
            writer.writerow(
                item
            )

    print(
        "Audit report saved to: {}"
        .format(
            report_path
        )
    )

    return report_path


# =============================================================================
# MAIN
# =============================================================================

def main():
    global_start = time.time()

    if not os.path.isfile(
        vars.RESOURCE_LIST
    ):
        print(
            "ERROR: Resource List not found: {}"
            .format(
                vars.RESOURCE_LIST
            )
        )

        return 1

    if int(
        vars.MAX_WORKERS
    ) < 1:
        print(
            "ERROR: MAX_WORKERS must be >= 1."
        )

        return 1

    print(
        "=" * 100
    )
    print(
        "STANDALONE DNS CONFIGURATION"
    )
    print(
        "=" * 100
    )
    print(
        "Resource list       : {}"
        .format(
            vars.RESOURCE_LIST
        )
    )
    print(
        "Sheet               : {}"
        .format(
            vars.SHEET_NAME
        )
    )
    print(
        "Excel rows          : {} - {}"
        .format(
            vars.START_ROW,
            vars.END_ROW,
        )
    )
    print(
        "Max workers         : {}"
        .format(
            vars.MAX_WORKERS
        )
    )
    print(
        "DNS server          : {}"
        .format(
            vars.DNS_SERVER
        )
    )
    print(
        "=" * 100
    )

    try:
        vm_list = load_excel_range(
            vars.RESOURCE_LIST,
            vars.SHEET_NAME,
            vars.START_ROW,
            vars.END_ROW,
            vars.EXCEL_EMPTY_ROW_STOP,
        )

    except Exception as exc:
        print(
            "ERROR loading Excel: {}"
            .format(exc)
        )

        return 1

    if not vm_list:
        print(
            "No populated rows found "
            "in the configured Excel range."
        )

        return 0

    total = len(
        vm_list
    )

    workers = min(
        int(
            vars.MAX_WORKERS
        ),
        total,
    )

    print(
        "\nFound {} populated row(s). "
        "Starting DNS processing with "
        "{} worker(s)...\n".format(
            total,
            workers,
        )
    )

    results = []

    with concurrent.futures.ThreadPoolExecutor(
        max_workers=workers
    ) as executor:
        future_map = {
            executor.submit(
                process_dns_row,
                vm,
            ): vm

            for vm in vm_list
        }

        for future in (
            concurrent.futures
            .as_completed(
                future_map
            )
        ):
            vm = future_map[
                future
            ]

            try:
                result = (
                    future.result()
                )

            except Exception as exc:
                result = {
                    "ExcelRow": vm.get(
                        "_excel_row",
                        "Unknown",
                    ),
                    "EquipmentType": _display_value(
                        vm,
                        "equipment_type",
                    ),
                    "Hostname": _display_value(
                        vm,
                        "hostname",
                        "logical_name",
                    ),
                    "LogicalName": _display_value(
                        vm,
                        "logical_name",
                        "hostname",
                    ),
                    "Status": "Failed",
                    "DNSRecords": 0,
                    "ReturnCode": "",
                    "Details": (
                        "Unhandled worker exception: {}"
                        .format(exc)
                    ),
                    "TimeSeconds": 0.0,
                }

            results.append(
                result
            )

    total_time = (
        time.time()
        - global_start
    )

    failed = print_final_report(
        results,
        total_time,
    )

    write_audit_report(
        results
    )

    return (
        1
        if failed
        else 0
    )


if __name__ == "__main__":
    sys.exit(
        main()
    )
