#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# BUILD_MARKER: SET_SERIAL_BL_V6_QUICK_RECOVERY_IML_OFFLINE_FALLBACK_20260901

import concurrent.futures
import csv
import hashlib
import os
import re
import sys
import threading
import time
import warnings
from datetime import datetime

import requests
import urllib3
from openpyxl import load_workbook

from functions import vars
from functions.output_log import run_logged_main

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
PRINT_LOCK = threading.RLock()
OA_EFUSE_LOCK = threading.RLock()
SUCCESS_CODES = (200, 201, 202, 204)
RETRY_CODES = (409, 429, 500, 502, 503, 504)

# V6 deliberately stops after one bounded HPE REST-state recovery attempt.
# These defaults live here so V6 works with the existing vars.py unchanged.
QUICK_RECOVERY_INITIAL_WAIT_SECONDS = getattr(
    vars, "BL_SERIAL_QUICK_RECOVERY_INITIAL_WAIT_SECONDS", 30
)
QUICK_RECOVERY_TIMEOUT_SECONDS = getattr(
    vars, "BL_SERIAL_QUICK_RECOVERY_TIMEOUT_SECONDS", 180
)
QUICK_RECOVERY_POLL_INTERVAL_SECONDS = getattr(
    vars, "BL_SERIAL_QUICK_RECOVERY_POLL_INTERVAL_SECONDS", 10
)
IML_SSH_TIMEOUT_SECONDS = getattr(vars, "BL_SERIAL_IML_SSH_TIMEOUT_SECONDS", 35)
IML_SCAN_LIMIT = getattr(vars, "BL_SERIAL_IML_SCAN_LIMIT", 80)
IML_MAX_RELEVANT = getattr(vars, "BL_SERIAL_IML_MAX_RELEVANT", 15)
IML_EVENT_NUMBERS = {266, 267, 268, 269, 282, 333, 334, 335, 336, 337, 338}
IML_KEYWORDS = (
    "bios", "rest", "restful", "nvram", "non-volatile", "rom", "post",
    "serial", "product id", "intelligent provisioning", "configuration",
)



class OfflineFallbackRequired(RuntimeError):
    """Raised when remote BIOS REST recovery failed and offline work is required."""

    def __init__(self, message, iml_summary="", iml_file=""):
        super().__init__(message)
        self.iml_summary = iml_summary
        self.iml_file = iml_file


def cfg(name, default):
    return getattr(vars, name, default)


def text(value):
    return "" if value is None else str(value).strip()


def is_invalid(value):
    return text(value).upper() in vars.INVALID_VALUES


def same(left, right):
    return text(left).upper() == text(right).upper()


def normalize_slot(value):
    try:
        return int(float(text(value)))
    except (TypeError, ValueError):
        return 999


def slot_text(slot):
    return "{:02}".format(slot) if slot != 999 else "??"


def shown(value):
    return text(value) or "<empty>"


def log(slot, ilo_ip, message):
    with PRINT_LOCK:
        print(
            "[Slot {} | {}] {}".format(
                slot_text(slot), ilo_ip or "NO-iLO-IP", message
            ),
            flush=True,
        )


def response_message(response):
    try:
        data = response.json()
        error = data.get("error", {})
        extended = error.get("@Message.ExtendedInfo", [])
        if extended:
            item = extended[0]
            parts = [
                item.get("MessageId") or item.get("MessageID") or "",
                item.get("Message") or "",
            ]
            message = " | ".join(part for part in parts if part)
            if message:
                return message
        if error.get("message"):
            return text(error["message"])
        messages = data.get("Messages", [])
        if messages:
            item = messages[0]
            parts = [
                item.get("MessageId") or item.get("MessageID") or "",
                item.get("Message") or "",
            ]
            message = " | ".join(part for part in parts if part)
            if message:
                return message
    except Exception:
        pass
    return text(response.text)[:700] or "HTTP {}".format(response.status_code)


def extract_link(value):
    if isinstance(value, str):
        return value if value.startswith(("/", "http://", "https://")) else None
    if not isinstance(value, dict):
        return None
    for key in ("@odata.id", "href"):
        link = value.get(key)
        if isinstance(link, str) and link:
            return link
    for nested in value.values():
        link = extract_link(nested)
        if link:
            return link
    return None


def find_link(data, names):
    wanted = {name.lower() for name in names}
    if isinstance(data, dict):
        for key, value in data.items():
            if key.lower() in wanted:
                link = extract_link(value)
                if link:
                    return link
        for value in data.values():
            link = find_link(value, names)
            if link:
                return link
    elif isinstance(data, list):
        for value in data:
            link = find_link(value, names)
            if link:
                return link
    return None


def collection_links(data):
    for key in ("Members", "Items"):
        items = data.get(key, []) if isinstance(data, dict) else []
        if isinstance(items, list):
            links = [extract_link(item) for item in items]
            links = [link for link in links if link]
            if links:
                return links
    return []


def find_reset_target(system_data):
    actions = system_data.get("Actions", {}) if isinstance(system_data, dict) else {}
    if not isinstance(actions, dict):
        return None
    for name, definition in actions.items():
        if "computersystem.reset" not in text(name).lower():
            continue
        if isinstance(definition, dict):
            target = definition.get("target") or definition.get("Target")
            if target:
                return text(target)
    return None


def find_action_target(data, action_token):
    """Find an advertised Redfish action target recursively by action name."""
    token = text(action_token).lower()
    if isinstance(data, dict):
        for key, value in data.items():
            if token in text(key).lower() and isinstance(value, dict):
                target = value.get("target") or value.get("Target")
                if target:
                    return text(target)
            target = find_action_target(value, token)
            if target:
                return target
    elif isinstance(data, list):
        for value in data:
            target = find_action_target(value, token)
            if target:
                return target
    return None


def nested_value(data, names):
    wanted = {text(name).lower() for name in names}
    if isinstance(data, dict):
        for key, value in data.items():
            if text(key).lower() in wanted and not isinstance(value, (dict, list)):
                return value
        for value in data.values():
            found = nested_value(value, names)
            if found not in (None, ""):
                return found
    elif isinstance(data, list):
        for value in data:
            found = nested_value(value, names)
            if found not in (None, ""):
                return found
    return None


def bios_values(data):
    source = data.get("Attributes", {}) if isinstance(data, dict) else {}
    if not isinstance(source, dict) or not source:
        source = data if isinstance(data, dict) else {}
    return {
        "SerialNumber": text(source.get("SerialNumber")),
        "ProductId": text(source.get("ProductId")),
    }


def system_values(data):
    if not isinstance(data, dict):
        return {"SerialNumber": "", "ProductId": ""}

    product = ""
    for field in vars.BL_SERIAL_SYSTEM_PRODUCT_FIELDS:
        product = text(data.get(field))
        if product:
            break

    return {
        "SerialNumber": text(data.get("SerialNumber")),
        "ProductId": product,
    }


def load_blades(enclosure_name):
    """Load blades and enclosure OA addresses in one optimized worksheet scan."""
    if not os.path.isfile(vars.RESOURCE_LIST):
        raise RuntimeError("Resource list not found: {}".format(vars.RESOURCE_LIST))

    workbook = load_workbook(vars.RESOURCE_LIST, read_only=True, data_only=True)
    try:
        if vars.SHEET_NAME not in workbook.sheetnames:
            raise RuntimeError("Worksheet '{}' not found".format(vars.SHEET_NAME))
        worksheet = workbook[vars.SHEET_NAME]
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
        )
        if not header_row:
            raise RuntimeError("Excel header row is empty")

        headers = {
            text(value).lower(): index
            for index, value in enumerate(header_row)
            if text(value)
        }
        missing = [
            column
            for column in vars.BL_SERIAL_REQUIRED_COLUMNS
            if column.lower() not in headers
        ]
        if missing:
            raise RuntimeError(
                "Required Excel columns missing: {}".format(", ".join(missing))
            )

        target = text(enclosure_name).lower()
        valid_types = {
            text(value).lower() for value in vars.BL_SERIAL_VALID_EQUIPMENT_TYPES
        }
        max_col = max(
            headers[column.lower()] for column in vars.BL_SERIAL_REQUIRED_COLUMNS
        ) + 1
        blades = []
        oa_by_slot = {}
        empty_count = 0
        data_started = False

        for excel_row, row in enumerate(
            worksheet.iter_rows(min_row=2, max_col=max_col, values_only=True), 2
        ):
            populated = any(value is not None and text(value) for value in row)
            if not populated:
                if data_started:
                    empty_count += 1
                    if empty_count >= vars.BL_SERIAL_EXCEL_EMPTY_ROW_STOP:
                        break
                continue

            data_started = True
            empty_count = 0
            enclosure = text(row[headers["enclosure_physical_name"]])
            if enclosure.lower() != target:
                continue

            equipment_type = text(row[headers["equipment_type"]])
            slot = normalize_slot(row[headers["enclosure_slot"]])
            ilo_ip = text(row[headers["ilo_ip"]])

            if equipment_type.lower() == "enclosure oa":
                if slot in (1, 2) and not is_invalid(ilo_ip):
                    oa_by_slot[slot] = ilo_ip
                continue

            if equipment_type.lower() not in valid_types:
                continue

            blades.append({
                "excel_row": excel_row,
                "server_slot": slot,
                "ilo_ip": ilo_ip,
                "serial_excel": text(
                    row[headers[vars.BL_SERIAL_NUMBER_COLUMN.lower()]]
                ),
            })
    finally:
        workbook.close()

    oa_ips = [oa_by_slot[key] for key in sorted(oa_by_slot)]
    unique = []
    seen = set()
    for blade in sorted(blades, key=lambda item: item["server_slot"]):
        key = (blade["server_slot"], blade["ilo_ip"].lower())
        if key in seen:
            log(
                blade["server_slot"],
                blade["ilo_ip"],
                "Duplicate Excel row {} ignored.".format(blade["excel_row"]),
            )
            continue
        seen.add(key)
        blade["oa_ips"] = tuple(oa_ips)
        unique.append(blade)
    return unique, oa_ips


def _oa_recv_until(channel, tokens, timeout_seconds):
    output = ""
    deadline = time.time() + timeout_seconds
    tokens = tuple(text(token).lower() for token in tokens)
    while time.time() < deadline:
        if channel.recv_ready():
            data = channel.recv(65535)
            if not data:
                break
            output += data.decode("utf-8", errors="replace")
            lower = output.lower()
            if any(token in lower for token in tokens):
                return output
        else:
            time.sleep(0.1)
    return output


def _oa_read_until_quiet(channel, timeout_seconds, quiet_seconds=0.75):
    """Read a complete OA CLI response until no new data arrives briefly."""
    output = ""
    deadline = time.time() + timeout_seconds
    last_data = None
    while time.time() < deadline:
        if channel.recv_ready():
            data = channel.recv(65535)
            if not data:
                break
            output += data.decode("utf-8", errors="replace")
            last_data = time.time()
            continue
        if output and last_data is not None:
            if time.time() - last_data >= quiet_seconds:
                break
        time.sleep(0.1)
    return output


def oa_efuse_reset(oa_ips, slot, ilo_ip):
    """Trip the blade E-fuse through the ACTIVE enclosure OA."""
    with OA_EFUSE_LOCK:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                import paramiko
            except Exception as exc:
                raise RuntimeError("Paramiko unavailable for OA recovery: {}".format(exc))

        errors = []
        for oa_ip in oa_ips:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            try:
                log(slot, ilo_ip, "OA recovery: connecting to {}...".format(oa_ip))
                ssh.connect(
                    hostname=oa_ip,
                    port=vars.ILO_BL_SSH_PORT,
                    username=vars.OA_USERNAME,
                    password=vars.OA_PASSWORD,
                    look_for_keys=False,
                    allow_agent=False,
                    timeout=vars.BL_SERIAL_OA_CONNECT_TIMEOUT_SECONDS,
                    banner_timeout=vars.BL_SERIAL_OA_CONNECT_TIMEOUT_SECONDS,
                    auth_timeout=vars.BL_SERIAL_OA_CONNECT_TIMEOUT_SECONDS,
                )
                transport = ssh.get_transport()
                if transport:
                    transport.set_keepalive(vars.BL_SERIAL_OA_KEEPALIVE_SECONDS)
                channel = ssh.invoke_shell()
                time.sleep(1)
                while channel.recv_ready():
                    channel.recv(65535)

                channel.send("SHOW OA STATUS\n")
                status = _oa_read_until_quiet(
                    channel, vars.BL_SERIAL_OA_COMMAND_TIMEOUT_SECONDS
                )
                role_match = re.search(
                    r"(?im)^\s*Role\s*:\s*(Active|Standby)\b",
                    status,
                )
                role = role_match.group(1).lower() if role_match else ""
                if role != "active":
                    compact_status = " ".join(status.split())[-700:]
                    errors.append(
                        "{} is not ACTIVE OA (detected role={}; SHOW OA STATUS: {})"
                        .format(
                            oa_ip,
                            role or "unknown",
                            compact_status or "<no output>",
                        )
                    )
                    continue
                log(
                    slot,
                    ilo_ip,
                    "OA recovery: {} confirmed ACTIVE by SHOW OA STATUS.".format(
                        oa_ip
                    ),
                )

                log(
                    slot,
                    ilo_ip,
                    "OA E-fuse recovery: RESET SERVER {} through ACTIVE OA {}.".format(
                        slot, oa_ip
                    ),
                )
                channel.send("RESET SERVER {}\n".format(slot))
                prompt = _oa_recv_until(
                    channel,
                    (
                        "do you want to continue",
                        "successfully reset the e-fuse",
                        "invalid",
                        "error",
                    ),
                    vars.BL_SERIAL_OA_COMMAND_TIMEOUT_SECONDS,
                )
                if "do you want to continue" in prompt.lower():
                    channel.send("YES\n")
                    prompt += _oa_recv_until(
                        channel,
                        ("successfully reset the e-fuse", "failed", "error", ">"),
                        vars.BL_SERIAL_OA_COMMAND_TIMEOUT_SECONDS,
                    )

                lower = prompt.lower()
                if "successfully reset the e-fuse" not in lower:
                    raise RuntimeError(
                        "OA RESET SERVER did not confirm E-fuse reset: {}".format(
                            " ".join(prompt.split())[-700:]
                        )
                    )
                log(slot, ilo_ip, "OA confirmed successful blade E-fuse reset.")
                return oa_ip
            except Exception as exc:
                errors.append("{}: {}".format(oa_ip, exc))
            finally:
                try:
                    ssh.close()
                except Exception:
                    pass

        raise RuntimeError(
            "Unable to perform OA E-fuse recovery: {}".format(" | ".join(errors))
        )


class IloRedfishClient:
    """Direct iLO 4 client with system-first identity and BIOS recovery."""

    def __init__(self, ilo_ip):
        self.ilo_ip = ilo_ip
        self.base_url = "https://{}:{}".format(ilo_ip, vars.ILO_REDFISH_PORT)
        self.session = requests.Session()
        self.session.auth = (vars.ILO_BL_USERNAME, vars.ILO_BL_PASSWORD)
        self.session.verify = False
        self.session.headers.update({
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Connection": "close",
        })
        self.api_prefix = None
        self.root_data = {}
        self.system_uri = None
        self.system_data = {}
        self.manager_uri = None
        self.manager_data = {}
        self.bios_uri = None
        self.settings_uri = None
        self.bios_data = {}
        self.settings_data = {}
        self.recovery_notes = []

    def url(self, uri):
        if uri.startswith(("http://", "https://")):
            return uri
        return self.base_url + (uri if uri.startswith("/") else "/" + uri)

    def get(self, uri, retries=3):
        last_error = None
        for attempt in range(1, retries + 1):
            try:
                response = self.session.get(
                    self.url(uri), timeout=vars.BL_SERIAL_REQUEST_TIMEOUT_SECONDS
                )
                if response.status_code == 200:
                    return response
                last_error = "GET {} -> HTTP {}: {}".format(
                    uri, response.status_code, response_message(response)
                )
                if response.status_code not in (429, 500, 502, 503, 504):
                    break
            except requests.exceptions.RequestException as exc:
                last_error = "GET {} -> {}".format(uri, exc)
            if attempt < retries:
                time.sleep(2)
        raise RuntimeError(last_error or "GET {} failed".format(uri))

    def post(self, uri, payload):
        try:
            response = self.session.post(
                self.url(uri),
                json=payload,
                timeout=vars.BL_SERIAL_REQUEST_TIMEOUT_SECONDS,
            )
        except requests.exceptions.RequestException as exc:
            raise RuntimeError("POST {} -> {}".format(uri, exc))
        if response.status_code not in SUCCESS_CODES:
            raise RuntimeError(
                "POST {} -> HTTP {}: {}".format(
                    uri, response.status_code, response_message(response)
                )
            )
        return response

    def _first_member(self, collection_uri):
        collection = self.get(collection_uri, retries=1).json()
        members = collection_links(collection)
        if not members:
            raise RuntimeError(
                "No collection member returned by {}".format(collection_uri)
            )
        return members[0]

    def discover_system(self, slot, quiet=False):
        errors = []
        for prefix in vars.BL_SERIAL_API_PREFIXES:
            prefix = prefix.rstrip("/")
            try:
                root_data = self.get(prefix + "/", retries=1).json()
                systems_uri = find_link(root_data, ("Systems",)) or prefix + "/Systems/"
                system_uri = self._first_member(systems_uri)
                system_data = self.get(system_uri, retries=1).json()

                manager_uri = None
                manager_data = {}
                managers_uri = find_link(root_data, ("Managers",))
                if managers_uri:
                    try:
                        manager_uri = self._first_member(managers_uri)
                        manager_data = self.get(manager_uri, retries=1).json()
                    except Exception:
                        manager_uri = None
                        manager_data = {}

                self.api_prefix = prefix
                self.root_data = root_data
                self.system_uri = system_uri
                self.system_data = system_data
                self.manager_uri = manager_uri
                self.manager_data = manager_data
                if not quiet:
                    log(
                        slot,
                        self.ilo_ip,
                        "ComputerSystem discovered via {}: {}".format(
                            prefix, system_uri
                        ),
                    )
                return
            except Exception as exc:
                errors.append("{}: {}".format(prefix, exc))

        raise RuntimeError(
            "ComputerSystem discovery failed: {}".format(" | ".join(errors))
        )

    def refresh_system(self):
        self.system_data = self.get(self.system_uri).json()
        return self.system_data

    def active_values(self):
        self.refresh_system()
        return system_values(self.system_data)

    def _settings_candidates(self, bios_candidates):
        candidates = []

        def add(uri):
            uri = text(uri)
            if uri and uri.lower() not in {item.lower() for item in candidates}:
                candidates.append(uri)

        for bios_uri in bios_candidates:
            add(bios_uri.rstrip("/") + "/Settings")

        system_base = self.system_uri.rstrip("/")
        add(system_base + "/Bios/Settings")
        add(system_base + "/BIOS/Settings")

        lower = system_base.lower()
        marker = "/redfish/v1/"
        index = lower.find(marker)
        if index >= 0:
            legacy_base = (
                system_base[:index] + "/rest/v1/" + system_base[index + len(marker):]
            )
            add(legacy_base + "/Bios/Settings")
            add(legacy_base + "/BIOS/Settings")

        marker = "/rest/v1/"
        index = lower.find(marker)
        if index >= 0:
            redfish_base = (
                system_base[:index]
                + "/redfish/v1/"
                + system_base[index + len(marker):]
            )
            add(redfish_base + "/Bios/Settings")
            add(redfish_base + "/BIOS/Settings")
        return candidates

    def _discover_bios_from_current_system(self):
        """Discover writable BIOS Settings even if the parent BIOS GET is broken."""
        self.refresh_system()
        linked_bios = find_link(self.system_data, ("Bios", "BIOS"))
        bios_candidates = []
        if linked_bios:
            bios_candidates.append(linked_bios)
        bios_candidates.extend([
            self.system_uri.rstrip("/") + "/Bios",
            self.system_uri.rstrip("/") + "/BIOS",
        ])

        errors = []
        bios_uri = None
        bios_data = {}
        settings_candidates = []
        seen = set()

        for candidate in bios_candidates:
            key = text(candidate).lower()
            if not key or key in seen:
                continue
            seen.add(key)
            try:
                candidate_data = self.get(candidate, retries=1).json()
                if bios_uri is None:
                    bios_uri = candidate
                    bios_data = candidate_data
                linked_settings = find_link(
                    candidate_data, ("SettingsObject", "Settings", "settings")
                )
                if linked_settings:
                    settings_candidates.append(linked_settings)
            except Exception as exc:
                errors.append(str(exc))

        # Critical iLO 4 compatibility behavior: a server can expose the active
        # ComputerSystem and the writable HpBios Settings resource even when the
        # parent BIOS resource returns ResourceMissingAtURI. Probe Settings
        # directly before clearing REST state or power cycling the blade.
        settings_candidates.extend(self._settings_candidates(bios_candidates))
        settings_seen = set()
        for settings_uri in settings_candidates:
            key = text(settings_uri).lower()
            if not key or key in settings_seen:
                continue
            settings_seen.add(key)
            try:
                settings_data = self.get(settings_uri, retries=1).json()
                return bios_uri, bios_data, settings_uri, settings_data
            except Exception as exc:
                errors.append(str(exc))

        raise RuntimeError(" | ".join(errors) or "No BIOS Settings resource discovered")

    def _log_bios_discovery(self, slot, recovered=False):
        prefix = "BIOS API recovered" if recovered else "BIOS API discovered"
        if self.bios_uri:
            message = "{}: current={} | pending={}".format(
                prefix, self.bios_uri, self.settings_uri
            )
        else:
            message = (
                "{} directly through writable Settings: pending={} "
                "(parent BIOS resource unavailable)"
            ).format(prefix, self.settings_uri)
        log(slot, self.ilo_ip, message)

    def _poll_bios_settings(
        self, slot, timeout_seconds, initial_wait, recovered_label, poll_interval=None
    ):
        deadline = time.monotonic() + timeout_seconds
        last_error = ""
        if initial_wait:
            time.sleep(initial_wait)
        next_progress = time.monotonic()
        while time.monotonic() < deadline:
            try:
                self.discover_system(slot, quiet=True)
                (
                    self.bios_uri,
                    self.bios_data,
                    self.settings_uri,
                    self.settings_data,
                ) = self._discover_bios_from_current_system()
                self._log_bios_discovery(slot, recovered=True)
                self.recovery_notes.append(recovered_label)
                return True, ""
            except Exception as exc:
                last_error = str(exc)
                if time.monotonic() >= next_progress:
                    log(
                        slot,
                        self.ilo_ip,
                        "Waiting for BIOS Settings provider: {}".format(last_error),
                    )
                    next_progress = time.monotonic() + 60
                time.sleep(
                    poll_interval
                    if poll_interval is not None
                    else vars.BL_SERIAL_REST_API_RECOVERY_POLL_INTERVAL_SECONDS
                )
        return False, last_error

    def discover_bios(self, slot, oa_ips=()):
        """Discover BIOS Settings; perform only one bounded HPE recovery attempt."""
        del oa_ips  # V6 intentionally does not perform automatic OA E-fuse recovery.
        try:
            (
                self.bios_uri,
                self.bios_data,
                self.settings_uri,
                self.settings_data,
            ) = self._discover_bios_from_current_system()
            self._log_bios_discovery(slot)
            return
        except Exception as first_error:
            if not vars.BL_SERIAL_REST_API_RECOVERY_ON_BIOS_404:
                raise

            error_text = text(first_error)
            if "404" not in error_text and "resourcemissingaturi" not in error_text.lower():
                raise

            log(
                slot,
                self.ilo_ip,
                "Writable BIOS Settings resource is missing (HTTP 404); "
                "performing ONE standard HPE REST-state recovery attempt.",
            )
            self.recover_rest_api_state(slot)

            recovered, last_error = self._poll_bios_settings(
                slot,
                QUICK_RECOVERY_TIMEOUT_SECONDS,
                QUICK_RECOVERY_INITIAL_WAIT_SECONDS,
                "REST API state recovery",
                poll_interval=QUICK_RECOVERY_POLL_INTERVAL_SECONDS,
            )
            if recovered:
                return

            log(
                slot,
                self.ilo_ip,
                "BIOS Settings still unavailable after standard recovery; "
                "stopping remote recovery and collecting IML diagnostics.",
            )
            iml_summary, iml_file = self.collect_iml_diagnostics(slot)
            message = (
                "BIOS Settings unavailable after one standard HPE REST-state "
                "recovery. No OA E-fuse/factory reset was attempted. "
                "Offline F9/CONREP recovery required. Last BIOS error: {}"
            ).format(last_error or error_text)
            raise OfflineFallbackRequired(message, iml_summary, iml_file)

    def _relevant_iml_entry(self, entry):
        severity = text(entry.get("Severity")).lower()
        message = text(entry.get("Message") or entry.get("Description"))
        event_number = nested_value(
            entry, ("EventNumber", "RecordId", "EventCode", "Number")
        )
        try:
            event_int = int(str(event_number).strip())
        except (TypeError, ValueError):
            event_int = None
        lower = message.lower()
        return (
            severity in ("warning", "critical")
            or event_int in IML_EVENT_NUMBERS
            or any(keyword in lower for keyword in IML_KEYWORDS)
        )

    def _iml_entry_text(self, entry):
        severity = text(entry.get("Severity")) or "Unknown"
        created = text(entry.get("Created") or entry.get("Updated"))
        message = text(entry.get("Message") or entry.get("Description"))
        event_number = nested_value(
            entry, ("EventNumber", "RecordId", "EventCode", "Number")
        )
        parts = []
        if event_number not in (None, ""):
            parts.append("Event {}".format(event_number))
        parts.append(severity)
        if created:
            parts.append(created)
        if message:
            parts.append(message)
        return " | ".join(parts)[:1000]

    def collect_iml_redfish(self):
        """Best-effort IML collection through Redfish; diagnostics only."""
        if not self.system_uri:
            return [], ""
        base = self.system_uri.rstrip("/")
        candidates = [
            base + "/LogServices/IML/Entries",
            base + "/logservices/iml/entries",
        ]
        lower = base.lower()
        if "/redfish/v1/" in lower:
            idx = lower.find("/redfish/v1/")
            legacy = base[:idx] + "/rest/v1/" + base[idx + len("/redfish/v1/"):]
            candidates.extend([
                legacy + "/LogServices/IML/Entries",
                legacy + "/logservices/iml/entries",
            ])

        errors = []
        for uri in candidates:
            try:
                collection = self.get(uri, retries=1).json()
                links = collection_links(collection)
                if not links:
                    members = collection.get("Members", []) if isinstance(collection, dict) else []
                    entries = [item for item in members if isinstance(item, dict)]
                else:
                    entries = []
                    for link in links[-IML_SCAN_LIMIT:]:
                        try:
                            entries.append(self.get(link, retries=1).json())
                        except Exception as exc:
                            errors.append(str(exc))
                relevant = [entry for entry in entries if self._relevant_iml_entry(entry)]
                summaries = [self._iml_entry_text(entry) for entry in relevant[-IML_MAX_RELEVANT:]]
                raw = "\n".join(summaries)
                return summaries, raw
            except Exception as exc:
                errors.append(str(exc))
        return [], "Redfish IML collection failed: {}".format(" | ".join(errors))

    def collect_iml_ssh(self):
        """Collect IML using the iLO 4 SMASH CLP: show /system1/log1."""
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                import paramiko
            except Exception as exc:
                return [], "", "Paramiko unavailable: {}".format(exc)

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(
                hostname=self.ilo_ip,
                port=vars.BL_SERIAL_REST_API_RECOVERY_SSH_PORT,
                username=vars.ILO_BL_USERNAME,
                password=vars.ILO_BL_PASSWORD,
                look_for_keys=False,
                allow_agent=False,
                timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
                banner_timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
                auth_timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
            )
            channel = ssh.invoke_shell()
            time.sleep(0.75)
            while channel.recv_ready():
                channel.recv(65535)
            channel.send("show /system1/log1\n")
            raw = _oa_read_until_quiet(
                channel, IML_SSH_TIMEOUT_SECONDS, quiet_seconds=2.0
            )
            lines = [" ".join(line.split()) for line in raw.splitlines() if line.strip()]
            relevant = []
            for line in lines:
                lower = line.lower()
                if (
                    any(keyword in lower for keyword in IML_KEYWORDS)
                    or "severity=critical" in lower
                    or "severity=warning" in lower
                    or "severity: critical" in lower
                    or "severity: warning" in lower
                ):
                    relevant.append(line[:1000])
            return relevant[-IML_MAX_RELEVANT:], raw, ""
        except Exception as exc:
            return [], "", str(exc)
        finally:
            try:
                ssh.close()
            except Exception:
                pass

    def collect_iml_diagnostics(self, slot):
        """Capture IML, preferring the non-REST iLO SSH path."""
        if not os.path.isdir(vars.LOG_DIR):
            os.makedirs(vars.LOG_DIR)

        ssh_items, ssh_raw, ssh_error = self.collect_iml_ssh()
        redfish_items, redfish_raw = self.collect_iml_redfish()
        items = []
        seen = set()
        for item in ssh_items + redfish_items:
            key = text(item).lower()
            if key and key not in seen:
                seen.add(key)
                items.append(item)
        items = items[-IML_MAX_RELEVANT:]

        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        path = os.path.join(
            vars.LOG_DIR,
            "{}_Slot{}_IML_{}.txt".format(
                vars.BL_SERIAL_REPORT_PREFIX, slot_text(slot), stamp
            ),
        )
        with open(path, "w", encoding="utf-8") as handle:
            handle.write("iLO: {}\n".format(self.ilo_ip))
            handle.write("Slot: {}\n".format(slot_text(slot)))
            handle.write("Captured: {}\n\n".format(datetime.now().isoformat()))
            handle.write("=== RELEVANT IML SUMMARY ===\n")
            if items:
                for item in items:
                    handle.write("{}\n".format(item))
            else:
                handle.write("No BIOS/NVRAM/REST warning or critical lines matched.\n")
            handle.write("\n=== iLO SSH: show /system1/log1 ===\n")
            if ssh_raw:
                handle.write(ssh_raw)
            else:
                handle.write("SSH IML unavailable: {}\n".format(ssh_error or "no output"))
            handle.write("\n\n=== REDFISH IML SUMMARY / ERROR ===\n")
            handle.write(redfish_raw or "No Redfish IML output.\n")

        if items:
            log(
                slot,
                self.ilo_ip,
                "Relevant IML: {}".format(" || ".join(items[:3])[:1800]),
            )
        else:
            log(slot, self.ilo_ip, "No matching BIOS/NVRAM/REST IML events found.")
        log(slot, self.ilo_ip, "IML diagnostic saved: {}".format(path))
        return " || ".join(items), path

    def current_bios_values(self):
        if not self.bios_uri:
            return self.active_values()
        self.bios_data = self.get(self.bios_uri).json()
        return bios_values(self.bios_data)

    def pending_values(self):
        self.settings_data = self.get(self.settings_uri).json()
        return bios_values(self.settings_data)

    def patch_bios(self, changes, slot):
        attributes_mode = (
            isinstance(self.bios_data.get("Attributes"), dict)
            or isinstance(self.settings_data.get("Attributes"), dict)
        )
        payload = {"Attributes": changes} if attributes_mode else changes
        headers = {}
        bios_password = text(vars.BL_SERIAL_BIOS_PASSWORD)
        if bios_password:
            headers["X-HPRESTFULAPI-AuthToken"] = hashlib.sha256(
                bios_password.encode("utf-8")
            ).hexdigest().upper()

        last_error = None
        for attempt in range(1, vars.BL_SERIAL_WRITE_RETRIES + 1):
            try:
                response = self.session.patch(
                    self.url(self.settings_uri),
                    json=payload,
                    headers=headers or None,
                    timeout=vars.BL_SERIAL_REQUEST_TIMEOUT_SECONDS,
                )
                if response.status_code in SUCCESS_CODES:
                    return
                message = response_message(response)
                last_error = "PATCH {} -> HTTP {}: {}".format(
                    self.settings_uri, response.status_code, message
                )
                retryable = (
                    response.status_code in RETRY_CODES
                    or vars.ILO_BL_POST_ERROR_STRING.lower() in message.lower()
                )
                if not retryable:
                    break
            except requests.exceptions.RequestException as exc:
                last_error = "PATCH {} -> {}".format(self.settings_uri, exc)

            if attempt < vars.BL_SERIAL_WRITE_RETRIES:
                log(
                    slot,
                    self.ilo_ip,
                    "Write attempt {}/{} failed; retrying in {}s: {}".format(
                        attempt,
                        vars.BL_SERIAL_WRITE_RETRIES,
                        vars.BL_SERIAL_WRITE_RETRY_DELAY_SECONDS,
                        last_error,
                    ),
                )
                time.sleep(vars.BL_SERIAL_WRITE_RETRY_DELAY_SECONDS)
        raise RuntimeError(last_error or "BIOS Settings PATCH failed")

    def clear_rest_api_state_http(self):
        errors = []
        manager_data = self.manager_data
        if self.manager_uri:
            try:
                manager_data = self.get(self.manager_uri, retries=1).json()
                self.manager_data = manager_data
            except Exception as exc:
                errors.append(str(exc))

        action_target = find_action_target(manager_data, "clearrestapistate")
        if action_target:
            try:
                self.post(action_target, {})
                return "Redfish action {}".format(action_target)
            except Exception as exc:
                errors.append(str(exc))

        # HPE's documented iLO 4 compatibility form. Prefer /rest/v1 even when
        # the active ComputerSystem was discovered through /redfish/v1.
        payload = {"Action": "ClearRestApiState", "Target": "/Oem/Hp"}
        candidates = ["/rest/v1/managers/1"]
        if self.manager_uri:
            candidates.append(self.manager_uri)
        candidates.append("/redfish/v1/managers/1")

        seen = set()
        for uri in candidates:
            key = text(uri).lower().rstrip("/")
            if not key or key in seen:
                continue
            seen.add(key)
            try:
                self.post(uri, payload)
                return "legacy action {}".format(uri)
            except Exception as exc:
                errors.append(str(exc))
        raise RuntimeError(" | ".join(errors))

    def clear_rest_api_state_ssh(self):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                import paramiko
            except Exception as exc:
                raise RuntimeError("Paramiko unavailable: {}".format(exc))

        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        try:
            ssh.connect(
                hostname=self.ilo_ip,
                port=vars.BL_SERIAL_REST_API_RECOVERY_SSH_PORT,
                username=vars.ILO_BL_USERNAME,
                password=vars.ILO_BL_PASSWORD,
                look_for_keys=False,
                allow_agent=False,
                timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
                banner_timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
                auth_timeout=vars.BL_SERIAL_REST_API_RECOVERY_SSH_TIMEOUT_SECONDS,
            )
            channel = ssh.invoke_shell()
            channel.settimeout(2.0)
            time.sleep(1)
            while channel.recv_ready():
                channel.recv(65535)
            channel.send("oemhp_clearRESTAPIstate\n")

            output = ""
            deadline = time.time() + 45
            last_data = time.time()
            while time.time() < deadline:
                if channel.recv_ready():
                    data = channel.recv(65535)
                    if not data:
                        break
                    output += data.decode("utf-8", errors="replace")
                    last_data = time.time()
                elif time.time() - last_data >= 3:
                    break
                else:
                    time.sleep(0.1)

            lower = output.lower()
            failure_tokens = ("invalid", "unknown command", "error", "failed")
            if any(token in lower for token in failure_tokens):
                raise RuntimeError(
                    "iLO SSH clear REST API state returned: {}".format(
                        " ".join(output.split())[-700:]
                    )
                )
            return "iLO SSH oemhp_clearRESTAPIstate"
        finally:
            try:
                ssh.close()
            except Exception:
                pass

    def recover_rest_api_state(self, slot):
        errors = []
        method_used = None
        for method in vars.BL_SERIAL_REST_API_RECOVERY_METHOD_ORDER:
            method = text(method).upper()
            try:
                if method == "SSH" and vars.BL_SERIAL_REST_API_RECOVERY_SSH_FALLBACK:
                    method_used = self.clear_rest_api_state_ssh()
                elif method == "HTTP" and vars.BL_SERIAL_REST_API_RECOVERY_HTTP:
                    method_used = self.clear_rest_api_state_http()
                if method_used:
                    break
            except Exception as exc:
                errors.append("{} recovery: {}".format(method, exc))

        if not method_used:
            raise RuntimeError(
                "Unable to clear iLO REST API state: {}".format(" | ".join(errors))
            )

        log(
            slot,
            self.ilo_ip,
            "REST API state clear completed using {}. Waiting {}s before host reboot.".format(
                method_used, vars.BL_SERIAL_REST_API_CLEAR_SETTLE_SECONDS
            ),
        )
        time.sleep(vars.BL_SERIAL_REST_API_CLEAR_SETTLE_SECONDS)
        reset_type, reset_uri = self.reboot()
        self.recovery_notes.append("REST state cleared using {}".format(method_used))
        log(
            slot,
            self.ilo_ip,
            "REST recovery reboot requested: {} via {}.".format(
                reset_type, reset_uri
            ),
        )

    def reboot(self):
        self.refresh_system()
        power_state = text(self.system_data.get("PowerState"))
        reset_type = "On" if power_state.lower() == "off" else vars.BL_SERIAL_REBOOT_TYPE
        reset_target = find_reset_target(self.system_data)
        errors = []

        if reset_target:
            try:
                self.post(reset_target, {"ResetType": reset_type})
                return reset_type, reset_target
            except Exception as exc:
                errors.append(str(exc))

        legacy_payload = {"Action": "Reset", "ResetType": reset_type}
        try:
            self.post(self.system_uri, legacy_payload)
            return reset_type, self.system_uri
        except Exception as exc:
            errors.append(str(exc))

        raise RuntimeError("Server reboot failed: {}".format(" | ".join(errors)))

    def close(self):
        self.session.close()


def new_result(blade):
    return {
        "server_slot": slot_text(blade["server_slot"]),
        "SN in Excel": blade["serial_excel"],
        "SN on Server": "",
        "New SN on Server": "",
        "SN Status": "Pending",
        "PN on Server": "",
        "New PN on Server": "",
        "PN Status": "Pending",
        "Reboot Status": "Not Required",
        "Fallback": "",
        "IML Diagnostics": "",
        "IML File": "",
        "Details": "",
        "_slot": blade["server_slot"],
        "_ilo_ip": blade["ilo_ip"],
    }


def verify_pending(client, expected):
    latest = {"SerialNumber": "", "ProductId": ""}
    for attempt in range(1, vars.BL_SERIAL_VERIFY_RETRIES + 1):
        latest = client.pending_values()
        if all(same(latest.get(key), value) for key, value in expected.items()):
            return True, latest
        if attempt < vars.BL_SERIAL_VERIFY_RETRIES:
            time.sleep(vars.BL_SERIAL_VERIFY_DELAY_SECONDS)
    return False, latest


def verify_active_after_reboot(client, expected, slot):
    time.sleep(vars.BL_SERIAL_REBOOT_INITIAL_WAIT_SECONDS)
    deadline = time.monotonic() + vars.BL_SERIAL_REBOOT_VERIFY_TIMEOUT_SECONDS
    latest = {"SerialNumber": "", "ProductId": ""}
    last_error = ""
    next_progress = time.monotonic()

    while time.monotonic() < deadline:
        try:
            latest = client.active_values()
            if all(same(latest.get(key), value) for key, value in expected.items()):
                return True, latest, ""
            last_error = "ComputerSystem still reports SN={} | PN/SKU={}".format(
                shown(latest.get("SerialNumber")), shown(latest.get("ProductId"))
            )
        except Exception as exc:
            last_error = str(exc)

        if time.monotonic() >= next_progress:
            log(
                slot,
                client.ilo_ip,
                "Waiting for reboot/identity apply: {}".format(last_error),
            )
            next_progress = time.monotonic() + 30
        time.sleep(vars.BL_SERIAL_REBOOT_POLL_INTERVAL_SECONDS)

    return False, latest, last_error or "Timed out waiting for active identity"


def mark_changed_fields_failed(result, changes):
    if "SerialNumber" in changes:
        result["SN Status"] = "Failed"
    if "ProductId" in changes:
        result["PN Status"] = "Failed"


def process_blade(blade):
    result = new_result(blade)
    slot = blade["server_slot"]
    ilo_ip = blade["ilo_ip"]
    serial_excel = blade["serial_excel"]
    product_target = text(vars.BL_Part_Number)

    if is_invalid(ilo_ip):
        result.update({
            "SN Status": "Failed",
            "PN Status": "Failed",
            "Reboot Status": "Not Attempted",
            "Details": "iLO IP is missing or invalid",
        })
        log(slot, ilo_ip, result["Details"])
        return result

    client = IloRedfishClient(ilo_ip)
    try:
        log(slot, ilo_ip, "Connecting to iLO...")
        client.discover_system(slot)
        ilo_fw = text(
            client.manager_data.get("FirmwareVersion")
            or client.manager_data.get("Firmware", {})
            .get("Current", {})
            .get("VersionString")
        )
        rom_version = text(client.system_data.get("BiosVersion"))
        if ilo_fw or rom_version:
            log(
                slot,
                ilo_ip,
                "Firmware: iLO={} | System ROM={}".format(
                    shown(ilo_fw), shown(rom_version)
                ),
            )

        active = client.active_values()
        serial_server = active["SerialNumber"]
        product_server = active["ProductId"]
        result.update({
            "SN on Server": serial_server,
            "New SN on Server": serial_server,
            "PN on Server": product_server,
            "New PN on Server": product_server,
        })
        log(
            slot,
            ilo_ip,
            "Active identity (ComputerSystem): SN={} | PN/SKU={}".format(
                shown(serial_server), shown(product_server)
            ),
        )
        log(
            slot,
            ilo_ip,
            "Excel/target: SN={} | PN={}".format(
                shown(serial_excel), shown(product_target)
            ),
        )

        changes = {}
        details = []
        if is_invalid(serial_excel):
            result["SN Status"] = "Skipped"
            details.append("SN skipped: Excel serial is invalid")
            log(slot, ilo_ip, "Serial number skipped: Excel value is invalid.")
        elif same(serial_excel, serial_server):
            result["SN Status"] = "Skipped"
            details.append("SN skipped: active ComputerSystem serial already correct")
            log(slot, ilo_ip, "Serial number already correct; skipping.")
        else:
            changes["SerialNumber"] = serial_excel
            log(
                slot,
                ilo_ip,
                "Serial number differs; active {} -> target {}.".format(
                    shown(serial_server), serial_excel
                ),
            )

        if same(product_target, product_server):
            result["PN Status"] = "Skipped"
            details.append("PN skipped: active ComputerSystem SKU already correct")
            log(slot, ilo_ip, "Product ID already correct; skipping.")
        else:
            changes["ProductId"] = product_target
            log(
                slot,
                ilo_ip,
                "Product ID differs; active {} -> target {}.".format(
                    shown(product_server), product_target
                ),
            )

        # Critical behavior: do not require the BIOS provider when the active
        # identity is already correct. This prevents a BIOS-provider 404 from
        # being misreported as an empty serial/product ID.
        if not changes:
            result["Details"] = "; ".join(details)
            return result

        client.discover_bios(slot, blade.get("oa_ips", ()))
        if client.recovery_notes:
            details.extend(client.recovery_notes)
        current_bios = client.current_bios_values()
        pending_before = client.pending_values()
        log(
            slot,
            ilo_ip,
            "BIOS identity: current SN={} | PN={} ; pending SN={} | PN={}".format(
                shown(current_bios["SerialNumber"]),
                shown(current_bios["ProductId"]),
                shown(pending_before["SerialNumber"]),
                shown(pending_before["ProductId"]),
            ),
        )

        client.patch_bios(changes, slot)
        verified, pending = verify_pending(client, changes)
        if "SerialNumber" in changes:
            result["New SN on Server"] = pending["SerialNumber"]
        if "ProductId" in changes:
            result["New PN on Server"] = pending["ProductId"]

        if not verified:
            mark_changed_fields_failed(result, changes)
            result["Reboot Status"] = "Not Attempted"
            details.append("PATCH succeeded but pending BIOS values did not verify")
            result["Details"] = "; ".join(details)
            log(slot, ilo_ip, "BIOS Settings verification FAILED; not rebooting.")
            return result

        if "SerialNumber" in changes:
            result["SN Status"] = "Changed"
        if "ProductId" in changes:
            result["PN Status"] = "Changed"
        log(slot, ilo_ip, "BIOS Settings write verified.")

        if not vars.BL_SERIAL_REBOOT_ON_CHANGE:
            result["Reboot Status"] = "Disabled"
            details.append("BIOS Settings verified; reboot disabled by configuration")
            result["Details"] = "; ".join(details)
            log(slot, ilo_ip, "Automatic reboot disabled; changes remain pending.")
            return result

        result["Reboot Status"] = "Requested"
        reset_type, reset_uri = client.reboot()
        log(
            slot,
            ilo_ip,
            "{} requested through {}; verifying active ComputerSystem identity.".format(
                reset_type, reset_uri
            ),
        )
        active_ok, active_after, active_error = verify_active_after_reboot(
            client, changes, slot
        )
        if "SerialNumber" in changes:
            result["New SN on Server"] = active_after["SerialNumber"]
        if "ProductId" in changes:
            result["New PN on Server"] = active_after["ProductId"]

        if active_ok:
            result["Reboot Status"] = "Verified"
            details.append("Reboot completed; changed identity is active")
            log(
                slot,
                ilo_ip,
                "Reboot/apply verified. Active SN={} | PN/SKU={}".format(
                    shown(active_after["SerialNumber"]),
                    shown(active_after["ProductId"]),
                ),
            )
        else:
            mark_changed_fields_failed(result, changes)
            result["Reboot Status"] = "Failed"
            details.append(
                "Reboot requested but active identity verification failed: {}".format(
                    active_error
                )
            )
            log(
                slot,
                ilo_ip,
                "Reboot/apply verification FAILED: {}".format(active_error),
            )

        result["Details"] = "; ".join(details)
        return result
    except OfflineFallbackRequired as exc:
        if result["SN Status"] == "Pending":
            result["SN Status"] = "Offline"
        if result["PN Status"] == "Pending":
            result["PN Status"] = "Offline"
        result["Reboot Status"] = "Recovery Tried"
        result["Fallback"] = "F9/CONREP"
        result["IML Diagnostics"] = exc.iml_summary
        result["IML File"] = exc.iml_file
        result["Details"] = str(exc)
        log(
            slot,
            ilo_ip,
            "OFFLINE FALLBACK REQUIRED: F9/CONREP. {}".format(exc),
        )
        return result
    except Exception as exc:
        if result["SN Status"] == "Pending":
            result["SN Status"] = "Failed"
        if result["PN Status"] == "Pending":
            result["PN Status"] = "Failed"
        if result["Reboot Status"] == "Requested":
            result["Reboot Status"] = "Failed"
        elif result["Reboot Status"] == "Not Required":
            result["Reboot Status"] = "Not Attempted"
        result["Details"] = str(exc)
        log(slot, ilo_ip, "FAILED: {}".format(exc))
        return result
    finally:
        client.close()


def write_report(results):
    if not os.path.isdir(vars.LOG_DIR):
        os.makedirs(vars.LOG_DIR)
    report_path = os.path.join(
        vars.LOG_DIR,
        "{}_Report_{}.csv".format(
            vars.BL_SERIAL_REPORT_PREFIX,
            datetime.now().strftime("%Y%m%d-%H%M%S"),
        ),
    )
    fields = [
        "server_slot",
        "SN in Excel",
        "SN on Server",
        "New SN on Server",
        "SN Status",
        "PN on Server",
        "New PN on Server",
        "PN Status",
        "Reboot Status",
        "Fallback",
        "IML Diagnostics",
        "IML File",
        "Details",
    ]
    with open(report_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for result in results:
            writer.writerow({key: result.get(key, "") for key in fields})
    return report_path


def write_offline_fallback(enclosure_name, results):
    offline = [
        result for result in results
        if result.get("Fallback") == "F9/CONREP"
        or result.get("SN Status") == "Offline"
        or result.get("PN Status") == "Offline"
    ]
    if not offline:
        return "", ""

    if not os.path.isdir(vars.LOG_DIR):
        os.makedirs(vars.LOG_DIR)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    stem = "{}_Offline_Fallback_{}".format(vars.BL_SERIAL_REPORT_PREFIX, stamp)
    csv_path = os.path.join(vars.LOG_DIR, stem + ".csv")
    txt_path = os.path.join(vars.LOG_DIR, stem + ".txt")
    fields = [
        "enclosure", "server_slot", "ilo_ip", "current_serial", "target_serial",
        "current_product_id", "target_product_id", "fallback", "iml_file",
        "iml_diagnostics", "reason",
    ]
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for result in offline:
            writer.writerow({
                "enclosure": enclosure_name,
                "server_slot": result["server_slot"],
                "ilo_ip": result["_ilo_ip"],
                "current_serial": result["SN on Server"],
                "target_serial": result["SN in Excel"],
                "current_product_id": result["PN on Server"],
                "target_product_id": vars.BL_Part_Number,
                "fallback": result.get("Fallback", "F9/CONREP"),
                "iml_file": result.get("IML File", ""),
                "iml_diagnostics": result.get("IML Diagnostics", ""),
                "reason": result.get("Details", ""),
            })

    with open(txt_path, "w", encoding="utf-8") as handle:
        handle.write("BLADE SERIAL / PRODUCT ID OFFLINE RECOVERY HANDOFF\n")
        handle.write("Enclosure: {}\n".format(enclosure_name))
        handle.write("Generated: {}\n\n".format(datetime.now().isoformat()))
        handle.write(
            "Remote BIOS REST recovery was attempted once and stopped. "
            "No automatic OA E-fuse or iLO factory reset was performed.\n\n"
        )
        for result in offline:
            handle.write("Slot {} | iLO {}\n".format(
                result["server_slot"], result["_ilo_ip"]
            ))
            handle.write("  Current SN : {}\n".format(shown(result["SN on Server"])))
            handle.write("  Target SN  : {}\n".format(shown(result["SN in Excel"])))
            handle.write("  Current PN : {}\n".format(shown(result["PN on Server"])))
            handle.write("  Target PN  : {}\n".format(shown(vars.BL_Part_Number)))
            if result.get("IML File"):
                handle.write("  IML file   : {}\n".format(result["IML File"]))
            handle.write("\n")
        handle.write("NON-REST RECOVERY OPTIONS\n")
        handle.write(
            "1. F9 -> System Configuration -> BIOS/Platform Configuration "
            "(RBSU) -> Advanced Options -> Advanced System ROM Options.\n"
        )
        handle.write("2. Set Serial Number to the target value above.\n")
        handle.write("3. Set Product ID to {}.\n".format(vars.BL_Part_Number))
        handle.write("4. Press F10 to save and allow the server to reboot.\n")
        handle.write(
            "5. Alternative: use HPE Scripting Toolkit/CONREP offline with the "
            "platform-appropriate CONREP definition file; do not invent section "
            "names. Capture the server configuration first and modify only the "
            "supported identity fields.\n"
        )
        handle.write(
            "6. Re-run set_serial_BL.py afterward; it verifies the active identity "
            "from ComputerSystem before attempting any write.\n"
        )
    return csv_path, txt_path


def print_report(enclosure_name, results):
    line = "=" * 166
    print("\n" + line)
    print("FINAL SERIAL / PRODUCT ID REPORT: Enclosure '{}'".format(enclosure_name))
    print(line)
    print(
        "{:<5} | {:<16} | {:<16} | {:<16} | {:<9} | {:<12} | {:<12} | {:<9} | {:<13} | {:<10}".format(
            "SLOT", "SN IN EXCEL", "SN ON SERVER", "NEW SN", "SN STATUS",
            "PN ON SERVER", "NEW PN", "PN STATUS", "REBOOT", "FALLBACK",
        )
    )
    print("-" * 166)

    sn_changed = sn_skipped = sn_failed = sn_offline = 0
    pn_changed = pn_skipped = pn_failed = pn_offline = 0
    reboot_verified = reboot_failed = 0
    issues = []

    for result in results:
        print(
            "{:<5} | {:<16} | {:<16} | {:<16} | {:<9} | {:<12} | {:<12} | {:<9} | {:<13} | {:<10}".format(
                result["server_slot"], shown(result["SN in Excel"]),
                shown(result["SN on Server"]), shown(result["New SN on Server"]),
                result["SN Status"], shown(result["PN on Server"]),
                shown(result["New PN on Server"]), result["PN Status"],
                result["Reboot Status"][:13], result.get("Fallback", "")[:10],
            )
        )

        status = result["SN Status"]
        if status == "Changed": sn_changed += 1
        elif status == "Skipped": sn_skipped += 1
        elif status == "Failed": sn_failed += 1
        elif status == "Offline": sn_offline += 1

        status = result["PN Status"]
        if status == "Changed": pn_changed += 1
        elif status == "Skipped": pn_skipped += 1
        elif status == "Failed": pn_failed += 1
        elif status == "Offline": pn_offline += 1

        if result["Reboot Status"] == "Verified": reboot_verified += 1
        elif result["Reboot Status"] == "Failed": reboot_failed += 1

        if (
            result["SN Status"] in ("Failed", "Offline")
            or result["PN Status"] in ("Failed", "Offline")
        ):
            issues.append(result)

    print("-" * 166)
    print(
        "TOTAL: {} | SN Changed/Skipped/Failed/Offline: {}/{}/{}/{} | "
        "PN Changed/Skipped/Failed/Offline: {}/{}/{}/{} | "
        "Reboot Verified/Failed: {}/{}".format(
            len(results), sn_changed, sn_skipped, sn_failed, sn_offline,
            pn_changed, pn_skipped, pn_failed, pn_offline,
            reboot_verified, reboot_failed,
        )
    )
    print(line)

    if issues:
        print("\nFAILURE / OFFLINE FALLBACK DETAILS:")
        for result in issues:
            print(
                "  Slot {} | {} | {} | {}".format(
                    result["server_slot"], result["_ilo_ip"],
                    result.get("Fallback") or "Failed", result["Details"],
                )
            )
            if result.get("IML File"):
                print("    IML: {}".format(result["IML File"]))


def main():
    print("Python executable : {}".format(sys.executable))
    print("Project directory : {}".format(vars.PROJECT_DIR))
    print("Resource list     : {}".format(vars.RESOURCE_LIST))
    print("Sheet             : {}".format(vars.SHEET_NAME))
    print("Blade types       : {}".format(", ".join(vars.BL_SERIAL_VALID_EQUIPMENT_TYPES)))
    print("Target Product ID : {}".format(vars.BL_Part_Number))
    print("Parallel workers  : {}".format(vars.BL_SERIAL_MAX_WORKERS))
    print("Identity read     : ComputerSystem.SerialNumber + ComputerSystem.SKU")
    print("BIOS write lookup : direct Settings probe before parent BIOS/recovery")
    print("API preference    : {}".format(" -> ".join(vars.BL_SERIAL_API_PREFIXES)))
    print(
        "REST 404 recovery : {}".format(
            "ONE attempt, max {}s total ({}s initial wait)".format(
                QUICK_RECOVERY_TIMEOUT_SECONDS, QUICK_RECOVERY_INITIAL_WAIT_SECONDS
            )
            if vars.BL_SERIAL_REST_API_RECOVERY_ON_BIOS_404
            else "DISABLED"
        )
    )
    print("OA E-fuse fallback: DISABLED in V6")
    print("Offline fallback  : iLO SSH IML capture + F9/CONREP handoff")
    print(
        "Automatic reboot  : {}".format(
            "ENABLED - reboot on change ({})".format(vars.BL_SERIAL_REBOOT_TYPE)
            if vars.BL_SERIAL_REBOOT_ON_CHANGE
            else "DISABLED - changes apply at next server reboot"
        )
    )

    enclosure_name = input("\nEnter the Enclosure Name to process: ").strip()
    if not enclosure_name:
        print("No enclosure name entered.")
        return 1

    print("\nScanning Excel inventory for '{}'...".format(enclosure_name))
    blades, oa_ips = load_blades(enclosure_name)
    if not blades:
        print("No blade servers found for enclosure '{}'.".format(enclosure_name))
        return 0

    workers = min(vars.BL_SERIAL_MAX_WORKERS, len(blades))
    print("OA recovery IPs   : {}".format(", ".join(oa_ips) if oa_ips else "not found"))
    print("Found {} blade server(s); using {} worker(s).\n".format(len(blades), workers))

    results = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        futures = [executor.submit(process_blade, blade) for blade in blades]
        for future in concurrent.futures.as_completed(futures):
            results.append(future.result())

    results.sort(key=lambda item: item["_slot"])
    print_report(enclosure_name, results)
    report_path = write_report(results)
    fallback_csv, fallback_txt = write_offline_fallback(enclosure_name, results)
    print("\nCSV report: {}".format(report_path))
    if fallback_csv:
        print("Offline fallback CSV : {}".format(fallback_csv))
        print("Offline fallback guide: {}".format(fallback_txt))
    print(
        "NOTE: SN/PN on Server are read from the active ComputerSystem identity. "
        "V6 performs one bounded REST-state recovery only; if BIOS Settings remain "
        "unavailable it captures IML and hands the blade off to F9/CONREP."
    )

    unresolved = any(
        result["SN Status"] in ("Failed", "Offline")
        or result["PN Status"] in ("Failed", "Offline")
        for result in results
    )
    return 1 if unresolved else 0


if __name__ == "__main__":
    sys.exit(
        run_logged_main(
            main,
            log_prefix=vars.BL_SERIAL_REPORT_PREFIX,
            title="BLADE SERIAL / PRODUCT ID RESTORATION",
        )
    )
