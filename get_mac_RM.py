# BUILD_MARKER: GET_MAC_RM_CENTRAL_V1_20260828

import concurrent.futures
import gc
import os
import platform
import subprocess
import time
import warnings

import requests
import urllib3
import win32com.client as win32
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from functions import vars
from functions.output_log import run_logged_main
from functions.reporting import (
    make_summary_row,
    print_summary_report,
    write_summary_csv,
)


urllib3.disable_warnings(
    urllib3.exceptions.InsecureRequestWarning
)

warnings.filterwarnings(
    "ignore",
    category=DeprecationWarning,
)


# ============================================================================
# HELPERS
# ============================================================================

def print_debug(message):
    if vars.MAC_RM_DEBUG:
        print(
            "[DEBUG] {}".format(
                message
            )
        )


def check_file_writable(file_path):
    """Verify that Excel is not locking the workbook before we start."""
    if not os.path.exists(
        file_path
    ):
        return True

    try:
        with open(
            file_path,
            "a+",
        ):
            pass

        return True

    except PermissionError:
        return False


def get_robust_session():
    """Build one retry-enabled requests session for a worker."""
    session = requests.Session()
    session.verify = False

    retry_strategy = Retry(
        total=vars.MAC_RM_API_RETRIES,
        backoff_factor=(
            vars.MAC_RM_API_RETRY_BACKOFF
        ),
        status_forcelist=list(
            vars.MAC_RM_API_RETRY_STATUS_CODES
        ),
    )

    adapter = HTTPAdapter(
        max_retries=retry_strategy
    )

    session.mount(
        "https://",
        adapter,
    )

    session.mount(
        "http://",
        adapter,
    )

    return session


def ping_host(
    ip_address,
    count=None,
    timeout_ms=None,
):
    """Ping a rack-mount iLO before opening a Redfish session."""
    if count is None:
        count = vars.MAC_RM_PING_COUNT

    if timeout_ms is None:
        timeout_ms = (
            vars.MAC_RM_PING_TIMEOUT_MS
        )

    is_windows = (
        platform.system().lower()
        == "windows"
    )

    param_count = (
        "-n"
        if is_windows
        else "-c"
    )

    param_timeout = (
        "-w"
        if is_windows
        else "-W"
    )

    if is_windows:
        timeout_value = str(
            timeout_ms
        )
    else:
        timeout_value = str(
            max(
                1,
                int(
                    timeout_ms / 1000
                ),
            )
        )

    command = [
        "ping",
        param_count,
        str(
            count
        ),
        param_timeout,
        timeout_value,
        ip_address,
    ]

    print_debug(
        "Executing ping: {}".format(
            " ".join(
                command
            )
        )
    )

    try:
        output = subprocess.run(
            command,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )

        return (
            output.returncode
            == 0
        )

    except Exception as exc:
        print_debug(
            "Ping execution failed: {}".format(
                exc
            )
        )

        return False


def get_redfish_details(
    session,
    ilo_ip,
    user,
    password,
):
    """Read serial number and server NIC MAC addresses from iLO Redfish."""
    serial_number = (
        "Serial Number not found"
    )

    macs = []

    base_url = "https://{}".format(
        ilo_ip
    )

    auth = (
        user,
        password,
    )

    try:
        system_url = (
            "{}/redfish/v1/Systems/1"
            .format(
                base_url
            )
        )

        system_response = session.get(
            system_url,
            auth=auth,
            timeout=(
                vars.MAC_RM_API_TIMEOUT_SECONDS
            ),
        )

        if (
            system_response.status_code
            == 200
        ):
            serial_number = str(
                system_response.json().get(
                    "SerialNumber",
                    serial_number,
                )
            ).strip()

        ethernet_url = (
            "{}/redfish/v1/Systems/1/"
            "EthernetInterfaces"
            .format(
                base_url
            )
        )

        ethernet_response = session.get(
            ethernet_url,
            auth=auth,
            timeout=(
                vars.MAC_RM_API_TIMEOUT_SECONDS
            ),
        )

        if (
            ethernet_response.status_code
            == 200
        ):
            members = (
                ethernet_response.json().get(
                    "Members",
                    [],
                )
            )

            for member in members:
                interface_uri = (
                    member.get(
                        "@odata.id"
                    )
                )

                if not interface_uri:
                    continue

                interface_response = (
                    session.get(
                        "{}{}".format(
                            base_url,
                            interface_uri,
                        ),
                        auth=auth,
                        timeout=(
                            vars.MAC_RM_API_TIMEOUT_SECONDS
                        ),
                    )
                )

                if (
                    interface_response.status_code
                    != 200
                ):
                    continue

                mac = (
                    interface_response.json().get(
                        "MACAddress"
                    )
                )

                if not mac:
                    continue

                clean_mac = (
                    str(
                        mac
                    )
                    .replace(
                        "-",
                        ":",
                    )
                    .upper()
                )

                if clean_mac not in macs:
                    macs.append(
                        clean_mac
                    )

    except requests.exceptions.RequestException as exc:
        print_debug(
            "Connection error to {}: {}".format(
                ilo_ip,
                exc,
            )
        )

    macs = (
        macs
        + [""] * vars.MAC_MAX_ADDRESSES
    )[
        :vars.MAC_MAX_ADDRESSES
    ]

    return (
        serial_number,
        macs,
    )


def process_single_server(server):
    """ThreadPoolExecutor worker."""
    start_time = time.time()

    ilo_ip = server[
        "ilo_ip"
    ]

    if (
        not ilo_ip
        or ilo_ip.lower() == "none"
    ):
        result = dict(
            server
        )

        result.update({
            "status": "skipped",
            "msg": (
                "iLO IP is missing or invalid."
            ),
            "time_seconds": round(
                time.time() - start_time,
                2,
            ),
        })

        return result

    if not ping_host(
        ilo_ip
    ):
        result = dict(
            server
        )

        result.update({
            "status": "skipped",
            "msg": (
                "Unreachable via ping "
                "(Timeout: {}ms)."
                .format(
                    vars.MAC_RM_PING_TIMEOUT_MS
                )
            ),
            "time_seconds": round(
                time.time() - start_time,
                2,
            ),
        })

        return result

    with get_robust_session() as session:
        serial_no, macs = (
            get_redfish_details(
                session,
                ilo_ip,
                vars.ILO_RM_USERNAME,
                vars.ILO_RM_PASSWORD,
            )
        )

    result = dict(
        server
    )

    result.update({
        "status": "success",
        "serial_no": serial_no,
        "macs": macs,
        "time_seconds": round(
            time.time() - start_time,
            2,
        ),
    })

    return result


def refresh_excel_formulas(file_path):
    """Ask native Excel to rebuild cached formula values."""
    print(
        "\nCommanding Excel to recalculate formulas "
        "in the background..."
    )

    absolute_path = os.path.abspath(
        file_path
    )

    excel = None
    workbook = None

    try:
        excel = win32.DispatchEx(
            "Excel.Application"
        )

        excel.Visible = False
        excel.DisplayAlerts = False

        workbook = (
            excel.Workbooks.Open(
                absolute_path
            )
        )

        workbook.Save()

        print(
            "Formula cache successfully rebuilt!"
        )

    except Exception as exc:
        print(
            "WARNING: Could not trigger Excel to "
            "rebuild formulas: {}".format(
                exc
            )
        )

    finally:
        if workbook:
            try:
                workbook.Close(
                    SaveChanges=False
                )
            except Exception:
                pass

        if excel:
            try:
                excel.Quit()
            except Exception:
                pass

        del workbook
        del excel

        gc.collect()


# ============================================================================
# MAIN
# ============================================================================

def main():
    excel_path = (
        vars.RESOURCE_LIST
    )

    sheet_name = (
        vars.SHEET_NAME
    )

    if not check_file_writable(
        excel_path
    ):
        print(
            "\nERROR: '{}' is currently open in "
            "another program.".format(
                vars.EXCEL_FILENAME
            )
        )

        print(
            "Please close the file and run the script "
            "again to prevent permission errors."
        )

        return

    print(
        "\nScanning Excel file "
        "(Fast Read-Only Mode)..."
    )

    workbook_read = load_workbook(
        excel_path,
        data_only=True,
        read_only=True,
    )

    try:
        if (
            sheet_name
            not in workbook_read.sheetnames
        ):
            print(
                "Error: Sheet '{}' not found."
                .format(
                    sheet_name
                )
            )

            return

        worksheet_read = (
            workbook_read[
                sheet_name
            ]
        )

        header_row = next(
            worksheet_read.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not header_row:
            print(
                "Error: Excel header row is empty."
            )

            return

        headers = {}

        for index, cell_value in enumerate(
            header_row,
            1,
        ):
            if cell_value:
                headers[
                    str(
                        cell_value
                    ).strip()
                ] = index

        missing_columns = [
            column
            for column
            in vars.MAC_RM_REQUIRED_COLUMNS
            if column not in headers
        ]

        if missing_columns:
            print(
                "Error: Required columns missing: {}"
                .format(
                    ", ".join(
                        missing_columns
                    )
                )
            )

            return

        target_equipment = dict(
            vars.MAC_RM_TARGET_EQUIPMENT
        )

        servers_to_process = []

        max_row = min(
            int(
                vars.END_ROW
            ),
            worksheet_read.max_row,
        )

        print(
            "Identifying Target Equipment..."
        )

        for row_idx, row_values in enumerate(
            worksheet_read.iter_rows(
                min_row=int(
                    vars.START_ROW
                ),
                max_row=max_row,
                values_only=True,
            ),
            int(
                vars.START_ROW
            ),
        ):
            if not any(
                row_values
            ):
                continue

            raw_equipment = row_values[
                headers[
                    "equipment_type"
                ] - 1
            ]

            equipment_type = str(
                raw_equipment
                or ""
            ).strip().upper()

            if (
                equipment_type
                not in target_equipment
            ):
                continue

            ilo_ip = str(
                row_values[
                    headers[
                        "ilo_ip"
                    ] - 1
                ]
                or ""
            ).strip()

            servers_to_process.append({
                "row_idx": row_idx,
                "eq_type": equipment_type,
                "ilo_ip": ilo_ip,
                "expected_macs": (
                    target_equipment[
                        equipment_type
                    ]
                ),
            })

    finally:
        workbook_read.close()

    total_servers = len(
        servers_to_process
    )

    print(
        "\nFound {} rackmount server(s) "
        "matching criteria.\n{}"
        .format(
            total_servers,
            "=" * 60,
        )
    )

    if total_servers == 0:
        return

    max_workers = min(
        int(
            vars.MAC_RM_MAX_WORKERS
        ),
        total_servers,
    )

    print(
        "Starting parallel processing "
        "(Max {} concurrent threads)..."
        .format(
            max_workers
        )
    )

    results = []

    with concurrent.futures.ThreadPoolExecutor(
        max_workers=max_workers
    ) as executor:
        future_to_server = {
            executor.submit(
                process_single_server,
                server,
            ): server
            for server
            in servers_to_process
        }

        for count, future in enumerate(
            concurrent.futures.as_completed(
                future_to_server
            ),
            1,
        ):
            try:
                result = (
                    future.result()
                )

                results.append(
                    result
                )

                print(
                    "[{}/{}] Completed Row {} ({}) - {}"
                    .format(
                        count,
                        total_servers,
                        result[
                            "row_idx"
                        ],
                        result[
                            "ilo_ip"
                        ],
                        result[
                            "status"
                        ].upper(),
                    )
                )

            except Exception as exc:
                server = (
                    future_to_server[
                        future
                    ]
                )

                print(
                    "[{}/{}] Row {} generated an "
                    "exception: {}".format(
                        count,
                        total_servers,
                        server[
                            "row_idx"
                        ],
                        exc,
                    )
                )

                failed_result = dict(server)
                failed_result.update({
                    "status": "failed",
                    "msg": str(exc),
                    "serial_no": "",
                    "macs": [""] * vars.MAC_MAX_ADDRESSES,
                    "time_seconds": 0.0,
                })
                results.append(failed_result)

    print(
        "\nLoading Excel file "
        "(Writing Mode) to inject data..."
    )

    workbook_write = load_workbook(
        excel_path
    )

    worksheet_write = (
        workbook_write[
            sheet_name
        ]
    )

    results.sort(
        key=lambda item:
            item[
                "row_idx"
            ]
    )

    for result in results:
        row_idx = result[
            "row_idx"
        ]

        equipment_type = result[
            "eq_type"
        ]

        ilo_ip = result[
            "ilo_ip"
        ]

        print(
            "\n--- Row {}: {} "
            "(iLO IP: {}) ---"
            .format(
                row_idx,
                equipment_type,
                ilo_ip,
            )
        )

        if (
            result[
                "status"
            ]
            in ("skipped", "failed")
        ):
            print(
                "WARNING: {}".format(
                    result.get(
                        "msg",
                        "MAC collection failed",
                    )
                )
            )

            continue

        serial_no = result[
            "serial_no"
        ]

        macs = result[
            "macs"
        ]

        expected_macs = result[
            "expected_macs"
        ]

        worksheet_write.cell(
            row=row_idx,
            column=headers[
                vars.MAC_SERIAL_COLUMN
            ],
        ).value = serial_no

        for index, column_name in enumerate(
            vars.MAC_NIC_COLUMNS
        ):
            worksheet_write.cell(
                row=row_idx,
                column=headers[
                    column_name
                ],
            ).value = macs[
                index
            ]

        print(
            "  Serial Number : {}".format(
                serial_no
            )
        )

        found_mac_count = 0

        for index in range(
            expected_macs
        ):
            mac_value = (
                macs[
                    index
                ]
            )

            if not mac_value:
                print(
                    "  mac{}          : [MISSING]"
                    .format(
                        index + 1
                    )
                )

                print(
                    "  -> WARNING: Expected mac{} "
                    "was not found on this server."
                    .format(
                        index + 1
                    )
                )

            else:
                print(
                    "  mac{}          : {}"
                    .format(
                        index + 1,
                        mac_value,
                    )
                )

                found_mac_count += 1

        if (
            found_mac_count
            < expected_macs
        ):
            print(
                "  -> OVERALL WARNING: Expected {} "
                "MACs, but only found {}."
                .format(
                    expected_macs,
                    found_mac_count,
                )
            )

    print(
        "\nSaving updates to {}..."
        .format(
            excel_path
        )
    )

    workbook_write.save(
        excel_path
    )

    workbook_write.close()

    print(
        "Excel file successfully updated!"
    )

    refresh_excel_formulas(
        excel_path
    )

    summary_rows = []

    for result in results:
        raw_status = result.get("status", "failed")
        macs = result.get(
            "macs",
            [""] * vars.MAC_MAX_ADDRESSES,
        )
        found_count = len([mac for mac in macs if mac])
        expected_count = result.get("expected_macs", 0)

        if raw_status == "skipped":
            status = "Skipped"
        elif raw_status == "failed":
            status = "Failed"
        elif found_count < expected_count:
            status = "Partial"
        else:
            status = "Successful"

        details = result.get("msg", "")

        if raw_status == "success":
            details = (
                "Serial={}; MACs={}/{}"
                .format(
                    result.get("serial_no", ""),
                    found_count,
                    expected_count,
                )
            )

        summary_rows.append(
            make_summary_row(
                row=result.get("row_idx", "-"),
                item_type=result.get("eq_type", "Rackmount"),
                name=result.get("eq_type", "Rackmount"),
                target=result.get("ilo_ip", ""),
                status=status,
                time_seconds=result.get("time_seconds", 0.0),
                details=details,
            )
        )

    print_summary_report(
        summary_rows,
        title="FINAL RACK-MOUNT MAC COLLECTION SUMMARY",
    )

    write_summary_csv(
        summary_rows,
        vars.SCRIPT_ARTIFACT_PREFIXES[
            "get_mac_RM"
        ],
    )

    return (
        1
        if any(
            row["Status"] in ("Failed", "Partial")
            for row in summary_rows
        )
        else 0
    )


if __name__ == "__main__":
    import sys

    sys.exit(
        run_logged_main(
            main,
            log_prefix=vars.SCRIPT_ARTIFACT_PREFIXES[
                "get_mac_RM"
            ],
            title="RACK-MOUNT MAC COLLECTION",
        )
    )
