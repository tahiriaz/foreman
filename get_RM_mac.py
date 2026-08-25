import os
import re
import warnings
import platform
import subprocess
import requests
import gc
import concurrent.futures
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import load_workbook
import win32com.client as win32

# Suppress warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==========================================
# Global Configuration Variables
# ==========================================
USERNAME = "Administrator"
PASSWORD = "Th@les01"

# Multithreading Settings
MAX_CONCURRENT_SERVERS = 16

# Directory and File Settings
EXCEL_DIR = "Templates"
EXCEL_FILE = "Resource List-v7.6.xlsx"
SHEET_NAME = "General Resource List"

# Processing Scope
START_ROW = 1057
END_ROW = 1072  # Adjust as needed, or set to None to process to the end of the sheet

# Target Equipment Types and Expected MACs (Case-Insensitive)
TARGET_EQUIPMENT = {
    "ESXi": 4,
    "NVR": 4,
    "VCA": 4
}

# Required columns for a row to be processed
REQUIRED_COLUMNS = ['ilo_ip', 'equipment_type']

# Ping & Network Settings
PING_COUNT = 2
PING_TIMEOUT_MS = 300
API_TIMEOUT_SEC = 10
API_RETRIES = 3

# Debug Toggle
DEBUG = False
# ==========================================

def print_debug(msg):
    if DEBUG:
        print(f"[DEBUG] {msg}")

def check_file_writable(file_path):
    """Verifies the Excel file is not locked by another process before starting."""
    if not os.path.exists(file_path):
        return True
    try:
        with open(file_path, "a+"):
            pass
        return True
    except PermissionError:
        return False

def get_robust_session():
    """Configures a requests session with automatic retries for resilient API calls."""
    session = requests.Session()
    session.verify = False
    retry_strategy = Retry(
        total=API_RETRIES,
        backoff_factor=1,
        status_forcelist=[429, 500, 502, 503, 504]
    )
    adapter = HTTPAdapter(max_retries=retry_strategy)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session

def ping_host(ip_address, count=PING_COUNT, timeout_ms=PING_TIMEOUT_MS):
    """Pings a host to check if it's alive before attempting an API connection."""
    param_count = '-n' if platform.system().lower() == 'windows' else '-c'
    param_timeout = '-w' if platform.system().lower() == 'windows' else '-W'
    
    if platform.system().lower() != 'windows':
        timeout_val = str(max(1, int(timeout_ms / 1000)))
    else:
        timeout_val = str(timeout_ms)

    command = ['ping', param_count, str(count), param_timeout, timeout_val, ip_address]
    print_debug(f"Executing ping: {' '.join(command)}")
    
    try:
        # Hide output completely to keep console clean during multithreading
        output = subprocess.run(command, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        return output.returncode == 0
    except Exception as e:
        print_debug(f"Ping execution failed: {e}")
        return False

def get_redfish_details(session, ilo_ip, user, pwd):
    """Connects to the iLO Redfish API to get Serial Number and NIC MAC addresses."""
    serial_number = "Serial Number not found"
    macs = []
    base_url = f"https://{ilo_ip}"
    auth = (user, pwd)
    
    try:
        # 1. Get Serial Number
        sys_url = f"{base_url}/redfish/v1/Systems/1"
        sys_resp = session.get(sys_url, auth=auth, timeout=API_TIMEOUT_SEC)
        
        if sys_resp.status_code == 200:
            serial_number = sys_resp.json().get('SerialNumber', serial_number).strip()

        # 2. Get MAC Addresses
        eth_url = f"{base_url}/redfish/v1/Systems/1/EthernetInterfaces"
        eth_resp = session.get(eth_url, auth=auth, timeout=API_TIMEOUT_SEC)
        
        if eth_resp.status_code == 200:
            members = eth_resp.json().get('Members', [])
            for member in members:
                ifc_uri = member.get('@odata.id')
                if ifc_uri:
                    ifc_resp = session.get(f"{base_url}{ifc_uri}", auth=auth, timeout=API_TIMEOUT_SEC)
                    if ifc_resp.status_code == 200:
                        mac = ifc_resp.json().get('MACAddress')
                        if mac:
                            clean_mac = mac.replace('-', ':').upper()
                            if clean_mac not in macs:
                                macs.append(clean_mac)

    except requests.exceptions.RequestException as e:
        print_debug(f"Connection error to {ilo_ip}: {e}")

    # Pad to exactly 6 elements
    return serial_number, (macs + [""] * 6)[:6]

def process_single_server(server):
    """Worker function executed by ThreadPoolExecutor."""
    ilo_ip = server['ilo_ip']
    row_idx = server['row_idx']
    
    if not ilo_ip or ilo_ip.lower() == "none":
        return {**server, 'status': 'skipped', 'msg': 'iLO IP is missing or invalid.'}
        
    if not ping_host(ilo_ip):
        return {**server, 'status': 'skipped', 'msg': f'Unreachable via ping (Timeout: {PING_TIMEOUT_MS}ms).'}

    # Use a fresh robust session per thread
    with get_robust_session() as session:
        serial_no, macs = get_redfish_details(session, ilo_ip, USERNAME, PASSWORD)
        
    return {**server, 'status': 'success', 'serial_no': serial_no, 'macs': macs}

def refresh_excel_formulas(file_path):
    """Safely recalculates formulas and strictly guarantees Excel process termination via COM cleanup."""
    print("\nCommanding Excel to recalculate formulas in the background...")
    abs_path = os.path.abspath(file_path)
    excel = None
    wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_path)
        wb.Save()
        print("Formula cache successfully rebuilt!")
    except Exception as e:
        print(f"WARNING: Could not trigger Excel to rebuild formulas: {e}")
    finally:
        if wb:
            try: wb.Close(SaveChanges=False)
            except Exception: pass
        if excel:
            try: excel.Quit()
            except Exception: pass
        del wb
        del excel
        gc.collect() # Force COM pointer release

def main():
    global TARGET_EQUIPMENT
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, EXCEL_DIR, EXCEL_FILE)
    
    if not check_file_writable(excel_path):
        print(f"\nERROR: '{EXCEL_FILE}' is currently open in another program.")
        print("Please close the file and run the script again to prevent permission errors.")
        return

    print(f"\nScanning Excel file (Fast Read-Only Mode)...")
    # OPTIMIZATION: read_only=True cuts memory usage and speeds up parsing significantly
    wb_read = load_workbook(excel_path, data_only=True, read_only=True)
    
    if SHEET_NAME not in wb_read.sheetnames:
        print(f"Error: Sheet '{SHEET_NAME}' not found.")
        return
        
    ws_read = wb_read[SHEET_NAME]
    
    # Map headers to column indices (assuming row 1 holds headers)
    headers = {}
    for idx, cell_value in enumerate(next(ws_read.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
        if cell_value: headers[str(cell_value).strip()] = idx
    
    missing_cols = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing_cols:
        print(f"Error: Required columns missing: {', '.join(missing_cols)}")
        return

    TARGET_EQUIPMENT = {k.upper(): v for k, v in TARGET_EQUIPMENT.items()}
    servers_to_process = []
    
    max_row = END_ROW if END_ROW else ws_read.max_row
    max_row = min(max_row, ws_read.max_row)

    print("Identifying Target Equipment...")
    # iter_rows is highly optimized for read_only mode
    for row_idx, row_values in enumerate(ws_read.iter_rows(min_row=START_ROW, max_row=max_row, values_only=True), START_ROW):
        if not any(row_values): continue # Skip empty rows safely
        
        raw_eq = row_values[headers['equipment_type'] - 1]
        eq_type = str(raw_eq or "").strip().upper()
        
        if eq_type in TARGET_EQUIPMENT:
            ilo_ip = str(row_values[headers['ilo_ip'] - 1] or "").strip()
            servers_to_process.append({
                'row_idx': row_idx,
                'eq_type': eq_type,
                'ilo_ip': ilo_ip,
                'expected_macs': TARGET_EQUIPMENT[eq_type]
            })

    wb_read.close() # Free resources
    
    total_servers = len(servers_to_process)
    print(f"\nFound {total_servers} rackmount server(s) matching criteria.\n" + "="*60)
    
    if total_servers == 0:
        return

    # Multithreading Execution
    print(f"Starting parallel processing (Max {MAX_CONCURRENT_SERVERS} concurrent threads)...")
    results = []
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_CONCURRENT_SERVERS) as executor:
        # Submit all tasks
        future_to_server = {executor.submit(process_single_server, s): s for s in servers_to_process}
        
        # Process as they complete
        for count, future in enumerate(concurrent.futures.as_completed(future_to_server), 1):
            try:
                res = future.result()
                results.append(res)
                print(f"[{count}/{total_servers}] Completed Row {res['row_idx']} ({res['ilo_ip']}) - {res['status'].upper()}")
            except Exception as exc:
                server = future_to_server[future]
                print(f"[{count}/{total_servers}] Row {server['row_idx']} generated an exception: {exc}")

    # Sequential File Write (Thread-Safe)
    print(f"\nLoading Excel file (Writing Mode) to inject data...")
    wb_write = load_workbook(excel_path)
    ws_write = wb_write[SHEET_NAME]
    
    # Sort results by row index to print them in order
    results.sort(key=lambda x: x['row_idx'])
    
    for res in results:
        row_idx = res['row_idx']
        eq_type = res['eq_type']
        ilo_ip = res['ilo_ip']
        
        print(f"\n--- Row {row_idx}: {eq_type} (iLO IP: {ilo_ip}) ---")
        
        if res['status'] == 'skipped':
            print(f"WARNING: {res['msg']}")
            continue
            
        serial_no = res['serial_no']
        macs = res['macs']
        expected_macs = res['expected_macs']
        
        # Write to Formula-preserved workbook
        ws_write.cell(row=row_idx, column=headers['serial_no']).value = serial_no
        ws_write.cell(row=row_idx, column=headers['nic0_mac']).value = macs[0]
        ws_write.cell(row=row_idx, column=headers['nic1_mac']).value = macs[1]
        ws_write.cell(row=row_idx, column=headers['nic2_mac']).value = macs[2]
        ws_write.cell(row=row_idx, column=headers['nic3_mac']).value = macs[3]
        ws_write.cell(row=row_idx, column=headers['nic4_mac']).value = macs[4]
        ws_write.cell(row=row_idx, column=headers['nic5_mac']).value = macs[5]
        
        # Print outputs
        print(f"  Serial Number : {serial_no}")
        found_mac_count = 0
        
        for i in range(expected_macs):
            mac_val = macs[i]
            if not mac_val:
                print(f"  mac{i+1}          : [MISSING]")
                print(f"  -> WARNING: Expected mac{i+1} was not found on this server.")
            else:
                print(f"  mac{i+1}          : {mac_val}")
                found_mac_count += 1
                
        if found_mac_count < expected_macs:
            print(f"  -> OVERALL WARNING: Expected {expected_macs} MACs, but only found {found_mac_count}.")

    # Save Excel Changes
    print(f"\nSaving updates to {excel_path}...")
    wb_write.save(excel_path)
    print("Excel file successfully updated!")
    
    # Trigger COM cache rebuild natively
    refresh_excel_formulas(excel_path)

if __name__ == "__main__":
    main()