#!/usr/bin/env python3

import gc
import os
import shutil
import warnings
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

warnings.filterwarnings(
    "ignore",
    message="Data Validation extension is not supported.*"
)


# =============================================================================
# CONFIGURATION
# =============================================================================

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(PROJECT_DIR, "Templates")

SOURCE_FILE = os.path.join(
    TEMPLATE_DIR,
    "Shamiya Expansion - Inventory Tracking v1.11.xlsx"
)
SOURCE_SHEET = "HP Inventory Tracking"

RESOURCE_FILE = os.path.join(
    TEMPLATE_DIR,
    "Resource List-v7.7.xlsx"
)
RESOURCE_SHEET = "General Resource List"

SOURCE_ENCLOSURE_COLUMN = "Enclosure_Physical_Name"
SOURCE_SLOT_COLUMN = "ID"
SOURCE_SERIAL_COLUMN = "serial_no"

RESOURCE_ENCLOSURE_COLUMN = "enclosure_physical_name"
RESOURCE_SLOT_COLUMN = "enclosure_slot"
RESOURCE_EQUIPMENT_TYPE_COLUMN = "equipment_type"
RESOURCE_SERIAL_COLUMN = "serial_no"

VALID_EQUIPMENT_TYPES = {
    "NVR Blade Server",
    "VCA Blade Server",
}

INVALID_VALUES = {
    "",
    "none",
    "null",
    "n/a",
    "n.a",
    "n.a.",
    "na",
    "unknown",
    "invalid",
    "not available",
    "not applicable",
    "-",
    "--",
}

CREATE_BACKUP = True
SAVE_ONLY_IF_CHANGED = True
REBUILD_FORMULA_CACHE = True
PRINT_UPDATED_ROWS = True


# =============================================================================
# HELPERS
# =============================================================================

def clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value):
    return clean_text(value).upper()


def normalize_slot(value):
    value = clean_text(value)
    if not value:
        return ""

    try:
        number = float(value)
        if number.is_integer():
            return str(int(number))
    except (TypeError, ValueError):
        pass

    return value.upper()


def normalize_serial(value):
    return clean_text(value).upper()


def is_valid_serial(value):
    serial = clean_text(value)
    return bool(serial) and serial.lower() not in INVALID_VALUES


def get_header_map(ws):
    row = next(
        ws.iter_rows(
            min_row=1,
            max_row=1,
            values_only=True
        )
    )

    return {
        clean_text(value).lower(): column
        for column, value in enumerate(row, start=1)
        if clean_text(value)
    }


def require_columns(header_map, columns, sheet_name):
    missing = [
        column
        for column in columns
        if column.lower() not in header_map
    ]

    if missing:
        raise RuntimeError(
            "Missing required column(s) in sheet '{}': {}".format(
                sheet_name,
                ", ".join(missing)
            )
        )


def create_backup(file_path):
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    base, ext = os.path.splitext(file_path)

    backup_path = "{}_BACKUP_{}{}".format(
        base,
        timestamp,
        ext
    )

    shutil.copy2(file_path, backup_path)
    return backup_path


def get_resource_table_end_row(ws):
    tables = []

    for table_name in ws.tables.keys():
        table = ws.tables[table_name]

        try:
            _, min_row, _, max_row = range_boundaries(table.ref)
        except (TypeError, ValueError, AttributeError):
            continue

        if min_row == 1:
            tables.append(
                (max_row, table_name, table.ref)
            )

    if not tables:
        raise RuntimeError(
            "No Excel table found in sheet '{}'. "
            "The script will not use worksheet max_row because this "
            "workbook contains formatting near Excel's row limit.".format(
                RESOURCE_SHEET
            )
        )

    tables.sort(reverse=True)
    max_row, table_name, table_ref = tables[0]

    print(
        "Resource Excel table           : {}".format(
            table_name
        )
    )
    print(
        "Resource Excel table range     : {}".format(
            table_ref
        )
    )
    print(
        "Resource data ends at row      : {}".format(
            max_row
        )
    )

    return max_row


def refresh_excel_formulas(file_path):
    """
    Open the workbook using native Microsoft Excel, perform a complete
    recalculation, and save it.

    openpyxl preserves formula expressions but does not calculate formulas.
    Saving with openpyxl can therefore leave formula cached values empty.

    Native Excel rebuilds those cached results so future data_only=True
    reads return the calculated values.
    """
    print("\nRebuilding Excel formula cache...")

    try:
        import pythoncom
        import win32com.client as win32
    except ImportError:
        raise RuntimeError(
            "Formula cache rebuild requires pywin32. "
            "Install it with: pip install pywin32"
        )

    excel = None
    workbook = None
    pythoncom.CoInitialize()

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False

        workbook = excel.Workbooks.Open(
            os.path.abspath(file_path),
            UpdateLinks=0,
            ReadOnly=False
        )

        print("Running Excel full formula recalculation...")

        # xlCalculationAutomatic = -4105
        try:
            excel.Calculation = -4105
        except Exception:
            pass

        excel.CalculateFullRebuild()

        print("Saving recalculated workbook...")

        workbook.Save()

        print("Formula cache successfully rebuilt.")

    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass

        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass

        workbook = None
        excel = None

        gc.collect()

        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass


def verify_formula_cache(
        file_path,
        table_end_row,
        equipment_col,
        enclosure_col,
        slot_col):
    """
    Verify that formula-derived Resource List fields are readable through
    openpyxl data_only=True after native Excel recalculation.
    """
    print("Verifying rebuilt formula cache...")

    wb = load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )

    try:
        ws = wb[RESOURCE_SHEET]

        min_col = min(
            equipment_col,
            enclosure_col,
            slot_col
        )
        max_col = max(
            equipment_col,
            enclosure_col,
            slot_col
        )

        valid_types = {
            normalize_name(value)
            for value in VALID_EQUIPMENT_TYPES
        }

        eligible = 0
        indexed = 0

        for row in ws.iter_rows(
                min_row=2,
                max_row=table_end_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True):

            equipment_type = clean_text(
                row[equipment_col - min_col]
            )

            if normalize_name(equipment_type) not in valid_types:
                continue

            eligible += 1

            enclosure = normalize_name(
                row[enclosure_col - min_col]
            )
            slot = normalize_slot(
                row[slot_col - min_col]
            )

            if enclosure and slot:
                indexed += 1

        if eligible == 0:
            raise RuntimeError(
                "Formula cache verification failed: zero eligible "
                "NVR/VCA blade rows were readable after recalculation."
            )

        print(
            "Formula cache verification     : OK"
        )
        print(
            "Cached eligible blade rows     : {}".format(
                eligible
            )
        )
        print(
            "Cached indexed blade rows      : {}".format(
                indexed
            )
        )

        return eligible

    finally:
        wb.close()


# =============================================================================
# MAIN
# =============================================================================

def main():
    print("=" * 100)
    print("BLADE SERIAL NUMBER COPY")
    print("=" * 100)
    print("Source   : {}".format(SOURCE_FILE))
    print("Sheet    : {}".format(SOURCE_SHEET))
    print("Target   : {}".format(RESOURCE_FILE))
    print("Sheet    : {}".format(RESOURCE_SHEET))
    print("=" * 100)

    if not os.path.isfile(SOURCE_FILE):
        raise RuntimeError(
            "Source workbook not found: {}".format(
                SOURCE_FILE
            )
        )

    if not os.path.isfile(RESOURCE_FILE):
        raise RuntimeError(
            "Resource workbook not found: {}".format(
                RESOURCE_FILE
            )
        )

    # =========================================================================
    # LOAD SOURCE
    # =========================================================================

    print("\nLoading source workbook...")

    source_wb = load_workbook(
        SOURCE_FILE,
        read_only=True,
        data_only=True
    )

    try:
        if SOURCE_SHEET not in source_wb.sheetnames:
            raise RuntimeError(
                "Source sheet '{}' not found.".format(
                    SOURCE_SHEET
                )
            )

        source_ws = source_wb[SOURCE_SHEET]
        source_headers = get_header_map(source_ws)

        require_columns(
            source_headers,
            [
                SOURCE_ENCLOSURE_COLUMN,
                SOURCE_SLOT_COLUMN,
                SOURCE_SERIAL_COLUMN,
            ],
            SOURCE_SHEET
        )

        enclosure_col = source_headers[
            SOURCE_ENCLOSURE_COLUMN.lower()
        ]
        slot_col = source_headers[
            SOURCE_SLOT_COLUMN.lower()
        ]
        serial_col = source_headers[
            SOURCE_SERIAL_COLUMN.lower()
        ]

        read_min_col = min(
            enclosure_col,
            slot_col,
            serial_col
        )
        read_max_col = max(
            enclosure_col,
            slot_col,
            serial_col
        )

        source_serials = {}
        source_conflicts = {}
        invalid_serials = []
        source_rows_with_key = 0

        print(
            "Reading blade serial numbers from source workbook..."
        )

        for row_num, row in enumerate(
                source_ws.iter_rows(
                    min_row=2,
                    max_row=source_ws.max_row,
                    min_col=read_min_col,
                    max_col=read_max_col,
                    values_only=True
                ),
                start=2):

            enclosure = normalize_name(
                row[enclosure_col - read_min_col]
            )
            slot = normalize_slot(
                row[slot_col - read_min_col]
            )
            serial = normalize_serial(
                row[serial_col - read_min_col]
            )

            if not enclosure or not slot:
                continue

            source_rows_with_key += 1

            if not is_valid_serial(serial):
                invalid_serials.append({
                    "row": row_num,
                    "enclosure": enclosure,
                    "slot": slot,
                    "serial": serial,
                })
                continue

            key = (enclosure, slot)

            if key in source_conflicts:
                conflict = source_conflicts[key]

                if serial not in conflict["serials"]:
                    conflict["rows"].append(row_num)
                    conflict["serials"].append(serial)

                continue

            if key in source_serials:
                existing = source_serials[key]

                if existing["serial"] != serial:
                    source_conflicts[key] = {
                        "enclosure": enclosure,
                        "slot": slot,
                        "rows": [
                            existing["row"],
                            row_num
                        ],
                        "serials": [
                            existing["serial"],
                            serial
                        ],
                    }

                    del source_serials[key]

                continue

            source_serials[key] = {
                "row": row_num,
                "enclosure": enclosure,
                "slot": slot,
                "serial": serial,
            }

    finally:
        source_wb.close()

    print(
        "Source rows with enclosure/slot : {}".format(
            source_rows_with_key
        )
    )
    print(
        "Usable source serial numbers    : {}".format(
            len(source_serials)
        )
    )
    print(
        "Invalid source serials          : {}".format(
            len(invalid_serials)
        )
    )
    print(
        "Conflicting source blades       : {}".format(
            len(source_conflicts)
        )
    )

    # =========================================================================
    # BACKUP
    # =========================================================================

    if CREATE_BACKUP:
        backup_path = create_backup(RESOURCE_FILE)

        print(
            "Backup created: {}".format(
                backup_path
            )
        )

    # =========================================================================
    # LOAD RESOURCE LIST
    # =========================================================================

    print("\nLoading Resource List...")

    # This workbook is used for writing.
    # data_only MUST NOT be enabled here because formulas must be retained.
    resource_wb = load_workbook(RESOURCE_FILE)

    try:
        if RESOURCE_SHEET not in resource_wb.sheetnames:
            raise RuntimeError(
                "Resource sheet '{}' not found.".format(
                    RESOURCE_SHEET
                )
            )

        resource_ws = resource_wb[RESOURCE_SHEET]
        resource_headers = get_header_map(resource_ws)

        require_columns(
            resource_headers,
            [
                RESOURCE_ENCLOSURE_COLUMN,
                RESOURCE_SLOT_COLUMN,
                RESOURCE_EQUIPMENT_TYPE_COLUMN,
                RESOURCE_SERIAL_COLUMN,
            ],
            RESOURCE_SHEET
        )

        res_enclosure_col = resource_headers[
            RESOURCE_ENCLOSURE_COLUMN.lower()
        ]
        res_slot_col = resource_headers[
            RESOURCE_SLOT_COLUMN.lower()
        ]
        res_equipment_col = resource_headers[
            RESOURCE_EQUIPMENT_TYPE_COLUMN.lower()
        ]
        res_serial_col = resource_headers[
            RESOURCE_SERIAL_COLUMN.lower()
        ]

        resource_end_row = get_resource_table_end_row(
            resource_ws
        )

        # ---------------------------------------------------------------------
        # READ CACHED FORMULA VALUES
        # ---------------------------------------------------------------------

        resource_values_wb = load_workbook(
            RESOURCE_FILE,
            read_only=True,
            data_only=True
        )

        try:
            resource_values_ws = resource_values_wb[
                RESOURCE_SHEET
            ]

            read_min_col = min(
                res_equipment_col,
                res_enclosure_col,
                res_slot_col
            )
            read_max_col = max(
                res_equipment_col,
                res_enclosure_col,
                res_slot_col
            )

            valid_types = {
                normalize_name(value)
                for value in VALID_EQUIPMENT_TYPES
            }

            resource_index = {}
            duplicate_resource_keys = []
            eligible_resource_rows = 0

            print(
                "Reading calculated Resource List blade information..."
            )

            for row_num, row in enumerate(
                    resource_values_ws.iter_rows(
                        min_row=2,
                        max_row=resource_end_row,
                        min_col=read_min_col,
                        max_col=read_max_col,
                        values_only=True
                    ),
                    start=2):

                equipment_type = clean_text(
                    row[res_equipment_col - read_min_col]
                )

                if (
                    normalize_name(equipment_type)
                    not in valid_types
                ):
                    continue

                eligible_resource_rows += 1

                enclosure = normalize_name(
                    row[res_enclosure_col - read_min_col]
                )
                slot = normalize_slot(
                    row[res_slot_col - read_min_col]
                )

                if not enclosure or not slot:
                    continue

                key = (enclosure, slot)

                if key in resource_index:
                    duplicate_resource_keys.append({
                        "key": key,
                        "first_row": resource_index[key],
                        "second_row": row_num,
                    })
                    continue

                resource_index[key] = row_num

        finally:
            resource_values_wb.close()

        print(
            "Eligible Resource List blades   : {}".format(
                eligible_resource_rows
            )
        )
        print(
            "Indexed Resource List blades    : {}".format(
                len(resource_index)
            )
        )

        if eligible_resource_rows == 0:
            raise RuntimeError(
                "No NVR/VCA blades could be read from the cached "
                "Resource List formulas."
            )

        # =========================================================================
        # MATCH
        # =========================================================================

        print("\nMatching blade serial numbers...")

        inserted = []
        changed = []
        already_matching = []
        source_not_in_resource = []
        matched_keys = set()

        for key, source in source_serials.items():
            row_num = resource_index.get(key)

            if row_num is None:
                source_not_in_resource.append(source)
                continue

            matched_keys.add(key)

            serial_cell = resource_ws.cell(
                row=row_num,
                column=res_serial_col
            )

            old_serial = normalize_serial(
                serial_cell.value
            )
            new_serial = source["serial"]

            if old_serial == new_serial:
                already_matching.append({
                    "row": row_num,
                    "enclosure": key[0],
                    "slot": key[1],
                    "serial": new_serial,
                })
                continue

            if old_serial:
                changed.append({
                    "row": row_num,
                    "enclosure": key[0],
                    "slot": key[1],
                    "old_serial": old_serial,
                    "new_serial": new_serial,
                })
            else:
                inserted.append({
                    "row": row_num,
                    "enclosure": key[0],
                    "slot": key[1],
                    "serial": new_serial,
                })

            serial_cell.value = new_serial

        # =========================================================================
        # TARGET-ONLY ROWS
        # =========================================================================

        target_only = []
        target_only_with_serial = []
        target_only_blank_serial = []

        for key, row_num in resource_index.items():
            if key in matched_keys:
                continue

            old_serial = normalize_serial(
                resource_ws.cell(
                    row=row_num,
                    column=res_serial_col
                ).value
            )

            item = {
                "row": row_num,
                "enclosure": key[0],
                "slot": key[1],
                "serial": old_serial,
            }

            target_only.append(item)

            if old_serial:
                target_only_with_serial.append(item)
            else:
                target_only_blank_serial.append(item)

        total_updates = len(inserted) + len(changed)

        print(
            "Matched source blades          : {} / {}".format(
                len(matched_keys),
                len(source_serials)
            )
        )
        print(
            "Matching complete               : {} update(s) required".format(
                total_updates
            )
        )

        # =========================================================================
        # SAVE
        # =========================================================================

        saved = False

        if total_updates or not SAVE_ONLY_IF_CHANGED:
            print("Saving Resource List with openpyxl...")

            resource_wb.save(RESOURCE_FILE)
            saved = True

            print("Resource List saved successfully.")
        else:
            print(
                "No serial number changes detected. "
                "Resource List was not rewritten."
            )

    finally:
        resource_wb.close()

    # =========================================================================
    # REBUILD FORMULA CACHE USING NATIVE EXCEL
    # =========================================================================

    if saved and REBUILD_FORMULA_CACHE:
        refresh_excel_formulas(RESOURCE_FILE)

        cached_eligible = verify_formula_cache(
            RESOURCE_FILE,
            resource_end_row,
            res_equipment_col,
            res_enclosure_col,
            res_slot_col
        )

        if cached_eligible != eligible_resource_rows:
            raise RuntimeError(
                "Formula cache verification mismatch: "
                "expected {} eligible blades but read {} after "
                "Excel recalculation.".format(
                    eligible_resource_rows,
                    cached_eligible
                )
            )

    # =========================================================================
    # SUMMARY
    # =========================================================================

    print("\n" + "=" * 100)
    print("SUMMARY")
    print("=" * 100)
    print(
        "Eligible Resource List blades       : {}".format(
            eligible_resource_rows
        )
    )
    print(
        "Usable source enclosure/slot blades : {}".format(
            len(source_serials)
        )
    )
    print(
        "Source blades matched               : {} / {}".format(
            len(matched_keys),
            len(source_serials)
        )
    )
    print(
        "Existing serial already matched     : {}".format(
            len(already_matching)
        )
    )
    print(
        "Blank serial numbers populated      : {}".format(
            len(inserted)
        )
    )
    print(
        "Existing serial numbers changed     : {}".format(
            len(changed)
        )
    )
    print(
        "Source blades not in Resource       : {}".format(
            len(source_not_in_resource)
        )
    )
    print(
        "Target blades not present in source : {}".format(
            len(target_only)
        )
    )
    print(
        "  Target-only with existing serial  : {}".format(
            len(target_only_with_serial)
        )
    )
    print(
        "  Target-only with blank serial     : {}".format(
            len(target_only_blank_serial)
        )
    )
    print(
        "Invalid source serials skipped      : {}".format(
            len(invalid_serials)
        )
    )
    print(
        "Conflicting source blades           : {}".format(
            len(source_conflicts)
        )
    )
    print(
        "Duplicate Resource blade keys       : {}".format(
            len(duplicate_resource_keys)
        )
    )

    if PRINT_UPDATED_ROWS and changed:
        print("\n" + "=" * 100)
        print("SERIAL NUMBERS CHANGED")
        print("=" * 100)

        for item in changed:
            print(
                "{} | Slot {} | Resource row {}: {} -> {}".format(
                    item["enclosure"],
                    item["slot"],
                    item["row"],
                    item["old_serial"],
                    item["new_serial"]
                )
            )

    if PRINT_UPDATED_ROWS and inserted:
        print("\n" + "=" * 100)
        print("BLANK SERIAL NUMBERS POPULATED")
        print("=" * 100)

        for item in inserted:
            print(
                "{} | Slot {} | Resource row {}: {}".format(
                    item["enclosure"],
                    item["slot"],
                    item["row"],
                    item["serial"]
                )
            )

    if source_not_in_resource:
        print("\n" + "=" * 100)
        print("SOURCE BLADES NOT FOUND IN RESOURCE LIST")
        print("=" * 100)

        for item in source_not_in_resource:
            print(
                "{} | Slot {} | Source row {} | Serial {}".format(
                    item["enclosure"],
                    item["slot"],
                    item["row"],
                    item["serial"]
                )
            )

    if target_only:
        print("\n" + "=" * 100)
        print("RESOURCE LIST BLADES NOT PRESENT IN SOURCE INVENTORY")
        print("=" * 100)

        for item in target_only:
            serial = item["serial"] or "<BLANK>"

            print(
                "{} | Slot {} | Resource row {} | Serial {}".format(
                    item["enclosure"],
                    item["slot"],
                    item["row"],
                    serial
                )
            )

    if source_conflicts:
        print("\n" + "=" * 100)
        print("WARNING: CONFLICTING SERIAL NUMBERS IN SOURCE")
        print("=" * 100)

        for conflict in source_conflicts.values():
            values = " | ".join(
                "Row {} = {}".format(row, serial)
                for row, serial in zip(
                    conflict["rows"],
                    conflict["serials"]
                )
            )

            print(
                "{} | Slot {} | {}".format(
                    conflict["enclosure"],
                    conflict["slot"],
                    values
                )
            )

    print("\n" + "=" * 100)

    if total_updates:
        print(
            "COMPLETED: {} serial number(s) updated.".format(
                total_updates
            )
        )
    else:
        print(
            "COMPLETED: No serial numbers required updating."
        )

    if saved and REBUILD_FORMULA_CACHE:
        print(
            "FORMULA CACHE: Native Excel recalculation completed "
            "and verified."
        )

    print("=" * 100)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
    except Exception as exc:
        print("\n" + "=" * 100)
        print("ERROR: {}".format(exc))
        print("=" * 100)
        raise