import os
import re
import warnings
import time
import gc

# ==========================================
# Suppress Deprecation Warnings
# ==========================================
warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", message=".*Python 3.6 is no longer supported.*")
warnings.filterwarnings("ignore", module="cryptography")
warnings.filterwarnings("ignore", module="paramiko")

import paramiko
from openpyxl import load_workbook
import win32com.client as win32

# ==========================================
# Global Configuration
# ==========================================
USERNAME = "Administrator"
PASSWORD = "Siljeddah15"
# ==========================================

class OAConnection:
    """Robust Interactive SSH connection wrapper for HP Onboard Administrator."""
    def __init__(self, primary_ip, secondary_ip, user, pwd):
        self.ips = [ip for ip in (primary_ip, secondary_ip) if ip]
        self.user = user
        self.pwd = pwd
        self.ssh = None
        self.channel = None
        self.active_ip = None

    def connect(self):
        for ip in self.ips:
            print(f"\nAttempting connection to OA at {ip}...")
            self.ssh = paramiko.SSHClient()
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            try:
                self.ssh.connect(
                    hostname=ip, username=self.user, password=self.pwd, 
                    look_for_keys=False, allow_agent=False, timeout=10,
                    banner_timeout=15, auth_timeout=15
                )
                
                transport = self.ssh.get_transport()
                if transport:
                    transport.set_keepalive(5)
                
                # Switch to interactive shell for robust large-buffer handling
                self.channel = self.ssh.invoke_shell()
                time.sleep(1)
                self._drain_channel()

                # Check if Active
                status_out = self.send_command("SHOW OA STATUS")
                if re.search(r"Role:\s*Active", status_out, flags=re.IGNORECASE):
                    print(f"Successfully connected to ACTIVE OA ({ip}).")
                    self.active_ip = ip
                    return True
                else:
                    print(f"OA at {ip} is standby/inactive. Closing and trying next...")
                    self.close()
                    
            except Exception as e:
                print(f"Failed to connect to {ip}: {e}")
                self.close()
                
        return False

    def _drain_channel(self):
        """Clears the buffer of any leftover output."""
        output = ""
        while self.channel and self.channel.recv_ready():
            data = self.channel.recv(65535)
            if not data: break
            output += data.decode("utf-8", errors="replace")
        return output

    def send_command(self, cmd, wait_string=">"):
        """Sends a command and waits for the prompt to guarantee full output capture."""
        if not self.channel: return ""
        self._drain_channel()
        self.channel.send(cmd + "\n")
        
        output = ""
        deadline = time.time() + 60  # 60 second timeout for large outputs
        while time.time() < deadline:
            if self.channel.recv_ready():
                data = self.channel.recv(65535)
                if not data: break
                
                text = data.decode("utf-8", errors="replace")
                output += text
                
                # Handle OA Pagination ("Press any key to continue" or similar)
                if "any key" in text.lower() or "more" in text.lower():
                    self.channel.send(" ")
                
                if wait_string in output:
                    time.sleep(0.5) # Let trailing bytes settle
                    output += self._drain_channel()
                    break
            else:
                time.sleep(0.1)
        return output

    def close(self):
        if self.channel:
            try: self.channel.close()
            except Exception: pass
        if self.ssh:
            try: self.ssh.close()
            except Exception: pass

def check_file_writable(file_path):
    if not os.path.exists(file_path): return True
    try:
        with open(file_path, "a+"): pass
        return True
    except PermissionError:
        return False

def get_all_enclosure_details(oa_conn):
    """Fetches all blade details via Batch CLI and parses the output."""
    results = {}
    
    print("Fetching global serial numbers for all blades...")
    info_raw = oa_conn.send_command("SHOW SERVER INFO ALL")
    
    print("Fetching global port maps for all blades...")
    port_raw = oa_conn.send_command("SHOW SERVER PORT MAP ALL")

    # 1. Parse Serial Numbers
    current_bay = None
    # More permissive regex to catch variations like "Server Blade #1" or "Blade 1"
    bay_regex = re.compile(r'(?:Server\s+)?Blade\s*(?:#|Number)?\s*(\d+)', re.IGNORECASE)
    
    for line in info_raw.splitlines():
        bay_match = bay_regex.search(line)
        if bay_match:
            current_bay = int(bay_match.group(1))
            if current_bay not in results:
                results[current_bay] = {'serial': 'Serial Number not found', 'macs': [""] * 6}
                
        if current_bay:
            sn_match = re.search(r'Serial Number:\s*([A-Za-z0-9\-]+)', line, re.IGNORECASE)
            if sn_match:
                results[current_bay]['serial'] = sn_match.group(1).strip()

    # 2. Parse Port Maps (MAC Addresses)
    current_bay = None
    current_section = "other"
    flb_macs_dict = {}
    other_macs_dict = {}
    mac_pattern = re.compile(r'(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}')

    for line in port_raw.splitlines():
        bay_match = bay_regex.search(line)
        if bay_match:
            current_bay = int(bay_match.group(1))
            if current_bay not in flb_macs_dict:
                flb_macs_dict[current_bay] = []
                other_macs_dict[current_bay] = []
            current_section = "other"
            continue
            
        if current_bay:
            upper_line = line.upper()
            if 'LOM' in upper_line or 'FLB' in upper_line:
                current_section = "flb"
            elif 'MEZZ' in upper_line:
                current_section = "other"
                
            found_macs = mac_pattern.findall(line)
            for mac in found_macs:
                clean_mac = mac.replace('-', ':').upper()
                if current_section == "flb":
                    if clean_mac not in flb_macs_dict[current_bay]:
                        flb_macs_dict[current_bay].append(clean_mac)
                else:
                    if clean_mac not in other_macs_dict[current_bay]:
                        other_macs_dict[current_bay].append(clean_mac)

    # 3. Combine MACs for each bay found
    for bay, flb_list in flb_macs_dict.items():
        if bay not in results:
            results[bay] = {'serial': 'Serial Number not found', 'macs': [""] * 6}
            
        other_list = other_macs_dict.get(bay, [])
        combined = flb_list + [m for m in other_list if m not in flb_list]
        for i in range(min(len(combined), 6)):
            results[bay]['macs'][i] = combined[i]

    return results

def refresh_excel_formulas(file_path):
    """Uses the native Excel application in the background to rebuild formula caches safely."""
    print("\nCommanding Excel to recalculate formulas in the background...")
    abs_path = os.path.abspath(file_path)
    excel = None
    wb = None
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_path)
        wb.Save()
        print("Formula cache successfully rebuilt!")
    except Exception as e:
        print(f"WARNING: Could not trigger Excel to rebuild formulas: {e}")
    finally:
        if wb:
            try: wb.Close(SaveChanges=False)
            except Exception: pass
        if excel:
            try: excel.Quit()
            except Exception: pass
        del wb
        del excel
        gc.collect()

def scan_excel_for_blades(excel_path, sheet_name, enclosure_name, empty_row_stop=25):
    print(f"\nScanning Excel file (Fast Read-Only Mode)...")
    wb = load_workbook(excel_path, data_only=True, read_only=True)
    
    if sheet_name not in wb.sheetnames:
        return None, None, f"Error: Sheet '{sheet_name}' not found."
        
    ws = wb[sheet_name]
    
    headers = {}
    for idx, cell_value in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)), 1):
        if cell_value: headers[str(cell_value).strip()] = idx

    required_cols = [
        'equipment_physical_name', 'enclosure_physical_name', 'equipment_type', 
        'enclosure_slot', 'ilo_ip', 'serial_no', 
        'nic0_mac', 'nic1_mac', 'nic2_mac', 'nic3_mac', 'nic4_mac', 'nic5_mac'
    ]
    missing_cols = [col for col in required_cols if col not in headers]
    if missing_cols:
        return None, None, f"Error: Required columns missing: {', '.join(missing_cols)}"

    blade_rows = []
    primary_oa_ip, secondary_oa_ip = None, None
    empty_count = 0

    for row_idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row_values):
            empty_count += 1
            if empty_count >= empty_row_stop: break
            continue
        else:
            empty_count = 0 

        enc_val = row_values[headers['enclosure_physical_name'] - 1]
        eq_phys_val = row_values[headers['equipment_physical_name'] - 1]
        eq_type_val = row_values[headers['equipment_type'] - 1]
        slot_val = row_values[headers['enclosure_slot'] - 1]
        ip_val = row_values[headers['ilo_ip'] - 1]

        enc_str = str(enc_val).strip().upper() if enc_val else ""
        eq_phys_str = str(eq_phys_val).strip().upper() if eq_phys_val else ""
        
        is_match = False
        if enc_str == enclosure_name: is_match = True
        elif len(enclosure_name) >= 14 and eq_phys_str.startswith(enclosure_name[:14]): is_match = True

        if is_match:
            eq_type = str(eq_type_val or "").strip().upper()
            try: slot_num = int(float(str(slot_val).strip()))
            except (ValueError, TypeError): slot_num = None

            if "ENCLOSURE OA" in eq_type and ip_val:
                if slot_num == 1: primary_oa_ip = str(ip_val).strip()
                elif slot_num == 2: secondary_oa_ip = str(ip_val).strip()
            elif "BLADE SERVER" in eq_type:
                blade_rows.append({'row_idx': row_idx, 'slot': slot_num, 'eq_type': eq_type})
                
    wb.close()
    return (primary_oa_ip, secondary_oa_ip, blade_rows, headers), None

def main():
    enclosure_name = input("Enter the Enclosure Name: ").strip().upper()
    
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(script_dir, 'Templates', 'Resource List-v7.6.xlsx')
    sheet_name = "General Resource List"

    if not check_file_writable(excel_path):
        print(f"\nERROR: '{excel_path}' is currently open in Excel.")
        print("Please close the file and run the script again to prevent permission errors.")
        return

    result, error = scan_excel_for_blades(excel_path, sheet_name, enclosure_name, empty_row_stop=25)
    if error:
        print(error)
        return
        
    primary_oa_ip, secondary_oa_ip, blade_rows, headers = result

    if not primary_oa_ip and not secondary_oa_ip:
        print(f"Error: Could not find OA IPs for Enclosure '{enclosure_name}'.")
        return
        
    if not blade_rows:
        print(f"No Blade Servers found for Enclosure '{enclosure_name}'.")
        return

    print(f"Found Primary OA: {primary_oa_ip} | Secondary OA: {secondary_oa_ip}")
    print(f"Found {len(blade_rows)} Blade Server(s) in this enclosure.")

    # Instantiate the new interactive connection
    oa = OAConnection(primary_oa_ip, secondary_oa_ip, USERNAME, PASSWORD)
    if not oa.connect():
        return

    # Fetch ALL blade details via batch PTY session
    hardware_data = {}
    try:
        hardware_data = get_all_enclosure_details(oa)
    finally:
        oa.close()
        print("SSH connection closed.")

    print(f"\nLoading Excel file (Writing Mode) to inject data...")
    # Loading without data_only=True preserves Excel formulas during the write operation
    wb_write = load_workbook(excel_path)
    ws_write = wb_write[sheet_name]

    for blade in blade_rows:
        row_idx = blade['row_idx']
        bay_num = blade['slot']
        eq_type = blade['eq_type']
        
        print(f"\n--- Processing Row {row_idx}: {eq_type} (Bay {bay_num}) ---")
        
        if bay_num is None:
            print(f"WARNING: Skipping blade at row {row_idx}. Missing or invalid enclosure_slot.")
            continue
            
        bay_data = hardware_data.get(bay_num, {'serial': 'Data not found', 'macs': [""] * 6})
        serial_no = bay_data['serial']
        macs = bay_data['macs']

        ws_write.cell(row=row_idx, column=headers['serial_no']).value = serial_no
        ws_write.cell(row=row_idx, column=headers['nic0_mac']).value = macs[0]
        ws_write.cell(row=row_idx, column=headers['nic1_mac']).value = macs[1]
        ws_write.cell(row=row_idx, column=headers['nic2_mac']).value = macs[2]
        ws_write.cell(row=row_idx, column=headers['nic3_mac']).value = macs[3]
        ws_write.cell(row=row_idx, column=headers['nic4_mac']).value = macs[4]
        ws_write.cell(row=row_idx, column=headers['nic5_mac']).value = macs[5]

        print(f"  Serial Number : {serial_no}")
        for i, mac in enumerate(macs, 1):
            if not mac:
                print(f"  mac{i}          : [MISSING]")
            else:
                print(f"  mac{i}          : {mac}")

    print(f"\nSaving updates to {excel_path}...")
    wb_write.save(excel_path)
    print("Excel file successfully updated!")
    
    # Run the COM logic to rebuild all formula caches natively in Excel
    refresh_excel_formulas(excel_path)

if __name__ == "__main__":
    main()