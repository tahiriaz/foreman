#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# BUILD_MARKER: RM_CONDITIONAL_ILO_RESET_V2_20260828

import os
import sys
import time
import json
import subprocess
import urllib3
import requests
import openpyxl
import pandas as pd
import concurrent.futures

from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from functions import vars


# Disable insecure request warnings for self-signed iLO certificates
urllib3.disable_warnings(
    urllib3.exceptions.InsecureRequestWarning
)


# =============================================================================
# CENTRALIZED CONFIGURATION
# =============================================================================

# Credentials, Excel settings, scope settings, concurrency, licensing, paths,
# timeouts and other reusable settings are maintained in functions/vars.py.

ILO_LOGIN = vars.ILO_RM_USERNAME
ILO_PASSWORD = vars.ILO_RM_PASSWORD

ILO_ADVANCED_LICENSE_KEY = vars.ILO_ADVANCED_LICENSE_KEY

FENCE_USER_NAME = vars.FENCE_USER_NAME
FENCE_USER_LOGIN = vars.FENCE_USER_LOGIN
FENCE_USER_PASSWORD = vars.FENCE_USER_PASSWORD

IPMI_PORT = vars.ILO_IPMI_PORT
ILO_DOMAIN = vars.ILO_DOMAIN

DIR_USER_CONTEXT_1 = vars.LDAP_USER_CONTEXTS[0]
DIR_USER_CONTEXT_2 = vars.LDAP_USER_CONTEXTS[1]
DIR_USER_CONTEXT_3 = vars.LDAP_USER_CONTEXTS[2]
DIR_USER_CONTEXT_4 = vars.LDAP_USER_CONTEXTS[3]
AD_GROUP_DN = vars.LDAP_GROUP_NAME

EQUIPMENT_TYPES = list(vars.ILO_RM_EQUIPMENT_TYPES)
START_ROW = vars.START_ROW
END_ROW = vars.END_ROW
MAX_WORKERS = vars.ILO_RM_MAX_WORKERS
SHEET_NAME = vars.SHEET_NAME

SCOPE_SETTINGS = vars.ILO_SCOPE_SETTINGS
TIMEZONE_MATCH = vars.ILO_TIMEZONE_SEARCH

DEBUG_MODE = vars.ILO_RM_DEBUG_MODE

BASE_DIR = Path(vars.PROJECT_DIR)
TEMPLATE_DIR = Path(vars.TEMPLATES_DIR)
EXCEL_FILE = Path(vars.RESOURCE_LIST)
CERT_FILE = Path(vars.ILO_LDAP_CERT_FILE)
LOG_DIR_NAME = os.path.basename(vars.LOG_DIR)
REPORT_PREFIX = vars.ILO_RM_REPORT_PREFIX

REDFISH_PORT = vars.ILO_REDFISH_PORT
REQUEST_TIMEOUT = vars.ILO_RM_REQUEST_TIMEOUT

ILO_RESET_INITIAL_WAIT_SECONDS = vars.ILO_RM_RESET_INITIAL_WAIT_SECONDS
ILO_RESET_RETRY_INTERVAL_SECONDS = vars.ILO_RM_RESET_RETRY_INTERVAL_SECONDS
ILO_RESET_MAX_WAIT_SECONDS = vars.ILO_RM_RESET_MAX_WAIT_SECONDS


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def log(server_name, ip, message):
    """
    Thread-safe formatted logger.
    """

    print(
        f"[{server_name} | {ip}] {message}",
        flush=True,
    )


def debug_log(server_name, ip, message):
    """
    Print additional diagnostics when DEBUG_MODE=True.
    """

    if DEBUG_MODE:

        log(
            server_name,
            ip,
            f"DEBUG: {message}",
        )


def is_reachable(ip, server_name):
    """
    Check basic network reachability before opening Redfish session.
    """

    log(
        server_name,
        ip,
        "Pinging...",
    )

    if os.name == "nt":

        command = [
            "ping",
            "-n",
            "2",
            "-w",
            "500",
            str(ip),
        ]

    else:

        command = [
            "ping",
            "-c",
            "2",
            "-W",
            "1",
            str(ip),
        ]

    try:

        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

        return result.returncode == 0

    except Exception:

        return False


def response_is_success(response):
    """
    Determine whether an iLO Redfish response should be considered
    successful.

    HPE can return ExtendedInfo messages such as ResetRequired together
    with an otherwise successful operation.
    """

    if response is None:

        return False

    if (
        response.status_code < 200
        or response.status_code >= 300
    ):

        return False

    try:

        data = response.json()

    except Exception:

        # Successful HTTP response with no JSON body
        return True

    error = data.get("error")

    if not error:

        return True

    extended_info = error.get(
        "@Message.ExtendedInfo",
        [],
    )

    for item in extended_info:

        message_id = str(
            item.get(
                "MessageId",
                ""
            )
        )

        if (
            message_id.endswith(".Success")
            or "ResetRequired" in message_id
            or "ResetInProgress" in message_id
            or "SystemResetRequired" in message_id
        ):

            return True

    return False


def get_extended_error_text(response):
    """
    Return useful Redfish response information for logging.
    """

    if response is None:

        return "No response"

    try:

        return json.dumps(
            response.json(),
            indent=2,
        )

    except Exception:

        return (
            response.text.strip()
            if response.text
            else ""
        )


def check_reset_required(response):
    """
    Detect common HPE reset-required indicators.
    """

    if response is None:

        return False

    text = (
        response.text
        if response.text
        else ""
    )

    reset_tokens = [
        "ResetRequired",
        "ResetInProgress",
        "SystemResetRequired",
        "SomePendingReset",
        "PendingReset",
    ]

    return any(
        token in text
        for token in reset_tokens
    )


def clean_string_list(values):
    """
    Normalize a list of values for comparison.
    """

    if not values:
        return []

    return [
        str(value).strip()
        for value in values
        if value is not None
        and str(value).strip() != ""
    ]


def configuration_requires_reset(data):
    """
    Return True only when an iLO resource explicitly reports pending/reset state.
    """
    if not isinstance(data, dict):
        return False

    values = []
    for key in (
        "ConfigurationSettings",
        "SettingsResult",
        "@Redfish.Settings",
    ):
        if key in data:
            values.append(data.get(key))

    try:
        searchable = json.dumps(values).lower()
    except Exception:
        searchable = str(values).lower()

    reset_tokens = (
        "resetrequired",
        "somependingreset",
        "pendingreset",
        "pending reset",
        "systemresetrequired",
    )

    return any(token in searchable for token in reset_tokens)

# =============================================================================
# EXCEL LOADER
# =============================================================================

def load_resource_excel_fast(
    filepath,
    sheet_name,
    start_row,
    end_row,
):
    """
    Read only the requested Excel row range using read-only streaming mode.
    """

    wb = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=True,
    )

    try:

        if sheet_name not in wb.sheetnames:

            raise ValueError(
                f"Sheet '{sheet_name}' not found in workbook."
            )

        ws = wb[sheet_name]

        header_row = next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            )
        )

        headers = [

            (
                str(cell).strip().lower()
                if cell
                else f"unnamed_{index}"
            )

            for index, cell
            in enumerate(header_row)
        ]

        data = []

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            values_only=True,
        ):

            populated = any(
                cell is not None
                and str(cell).strip() != ""
                for cell in row
            )

            if not populated:
                continue

            data.append(row)

        return pd.DataFrame(
            data,
            columns=headers,
        )

    finally:

        wb.close()


# =============================================================================
# REDFISH CLIENT
# =============================================================================

class RedfishClient:

    def __init__(
        self,
        ip,
        username,
        password,
    ):

        self.ip = ip

        self.username = username
        self.password = password

        self.base_url = (
            f"https://{ip}:"
            f"{REDFISH_PORT}/redfish/v1"
        )

        self.session = requests.Session()

        self.session.verify = False

        self.session.headers.update(
            {
                "Accept": "application/json",
                "Content-Type": "application/json",
                "OData-Version": "4.0",
                "Connection": "close",
            }
        )

        # Automatically retry only safe/read operations.
        #
        # Do NOT automatically repeat configuration PATCH/POST operations.
        retries = Retry(
            total=3,
            connect=3,
            read=3,
            status=3,
            backoff_factor=2,
            status_forcelist=[
                500,
                502,
                503,
                504,
            ],
            allowed_methods=frozenset(
                [
                    "GET",
                    "HEAD",
                    "OPTIONS",
                ]
            ),
            raise_on_status=False,
        )

        adapter = HTTPAdapter(
            max_retries=retries,
        )

        self.session.mount(
            "https://",
            adapter,
        )

        self.token = None

        self.manager_uri = None
        self.system_uri = None


    # =========================================================================
    # AUTHENTICATION
    # =========================================================================

    def login(self):
        """
        Authenticate using Redfish SessionService.

        Fall back to HTTP Basic authentication if SessionService
        authentication fails.
        """

        payload = {
            "UserName": self.username,
            "Password": self.password,
        }

        url = (
            f"{self.base_url}/"
            "SessionService/Sessions"
        )

        try:

            response = self.session.post(
                url,
                json=payload,
                timeout=REQUEST_TIMEOUT,
            )

            if response.status_code in [
                200,
                201,
            ]:

                self.token = (
                    response.headers.get(
                        "X-Auth-Token"
                    )
                )

                if self.token:

                    self.session.headers.update(
                        {
                            "X-Auth-Token":
                                self.token
                        }
                    )

                    self.session.auth = None

                return True

            # -----------------------------------------------------------------
            # Basic Authentication fallback
            # -----------------------------------------------------------------

            self.session.auth = (
                self.username,
                self.password,
            )

            response = self.session.get(
                f"{self.base_url}/",
                timeout=REQUEST_TIMEOUT,
            )

            return (
                response.status_code == 200
            )

        except requests.exceptions.RequestException:

            return False


    def logout(self):
        """
        Close HTTP connection pool.
        """

        try:

            self.session.close()

        except Exception:

            pass


    # =========================================================================
    # HTTP WRAPPER
    # =========================================================================

    def _request(
        self,
        method,
        endpoint,
        payload=None,
    ):

        if endpoint.startswith(
            "/redfish/v1"
        ):

            url = (
                f"https://{self.ip}:"
                f"{REDFISH_PORT}"
                f"{endpoint}"
            )

        else:

            url = (
                f"{self.base_url}/"
                f"{endpoint.lstrip('/')}"
            )

        kwargs = {
            "timeout": REQUEST_TIMEOUT
        }

        if payload is not None:

            kwargs["json"] = payload

        return self.session.request(
            method,
            url,
            **kwargs,
        )


    def get(self, endpoint):

        return self._request(
            "GET",
            endpoint,
        )


    def post(
        self,
        endpoint,
        payload,
    ):

        return self._request(
            "POST",
            endpoint,
            payload=payload,
        )


    def patch(
        self,
        endpoint,
        payload,
    ):

        return self._request(
            "PATCH",
            endpoint,
            payload=payload,
        )


    # =========================================================================
    # RESOURCE DISCOVERY
    # =========================================================================

    def discover_resources(self):
        """
        Dynamically discover Manager and ComputerSystem resources.
        """

        root_response = self.get("/")

        root_response.raise_for_status()

        data = root_response.json()

        # ---------------------------------------------------------------------
        # Manager
        # ---------------------------------------------------------------------

        managers_uri = (
            data.get(
                "Managers",
                {}
            )
            .get(
                "@odata.id"
            )
        )

        if not managers_uri:

            raise Exception(
                "Managers resource not found in Redfish root."
            )

        managers_response = self.get(
            managers_uri
        )

        managers_response.raise_for_status()

        managers = (
            managers_response
            .json()
            .get(
                "Members",
                []
            )
        )

        if not managers:

            raise Exception(
                "No Redfish Manager resource found."
            )

        self.manager_uri = (
            managers[0]
            .get(
                "@odata.id"
            )
        )

        if not self.manager_uri:

            raise Exception(
                "Manager @odata.id is missing."
            )

        # ---------------------------------------------------------------------
        # ComputerSystem
        # ---------------------------------------------------------------------

        systems_uri = (
            data.get(
                "Systems",
                {}
            )
            .get(
                "@odata.id"
            )
        )

        if not systems_uri:

            raise Exception(
                "Systems resource not found in Redfish root."
            )

        systems_response = self.get(
            systems_uri
        )

        systems_response.raise_for_status()

        systems = (
            systems_response
            .json()
            .get(
                "Members",
                []
            )
        )

        if not systems:

            raise Exception(
                "No Redfish ComputerSystem resource found."
            )

        self.system_uri = (
            systems[0]
            .get(
                "@odata.id"
            )
        )

        if not self.system_uri:

            raise Exception(
                "ComputerSystem @odata.id is missing."
            )


# =============================================================================
# iLO RESET / RECONNECT
# =============================================================================

def reset_ilo_and_reconnect(
    client,
    log_prefix,
):
    """
    Reset iLO and establish a fresh Redfish session when it becomes available.

    Timing:

        ILO_RESET_INITIAL_WAIT_SECONDS
            Fixed wait immediately following Manager.Reset.

        ILO_RESET_RETRY_INTERVAL_SECONDS
            Delay between reconnect attempts.

        ILO_RESET_MAX_WAIT_SECONDS
            Maximum reconnect period AFTER initial wait.
    """

    server_name, ip = log_prefix

    if not client.manager_uri:

        raise Exception(
            "Manager URI is not available for iLO reset."
        )

    reset_uri = (
        client.manager_uri.rstrip("/")
        + "/Actions/Manager.Reset"
    )

    log(
        server_name,
        ip,
        "Triggering iLO reset...",
    )

    try:

        response = client.post(
            reset_uri,
            {
                "ResetType":
                    "ForceRestart"
            },
        )

        if (
            response is not None
            and response.status_code
            not in [
                200,
                202,
                204,
            ]
        ):

            raise Exception(
                "iLO reset request failed. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            )

    except requests.exceptions.ConnectionError:

        # Some iLO releases terminate HTTPS immediately after accepting
        # the management processor reset.
        pass

    except requests.exceptions.ReadTimeout:

        # The reset may interrupt the HTTP response.
        pass

    try:

        client.logout()

    except Exception:

        pass

    # -------------------------------------------------------------------------
    # Initial wait
    # -------------------------------------------------------------------------

    log(
        server_name,
        ip,
        (
            f"Waiting {ILO_RESET_INITIAL_WAIT_SECONDS}s "
            "before checking iLO..."
        ),
    )

    time.sleep(
        ILO_RESET_INITIAL_WAIT_SECONDS
    )

    reconnect_start = (
        time.time()
    )

    attempt = 0

    # -------------------------------------------------------------------------
    # Reconnect loop
    # -------------------------------------------------------------------------

    while (
        time.time()
        - reconnect_start
        < ILO_RESET_MAX_WAIT_SECONDS
    ):

        attempt += 1

        elapsed = int(
            time.time()
            - reconnect_start
        )

        new_client = RedfishClient(
            ip=ip,
            username=ILO_LOGIN,
            password=ILO_PASSWORD,
        )

        try:

            if new_client.login():

                new_client.discover_resources()

                total_elapsed = (
                    ILO_RESET_INITIAL_WAIT_SECONDS
                    + elapsed
                )

                log(
                    server_name,
                    ip,
                    (
                        "iLO is online and Redfish "
                        "authentication succeeded "
                        f"after approximately {total_elapsed}s."
                    ),
                )

                return new_client

        except Exception as exc:

            debug_log(
                server_name,
                ip,
                (
                    f"Reconnect attempt {attempt} "
                    f"failed: {exc}"
                ),
            )

        try:

            new_client.logout()

        except Exception:

            pass

        elapsed = int(
            time.time()
            - reconnect_start
        )

        remaining = max(
            0,
            ILO_RESET_MAX_WAIT_SECONDS
            - elapsed,
        )

        if remaining <= 0:
            break

        log(
            server_name,
            ip,
            (
                "iLO not ready yet. "
                f"Retrying in "
                f"{ILO_RESET_RETRY_INTERVAL_SECONDS}s "
                f"(attempt {attempt}, "
                f"remaining timeout ~{remaining}s)..."
            ),
        )

        sleep_time = min(
            ILO_RESET_RETRY_INTERVAL_SECONDS,
            remaining,
        )

        time.sleep(
            sleep_time
        )

    raise Exception(
        "iLO did not become available within "
        f"{ILO_RESET_MAX_WAIT_SECONDS} seconds "
        "after the initial "
        f"{ILO_RESET_INITIAL_WAIT_SECONDS}-second wait."
    )


# =============================================================================
# LICENSE
# =============================================================================

def apply_license(
    client,
    log_prefix,
):

    endpoint = (
        client.manager_uri.rstrip("/")
        + "/LicenseService/"
    )

    response = client.get(
        endpoint
    )

    if response.status_code == 200:

        for member in (
            response.json()
            .get(
                "Members",
                []
            )
        ):

            member_uri = (
                member.get(
                    "@odata.id"
                )
            )

            if not member_uri:
                continue

            lic_res = client.get(
                member_uri
            )

            if (
                lic_res.status_code == 200
                and "Advanced"
                in str(
                    lic_res.json()
                    .get(
                        "License",
                        ""
                    )
                )
            ):

                log(
                    log_prefix[0],
                    log_prefix[1],
                    "iLO Advanced license already installed.",
                )

                return "SKIP", False

    log(
        log_prefix[0],
        log_prefix[1],
        "Applying iLO Advanced license...",
    )

    response = client.post(
        endpoint,
        {
            "LicenseKey":
                ILO_ADVANCED_LICENSE_KEY
        },
    )

    if response_is_success(
        response
    ):

        time.sleep(5)

        return (
            "OK",
            check_reset_required(response),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "License installation failed: "
            f"{get_extended_error_text(response)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# FENCING USER
# =============================================================================

def apply_fencing_user(
    client,
    log_prefix,
):

    response = client.get(
        "/AccountService/Accounts/"
    )

    if response.status_code == 200:

        for member in (
            response.json()
            .get(
                "Members",
                []
            )
        ):

            member_uri = (
                member.get(
                    "@odata.id"
                )
            )

            if not member_uri:
                continue

            acc_res = client.get(
                member_uri
            )

            if (
                acc_res.status_code == 200
                and acc_res.json()
                .get(
                    "UserName"
                )
                == FENCE_USER_LOGIN
            ):

                log(
                    log_prefix[0],
                    log_prefix[1],
                    (
                        f"Fencing user '{FENCE_USER_LOGIN}' "
                        "already exists."
                    ),
                )

                return "SKIP", False

    log(
        log_prefix[0],
        log_prefix[1],
        "Creating fencing user...",
    )

    payload = {

        "UserName":
            FENCE_USER_LOGIN,

        "Password":
            FENCE_USER_PASSWORD,

        "Oem": {
            "Hpe": {

                "LoginName":
                    FENCE_USER_LOGIN,

                "Privileges": {

                    "LoginPriv":
                        True,

                    "VirtualPowerAndResetPriv":
                        True,

                    "RemoteConsolePriv":
                        False,

                    "VirtualMediaPriv":
                        False,

                    "UserConfigPriv":
                        False,

                    "iLOConfigPriv":
                        False,
                },
            }
        },
    }

    response = client.post(
        "/AccountService/Accounts/",
        payload,
    )

    if response_is_success(
        response
    ):

        return (
            "OK",
            check_reset_required(response),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "Fencing user creation failed: "
            f"{get_extended_error_text(response)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# LDAP
# =============================================================================

def configure_ldap(
    client,
    scope_data,
    log_prefix,
):

    log(
        log_prefix[0],
        log_prefix[1],
        "Configuring Generic LDAP...",
    )

    ldap_payload = {

        "ActiveDirectory": {
            "ServiceEnabled":
                False
        },

        "LDAP": {

            "ServiceEnabled":
                True,

            "ServiceAddresses": [
                scope_data[
                    "DIRECTORY_SERVER"
                ]
            ],

            "Authentication": {
                "AuthenticationType":
                    "UsernameAndPassword"
            },

            "LDAPService": {

                "SearchSettings": {

                    "BaseDistinguishedNames": [
                        DIR_USER_CONTEXT_1,
                        DIR_USER_CONTEXT_2,
                        DIR_USER_CONTEXT_3,
                        DIR_USER_CONTEXT_4,
                    ]
                }
            },

            "RemoteRoleMapping": [
                {
                    "LocalRole":
                        "Administrator",

                    "RemoteGroup":
                        AD_GROUP_DN,
                }
            ],
        },

        "Oem": {

            "Hpe": {

                "DirectorySettings": {

                    "LdapAuthenticationMode":
                        "DefaultSchema"
                }
            }
        },
    }

    response = client.patch(
        "/AccountService/",
        ldap_payload,
    )

    if response_is_success(
        response
    ):

        return (
            "OK",
            check_reset_required(response),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "LDAP configuration failed: "
            f"{get_extended_error_text(response)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# LDAP CERTIFICATE
# =============================================================================

def configure_ldap_certificate(
    client,
    log_prefix,
):

    if not CERT_FILE.is_file():

        log(
            log_prefix[0],
            log_prefix[1],
            (
                "Certificate file missing: "
                f"{CERT_FILE}"
            ),
        )

        return "SKIP", False

    with open(
        CERT_FILE,
        "r",
        encoding="utf-8",
    ) as cert_handle:

        certificate = (
            cert_handle
            .read()
            .strip()
        )

    # -------------------------------------------------------------------------
    # Standard Redfish method
    # -------------------------------------------------------------------------

    cert_res = client.post(
        (
            "/AccountService/"
            "ExternalAccountProviders/"
            "LDAP/Certificates/"
        ),
        {
            "CertificateString":
                certificate,

            "CertificateType":
                "PEM",
        },
    )

    if response_is_success(
        cert_res
    ):

        log(
            log_prefix[0],
            log_prefix[1],
            "LDAP CA certificate imported (Standard).",
        )

        return (
            "OK",
            check_reset_required(cert_res),
        )

    # -------------------------------------------------------------------------
    # HPE OEM fallback
    # -------------------------------------------------------------------------

    cert_res_oem = client.post(

        client.manager_uri.rstrip("/")
        + (
            "/SecurityService/"
            "Actions/Oem/Hpe/"
            "HpeiLOSecurityService."
            "ImportDirectoryCACertificate"
        ),

        {
            "Certificate":
                certificate
        },
    )

    if response_is_success(
        cert_res_oem
    ):

        log(
            log_prefix[0],
            log_prefix[1],
            "LDAP CA certificate imported (OEM).",
        )

        return (
            "OK",
            check_reset_required(cert_res_oem),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "LDAP certificate import failed. "
            "Standard response: "
            f"{get_extended_error_text(cert_res)}; "
            "OEM response: "
            f"{get_extended_error_text(cert_res_oem)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# SERVER IDENTIFICATION
# =============================================================================

def configure_server_identification(
    client,
    server_name,
    server_fqdn,
    log_prefix,
):

    payload = {

        "HostName":
            server_name,

        "Oem": {

            "Hpe": {

                "ServerFQDN":
                    server_fqdn
            }
        },
    }

    response = client.patch(
        client.system_uri.rstrip("/")
        + "/",
        payload,
    )

    if response_is_success(
        response
    ):

        return (
            "OK",
            check_reset_required(response),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "Server identification configuration failed: "
            f"{get_extended_error_text(response)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# iLO NETWORK
# =============================================================================

def configure_ilo_network(
    client,
    server_fqdn,
    scope_data,
    log_prefix,
):
    """
    Configure iLO hostname, domain, DNS, and DHCP ownership settings.

    Current values are read first and only differences are PATCHed.

    An intermediate iLO reset is requested only when:
      - iLO explicitly reports that a network change requires reset; or
      - DHCP ownership of NTP was actually changed, because iLO 5 must apply
        that ownership change before StaticNTPServers can be modified.
    """
    server_name, ip = log_prefix
    ethernet_uri = (
        client.manager_uri.rstrip("/")
        + "/EthernetInterfaces/1/"
    )

    desired_hostname = server_fqdn.split(".")[0]
    desired_dns = [
        str(scope_data["PRIMARY_DNS"]).strip(),
        str(scope_data["SECONDARY_DNS"]).strip(),
    ]
    desired_domain = str(ILO_DOMAIN).strip()

    try:
        response = client.get(ethernet_uri)
    except Exception as exc:
        log(
            server_name,
            ip,
            "Unable to read current iLO network configuration: {}".format(exc),
        )
        return "FAIL", False

    if response.status_code != 200:
        log(
            server_name,
            ip,
            (
                "Unable to read current iLO network configuration. "
                "HTTP {}: {}"
            ).format(
                response.status_code,
                get_extended_error_text(response),
            ),
        )
        return "FAIL", False

    try:
        current = response.json()
    except Exception:
        log(
            server_name,
            ip,
            "iLO Ethernet resource returned invalid JSON.",
        )
        return "FAIL", False

    oem_hpe = (
        current.get("Oem", {})
        .get("Hpe", {})
        or {}
    )
    dhcp_v4 = oem_hpe.get("DHCPv4", {}) or {}
    dhcp_v6 = oem_hpe.get("DHCPv6", {}) or {}

    current_hostname = str(current.get("HostName", "") or "").strip()
    current_dns = clean_string_list(
        current.get("StaticNameServers", []) or []
    )
    current_domain = str(
        oem_hpe.get("DomainName", "") or ""
    ).strip()

    changes_made = False
    needs_reset = False

    dhcp_flags = (
        "UseDNSServers",
        "UseDomainName",
        "UseNTPServers",
    )

    dhcp_v4_needs_change = any(
        dhcp_v4.get(flag) is not False
        for flag in dhcp_flags
    )
    dhcp_v6_needs_change = any(
        dhcp_v6.get(flag) is not False
        for flag in dhcp_flags
    )

    ntp_ownership_changed = (
        dhcp_v4.get("UseNTPServers") is not False
        or dhcp_v6.get("UseNTPServers") is not False
    )

    if dhcp_v4_needs_change or dhcp_v6_needs_change:
        log(
            server_name,
            ip,
            "Disabling DHCP supplied DNS/Domain/NTP settings...",
        )

        dhcp_payload = {
            "Oem": {
                "Hpe": {
                    "DHCPv4": {
                        "UseDNSServers": False,
                        "UseDomainName": False,
                        "UseNTPServers": False,
                    },
                    "DHCPv6": {
                        "UseDNSServers": False,
                        "UseDomainName": False,
                        "UseNTPServers": False,
                    },
                }
            }
        }

        response = client.patch(
            ethernet_uri,
            dhcp_payload,
        )

        if not response_is_success(response):
            log(
                server_name,
                ip,
                (
                    "Failed to disable DHCP supplied DNS/Domain/NTP "
                    "settings. HTTP {}: {}"
                ).format(
                    response.status_code,
                    get_extended_error_text(response),
                ),
            )
            return "FAIL", False

        changes_made = True

        if check_reset_required(response):
            needs_reset = True

        if ntp_ownership_changed:
            needs_reset = True
            log(
                server_name,
                ip,
                (
                    "DHCP ownership of NTP changed. An intermediate iLO "
                    "reset is required before static SNTP configuration."
                ),
            )
    else:
        log(
            server_name,
            ip,
            "DHCP supplied DNS/Domain/NTP settings already disabled.",
        )

    hostname_matches = (
        current_hostname.lower()
        == desired_hostname.lower()
    )
    dns_matches = (
        [value.lower() for value in current_dns]
        == [value.lower() for value in desired_dns]
    )
    domain_matches = (
        current_domain.lower()
        == desired_domain.lower()
    )

    if hostname_matches and dns_matches and domain_matches:
        log(
            server_name,
            ip,
            (
                "iLO hostname, domain, and static DNS are already "
                "configured correctly."
            ),
        )
    else:
        network_payload = {}

        if not hostname_matches:
            network_payload["HostName"] = desired_hostname

        if not dns_matches:
            network_payload["StaticNameServers"] = desired_dns

        if not domain_matches:
            network_payload["Oem"] = {
                "Hpe": {
                    "DomainName": desired_domain
                }
            }

        log(
            server_name,
            ip,
            (
                "Updating iLO network settings: "
                "hostname_change={}, dns_change={}, domain_change={}"
            ).format(
                not hostname_matches,
                not dns_matches,
                not domain_matches,
            ),
        )

        response = client.patch(
            ethernet_uri,
            network_payload,
        )

        if not response_is_success(response):
            log(
                server_name,
                ip,
                (
                    "iLO network configuration failed. "
                    "HTTP {}: {}"
                ).format(
                    response.status_code,
                    get_extended_error_text(response),
                ),
            )
            return "FAIL", needs_reset

        changes_made = True

        if check_reset_required(response):
            needs_reset = True

    if not changes_made:
        return "SKIP", False

    return "OK", needs_reset

# =============================================================================
# TIMEZONE DISCOVERY
# =============================================================================

def find_matching_timezone(
    timezone_list,
    match_text,
):
    """
    Search TimeZoneList for any entry containing match_text.

    Searches all simple scalar fields instead of relying only on Name because
    firmware revisions can expose slightly different field names.

    Returns:
        matching timezone dict or None
    """

    search_text = (
        str(match_text)
        .strip()
        .lower()
    )

    if not search_text:

        return None

    for timezone in timezone_list:

        if not isinstance(
            timezone,
            dict,
        ):
            continue

        searchable_parts = []

        for key, value in timezone.items():

            if isinstance(
                value,
                (
                    str,
                    int,
                    float,
                ),
            ):

                searchable_parts.append(
                    str(value)
                )

        searchable = (
            " ".join(
                searchable_parts
            )
            .lower()
        )

        if search_text in searchable:

            return timezone

    return None


# =============================================================================
# TIMEZONE CONFIGURATION
# =============================================================================

def configure_timezone(
    client,
    log_prefix,
):
    """
    Configure iLO timezone by dynamically searching TimeZoneList for
    TIMEZONE_MATCH.

    Example:

        TIMEZONE_MATCH = "Riyadh"

    No timezone Index is hard-coded.
    """

    server_name, ip = log_prefix

    datetime_uri = (
        client.manager_uri.rstrip("/")
        + "/DateTime/"
    )

    log(
        server_name,
        ip,
        (
            "Reading iLO timezone configuration "
            f"and searching for '{TIMEZONE_MATCH}'..."
        ),
    )

    try:

        response = client.get(
            datetime_uri
        )

    except Exception as exc:

        log(
            server_name,
            ip,
            (
                "Unable to read DateTime resource "
                f"for timezone configuration: {exc}"
            ),
        )

        return "FAIL", False

    if response.status_code != 200:

        log(
            server_name,
            ip,
            (
                "Unable to read DateTime resource "
                "for timezone configuration. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            ),
        )

        return "FAIL", False

    try:

        data = response.json()

    except Exception:

        log(
            server_name,
            ip,
            "DateTime resource returned invalid JSON.",
        )

        return "FAIL", False

    timezone_list = (
        data.get(
            "TimeZoneList",
            []
        )
        or []
    )

    current_timezone = (
        data.get(
            "TimeZone",
            {}
        )
        or {}
    )

    if not timezone_list:

        log(
            server_name,
            ip,
            "iLO DateTime resource returned no TimeZoneList.",
        )

        return "FAIL", False

    # -------------------------------------------------------------------------
    # Dynamic Riyadh search
    # -------------------------------------------------------------------------

    target_timezone = (
        find_matching_timezone(
            timezone_list,
            TIMEZONE_MATCH,
        )
    )

    if target_timezone is None:

        log(
            server_name,
            ip,
            (
                "No supported timezone containing "
                f"'{TIMEZONE_MATCH}' was found."
            ),
        )

        if DEBUG_MODE:

            for timezone in timezone_list:

                debug_log(
                    server_name,
                    ip,
                    (
                        "Timezone candidate: "
                        f"{timezone}"
                    ),
                )

        return "FAIL", False

    target_index = (
        target_timezone.get(
            "Index"
        )
    )

    target_name = str(
        target_timezone.get(
            "Name",
            target_timezone.get(
                "Value",
                target_timezone,
            ),
        )
    )

    target_offset = str(
        target_timezone.get(
            "UtcOffset",
            target_timezone.get(
                "UTCOffset",
                "Unknown",
            ),
        )
    )

    if target_index is None:

        log(
            server_name,
            ip,
            (
                f"Timezone matching '{TIMEZONE_MATCH}' "
                "was found but has no Index. "
                f"Entry={target_timezone}"
            ),
        )

        return "FAIL", False

    # -------------------------------------------------------------------------
    # Current timezone diagnostics
    # -------------------------------------------------------------------------

    current_index = None
    current_name = ""

    if isinstance(
        current_timezone,
        dict,
    ):

        current_index = (
            current_timezone.get(
                "Index"
            )
        )

        current_name = str(
            current_timezone.get(
                "Name",
                current_timezone.get(
                    "Value",
                    "",
                ),
            )
        )

    else:

        current_name = str(
            current_timezone
        )

    log(
        server_name,
        ip,
        (
            "Current timezone: "
            f"Index={current_index}, "
            f"Name={current_name}; "
            "Target timezone: "
            f"Index={target_index}, "
            f"Name={target_name}, "
            f"UTC Offset={target_offset}"
        ),
    )

    # -------------------------------------------------------------------------
    # Already correct?
    # -------------------------------------------------------------------------

    if (
        current_index is not None
        and str(current_index)
        == str(target_index)
    ):

        log(
            server_name,
            ip,
            (
                "Timezone already configured correctly: "
                f"{target_name}"
            ),
        )

        return "SKIP", False

    # A firmware may not expose Index in current TimeZone but may expose Name.
    if (
        TIMEZONE_MATCH.lower()
        in current_name.lower()
    ):

        log(
            server_name,
            ip,
            (
                "Current timezone already matches "
                f"'{TIMEZONE_MATCH}': {current_name}"
            ),
        )

        return "SKIP", False

    # -------------------------------------------------------------------------
    # Configure timezone
    # -------------------------------------------------------------------------

    log(
        server_name,
        ip,
        (
            "Configuring iLO timezone: "
            f"{target_name} "
            f"(Index={target_index}, "
            f"UTC Offset={target_offset})"
        ),
    )

    payload = {

        "TimeZone": {

            "Index":
                target_index
        }
    }

    response = client.patch(
        datetime_uri,
        payload,
    )

    if not response_is_success(
        response
    ):

        log(
            server_name,
            ip,
            (
                "Timezone configuration FAILED. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            ),
        )

        return "FAIL", False

    needs_reset = check_reset_required(response)

    try:
        readback = client.get(datetime_uri)

        if readback.status_code == 200:
            readback_data = readback.json()

            if configuration_requires_reset(readback_data):
                needs_reset = True

    except Exception as exc:
        debug_log(
            server_name,
            ip,
            (
                "Unable to perform timezone reset-state read-back: {}"
            ).format(exc),
        )

    if needs_reset:
        log(
            server_name,
            ip,
            (
                "Timezone configuration accepted and iLO reports that "
                "a reset is required."
            ),
        )
    else:
        log(
            server_name,
            ip,
            (
                "Timezone configuration accepted; iLO does not report "
                "a reset requirement."
            ),
        )

    return "OK", needs_reset


# =============================================================================
# TIMEZONE VERIFICATION
# =============================================================================

def verify_timezone(
    client,
    log_prefix,
):
    """
    Verify that the active timezone matches TIMEZONE_MATCH.

    If current TimeZone includes an Index, the target index is rediscovered
    from TimeZoneList and compared as well.
    """

    server_name, ip = log_prefix

    datetime_uri = (
        client.manager_uri.rstrip("/")
        + "/DateTime/"
    )

    try:

        response = client.get(
            datetime_uri
        )

    except Exception as exc:

        log(
            server_name,
            ip,
            (
                "Timezone verification failed "
                f"while reading DateTime: {exc}"
            ),
        )

        return False

    if response.status_code != 200:

        log(
            server_name,
            ip,
            (
                "Timezone verification failed. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            ),
        )

        return False

    try:

        data = response.json()

    except Exception:

        log(
            server_name,
            ip,
            (
                "Timezone verification failed: "
                "DateTime response was not valid JSON."
            ),
        )

        return False

    current_timezone = (
        data.get(
            "TimeZone",
            {}
        )
        or {}
    )

    timezone_list = (
        data.get(
            "TimeZoneList",
            []
        )
        or []
    )

    target_timezone = (
        find_matching_timezone(
            timezone_list,
            TIMEZONE_MATCH,
        )
    )

    current_index = None
    current_name = ""
    current_offset = ""

    if isinstance(
        current_timezone,
        dict,
    ):

        current_index = (
            current_timezone.get(
                "Index"
            )
        )

        current_name = str(
            current_timezone.get(
                "Name",
                current_timezone.get(
                    "Value",
                    "",
                ),
            )
        )

        current_offset = str(
            current_timezone.get(
                "UtcOffset",
                current_timezone.get(
                    "UTCOffset",
                    "",
                ),
            )
        )

        searchable = (
            " ".join(
                str(value)
                for value
                in current_timezone.values()
                if isinstance(
                    value,
                    (
                        str,
                        int,
                        float,
                    ),
                )
            )
            .lower()
        )

    else:

        current_name = str(
            current_timezone
        )

        searchable = (
            current_name.lower()
        )

    # -------------------------------------------------------------------------
    # Direct name/value match
    # -------------------------------------------------------------------------

    if (
        TIMEZONE_MATCH.lower()
        in searchable
    ):

        log(
            server_name,
            ip,
            (
                "TIMEZONE VERIFIED successfully. "
                f"Index={current_index}; "
                f"Name={current_name}; "
                f"UTC Offset={current_offset}"
            ),
        )

        return True

    # -------------------------------------------------------------------------
    # Compare dynamically discovered index if current TimeZone only exposes
    # an index without a useful display name.
    # -------------------------------------------------------------------------

    if (
        target_timezone is not None
        and current_index is not None
    ):

        target_index = (
            target_timezone.get(
                "Index"
            )
        )

        if (
            target_index is not None
            and str(current_index)
            == str(target_index)
        ):

            log(
                server_name,
                ip,
                (
                    "TIMEZONE VERIFIED successfully "
                    "by timezone Index. "
                    f"Index={current_index}; "
                    f"Target={target_timezone}"
                ),
            )

            return True

    log(
        server_name,
        ip,
        (
            "Timezone verification FAILED. "
            f"Expected match='{TIMEZONE_MATCH}'; "
            f"Current TimeZone={current_timezone}"
        ),
    )

    return False


# =============================================================================
# SNTP / NTP CONFIGURATION
# =============================================================================

def configure_sntp(
    client,
    scope_data,
    log_prefix,
):
    """
    Configure static SNTP/NTP servers through the iLO 5 DateTime resource.

    This runs AFTER DHCPv4.UseNTPServers and DHCPv6.UseNTPServers have been
    disabled and the intermediate iLO reset has completed.
    """

    server_name, ip = log_prefix

    primary_ntp = str(
        scope_data[
            "PRIMARY_NTP"
        ]
    ).strip()

    secondary_ntp = str(
        scope_data[
            "SECONDARY_NTP"
        ]
    ).strip()

    desired_servers = [
        primary_ntp,
        secondary_ntp,
    ]

    datetime_uri = (
        client.manager_uri.rstrip("/")
        + "/DateTime/"
    )

    log(
        server_name,
        ip,
        "Reading current iLO NTP configuration...",
    )

    try:

        response = client.get(
            datetime_uri
        )

    except Exception as exc:

        log(
            server_name,
            ip,
            (
                "Unable to read DateTime resource: "
                f"{exc}"
            ),
        )

        return "FAIL", False

    if response.status_code != 200:

        log(
            server_name,
            ip,
            (
                "Unable to read iLO DateTime resource. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            ),
        )

        return "FAIL", False

    try:

        current_data = (
            response.json()
        )

    except Exception:

        log(
            server_name,
            ip,
            "DateTime resource returned invalid JSON.",
        )

        return "FAIL", False

    current_ntp = (
        current_data.get(
            "NTPServers",
            []
        )
        or []
    )

    current_static_ntp = (
        current_data.get(
            "StaticNTPServers",
            []
        )
        or []
    )

    configuration_state = (
        current_data.get(
            "ConfigurationSettings",
            "Unknown",
        )
    )

    log(
        server_name,
        ip,
        (
            f"Current NTPServers={current_ntp}; "
            f"StaticNTPServers={current_static_ntp}; "
            f"ConfigurationSettings={configuration_state}"
        ),
    )

    if (
        "StaticNTPServers"
        not in current_data
    ):

        log(
            server_name,
            ip,
            (
                "StaticNTPServers property is not exposed "
                "by this iLO DateTime resource."
            ),
        )

        return "FAIL", False

    current_static_clean = (
        clean_string_list(
            current_static_ntp
        )
    )

    desired_clean = (
        clean_string_list(
            desired_servers
        )
    )

    # -------------------------------------------------------------------------
    # Already configured?
    # -------------------------------------------------------------------------

    if (
        len(current_static_clean)
        == len(desired_clean)
        and set(current_static_clean)
        == set(desired_clean)
    ):

        log(
            server_name,
            ip,
            "Static NTP servers already configured correctly.",
        )

        return "SKIP", False

    # -------------------------------------------------------------------------
    # Configure NTP
    # -------------------------------------------------------------------------

    log(
        server_name,
        ip,
        (
            "Configuring static NTP servers: "
            f"{primary_ntp}, "
            f"{secondary_ntp}"
        ),
    )

    payload = {

        "StaticNTPServers": [
            primary_ntp,
            secondary_ntp,
        ]
    }

    response = client.patch(
        datetime_uri,
        payload,
    )

    if not response_is_success(
        response
    ):

        error_text = (
            get_extended_error_text(
                response
            )
        )

        log(
            server_name,
            ip,
            (
                "SNTP/NTP configuration FAILED. "
                f"HTTP {response.status_code}: "
                f"{error_text}"
            ),
        )

        if (
            "SNTPConfigurationManagedByDHCPAndIsReadOnly"
            in error_text
            or "ManagedByDHCP"
            in error_text
        ):

            log(
                server_name,
                ip,
                (
                    "iLO still considers NTP managed by DHCP. "
                    "Verify DHCPv4.UseNTPServers=False and "
                    "DHCPv6.UseNTPServers=False."
                ),
            )

        return "FAIL", False

    needs_reset = check_reset_required(response)

    # -------------------------------------------------------------------------
    # Immediate readback
    # -------------------------------------------------------------------------
    try:
        verify_response = client.get(
            datetime_uri
        )

        if verify_response.status_code == 200:
            verify_data = verify_response.json()

            verify_static = clean_string_list(
                verify_data.get(
                    "StaticNTPServers",
                    [],
                )
                or []
            )

            verify_active = clean_string_list(
                verify_data.get(
                    "NTPServers",
                    [],
                )
                or []
            )

            static_match = (
                len(verify_static) == len(desired_clean)
                and set(verify_static) == set(desired_clean)
            )

            active_match = (
                len(verify_active) == len(desired_clean)
                and set(verify_active) == set(desired_clean)
            )

            if configuration_requires_reset(verify_data):
                needs_reset = True

            if not static_match and not active_match:
                needs_reset = True

            log(
                server_name,
                ip,
                (
                    "SNTP PATCH accepted. StaticNTPServers={}; "
                    "NTPServers={}; ConfigurationSettings={}; "
                    "ResetRequired={}"
                ).format(
                    verify_data.get("StaticNTPServers", []),
                    verify_data.get("NTPServers", []),
                    verify_data.get(
                        "ConfigurationSettings",
                        "Unknown",
                    ),
                    needs_reset,
                ),
            )

    except Exception as exc:
        debug_log(
            server_name,
            ip,
            (
                "Unable to perform immediate SNTP read-back: {}"
            ).format(exc),
        )

    if needs_reset:
        log(
            server_name,
            ip,
            (
                "Static SNTP configuration changed and iLO indicates "
                "that a reset is required."
            ),
        )
    else:
        log(
            server_name,
            ip,
            (
                "Static SNTP configuration changed successfully; "
                "no iLO reset is required."
            ),
        )

    return "OK", needs_reset


# =============================================================================
# SNTP VERIFICATION
# =============================================================================

def verify_sntp(
    client,
    scope_data,
    log_prefix,
):
    """
    Verify NTP configuration after final iLO reset.
    """

    server_name, ip = log_prefix

    datetime_uri = (
        client.manager_uri.rstrip("/")
        + "/DateTime/"
    )

    desired_servers = [
        str(
            scope_data[
                "PRIMARY_NTP"
            ]
        ).strip(),

        str(
            scope_data[
                "SECONDARY_NTP"
            ]
        ).strip(),
    ]

    try:

        response = client.get(
            datetime_uri
        )

    except Exception as exc:

        log(
            server_name,
            ip,
            (
                "SNTP verification failed "
                f"while reading DateTime: {exc}"
            ),
        )

        return False

    if response.status_code != 200:

        log(
            server_name,
            ip,
            (
                "SNTP verification failed. "
                f"HTTP {response.status_code}: "
                f"{get_extended_error_text(response)}"
            ),
        )

        return False

    try:

        data = response.json()

    except Exception:

        log(
            server_name,
            ip,
            (
                "SNTP verification failed: "
                "DateTime response was not valid JSON."
            ),
        )

        return False

    static_servers = (
        data.get(
            "StaticNTPServers",
            []
        )
        or []
    )

    active_servers = (
        data.get(
            "NTPServers",
            []
        )
        or []
    )

    static_clean = (
        clean_string_list(
            static_servers
        )
    )

    active_clean = (
        clean_string_list(
            active_servers
        )
    )

    desired_clean = (
        clean_string_list(
            desired_servers
        )
    )

    static_match = (
        len(static_clean)
        == len(desired_clean)
        and set(static_clean)
        == set(desired_clean)
    )

    active_match = (
        len(active_clean)
        == len(desired_clean)
        and set(active_clean)
        == set(desired_clean)
    )

    configuration_state = (
        data.get(
            "ConfigurationSettings",
            "Unknown",
        )
    )

    if (
        static_match
        or active_match
    ):

        log(
            server_name,
            ip,
            (
                "SNTP VERIFIED successfully. "
                f"StaticNTPServers={static_servers}; "
                f"NTPServers={active_servers}; "
                f"ConfigurationSettings="
                f"{configuration_state}"
            ),
        )

        return True

    log(
        server_name,
        ip,
        (
            "SNTP verification FAILED. "
            f"Expected={desired_servers}; "
            f"StaticNTPServers={static_servers}; "
            f"NTPServers={active_servers}; "
            f"ConfigurationSettings={configuration_state}"
        ),
    )

    return False


# =============================================================================
# IPMI
# =============================================================================

def configure_ipmi(
    client,
    log_prefix,
):
    """
    Configure IPMI only when the active iLO setting differs.

    Previously this function PATCHed NetworkProtocol on every execution.
    iLO 5 can return ResetRequired for that PATCH even when IPMI is already
    enabled on the requested port, causing an unnecessary iLO reset.
    """
    server_name, ip = log_prefix
    endpoint = (
        client.manager_uri.rstrip("/")
        + "/NetworkProtocol/"
    )

    try:
        response = client.get(endpoint)
    except Exception as exc:
        log(
            server_name,
            ip,
            "Unable to read current IPMI configuration: {}".format(exc),
        )
        return "FAIL", False

    if response.status_code != 200:
        log(
            server_name,
            ip,
            (
                "Unable to read current IPMI configuration. "
                "HTTP {}: {}"
            ).format(
                response.status_code,
                get_extended_error_text(response),
            ),
        )
        return "FAIL", False

    try:
        data = response.json()
    except Exception:
        log(
            server_name,
            ip,
            "NetworkProtocol resource returned invalid JSON.",
        )
        return "FAIL", False

    current = data.get("IPMI", {}) or {}
    current_enabled = current.get("ProtocolEnabled")
    enabled_property = "ProtocolEnabled"

    # Some iLO firmware revisions expose Enabled instead of ProtocolEnabled.
    if current_enabled is None:
        current_enabled = current.get("Enabled")
        if "Enabled" in current:
            enabled_property = "Enabled"

    current_port = current.get("Port")

    try:
        port_matches = int(current_port) == int(IPMI_PORT)
    except (TypeError, ValueError):
        port_matches = str(current_port).strip() == str(IPMI_PORT).strip()

    if current_enabled is True and port_matches:
        log(
            server_name,
            ip,
            "IPMI already configured correctly: enabled=True, port={}.".format(
                IPMI_PORT
            ),
        )
        return "SKIP", False

    log(
        server_name,
        ip,
        (
            "Updating IPMI configuration: current_enabled={!r}, "
            "current_port={!r}, target_enabled=True, target_port={}"
        ).format(
            current_enabled,
            current_port,
            IPMI_PORT,
        ),
    )

    payload = {
        "IPMI": {
            enabled_property: True,
            "Port": IPMI_PORT,
        }
    }

    response = client.patch(
        endpoint,
        payload,
    )

    # If the firmware rejected Enabled, retry once with ProtocolEnabled.
    if (
        not response_is_success(response)
        and enabled_property == "Enabled"
    ):
        payload = {
            "IPMI": {
                "ProtocolEnabled": True,
                "Port": IPMI_PORT,
            }
        }
        response = client.patch(
            endpoint,
            payload,
        )

    if not response_is_success(response):
        log(
            server_name,
            ip,
            (
                "IPMI configuration failed: {}"
            ).format(
                get_extended_error_text(response)
            ),
        )
        return "FAIL", False

    needs_reset = check_reset_required(response)

    # Immediate read-back provides an audit trail. If the setting is already
    # active and iLO did not explicitly request a reset, no reset is needed.
    try:
        verify_response = client.get(endpoint)
        if verify_response.status_code == 200:
            verify_ipmi = verify_response.json().get("IPMI", {}) or {}
            verify_enabled = verify_ipmi.get("ProtocolEnabled")
            if verify_enabled is None:
                verify_enabled = verify_ipmi.get("Enabled")
            verify_port = verify_ipmi.get("Port")

            log(
                server_name,
                ip,
                (
                    "IPMI PATCH accepted. Read-back: "
                    "enabled={!r}, port={!r}, reset_required={}"
                ).format(
                    verify_enabled,
                    verify_port,
                    needs_reset,
                ),
            )
    except Exception as exc:
        debug_log(
            server_name,
            ip,
            "Unable to perform IPMI read-back: {}".format(exc),
        )

    return "OK", needs_reset


# =============================================================================
# BOOT ORDER
# =============================================================================

def set_boot_order(
    client,
    log_prefix,
):

    response = client.patch(

        client.system_uri.rstrip("/")
        + "/",

        {
            "Boot": {

                "BootSourceOverrideTarget":
                    "Pxe",

                "BootSourceOverrideEnabled":
                    "Once",
            }
        },
    )

    if response_is_success(
        response
    ):

        return (
            "OK",
            check_reset_required(response),
        )

    log(
        log_prefix[0],
        log_prefix[1],
        (
            "Boot override configuration failed: "
            f"{get_extended_error_text(response)}"
        ),
    )

    return "FAIL", False


# =============================================================================
# PARALLEL WORKER
# =============================================================================

def process_server(
    row_dict,
):

    start_time = (
        time.time()
    )

    server_name = str(
        row_dict.get(
            "hostname",
            "Unknown",
        )
    ).strip()

    ip = str(
        row_dict.get(
            "ilo_ip",
            "",
        )
    ).strip()

    scope = str(
        row_dict.get(
            "scope",
            "",
        )
    ).strip().upper()

    server_fqdn = str(
        row_dict.get(
            "ilo_hostname",
            "",
        )
    ).strip()

    # -------------------------------------------------------------------------
    # Report structure
    # -------------------------------------------------------------------------

    result = {

        "Hostname":
            server_name,

        "IP":
            ip,

        "Status":
            "Failed",

        "AUTH":
            "-",

        "LIC":
            "-",

        "USR":
            "-",

        "NET":
            "-",

        "IPMI":
            "-",

        "ID":
            "-",

        "LDAP":
            "-",

        "CERT":
            "-",

        "TZ":
            "-",

        "SNTP":
            "-",

        "BOOT":
            "-",

        "Time":
            0.0,
    }

    # -------------------------------------------------------------------------
    # Required values
    # -------------------------------------------------------------------------

    if (
        not ip
        or not server_name
        or not scope
        or not server_fqdn
    ):

        result[
            "Status"
        ] = "Skipped"

        result[
            "Time"
        ] = (
            time.time()
            - start_time
        )

        return result

    if scope not in SCOPE_SETTINGS:

        result[
            "Status"
        ] = "Skipped"

        result[
            "Time"
        ] = (
            time.time()
            - start_time
        )

        return result

    scope_data = (
        SCOPE_SETTINGS[
            scope
        ]
    )

    log_prefix = (
        server_name,
        ip,
    )

    # -------------------------------------------------------------------------
    # Reachability
    # -------------------------------------------------------------------------

    if not is_reachable(
        ip,
        server_name,
    ):

        log(
            server_name,
            ip,
            "iLO is not reachable.",
        )

        result[
            "Time"
        ] = (
            time.time()
            - start_time
        )

        return result

    client = RedfishClient(
        ip=ip,
        username=ILO_LOGIN,
        password=ILO_PASSWORD,
    )

    server_has_errors = False

    phase1_needs_reset = False
    phase1_reset_tasks = []

    timezone_needs_reset = False
    sntp_needs_reset = False

    try:

        # =====================================================================
        # AUTHENTICATION
        # =====================================================================

        log(
            *log_prefix,
            "Authenticating...",
        )

        if not client.login():

            result[
                "AUTH"
            ] = "FAIL"

            return result

        result[
            "AUTH"
        ] = "OK"

        client.discover_resources()

        # =====================================================================
        # PHASE 1
        #
        # Configure everything except TimeZone and StaticNTPServers.
        #
        # NET disables DHCP-provided NTP.
        # =====================================================================

        phase1_tasks = [

            (
                "LIC",
                apply_license,
                (
                    client,
                    log_prefix,
                ),
            ),

            (
                "USR",
                apply_fencing_user,
                (
                    client,
                    log_prefix,
                ),
            ),

            (
                "NET",
                configure_ilo_network,
                (
                    client,
                    server_fqdn,
                    scope_data,
                    log_prefix,
                ),
            ),

            (
                "IPMI",
                configure_ipmi,
                (
                    client,
                    log_prefix,
                ),
            ),

            (
                "ID",
                configure_server_identification,
                (
                    client,
                    server_name,
                    server_fqdn,
                    log_prefix,
                ),
            ),

            (
                "LDAP",
                configure_ldap,
                (
                    client,
                    scope_data,
                    log_prefix,
                ),
            ),

            (
                "CERT",
                configure_ldap_certificate,
                (
                    client,
                    log_prefix,
                ),
            ),

            (
                "BOOT",
                set_boot_order,
                (
                    client,
                    log_prefix,
                ),
            ),
        ]

        for (
            task_name,
            task_func,
            args,
        ) in phase1_tasks:

            try:

                status, needs_reset = (
                    task_func(
                        *args
                    )
                )

                result[
                    task_name
                ] = status

                if status == "FAIL":

                    server_has_errors = True

                if needs_reset:
                    phase1_needs_reset = True
                    phase1_reset_tasks.append(task_name)

            except Exception as exc:

                log(
                    *log_prefix,
                    (
                        f"Task {task_name} "
                        f"Exception: {exc}"
                    ),
                )

                result[
                    task_name
                ] = "FAIL"

                server_has_errors = True

        # =====================================================================
        # INTERMEDIATE iLO RESET
        #
        # Required before StaticNTPServers to make the DHCP ownership change
        # active on iLO 5.
        # =====================================================================

        if phase1_needs_reset:

            log(
                *log_prefix,
                (
                    "Phase-1 configuration requires iLO reset "
                    "before timezone/SNTP. Requested by: {}"
                ).format(
                    ", ".join(phase1_reset_tasks)
                    if phase1_reset_tasks
                    else "Unknown"
                ),
            )

            client = (
                reset_ilo_and_reconnect(
                    client,
                    log_prefix,
                )
            )
        else:
            log(
                *log_prefix,
                (
                    "No phase-1 setting requires an iLO reset. "
                    "Continuing directly to timezone/SNTP."
                ),
            )

        # =====================================================================
        # PHASE 2A - TIMEZONE
        # =====================================================================

        try:

            (
                timezone_status,
                timezone_needs_reset,
            ) = (
                configure_timezone(
                    client,
                    log_prefix,
                )
            )

            result[
                "TZ"
            ] = timezone_status

            if (
                timezone_status
                == "FAIL"
            ):

                server_has_errors = True

        except Exception as exc:

            log(
                *log_prefix,
                (
                    "Task TZ Exception: "
                    f"{exc}"
                ),
            )

            result[
                "TZ"
            ] = "FAIL"

            timezone_needs_reset = False

            server_has_errors = True

        # =====================================================================
        # PHASE 2B - SNTP
        # =====================================================================

        try:

            (
                sntp_status,
                sntp_needs_reset,
            ) = (
                configure_sntp(
                    client,
                    scope_data,
                    log_prefix,
                )
            )

            result[
                "SNTP"
            ] = sntp_status

            if (
                sntp_status
                == "FAIL"
            ):

                server_has_errors = True

        except Exception as exc:

            log(
                *log_prefix,
                (
                    "Task SNTP Exception: "
                    f"{exc}"
                ),
            )

            result[
                "SNTP"
            ] = "FAIL"

            sntp_needs_reset = False

            server_has_errors = True

        # =====================================================================
        # FINAL DATE/TIME RESET
        #
        # ONE reset covers both timezone and SNTP changes.
        # =====================================================================

        if (
            timezone_needs_reset
            or sntp_needs_reset
        ):
            final_reset_reasons = []

            if timezone_needs_reset:
                final_reset_reasons.append("Timezone")

            if sntp_needs_reset:
                final_reset_reasons.append("SNTP")

            log(
                *log_prefix,
                (
                    "Resetting iLO because the following changed "
                    "setting(s) require it: {}"
                ).format(
                    ", ".join(final_reset_reasons)
                ),
            )

            client = (
                reset_ilo_and_reconnect(
                    client,
                    log_prefix,
                )
            )
        else:
            log(
                *log_prefix,
                "Timezone/SNTP do not require an iLO reset.",
            )

        # =====================================================================
        # TIMEZONE VERIFICATION
        # =====================================================================

        if result[
            "TZ"
        ] in [
            "OK",
            "SKIP",
        ]:

            log(
                *log_prefix,
                "Verifying active timezone configuration...",
            )

            if not verify_timezone(
                client,
                log_prefix,
            ):

                result[
                    "TZ"
                ] = "FAIL"

                server_has_errors = True

        # =====================================================================
        # SNTP VERIFICATION
        # =====================================================================

        if result[
            "SNTP"
        ] in [
            "OK",
            "SKIP",
        ]:

            log(
                *log_prefix,
                "Verifying active SNTP/NTP configuration...",
            )

            if not verify_sntp(
                client,
                scope_data,
                log_prefix,
            ):

                result[
                    "SNTP"
                ] = "FAIL"

                server_has_errors = True

        # =====================================================================
        # FINAL STATUS
        # =====================================================================

        result[
            "Status"
        ] = (

            "Completed w/ Errors"

            if server_has_errors

            else "Successful"
        )

    except Exception as exc:

        log(
            *log_prefix,
            (
                "Global Exception: "
                f"{exc}"
            ),
        )

        result[
            "Status"
        ] = "Failed"

    finally:

        try:

            client.logout()

        except Exception:

            pass

        result[
            "Time"
        ] = (
            time.time()
            - start_time
        )

    return result


# =============================================================================
# MAIN
# =============================================================================

def main():

    t_start_global = (
        time.time()
    )

    print(
        "\n"
        + "=" * 90
    )

    print(
        "HPE ProLiant Gen10+ / "
        "iLO 5 Redfish Parallel Provisioner"
    )

    print(
        "=" * 90
        + "\n"
    )

    # -------------------------------------------------------------------------
    # Configuration summary
    # -------------------------------------------------------------------------

    print(
        "Configuration:"
    )

    print(
        (
            "  Equipment types         : "
            f"{', '.join(EQUIPMENT_TYPES)}"
        )
    )

    print(
        (
            "  Excel rows              : "
            f"{START_ROW} - {END_ROW}"
        )
    )

    print(
        (
            "  Max concurrent sessions : "
            f"{MAX_WORKERS}"
        )
    )

    print(
        (
            "  Timezone match          : "
            f"{TIMEZONE_MATCH}"
        )
    )

    print()

    print(
        "iLO Reset / Reconnect Timing:"
    )

    print(
        (
            "  Initial wait            : "
            f"{ILO_RESET_INITIAL_WAIT_SECONDS} seconds"
        )
    )

    print(
        (
            "  Retry interval          : "
            f"{ILO_RESET_RETRY_INTERVAL_SECONDS} seconds"
        )
    )

    print(
        (
            "  Maximum retry window    : "
            f"{ILO_RESET_MAX_WAIT_SECONDS} seconds"
        )
    )

    print(
        (
            "  Maximum theoretical wait: "
            f"{ILO_RESET_INITIAL_WAIT_SECONDS + ILO_RESET_MAX_WAIT_SECONDS} "
            "seconds"
        )
    )

    print()

    # -------------------------------------------------------------------------
    # Excel existence
    # -------------------------------------------------------------------------

    if not EXCEL_FILE.is_file():

        print(
            (
                "ERROR: Excel file not found at: "
                f"{EXCEL_FILE}"
            )
        )

        return 1

    # -------------------------------------------------------------------------
    # Load Excel range
    # -------------------------------------------------------------------------

    try:

        t_load_start = (
            time.time()
        )

        print(
            (
                "Loading inventory "
                f"(Slicing Rows {START_ROW} "
                f"to {END_ROW})..."
            )
        )

        df = (
            load_resource_excel_fast(
                EXCEL_FILE,
                SHEET_NAME,
                START_ROW,
                END_ROW,
            )
        )

        time_excel_load = (
            time.time()
            - t_load_start
        )

    except Exception as exc:

        print(
            (
                "ERROR: Failed to read Excel slice: "
                f"{exc}"
            )
        )

        return 1

    # -------------------------------------------------------------------------
    # Validate equipment_type
    # -------------------------------------------------------------------------

    if (
        df.empty
        or "equipment_type"
        not in df.columns
    ):

        print(
            (
                "ERROR: No valid rows or "
                "'equipment_type' column missing."
            )
        )

        return 1

    target_types = [
        equipment_type.upper()
        for equipment_type
        in EQUIPMENT_TYPES
    ]

    target_df = df[

        df[
            "equipment_type"
        ]
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(
            target_types
        )
    ]

    if target_df.empty:

        print(
            (
                "No servers matching equipment "
                "types in specified row range."
            )
        )

        return 0

    servers_to_process = (
        target_df.to_dict(
            "records"
        )
    )

    total_servers = len(
        servers_to_process
    )

    safe_workers = min(
        MAX_WORKERS,
        total_servers,
    )

    final_report = []

    # =========================================================================
    # PARALLEL REDFISH PROCESSING
    # =========================================================================

    t_redfish_start = (
        time.time()
    )

    print(
        (
            f"Found {total_servers} matching server(s). "
            "Starting parallel Redfish phase "
            f"({safe_workers} workers)...\n"
        )
    )

    with concurrent.futures.ThreadPoolExecutor(
        max_workers=safe_workers
    ) as executor:

        futures = {

            executor.submit(
                process_server,
                server,
            ):
                server

            for server
            in servers_to_process
        }

        for future in (
            concurrent.futures
            .as_completed(
                futures
            )
        ):

            server_data = (
                futures[
                    future
                ]
            )

            try:

                result = (
                    future.result()
                )

            except Exception as exc:

                hostname = str(
                    server_data.get(
                        "hostname",
                        "Unknown",
                    )
                )

                ip = str(
                    server_data.get(
                        "ilo_ip",
                        "Unknown",
                    )
                )

                log(
                    hostname,
                    ip,
                    (
                        "Unhandled worker exception: "
                        f"{exc}"
                    ),
                )

                result = {

                    "Hostname":
                        hostname,

                    "IP":
                        ip,

                    "Status":
                        "Failed",

                    "AUTH":
                        "-",

                    "LIC":
                        "-",

                    "USR":
                        "-",

                    "NET":
                        "-",

                    "IPMI":
                        "-",

                    "ID":
                        "-",

                    "LDAP":
                        "-",

                    "CERT":
                        "-",

                    "TZ":
                        "-",

                    "SNTP":
                        "-",

                    "BOOT":
                        "-",

                    "Time":
                        0.0,
                }

            final_report.append(
                result
            )

    time_redfish_phase = (
        time.time()
        - t_redfish_start
    )

    # =========================================================================
    # FINAL REPORT
    # =========================================================================

    t_report_start = (
        time.time()
    )

    sorted_report = sorted(
        final_report,
        key=lambda item:
            item[
                "Hostname"
            ],
    )

    successful = sum(

        1

        for item
        in final_report

        if item[
            "Status"
        ] == "Successful"
    )

    skipped = sum(

        1

        for item
        in final_report

        if item[
            "Status"
        ] == "Skipped"
    )

    failed = (
        total_servers
        - successful
        - skipped
    )

    print(
        "\nFINAL EXECUTION REPORT:"
    )

    print(
        "=" * 162
    )

    print(
        f"{'HOSTNAME':<18} | "
        f"{'iLO IP':<15} | "
        f"{'OVERALL STATUS':<19} | "
        f"AUTH | "
        f"LIC  | "
        f"USR  | "
        f"NET  | "
        f"IPMI | "
        f"ID   | "
        f"LDAP | "
        f"CERT | "
        f"TZ   | "
        f"SNTP | "
        f"BOOT | "
        f"Time (s)"
    )

    print(
        "-" * 162
    )

    for srv in sorted_report:

        print(
            f"{srv['Hostname']:<18} | "
            f"{srv['IP']:<15} | "
            f"{srv['Status']:<19} | "
            f"{srv['AUTH']:<4} | "
            f"{srv['LIC']:<4} | "
            f"{srv['USR']:<4} | "
            f"{srv['NET']:<4} | "
            f"{srv['IPMI']:<4} | "
            f"{srv['ID']:<4} | "
            f"{srv['LDAP']:<4} | "
            f"{srv['CERT']:<4} | "
            f"{srv['TZ']:<4} | "
            f"{srv['SNTP']:<4} | "
            f"{srv['BOOT']:<4} | "
            f"{srv['Time']:>8.2f}"
        )

    print("\n")

    print(
        f"TOTAL FOUND: {total_servers} | "
        f"SUCCESSFUL: {successful} | "
        f"FAILED: {failed} | "
        f"SKIPPED: {skipped}"
    )

    print(
        "=" * 162
        + "\n"
    )

    # =========================================================================
    # PERFORMANCE SUMMARY
    # =========================================================================

    time_report_phase = (
        time.time()
        - t_report_start
    )

    total_execution_time = (
        time.time()
        - t_start_global
    )

    print(
        "=" * 70
    )

    print(
        "PERFORMANCE SUMMARY"
    )

    print(
        "=" * 70
    )

    print(
        (
            "Excel load/parse              : "
            f"{time_excel_load:>9.2f} sec"
        )
    )

    print(
        (
            "Parallel Redfish phase        : "
            f"{time_redfish_phase:>9.2f} sec"
        )
    )

    print(
        (
            "Final report                  : "
            f"{time_report_phase:>9.2f} sec"
        )
    )

    print(
        "-" * 70
    )

    print(
        (
            "TOTAL EXECUTION TIME          : "
            f"{total_execution_time:>9.2f} sec"
        )
    )

    print(
        "=" * 70
        + "\n"
    )

    print(
        "Concurrency used:"
    )

    print(
        (
            f"  Total servers                    : "
            f"{total_servers}"
        )
    )

    print(
        (
            "  REDFISH_CONCURRENT_SESSIONS      : "
            f"{safe_workers}"
        )
    )

    print()

    # =========================================================================
    # CSV AUDIT LOG
    # =========================================================================

    if final_report:

        log_dir = (
            BASE_DIR
            / LOG_DIR_NAME
        )

        log_dir.mkdir(
            parents=True,
            exist_ok=True,
        )

        timestamp = (
            time.strftime(
                "%Y%m%d-%H%M%S"
            )
        )

        report_filename = (
            f"{REPORT_PREFIX}_"
            f"{timestamp}.csv"
        )

        report_path = (
            log_dir
            / report_filename
        )

        pd.DataFrame(
            sorted_report
        ).to_csv(
            report_path,
            index=False,
        )

        print(
            (
                "Audit log saved to: "
                f"{report_path}\n"
            )
        )

    return 0


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":

    sys.exit(
        main()
    )