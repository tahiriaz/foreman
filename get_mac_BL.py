# BUILD_MARKER: GET_MAC_BL_CENTRAL_V3_20260901_ENCLOSURE_COMPONENT_SERIALS

import gc
import os
import re
import time
import warnings

warnings.filterwarnings("ignore", category=DeprecationWarning)
warnings.filterwarnings("ignore", message=".*Python 3.6 is no longer supported.*")
warnings.filterwarnings("ignore", module="cryptography")
warnings.filterwarnings("ignore", module="paramiko")

import paramiko
import win32com.client as win32
from openpyxl import load_workbook

from functions import vars
from functions.output_log import run_logged_main
from functions.reporting import make_summary_row, print_summary_report, write_summary_csv


# ============================================================================
# OA CONNECTION
# ============================================================================

class OAConnection:
    """Interactive SSH connection wrapper for HPE Onboard Administrator."""

    def __init__(self, primary_ip, secondary_ip, user, password):
        self.ips = [ip for ip in (primary_ip, secondary_ip) if ip]
        self.user = user
        self.password = password
        self.ssh = None
        self.channel = None
        self.active_ip = None

    def connect(self):
        for ip in self.ips:
            print("\nAttempting connection to OA at {}...".format(ip))
            self.ssh = paramiko.SSHClient()
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            try:
                self.ssh.connect(
                    hostname=ip,
                    username=self.user,
                    password=self.password,
                    look_for_keys=False,
                    allow_agent=False,
                    timeout=vars.MAC_BL_OA_CONNECT_TIMEOUT,
                    banner_timeout=vars.MAC_BL_OA_BANNER_TIMEOUT,
                    auth_timeout=vars.MAC_BL_OA_AUTH_TIMEOUT,
                )
                transport = self.ssh.get_transport()
                if transport:
                    transport.set_keepalive(vars.MAC_BL_OA_KEEPALIVE_INTERVAL)

                self.channel = self.ssh.invoke_shell()
                time.sleep(vars.MAC_BL_OA_SHELL_SETTLE_SECONDS)
                self._drain_channel()
                status_output = self.send_command("SHOW OA STATUS")

                if re.search(r"Role:\s*Active", status_output, flags=re.IGNORECASE):
                    print("Successfully connected to ACTIVE OA ({}).".format(ip))
                    self.active_ip = ip
                    return True

                print("OA at {} is standby/inactive. Closing and trying next...".format(ip))
                self.close()
            except Exception as exc:
                print("Failed to connect to {}: {}".format(ip, exc))
                self.close()
        return False

    def _drain_channel(self):
        output = ""
        while self.channel and self.channel.recv_ready():
            data = self.channel.recv(65535)
            if not data:
                break
            output += data.decode("utf-8", errors="replace")
        return output

    def send_command(self, command, wait_string=">"):
        """Send one OA command and return the complete interactive output."""
        if not self.channel:
            return ""

        self._drain_channel()
        self.channel.send(command + "\n")
        output = ""
        deadline = time.time() + vars.MAC_BL_OA_COMMAND_TIMEOUT_SECONDS

        while time.time() < deadline:
            if self.channel.recv_ready():
                data = self.channel.recv(65535)
                if not data:
                    break

                text = data.decode("utf-8", errors="replace")
                output += text
                lower_text = text.lower()
                if "any key" in lower_text or "more" in lower_text:
                    self.channel.send(" ")

                if wait_string in output:
                    time.sleep(vars.MAC_BL_OA_COMMAND_SETTLE_SECONDS)
                    output += self._drain_channel()
                    break
            else:
                time.sleep(vars.MAC_BL_OA_POLL_INTERVAL_SECONDS)
        return output

    def close(self):
        if self.channel:
            try:
                self.channel.close()
            except Exception:
                pass
        self.channel = None

        if self.ssh:
            try:
                self.ssh.close()
            except Exception:
                pass
        self.ssh = None


# ============================================================================
# HELPERS
# ============================================================================

def check_file_writable(file_path):
    if not os.path.exists(file_path):
        return True
    try:
        with open(file_path, "a+"):
            pass
        return True
    except PermissionError:
        return False


def normalize_text(value):
    return str(value or "").strip()


def normalize_type(value):
    return normalize_text(value).upper()


def parse_slot(value):
    try:
        return int(float(normalize_text(value)))
    except (ValueError, TypeError):
        return None


def is_valid_serial(value):
    serial = normalize_text(value)
    if not serial:
        return False
    invalid_values = {str(item).strip().upper() for item in vars.MAC_BL_SERIAL_INVALID_VALUES}
    return serial.upper() not in invalid_values


def extract_first_serial(output):
    """Return the first valid Serial Number field from one OA command response."""
    for line in (output or "").splitlines():
        match = re.search(r"(?:Product\s+)?Serial Number\s*:\s*(.+?)\s*$", line, re.IGNORECASE)
        if not match:
            continue
        serial = match.group(1).strip()
        if is_valid_serial(serial):
            return serial
    return ""


def get_component_serials(oa_connection, inventory):
    """Fetch enclosure, OA, and interconnect serials for rows present in Excel."""
    serials = {"enclosure": "", "oa": {}, "switch": {}}

    if inventory["enclosure_rows"]:
        print("Fetching enclosure serial number...")
        serials["enclosure"] = extract_first_serial(
            oa_connection.send_command("SHOW ENCLOSURE INFO")
        )

    for slot in sorted(inventory["oa_rows"]):
        if slot not in vars.MAC_BL_OA_SLOTS:
            continue
        print("Fetching OA {} serial number...".format(slot))
        serials["oa"][slot] = extract_first_serial(
            oa_connection.send_command("SHOW OA INFO {}".format(slot))
        )

    for slot in sorted(inventory["switch_rows"]):
        if slot not in vars.MAC_BL_SWITCH_SLOTS:
            continue
        print("Fetching blade switch {} serial number...".format(slot))
        serials["switch"][slot] = extract_first_serial(
            oa_connection.send_command("SHOW INTERCONNECT INFO {}".format(slot))
        )

    return serials


def get_all_blade_details(oa_connection):
    """Fetch serial numbers and MAC addresses for every blade in one enclosure."""
    results = {}
    print("Fetching global serial numbers for all blades...")
    info_raw = oa_connection.send_command("SHOW SERVER INFO ALL")
    print("Fetching global port maps for all blades...")
    port_raw = oa_connection.send_command("SHOW SERVER PORT MAP ALL")

    bay_regex = re.compile(
        r"(?:Server\s+)?Blade\s*(?:#|Number)?\s*(\d+)",
        re.IGNORECASE,
    )
    current_bay = None

    for line in info_raw.splitlines():
        bay_match = bay_regex.search(line)
        if bay_match:
            current_bay = int(bay_match.group(1))
            results.setdefault(
                current_bay,
                {"serial": "", "macs": [""] * vars.MAC_MAX_ADDRESSES},
            )

        if current_bay:
            serial_match = re.search(
                r"Serial Number:\s*([A-Za-z0-9\-]+)",
                line,
                re.IGNORECASE,
            )
            if serial_match:
                serial = serial_match.group(1).strip()
                if is_valid_serial(serial):
                    results[current_bay]["serial"] = serial

    current_bay = None
    current_section = "other"
    flb_macs = {}
    other_macs = {}
    mac_pattern = re.compile(r"(?:[0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}")

    for line in port_raw.splitlines():
        bay_match = bay_regex.search(line)
        if bay_match:
            current_bay = int(bay_match.group(1))
            flb_macs.setdefault(current_bay, [])
            other_macs.setdefault(current_bay, [])
            current_section = "other"
            continue

        if not current_bay:
            continue

        upper_line = line.upper()
        if "LOM" in upper_line or "FLB" in upper_line:
            current_section = "flb"
        elif "MEZZ" in upper_line:
            current_section = "other"

        for mac in mac_pattern.findall(line):
            clean_mac = mac.replace("-", ":").upper()
            target = flb_macs[current_bay] if current_section == "flb" else other_macs[current_bay]
            if clean_mac not in target:
                target.append(clean_mac)

    for bay, flb_list in flb_macs.items():
        results.setdefault(bay, {"serial": "", "macs": [""] * vars.MAC_MAX_ADDRESSES})
        other_list = other_macs.get(bay, [])
        combined = flb_list + [mac for mac in other_list if mac not in flb_list]
        for index, mac in enumerate(combined[:vars.MAC_MAX_ADDRESSES]):
            results[bay]["macs"][index] = mac

    return results


def refresh_excel_formulas(file_path):
    """Use native Excel to recalculate formulas and rebuild cached values."""
    print("\nCommanding Excel to perform a full formula recalculation...")
    absolute_path = os.path.abspath(file_path)
    excel = None
    workbook = None

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.AskToUpdateLinks = False
        workbook = excel.Workbooks.Open(
            absolute_path,
            UpdateLinks=0,
            ReadOnly=False,
        )
        try:
            excel.Calculation = -4105  # xlCalculationAutomatic
        except Exception:
            pass
        excel.CalculateFullRebuild()
        workbook.Save()
        print("Formula cache successfully rebuilt!")
    except Exception as exc:
        print("WARNING: Could not trigger Excel to rebuild formulas: {}".format(exc))
    finally:
        if workbook:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if excel:
            try:
                excel.Quit()
            except Exception:
                pass
        workbook = None
        excel = None
        gc.collect()


def row_matches_enclosure(enclosure_name, enclosure_value, equipment_physical_value):
    enclosure_text = normalize_type(enclosure_value)
    equipment_physical_text = normalize_type(equipment_physical_value)
    if enclosure_text == enclosure_name:
        return True
    return (
        len(enclosure_name) >= 14
        and equipment_physical_text.startswith(enclosure_name[:14])
    )


def scan_excel_for_enclosure(excel_path, sheet_name, enclosure_name):
    """Locate enclosure component rows, blade rows, and both OA management IPs."""
    print("\nScanning Excel file (Fast Read-Only Mode)...")
    workbook = load_workbook(excel_path, data_only=True, read_only=True)

    try:
        if sheet_name not in workbook.sheetnames:
            return None, "Error: Sheet '{}' not found.".format(sheet_name)

        worksheet = workbook[sheet_name]
        header_row = next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
            None,
        )
        if not header_row:
            return None, "Error: Excel header row is empty."

        headers = {
            str(cell_value).strip(): index
            for index, cell_value in enumerate(header_row, 1)
            if cell_value
        }
        missing_columns = [
            column for column in vars.MAC_BL_REQUIRED_COLUMNS if column not in headers
        ]
        if missing_columns:
            return None, "Error: Required columns missing: {}".format(
                ", ".join(missing_columns)
            )

        inventory = {
            "primary_oa_ip": None,
            "secondary_oa_ip": None,
            "enclosure_rows": [],
            "oa_rows": {},
            "switch_rows": {},
            "blade_rows": [],
            "headers": headers,
        }
        empty_count = 0

        for row_idx, row_values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True),
            2,
        ):
            if not any(row_values):
                empty_count += 1
                if empty_count >= vars.MAC_BL_EXCEL_EMPTY_ROW_STOP:
                    break
                continue
            empty_count = 0

            enclosure_value = row_values[headers["enclosure_physical_name"] - 1]
            physical_value = row_values[headers["equipment_physical_name"] - 1]
            if not row_matches_enclosure(enclosure_name, enclosure_value, physical_value):
                continue

            equipment_type = normalize_type(
                row_values[headers["equipment_type"] - 1]
            )
            slot_number = parse_slot(row_values[headers["enclosure_slot"] - 1])
            ip_value = normalize_text(row_values[headers["ilo_ip"] - 1])
            row_info = {
                "row_idx": row_idx,
                "slot": slot_number,
                "eq_type": equipment_type,
            }

            if equipment_type == vars.OA_SERVER_ENCLOSURE_TYPE.upper():
                inventory["enclosure_rows"].append(row_info)
            elif equipment_type == vars.OA_ENCLOSURE_OA_TYPE.upper():
                if slot_number in vars.MAC_BL_OA_SLOTS:
                    inventory["oa_rows"][slot_number] = row_info
                    if slot_number == 1 and ip_value:
                        inventory["primary_oa_ip"] = ip_value
                    elif slot_number == 2 and ip_value:
                        inventory["secondary_oa_ip"] = ip_value
            elif equipment_type == vars.OA_ENCLOSURE_SWITCH_TYPE.upper():
                if slot_number in vars.MAC_BL_SWITCH_SLOTS:
                    inventory["switch_rows"][slot_number] = row_info
            elif equipment_type in {
                item.upper() for item in vars.OA_BLADE_EQUIPMENT_TYPES
            }:
                inventory["blade_rows"].append(row_info)

        return inventory, None
    finally:
        workbook.close()


def serial_location_label(equipment_type, slot):
    if equipment_type == vars.OA_SERVER_ENCLOSURE_TYPE.upper():
        return "Enclosure"
    if equipment_type == vars.OA_ENCLOSURE_OA_TYPE.upper():
        return "OA {}".format(slot if slot is not None else "Unknown")
    if equipment_type == vars.OA_ENCLOSURE_SWITCH_TYPE.upper():
        return "Switch {}".format(slot if slot is not None else "Unknown")
    return "Blade {}".format(slot if slot is not None else "Unknown")


def update_serial_cell(
    worksheet,
    headers,
    row_info,
    serial_no,
    enclosure_name,
    serial_changes,
    serial_additions,
):
    """Write a non-empty serial only when Excel is blank or the value changed."""
    row_idx = row_info["row_idx"]
    equipment_type = row_info["eq_type"]
    slot = row_info["slot"]
    label = serial_location_label(equipment_type, slot)
    serial_cell = worksheet.cell(row=row_idx, column=headers[vars.MAC_SERIAL_COLUMN])
    existing_serial = normalize_text(serial_cell.value)

    if not is_valid_serial(serial_no):
        print("  Serial Number : [NOT FOUND - EXCEL UNCHANGED]")
        if existing_serial:
            print("  Existing Excel: {}".format(existing_serial))
        return "missing"

    serial_no = normalize_text(serial_no)
    if existing_serial and existing_serial.upper() == serial_no.upper():
        print("  Serial Number : {} [UNCHANGED]".format(serial_no))
        return "unchanged"

    if is_valid_serial(existing_serial):
        serial_cell.value = serial_no
        serial_changes.append({
            "row": row_idx,
            "enclosure": enclosure_name,
            "item": label,
            "equipment_type": equipment_type,
            "slot": slot,
            "old_serial": existing_serial,
            "new_serial": serial_no,
        })
        print("  Serial Number : {} -> {} [UPDATED]".format(existing_serial, serial_no))
        return "changed"

    serial_cell.value = serial_no
    serial_additions.append({
        "row": row_idx,
        "enclosure": enclosure_name,
        "item": label,
        "equipment_type": equipment_type,
        "slot": slot,
        "serial": serial_no,
    })
    print("  Serial Number : {} [ADDED]".format(serial_no))
    return "added"


def append_component_summary(
    summary_rows,
    row_info,
    enclosure_name,
    serial_no,
    serial_action,
    active_oa_ip,
    item_start,
):
    label = serial_location_label(row_info["eq_type"], row_info["slot"])
    status = "Partial" if serial_action == "missing" else "Successful"
    summary_rows.append(
        make_summary_row(
            row=row_info["row_idx"],
            item_type=row_info["eq_type"],
            name=label,
            target=enclosure_name,
            status=status,
            time_seconds=time.time() - item_start,
            details="Serial={}; Excel={}; ActiveOA={}".format(
                normalize_text(serial_no) if is_valid_serial(serial_no) else "NOT FOUND",
                serial_action.upper(),
                active_oa_ip or "Unknown",
            ),
        )
    )


# ============================================================================
# MAIN
# ============================================================================

def main():
    enclosure_name = input("Enter the Enclosure Name: ").strip().upper()
    excel_path = vars.RESOURCE_LIST
    sheet_name = vars.SHEET_NAME

    if not check_file_writable(excel_path):
        print("\nERROR: '{}' is currently open in Excel.".format(excel_path))
        print("Please close the file and run the script again to prevent permission errors.")
        return 1

    inventory, error = scan_excel_for_enclosure(excel_path, sheet_name, enclosure_name)
    if error:
        print(error)
        return 1

    primary_oa_ip = inventory["primary_oa_ip"]
    secondary_oa_ip = inventory["secondary_oa_ip"]
    if not primary_oa_ip and not secondary_oa_ip:
        print("Error: Could not find OA IPs for Enclosure '{}'.".format(enclosure_name))
        return 1

    total_components = (
        len(inventory["enclosure_rows"])
        + len(inventory["oa_rows"])
        + len(inventory["switch_rows"])
        + len(inventory["blade_rows"])
    )
    if not total_components:
        print("No enclosure components found for '{}'.".format(enclosure_name))
        return 1

    print(
        "Found Primary OA: {} | Secondary OA: {}".format(
            primary_oa_ip or "Not found",
            secondary_oa_ip or "Not found",
        )
    )
    print(
        "Found {} enclosure row(s), {} OA(s), {} switch(es), and {} blade server(s).".format(
            len(inventory["enclosure_rows"]),
            len(inventory["oa_rows"]),
            len(inventory["switch_rows"]),
            len(inventory["blade_rows"]),
        )
    )

    oa = OAConnection(
        primary_oa_ip,
        secondary_oa_ip,
        vars.OA_USERNAME,
        vars.OA_PASSWORD,
    )
    if not oa.connect():
        return 1

    component_serials = {}
    blade_data = {}
    try:
        component_serials = get_component_serials(oa, inventory)
        if inventory["blade_rows"]:
            blade_data = get_all_blade_details(oa)
    finally:
        oa.close()
        print("SSH connection closed.")

    print("\nLoading Excel file (Writing Mode) to inject data...")
    workbook_write = load_workbook(excel_path)
    worksheet_write = workbook_write[sheet_name]
    headers = inventory["headers"]
    summary_rows = []
    serial_changes = []
    serial_additions = []

    for row_info in inventory["enclosure_rows"]:
        item_start = time.time()
        print("\n--- Processing Row {}: Server Enclosure ---".format(row_info["row_idx"]))
        serial_no = component_serials.get("enclosure", "")
        action = update_serial_cell(
            worksheet_write,
            headers,
            row_info,
            serial_no,
            enclosure_name,
            serial_changes,
            serial_additions,
        )
        append_component_summary(
            summary_rows,
            row_info,
            enclosure_name,
            serial_no,
            action,
            oa.active_ip,
            item_start,
        )

    for slot in sorted(inventory["oa_rows"]):
        row_info = inventory["oa_rows"][slot]
        item_start = time.time()
        print("\n--- Processing Row {}: Enclosure OA (Slot {}) ---".format(row_info["row_idx"], slot))
        serial_no = component_serials.get("oa", {}).get(slot, "")
        action = update_serial_cell(
            worksheet_write,
            headers,
            row_info,
            serial_no,
            enclosure_name,
            serial_changes,
            serial_additions,
        )
        append_component_summary(
            summary_rows,
            row_info,
            enclosure_name,
            serial_no,
            action,
            oa.active_ip,
            item_start,
        )

    for slot in sorted(inventory["switch_rows"]):
        row_info = inventory["switch_rows"][slot]
        item_start = time.time()
        print("\n--- Processing Row {}: Enclosure Switch (Slot {}) ---".format(row_info["row_idx"], slot))
        serial_no = component_serials.get("switch", {}).get(slot, "")
        action = update_serial_cell(
            worksheet_write,
            headers,
            row_info,
            serial_no,
            enclosure_name,
            serial_changes,
            serial_additions,
        )
        append_component_summary(
            summary_rows,
            row_info,
            enclosure_name,
            serial_no,
            action,
            oa.active_ip,
            item_start,
        )

    for blade in inventory["blade_rows"]:
        item_start = time.time()
        row_idx = blade["row_idx"]
        bay_number = blade["slot"]
        equipment_type = blade["eq_type"]
        print(
            "\n--- Processing Row {}: {} (Bay {}) ---".format(
                row_idx,
                equipment_type,
                bay_number,
            )
        )

        if bay_number is None:
            print("WARNING: Skipping blade at row {}. Missing or invalid enclosure_slot.".format(row_idx))
            summary_rows.append(
                make_summary_row(
                    row=row_idx,
                    item_type=equipment_type,
                    name="Unknown bay",
                    target=enclosure_name,
                    status="Skipped",
                    time_seconds=time.time() - item_start,
                    details="Missing or invalid enclosure_slot",
                )
            )
            continue

        bay_data = blade_data.get(
            bay_number,
            {"serial": "", "macs": [""] * vars.MAC_MAX_ADDRESSES},
        )
        serial_no = normalize_text(bay_data.get("serial"))
        macs = list(bay_data.get("macs") or [])[:vars.MAC_MAX_ADDRESSES]
        macs += [""] * (vars.MAC_MAX_ADDRESSES - len(macs))
        serial_action = update_serial_cell(
            worksheet_write,
            headers,
            blade,
            serial_no,
            enclosure_name,
            serial_changes,
            serial_additions,
        )

        for index, column_name in enumerate(vars.MAC_NIC_COLUMNS):
            worksheet_write.cell(
                row=row_idx,
                column=headers[column_name],
            ).value = macs[index]

        for index, mac in enumerate(macs, 1):
            print("  mac{}          : {}".format(index, mac if mac else "[MISSING]"))

        found_count = len([mac for mac in macs if mac])
        summary_status = (
            "Partial"
            if serial_action == "missing" or found_count < vars.MAC_MAX_ADDRESSES
            else "Successful"
        )
        summary_rows.append(
            make_summary_row(
                row=row_idx,
                item_type=equipment_type,
                name="Bay {:02}".format(bay_number),
                target=enclosure_name,
                status=summary_status,
                time_seconds=time.time() - item_start,
                details="Serial={}; Excel={}; MACs={}/{}; ActiveOA={}".format(
                    serial_no if is_valid_serial(serial_no) else "NOT FOUND",
                    serial_action.upper(),
                    found_count,
                    vars.MAC_MAX_ADDRESSES,
                    oa.active_ip or "Unknown",
                ),
            )
        )

    print("\nSaving updates to {}...".format(excel_path))
    workbook_write.save(excel_path)
    workbook_write.close()
    print("Excel file successfully updated!")
    refresh_excel_formulas(excel_path)

    if serial_changes:
        print("\n" + "=" * 100)
        print("SERIAL NUMBERS CHANGED")
        print("=" * 100)
        for change in serial_changes:
            print(
                "{} | {} | Resource row {}: {} -> {}".format(
                    change["enclosure"],
                    change["item"],
                    change["row"],
                    change["old_serial"],
                    change["new_serial"],
                )
            )
        print("=" * 100)

    if serial_additions:
        print(
            "\nSerial numbers added to blank/invalid Excel cells: {}".format(
                len(serial_additions)
            )
        )

    print_summary_report(
        summary_rows,
        title="FINAL BLADE / ENCLOSURE SERIAL & MAC COLLECTION SUMMARY",
    )
    write_summary_csv(summary_rows, vars.SCRIPT_ARTIFACT_PREFIXES["get_mac_BL"])

    return 1 if any(row["Status"] in ("Failed", "Partial") for row in summary_rows) else 0


if __name__ == "__main__":
    import sys

    sys.exit(
        run_logged_main(
            main,
            log_prefix=vars.SCRIPT_ARTIFACT_PREFIXES["get_mac_BL"],
            title="BLADE / ENCLOSURE SERIAL & MAC COLLECTION",
        )
    )
