# BUILD_MARKER: OA_CONFIG_CENTRAL_V2_FULL_WORKBOOK_20260828

import os
import sys
import time

import openpyxl
from jinja2 import Environment, FileSystemLoader

from functions import vars
from functions.output_log import run_logged_main
from functions.reporting import (
    make_summary_row,
    print_summary_report,
    write_summary_csv,
)


# ============================================================================
# HELPERS
# ============================================================================

def generate_rack_name(enc_name):
    """Generate rack name from the enclosure physical name."""
    try:
        parts = str(enc_name).strip().split("-")

        if (
            len(parts) >= 3
            and parts[0] == "SV"
        ):
            site = parts[1]
            num_part = parts[2]

            if len(num_part) == 4:
                return "RA-{}-A{}-{}".format(
                    site,
                    num_part[1],
                    num_part[2:4],
                )

    except Exception as exc:
        print(
            "Could not parse rack name for {}: {}"
            .format(
                enc_name,
                exc,
            )
        )

    return "UNKNOWN_RACK"


def is_empty(value):
    """Return True when a value matches the project's global empty values."""
    if value is None:
        return True

    return (
        str(value)
        .strip()
        .upper()
        in vars.INVALID_VALUES
    )


def get_val(row, col_name):
    """Safely retrieve one normalized Excel value."""
    value = row.get(
        col_name,
        "",
    )

    if is_empty(value):
        return ""

    return str(
        value
    ).strip()


def load_certificate(file_path):
    """Load a PEM certificate/template fragment if present."""
    if not os.path.isfile(
        file_path
    ):
        print(
            "Warning: Certificate file not found: {}"
            .format(
                file_path
            )
        )

        return ""

    with open(
        file_path,
        "r",
        encoding="utf-8",
    ) as handle:
        return handle.read().strip()


def load_complete_excel(
    file_path,
    sheet_name,
):
    """
    Load the complete configured worksheet using openpyxl read-only mode.

    Excel row 1 contains headers. Every populated row from Excel row 2 through
    the end of the worksheet is considered. Empty rows are skipped, but they do
    not stop the scan, so enclosures separated by blank areas are still found.
    """
    workbook = openpyxl.load_workbook(
        file_path,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                "Sheet '{}' not found in workbook."
                .format(
                    sheet_name
                )
            )

        worksheet = workbook[
            sheet_name
        ]

        header_values = next(
            worksheet.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not header_values:
            return []

        headers = []

        for index, cell in enumerate(
            header_values
        ):
            if (
                cell is None
                or str(cell).strip() == ""
            ):
                header = (
                    "Unnamed_{}"
                    .format(
                        index
                    )
                )
            else:
                header = str(
                    cell
                ).strip()

            headers.append(
                header
            )

        records = []

        for excel_row, values in enumerate(
            worksheet.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):
            if all(
                value is None
                or str(value).strip() == ""
                for value in values
            ):
                continue

            row_values = list(
                values[:len(headers)]
            )

            if len(row_values) < len(headers):
                row_values.extend(
                    [None]
                    * (
                        len(headers)
                        - len(row_values)
                    )
                )

            record = dict(
                zip(
                    headers,
                    row_values,
                )
            )

            record[
                "_excel_row"
            ] = excel_row

            records.append(
                record
            )

        return records

    finally:
        workbook.close()


def validate_columns(records):
    """Verify that the workbook includes all columns required by OA generation."""
    if not records:
        return []

    present = set(
        records[0].keys()
    )

    return [
        column
        for column in vars.OA_REQUIRED_COLUMNS
        if column not in present
    ]


def get_scope_template_vars(scope_name):
    """
    Convert shared ILO_SCOPE_SETTINGS names to the variable names expected by
    OACONFIG.j2.
    """
    scope = str(
        scope_name
    ).strip().upper()

    scope_data = vars.ILO_SCOPE_SETTINGS.get(
        scope
    )

    if not scope_data:
        return {}

    return {
        "gateway": scope_data.get(
            "GATEWAY",
            "",
        ),
        "mask": scope_data.get(
            "SUBNET_MASK",
            "",
        ),
        "dns1": scope_data.get(
            "PRIMARY_DNS",
            "",
        ),
        "dns2": scope_data.get(
            "SECONDARY_DNS",
            "",
        ),
        "ntp1": scope_data.get(
            "PRIMARY_NTP",
            "",
        ),
        "ntp2": scope_data.get(
            "SECONDARY_NTP",
            "",
        ),
        "domain_controller": scope_data.get(
            "DIRECTORY_SERVER",
            "",
        ),
    }


def new_enclosure(
    enc_name,
    scope,
    excel_row=None,
):
    """Create a complete enclosure dictionary with all Jinja slot keys."""
    enclosure = {
        "enc_name": enc_name,
        "rack_name": generate_rack_name(
            enc_name
        ),
        "scope": scope,
        "_excel_row": excel_row,
        "oa01_name": "",
        "oa01_ip": "",
        "oa02_name": "",
        "oa02_ip": "",
    }

    for slot in range(
        1,
        vars.OA_BLADE_SLOT_COUNT + 1,
    ):
        slot_text = str(
            slot
        ).zfill(
            2
        )

        enclosure[
            "name_blade_{}".format(
                slot_text
            )
        ] = ""

        enclosure[
            "ip_blade_{}".format(
                slot_text
            )
        ] = ""

    for slot in range(
        1,
        vars.OA_INTERCONNECT_SLOT_COUNT + 1,
    ):
        slot_text = str(
            slot
        ).zfill(
            2
        )

        enclosure[
            "name_interconnect_{}".format(
                slot_text
            )
        ] = ""

        enclosure[
            "interconnect_{}".format(
                slot_text
            )
        ] = ""

    return enclosure


def normalize_slot(value):
    """Normalize Excel slot values such as 1.0 to 01."""
    text = str(
        value
    ).strip()

    try:
        return str(
            int(
                float(
                    text
                )
            )
        ).zfill(
            2
        )

    except Exception:
        return text


# ============================================================================
# INVENTORY PROCESSING
# ============================================================================

def build_enclosure_data(records):
    """Build Jinja enclosure objects from the selected Excel rows."""
    enclosures_data = []
    current_enclosure = None

    for row in records:
        excel_row = row.get(
            "_excel_row",
            "Unknown",
        )

        eq_type = get_val(
            row,
            "equipment_type",
        )

        enc_phys_name = get_val(
            row,
            "enclosure_physical_name",
        )

        current_scope = get_val(
            row,
            "scope",
        ).upper()

        if (
            eq_type
            == vars.OA_SERVER_ENCLOSURE_TYPE
        ):
            if not current_scope:
                print(
                    "  [!] WARNING (Row {}): Missing 'scope' for "
                    "Server Enclosure '{}'. Network variables will "
                    "not load.".format(
                        excel_row,
                        enc_phys_name,
                    )
                )

            current_enclosure = new_enclosure(
                enc_phys_name,
                current_scope,
                excel_row=excel_row,
            )

            enclosures_data.append(
                current_enclosure
            )

            continue

        if (
            eq_type
            not in vars.OA_COMPONENT_EQUIPMENT_TYPES
        ):
            continue

        if current_enclosure is None:
            print(
                "  [!] WARNING (Row {}): Component '{}' appears before "
                "a Server Enclosure row inside the configured Excel range; "
                "skipping it.".format(
                    excel_row,
                    eq_type,
                )
            )

            continue

        slot = get_val(
            row,
            "enclosure_slot",
        )

        ip = get_val(
            row,
            "ilo_ip",
        )

        hostname = get_val(
            row,
            "hostname",
        )

        if not slot:
            print(
                "  [!] WARNING (Row {}): Missing 'enclosure_slot' for {} "
                "in enclosure '{}'.".format(
                    excel_row,
                    eq_type,
                    current_enclosure[
                        "enc_name"
                    ],
                )
            )

        if not ip:
            print(
                "  [!] WARNING (Row {}): Missing 'ilo_ip' for {} "
                "(Slot: {}) in enclosure '{}'.".format(
                    excel_row,
                    eq_type,
                    slot,
                    current_enclosure[
                        "enc_name"
                    ],
                )
            )

        if not hostname:
            print(
                "  [!] WARNING (Row {}): Missing 'hostname' for {} "
                "(Slot: {}) in enclosure '{}'.".format(
                    excel_row,
                    eq_type,
                    slot,
                    current_enclosure[
                        "enc_name"
                    ],
                )
            )

        if not slot:
            continue

        slot_clean = normalize_slot(
            slot
        )

        if (
            eq_type
            == vars.OA_ENCLOSURE_OA_TYPE
        ):
            if slot_clean == "01":
                current_enclosure[
                    "oa01_name"
                ] = hostname

                current_enclosure[
                    "oa01_ip"
                ] = ip

            elif slot_clean == "02":
                current_enclosure[
                    "oa02_name"
                ] = hostname

                current_enclosure[
                    "oa02_ip"
                ] = ip

        elif (
            eq_type
            in vars.OA_BLADE_EQUIPMENT_TYPES
        ):
            current_enclosure[
                "name_blade_{}".format(
                    slot_clean
                )
            ] = hostname

            current_enclosure[
                "ip_blade_{}".format(
                    slot_clean
                )
            ] = ip

        elif (
            eq_type
            == vars.OA_ENCLOSURE_SWITCH_TYPE
        ):
            current_enclosure[
                "name_interconnect_{}".format(
                    slot_clean
                )
            ] = hostname

            current_enclosure[
                "interconnect_{}".format(
                    slot_clean
                )
            ] = ip

    return enclosures_data


# ============================================================================
# TEMPLATE RENDERING
# ============================================================================

def render_configs(
    enclosures_data,
    cert_1_content,
    cert_2_content,
):
    """Render one OA configuration file per enclosure."""
    if not os.path.isdir(
        vars.OA_CONFIG_OUTPUT_DIR
    ):
        os.makedirs(
            vars.OA_CONFIG_OUTPUT_DIR
        )

    environment = Environment(
        loader=FileSystemLoader(
            vars.TEMPLATES_DIR
        )
    )

    template = environment.get_template(
        vars.OA_CONFIG_TEMPLATE_FILENAME
    )

    ldap_contexts = list(
        vars.LDAP_USER_CONTEXTS
    )

    while len(
        ldap_contexts
    ) < 3:
        ldap_contexts.append(
            ""
        )

    base_vars = {
        "domain": vars.DOMAIN_NAME,
        "ilo_admin_group": (
            vars.OA_ILO_ADMIN_GROUP
        ),
        "ldap_search_01": (
            ldap_contexts[0]
        ),
        "ldap_search_02": (
            ldap_contexts[1]
        ),
        "ldap_search_03": (
            ldap_contexts[2]
        ),
        "remote_syslog_server": (
            vars.OA_REMOTE_SYSLOG_SERVER
        ),
        "snmp_community": (
            vars.OA_SNMP_COMMUNITY
        ),
        "snmp_contact": (
            vars.OA_SNMP_CONTACT
        ),
        "snmp_location": (
            vars.OA_SNMP_LOCATION
        ),
        "cert_1": cert_1_content,
        "cert_2": cert_2_content,
    }

    print(
        "\nSetting up Jinja environment targeting: {}"
        .format(
            vars.TEMPLATES_DIR
        )
    )

    print(
        "Generating configuration files in: {}"
        .format(
            vars.OA_CONFIG_OUTPUT_DIR
        )
    )

    generated = 0
    overwritten = 0
    skipped = 0
    summary_rows = []

    for enclosure in enclosures_data:
        item_start = time.time()
        if not enclosure[
            "enc_name"
        ]:
            skipped += 1
            summary_rows.append(
                make_summary_row(
                    row=enclosure.get("_excel_row", "-"),
                    item_type="OA Config",
                    name="Unknown enclosure",
                    target="",
                    status="Skipped",
                    time_seconds=time.time() - item_start,
                    details="Missing enclosure name",
                )
            )
            continue

        enc_scope = enclosure.get(
            "scope",
            "",
        )

        network_vars = (
            get_scope_template_vars(
                enc_scope
            )
        )

        if not network_vars:
            print(
                "  -> Skipping generation: {} "
                "(Unknown or missing scope: '{}')"
                .format(
                    enclosure[
                        "enc_name"
                    ],
                    enc_scope,
                )
            )

            skipped += 1
            summary_rows.append(
                make_summary_row(
                    row=enclosure.get("_excel_row", "-"),
                    item_type="OA Config",
                    name=enclosure.get("enc_name", ""),
                    target="",
                    status="Skipped",
                    time_seconds=time.time() - item_start,
                    details=(
                        "Unknown or missing scope: '{}'"
                        .format(enc_scope)
                    ),
                )
            )
            continue

        render_vars = {}

        render_vars.update(
            base_vars
        )

        render_vars.update(
            network_vars
        )

        render_vars.update(
            enclosure
        )

        rendered_config = (
            template.render(
                render_vars
            )
        )

        output_filename = (
            "{}.txt".format(
                enclosure[
                    "enc_name"
                ]
            )
        )

        output_filepath = os.path.join(
            vars.OA_CONFIG_OUTPUT_DIR,
            output_filename,
        )

        file_exists = os.path.exists(
            output_filepath
        )

        with open(
            output_filepath,
            "w",
            encoding="utf-8",
        ) as handle:
            handle.write(
                rendered_config
            )

        if file_exists:
            overwritten += 1

            print(
                "  -> Successfully overwritten: {} [{}]"
                .format(
                    output_filename,
                    enc_scope,
                )
            )

        else:
            generated += 1

            print(
                "  -> Successfully generated: {} [{}]"
                .format(
                    output_filename,
                    enc_scope,
                )
            )

        action = (
            "Overwritten"
            if file_exists
            else "Generated"
        )

        summary_rows.append(
            make_summary_row(
                row=enclosure.get("_excel_row", "-"),
                item_type="OA Config",
                name=enclosure.get("enc_name", ""),
                target=output_filepath,
                status="Successful",
                time_seconds=time.time() - item_start,
                details="{} [{}]".format(action, enc_scope),
            )
        )

    return (
        generated,
        overwritten,
        skipped,
        summary_rows,
    )


# ============================================================================
# MAIN
# ============================================================================

def main():
    print(
        "=" * 100
    )
    print(
        "OA CONFIG GENERATOR"
    )
    print(
        "=" * 100
    )
    print(
        "Resource list       : {}"
        .format(
            vars.RESOURCE_LIST
        )
    )
    print(
        "Sheet               : {}"
        .format(
            vars.SHEET_NAME
        )
    )
    print(
        "Excel rows          : Complete worksheet"
    )
    print(
        "Template            : {}"
        .format(
            vars.OA_CONFIG_TEMPLATE_FILENAME
        )
    )
    print(
        "Output directory    : {}"
        .format(
            vars.OA_CONFIG_OUTPUT_DIR
        )
    )
    print(
        "=" * 100
    )

    if not os.path.isfile(
        vars.RESOURCE_LIST
    ):
        print(
            "ERROR: Resource List not found: {}"
            .format(
                vars.RESOURCE_LIST
            )
        )

        return 1

    template_path = os.path.join(
        vars.TEMPLATES_DIR,
        vars.OA_CONFIG_TEMPLATE_FILENAME,
    )

    if not os.path.isfile(
        template_path
    ):
        print(
            "ERROR: OA Jinja template not found: {}"
            .format(
                template_path
            )
        )

        return 1

    cert_1_content = load_certificate(
        vars.OA_APP_CERT_FILE
    )

    cert_2_content = load_certificate(
        vars.OA_NWS_CERT_FILE
    )

    try:
        records = load_complete_excel(
            vars.RESOURCE_LIST,
            vars.SHEET_NAME,
        )

    except Exception as exc:
        print(
            "ERROR loading Excel workbook: {}"
            .format(
                exc
            )
        )

        return 1

    if not records:
        print(
            "No populated rows found in the configured worksheet."
        )

        return 0

    missing_columns = validate_columns(
        records
    )

    if missing_columns:
        print(
            "\n[!] CRITICAL: Required Excel columns are missing: {}"
            .format(
                ", ".join(
                    missing_columns
                )
            )
        )

        return 1

    enclosures_data = build_enclosure_data(
        records
    )

    if not enclosures_data:
        print(
            "No Server Enclosure rows were found in the configured Excel range."
        )

        return 0

    (
        generated,
        overwritten,
        skipped,
        summary_rows,
    ) = render_configs(
        enclosures_data,
        cert_1_content,
        cert_2_content,
    )

    print(
        "\nDone. Generated: {} | Overwritten: {} | Skipped: {}"
        .format(
            generated,
            overwritten,
            skipped,
        )
    )

    print_summary_report(
        summary_rows,
        title="FINAL OA CONFIG GENERATION SUMMARY",
    )

    write_summary_csv(
        summary_rows,
        vars.SCRIPT_ARTIFACT_PREFIXES[
            "gen_oaconfig"
        ],
    )

    return 0


if __name__ == "__main__":
    sys.exit(
        run_logged_main(
            main,
            log_prefix=vars.SCRIPT_ARTIFACT_PREFIXES[
                "gen_oaconfig"
            ],
            title="OA CONFIG GENERATOR",
        )
    )
