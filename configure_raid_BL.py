import os
import sys
import time
import json
import urllib3
import requests
import openpyxl
import pandas as pd
import concurrent.futures
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from functions import vars

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# --- 1. Centralized Configuration ---
# Reusable settings are maintained in functions/vars.py. Local names below are
# aliases only; configuration values are not duplicated in this script.

EXCEL_PATH = vars.RESOURCE_LIST
SHEET_NAME = vars.SHEET_NAME
START_ROW = vars.START_ROW
END_ROW = vars.END_ROW
EXCEL_EMPTY_ROW_STOP = vars.RAID_EXCEL_EMPTY_ROW_STOP
LOG_DIR = vars.LOG_DIR

USERNAME = vars.ILO_BL_USERNAME
PASSWORD = vars.ILO_BL_PASSWORD
ISO_URL = vars.RAID_BL_ISO_URL

VALID_EQUIPMENT_TYPES = list(vars.RAID_BL_EQUIPMENT_TYPES)
REQUIRED_COLUMNS = list(vars.RAID_BL_REQUIRED_COLUMNS)

CONTROLLER_TIMEOUT_MINUTES = vars.RAID_CONTROLLER_TIMEOUT_MINUTES
RAID_TIMEOUT_MINUTES = vars.RAID_TIMEOUT_MINUTES
POLL_INTERVAL_SECONDS = vars.RAID_POLL_INTERVAL_SECONDS
MAX_WORKERS = vars.RAID_MAX_WORKERS

REQUEST_TIMEOUT_SECONDS = vars.RAID_BL_REQUEST_TIMEOUT_SECONDS
AUTH_RETRIES = vars.RAID_AUTH_RETRIES
AUTH_RETRY_DELAY_SECONDS = vars.RAID_AUTH_RETRY_DELAY_SECONDS
HTTP_RETRY_TOTAL = vars.RAID_HTTP_RETRY_TOTAL
HTTP_RETRY_BACKOFF_FACTOR = vars.RAID_HTTP_RETRY_BACKOFF_FACTOR
HTTP_RETRY_STATUS_CODES = vars.RAID_HTTP_RETRY_STATUS_CODES

REPORT_PREFIX = vars.RAID_BL_REPORT_PREFIX


# --- Helper: Thread-Safe Logger ---
def log(slot, ip, message):
    slot_str = f"{int(slot):02}" if isinstance(slot, (int, float)) else "??"
    print(f"[Slot {slot_str} | {ip}] {message}")

# --- Helper: Fast Excel Loader ---
def load_resource_excel_fast(
    filepath,
    sheet_name,
    start_row,
    end_row,
    empty_row_stop=25,
):
    """
    Read only the configured global Excel row range using openpyxl read-only
    streaming mode. START_ROW / END_ROW are actual Excel row numbers.
    """
    start_row = int(start_row)
    end_row = int(end_row) if end_row is not None else None

    if start_row < 2:
        raise ValueError("START_ROW must be >= 2.")

    if end_row is not None and end_row < start_row:
        raise ValueError("END_ROW must be >= START_ROW.")

    wb = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=True,
    )

    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in the workbook."
            )

        ws = wb[sheet_name]

        header_values = next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not header_values:
            return pd.DataFrame()

        headers = [
            (
                str(cell).strip()
                if cell is not None and str(cell).strip() != ""
                else f"Unnamed_{index}"
            )
            for index, cell in enumerate(header_values)
        ]

        data = []
        empty_count = 0

        for row in ws.iter_rows(
            min_row=start_row,
            max_row=end_row,
            values_only=True,
        ):
            normalized = tuple(row[:len(headers)])

            if len(normalized) < len(headers):
                normalized += (None,) * (
                    len(headers) - len(normalized)
                )

            if all(
                cell is None or str(cell).strip() == ""
                for cell in normalized
            ):
                empty_count += 1

                if empty_count >= empty_row_stop:
                    break

                continue

            empty_count = 0
            data.append(normalized)

        return pd.DataFrame(
            data,
            columns=headers,
        )

    finally:
        wb.close()

# --- Server Processing Class ---
class ServerProcessor:
    def __init__(self, ip, slot, username, password):
        self.ip = ip
        self.slot = slot
        self.base_url = f"https://{ip}"
        self.username = username
        self.password = password
        
        self.session = requests.Session()
        self.session.verify = False
        self.session.headers.update({"Content-Type": "application/json"})
        
        # Native Retry Adapter for HTTP 5xx errors and connection drops
        retries = Retry(
            total=HTTP_RETRY_TOTAL,
            backoff_factor=HTTP_RETRY_BACKOFF_FACTOR,
            status_forcelist=list(HTTP_RETRY_STATUS_CODES),
        )
        adapter = HTTPAdapter(max_retries=retries)
        self.session.mount("https://", adapter)
        self.session.mount("http://", adapter)

    def _request(self, method, url, **kwargs):
        """Wrapper to enforce default timeouts and catch hangs."""
        kwargs.setdefault('timeout', REQUEST_TIMEOUT_SECONDS)
        return self.session.request(method, url, **kwargs)

    def authenticate(self, retries=AUTH_RETRIES):
        for attempt in range(retries):
            try:
                auth_url = f"{self.base_url}/redfish/v1/SessionService/Sessions"
                resp = self._request("POST", auth_url, json={"UserName": self.username, "Password": self.password})
                resp.raise_for_status()
                self.session.headers.update({"X-Auth-Token": resp.headers.get("X-Auth-Token")})
                return
            except Exception as e:
                if attempt == retries - 1:
                    raise Exception(f"Auth failed after {retries} attempts: {e}")
                log(self.slot, self.ip, f"Auth busy or failed. Retrying in {AUTH_RETRY_DELAY_SECONDS}s... (Attempt {attempt+2}/{retries})")
                time.sleep(AUTH_RETRY_DELAY_SECONDS)

    def get_system_info(self):
        resp = self._request("GET", f"{self.base_url}/redfish/v1/Systems/1/")
        resp.raise_for_status()
        return resp.json().get("PowerState", "Unknown")

    def get_storage_inventory(self):
        ctrl_url = f"{self.base_url}/redfish/v1/Systems/1/SmartStorage/ArrayControllers/"
        try:
            controllers = self._request("GET", ctrl_url).json().get("Members", [])
            if not controllers: return None, []
            
            ctrl_uri = controllers[0].get("@odata.id", controllers[0].get("href"))
            ctrl_data = self._request("GET", f"{self.base_url}{ctrl_uri}").json()
            
            links = ctrl_data.get("links", ctrl_data.get("Links", {}))
            ld_href = links.get("LogicalDrives", {}).get("href")
            logical_drives = []
            if ld_href:
                ld_members = self._request("GET", f"{self.base_url}{ld_href}").json().get("Members", [])
                for ld in ld_members:
                    ld_uri = ld.get("@odata.id", ld.get("href"))
                    logical_drives.append(self._request("GET", f"{self.base_url}{ld_uri}").json())
            return ctrl_uri, logical_drives
        except Exception:
            return None, []

    def mount_virtual_media(self, iso_url):
        log(self.slot, self.ip, "Configuring Virtual Media...")
        vm_members = self._request("GET", f"{self.base_url}/redfish/v1/Managers/1/VirtualMedia/").json().get("Members", [])
        
        cd_vm_uri, insert_uri, eject_uri = None, None, None
        
        for member in vm_members:
            vm_uri = member.get("@odata.id", member.get("href"))
            vm_data = self._request("GET", f"{self.base_url}{vm_uri}").json()
            types = vm_data.get("MediaTypes", [])
            
            if "CD" in types or "DVD" in types or str(vm_data.get("Id")) == "2":
                cd_vm_uri = vm_uri
                all_actions = {}
                all_actions.update(vm_data.get("Actions", {}))
                all_actions.update(vm_data.get("Oem", {}).get("Hp", {}).get("Actions", {}))
                
                for a_name, a_info in all_actions.items():
                    if "Insert" in a_name: insert_uri = a_info.get("target")
                    if "Eject" in a_name: eject_uri = a_info.get("target")
                
                if vm_data.get("Inserted") or vm_data.get("Image"):
                    if eject_uri: self._request("POST", f"{self.base_url}{eject_uri}", json={})
                break

        if not insert_uri: raise Exception("Virtual CD/DVD action not found.")

        payload = {"Image": iso_url}
        resp = self._request("POST", f"{self.base_url}{insert_uri}", json=payload)
        
        if resp.status_code >= 400:
            resp = self._request("PATCH", f"{self.base_url}{cd_vm_uri}", json=payload)
            
        resp.raise_for_status()
        log(self.slot, self.ip, "ISO Mounted Successfully.")
        return eject_uri

    def eject_virtual_media(self, eject_uri):
        if not eject_uri: return
        try:
            self._request("POST", f"{self.base_url}{eject_uri}", json={})
        except: pass

    def set_boot_to_bios(self):
        self._request("PATCH", f"{self.base_url}/redfish/v1/Systems/1/", json={"Boot": {"BootSourceOverrideTarget": "BiosSetup"}})

    def set_boot_to_cdrom(self):
        self._request("PATCH", f"{self.base_url}/redfish/v1/Systems/1/", json={"Boot": {"BootSourceOverrideTarget": "Cd"}})

    def set_boot_to_normal(self):
        self._request("PATCH", f"{self.base_url}/redfish/v1/Systems/1/", json={"Boot": {"BootSourceOverrideTarget": "None"}})

    def reboot_server(self, power_state):
        reset_type = "ForceRestart" if power_state == "On" else "On"
        self._request("POST", f"{self.base_url}/redfish/v1/Systems/1/Actions/ComputerSystem.Reset", json={"ResetType": reset_type})

    def poll_for_new_raid(self, baseline_ids, timeout_minutes):
        max_attempts = max(1, int((timeout_minutes * 60) / POLL_INTERVAL_SECONDS)) 
        for i in range(max_attempts):
            time.sleep(POLL_INTERVAL_SECONDS)
            try:
                ctrl_uri, current_ld = self.get_storage_inventory()
                if not ctrl_uri: continue
                current_ids = [ld.get("LogicalDriveName", "Unknown") for ld in current_ld]
                
                if current_ids and set(current_ids) != set(baseline_ids):
                    log(self.slot, self.ip, "SUCCESS: New RAID volume detected!")
                    return True
            except:
                pass
            if (i % 2) == 0: log(self.slot, self.ip, f"Polling array... (Attempt {i+1}/{max_attempts})")
        return False

# --- 2. The Parallel Worker Function ---
def process_blade(server_data, overwrite_raid, ctrl_timeout, raid_timeout):
    ip = server_data['ilo_ip']
    slot = server_data['enclosure_slot']
    
    server = ServerProcessor(ip, slot, USERNAME, PASSWORD)
    eject_uri = None
    
    try:
        log(slot, ip, "Starting configuration process...")
        server.authenticate()
        power_state = server.get_system_info()
        
        # Initial check for controller
        ctrl_uri, baseline_ld = server.get_storage_inventory()
        
        # If no controller, boot to BIOS and wait for it
        if not ctrl_uri:
            log(slot, ip, "Controller hidden. Powering on / Rebooting to POST to initialize storage...")
            server.set_boot_to_bios()
            server.reboot_server(power_state)
            
            max_attempts = max(1, int((ctrl_timeout * 60) / POLL_INTERVAL_SECONDS))
            for i in range(max_attempts):
                time.sleep(POLL_INTERVAL_SECONDS)
                ctrl_uri, baseline_ld = server.get_storage_inventory()
                if ctrl_uri:
                    log(slot, ip, "Storage controller detected successfully.")
                    break
                if (i % 2) == 0:
                    log(slot, ip, f"Waiting for controller POST... (Attempt {i+1}/{max_attempts})")
                    
            if not ctrl_uri:
                return {"slot": slot, "ip": ip, "status": "Failed", "reason": f"Controller failed to initialize after {ctrl_timeout} mins"}

        baseline_ids = [ld.get("LogicalDriveName", "Unknown") for ld in baseline_ld]
        
        # Raid Evaluation Logic
        if baseline_ld:
            if not overwrite_raid:
                log(slot, ip, "Existing RAID found. Skipping (Overwrite=No).")
                server.set_boot_to_normal()
                server.reboot_server("On") 
                return {"slot": slot, "ip": ip, "status": "Skipped", "reason": "Existing RAID found"}
            else:
                log(slot, ip, f"Existing RAID found ({len(baseline_ld)} vols). Overwriting...")
        else:
            log(slot, ip, "No existing RAID. Proceeding with creation...")
            
        # Execute ISO mount and format
        eject_uri = server.mount_virtual_media(ISO_URL)
        server.set_boot_to_cdrom()
        server.reboot_server("On") 
        
        success = server.poll_for_new_raid(baseline_ids, raid_timeout)
        
        server.eject_virtual_media(eject_uri)
        server.set_boot_to_normal()
        
        if success:
            server.reboot_server("On")
            return {"slot": slot, "ip": ip, "status": "Successful", "reason": f"{vars.RAID_DISPLAY_NAME} Created Successfully"}
        else:
            return {"slot": slot, "ip": ip, "status": "Failed", "reason": "Timeout waiting for array rebuild"}
            
    except Exception as e:
        try:
            if eject_uri: server.eject_virtual_media(eject_uri)
            server.set_boot_to_normal()
        except: pass
        return {"slot": slot, "ip": ip, "status": "Failed", "reason": str(e)}
    finally:
        # ALWAYS close the connection to prevent iLO memory leaks and locks
        server.session.close()

# --- 3. Main Orchestration Logic ---
def main():
    excel_path = EXCEL_PATH
    
    if not os.path.exists(excel_path):
        print(f"ERROR: Could not find Excel file at {excel_path}")
        sys.exit(1)
        
    try:
        print(f"Loading data from {excel_path} (Sheet: '{SHEET_NAME}')...")
        end_text = END_ROW if END_ROW is not None else "end of data"
        print(
            f"Excel row range: {START_ROW} through "
            f"{end_text} (inclusive)"
        )
        print(
            "Optimized load enabled: within the selected range, "
            f"reading stops after {EXCEL_EMPTY_ROW_STOP} "
            "consecutive empty rows."
        )
        df = load_resource_excel_fast(
            excel_path,
            SHEET_NAME,
            START_ROW,
            END_ROW,
            EXCEL_EMPTY_ROW_STOP,
        )
    except Exception as e:
        print(f"Failed to read Excel file: {e}")
        sys.exit(1)
        
    target_enclosure = input("\nEnter the Enclosure Name to process: ").strip()
    
    # Avoid KeyErrors if file is totally empty or missing columns
    if 'enclosure_physical_name' not in df.columns or 'equipment_type' not in df.columns:
        print("ERROR: Excel file does not contain required identifier columns.")
        sys.exit(1)
        
    df_filtered = df[df['enclosure_physical_name'] == target_enclosure]
    if df_filtered.empty:
        print(f"\nNo servers found for enclosure: '{target_enclosure}'. Exiting.")
        sys.exit(0)
        
    df_filtered = df_filtered[df_filtered['equipment_type'].isin(VALID_EQUIPMENT_TYPES)]
    
    servers_to_process = []
    final_report = [] 
    
    for index, row in df_filtered.iterrows():
        raw_slot = row.get('enclosure_slot')
        raw_ip = row.get('ilo_ip')
        
        slot_num = int(raw_slot) if pd.notna(raw_slot) and str(raw_slot).isdigit() else 999 
        ip_str = str(raw_ip).strip() if pd.notna(raw_ip) else "Unknown"
        
        missing = [col for col in REQUIRED_COLUMNS if col not in row or pd.isna(row[col]) or str(row[col]).strip() == '']
        
        if missing:
            final_report.append({
                "slot": slot_num,
                "ip": ip_str,
                "status": "Skipped",
                "reason": f"Missing column values: {', '.join(missing)}"
            })
            log(slot_num, ip_str, f"SKIPPING: Missing required columns: {missing}")
        else:
            servers_to_process.append({
                "enclosure_slot": slot_num,
                "ilo_ip": ip_str
            })
            
    total_found = len(servers_to_process)
    print(f"\nTotal valid Gen9 Blade Servers found for processing: {total_found}")
    
    if total_found == 0:
        print("No valid servers left to process. Exiting.")
        sys.exit(0)
        
    user_input = input("\nDo you want to OVERWRITE servers with already configured RAID volumes? (yes/no): ").strip().lower()
    overwrite_raid = (user_input == 'yes')
    
    print("\n" + "="*80)
    print("STARTING PARALLEL BLADE RAID CONFIGURATION")
    print("="*80)
    
    # Cap parallel execution to prevent dropping requests on the switch or running out of RAM
    safe_workers = min(MAX_WORKERS, total_found)
    
    # Thread Pool Execution
    with concurrent.futures.ThreadPoolExecutor(max_workers=safe_workers) as executor:
        futures = {
            executor.submit(process_blade, srv, overwrite_raid, CONTROLLER_TIMEOUT_MINUTES, RAID_TIMEOUT_MINUTES): srv 
            for srv in servers_to_process
        }
        
        for future in concurrent.futures.as_completed(futures):
            res = future.result()
            final_report.append({
                "slot": res["slot"],
                "ip": res["ip"],
                "status": res["status"],
                "reason": res["reason"]
            })
            log(res["slot"], res["ip"], f"FINISHED - Status: {res['status']}")

    # --- 4. Final Sorted Summary Report ---
    final_report = sorted(final_report, key=lambda x: x["slot"])
    
    print("\n\n" + "="*100)
    print(f"FINAL EXECUTION REPORT: Enclosure '{target_enclosure}'")
    print("="*100)
    print(f"{'SLOT':<6} | {'iLO IP ADDRESS':<16} | {'STATUS':<12} | {'REASON / DETAILS'}")
    print("-" * 100)
    
    for srv in final_report:
        slot_disp = f"{srv['slot']:02}" if srv['slot'] != 999 else "??"
        ip_disp = srv['ip']
        status_disp = srv['status']
        reason_disp = srv['reason']
        print(f"{slot_disp:<6} | {ip_disp:<16} | {status_disp:<12} | {reason_disp}")
        
    print("="*100 + "\n")

    # --- Write execution report to CSV for auditing ---
    if final_report:
        # Create log directory relative to the script's location
        log_dir = LOG_DIR
        os.makedirs(log_dir, exist_ok=True)
        
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        report_filename = (
            f"{REPORT_PREFIX}_"
            f"{target_enclosure.replace('/', '_')}_"
            f"rows_{START_ROW}-"
            f"{END_ROW if END_ROW is not None else 'end'}_"
            f"{timestamp}.csv"
        )
        report_path = os.path.join(log_dir, report_filename)
        
        report_df = pd.DataFrame(final_report)
        report_df = report_df[["slot", "ip", "status", "reason"]]
        report_df.to_csv(report_path, index=False)
        
        print(f"Summary report saved to: {report_path}\n")

if __name__ == "__main__":
    main()