import os
import sys
import time
import re
import urllib3
import requests
import openpyxl
import pandas as pd
import concurrent.futures

# BUILD_MARKER: RAID_RM_HPE_MR_INITIALIZE_V8_20260830

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from functions import vars
from functions.output_log import run_logged_main
from functions.reporting import (
    make_summary_row,
    print_summary_report,
    write_summary_csv,
)


urllib3.disable_warnings(
    urllib3.exceptions.InsecureRequestWarning
)


SCRIPT_BUILD = "RAID_RM_HPE_MR_INITIALIZE_V8_20260830"


# =============================================================================
# 1. CENTRALIZED CONFIGURATION
# =============================================================================

# Reusable settings are maintained in functions/vars.py. Local names below are
# aliases only; configuration values are not duplicated in this script.

EXCEL_PATH = vars.RESOURCE_LIST
SHEET_NAME = vars.SHEET_NAME
EXCEL_EMPTY_ROW_STOP = vars.RAID_EXCEL_EMPTY_ROW_STOP
START_ROW = vars.START_ROW
END_ROW = vars.END_ROW
LOG_DIR = vars.LOG_DIR

USERNAME = vars.ILO_RM_USERNAME
PASSWORD = vars.ILO_RM_PASSWORD

REQUEST_TIMEOUT_SECONDS = vars.RAID_RM_REQUEST_TIMEOUT_SECONDS
WRITE_TIMEOUT_SECONDS = vars.RAID_RM_WRITE_TIMEOUT_SECONDS

VALID_EQUIPMENT_TYPES = list(vars.RAID_RM_EQUIPMENT_TYPES)
REQUIRED_COLUMNS = list(vars.RAID_RM_REQUIRED_COLUMNS)

RAID_LEVEL = vars.RAID_LEVEL
RAID_DISPLAY_NAME = vars.RAID_DISPLAY_NAME
RAID_DATA_DRIVE_COUNT = vars.RAID_DATA_DRIVE_COUNT
RAID_CAPACITY_GIB = vars.RAID_CAPACITY_GIB

MAX_WORKERS = vars.RAID_MAX_WORKERS

CONTROLLER_TIMEOUT_MINUTES = vars.RAID_CONTROLLER_TIMEOUT_MINUTES
RAID_TIMEOUT_MINUTES = vars.RAID_TIMEOUT_MINUTES
POLL_INTERVAL_SECONDS = vars.RAID_POLL_INTERVAL_SECONDS

AUTH_RETRIES = vars.RAID_AUTH_RETRIES
AUTH_RETRY_DELAY_SECONDS = vars.RAID_AUTH_RETRY_DELAY_SECONDS
HTTP_RETRY_TOTAL = vars.RAID_HTTP_RETRY_TOTAL
HTTP_RETRY_BACKOFF_FACTOR = vars.RAID_HTTP_RETRY_BACKOFF_FACTOR
HTTP_RETRY_STATUS_CODES = vars.RAID_HTTP_RETRY_STATUS_CODES

REBOOT_DETECTION_TIMEOUT_SECONDS = (
    vars.RAID_RM_REBOOT_DETECTION_TIMEOUT_SECONDS
)
REBOOT_DETECTION_POLL_SECONDS = (
    vars.RAID_RM_REBOOT_DETECTION_POLL_SECONDS
)
POWER_OFF_TIMEOUT_SECONDS = vars.RAID_RM_POWER_OFF_TIMEOUT_SECONDS
POWER_ON_INITIAL_WAIT_SECONDS = (
    vars.RAID_RM_POWER_ON_INITIAL_WAIT_SECONDS
)

BOOT_OVERRIDE_POST_RECOVERY = (
    vars.RAID_RM_BOOT_OVERRIDE_POST_RECOVERY
)
BOOT_OVERRIDE_RETRY_INTERVAL_SECONDS = (
    vars.RAID_RM_BOOT_OVERRIDE_RETRY_INTERVAL_SECONDS
)
BOOT_OVERRIDE_RETRY_TIMEOUT_SECONDS = (
    vars.RAID_RM_BOOT_OVERRIDE_RETRY_TIMEOUT_SECONDS
)

DMTF_APPLY_TIMEOUT_SECONDS = (
    vars.RAID_RM_DMTF_APPLY_TIMEOUT_SECONDS
)
DMTF_POLL_INTERVAL_SECONDS = (
    vars.RAID_RM_DMTF_POLL_INTERVAL_SECONDS
)
INITIALIZE_OVERWRITTEN_VOLUME = (
    vars.RAID_RM_INITIALIZE_OVERWRITTEN_VOLUME
)
INITIALIZE_METHOD = vars.RAID_RM_INITIALIZE_METHOD
INITIALIZE_TYPE = vars.RAID_RM_INITIALIZE_TYPE
INITIALIZE_ACTION_DISCOVERY_TIMEOUT_SECONDS = (
    vars.RAID_RM_INITIALIZE_ACTION_DISCOVERY_TIMEOUT_SECONDS
)
INITIALIZE_POLL_INTERVAL_SECONDS = (
    vars.RAID_RM_INITIALIZE_POLL_INTERVAL_SECONDS
)
INITIALIZE_TIMEOUT_SECONDS = (
    vars.RAID_RM_INITIALIZE_TIMEOUT_SECONDS
)
INITIALIZE_SETTLE_SECONDS = (
    vars.RAID_RM_INITIALIZE_SETTLE_SECONDS
)

REPORT_PREFIX = vars.RAID_RM_REPORT_PREFIX


# =============================================================================
# 2. HELPERS
# =============================================================================

def log(row_number, ip, message):

    row_str = (
        f"{int(row_number):04}"
        if isinstance(row_number, (int, float))
        else "????"
    )

    print(
        f"[Row {row_str} | {ip}] {message}",
        flush=True,
    )


def natural_sort_key(value):

    return [
        int(part) if part.isdigit() else part.lower()
        for part in re.split(
            r"(\d+)",
            str(value),
        )
    ]


def get_member_uri(member):

    if not isinstance(member, dict):
        return None

    return (
        member.get("@odata.id")
        or member.get("href")
    )


def get_link_uri(data, key):

    if not isinstance(data, dict):
        return None

    # Standard Links
    for links_key in (
        "Links",
        "links",
    ):

        links = data.get(
            links_key,
            {},
        )

        if not isinstance(links, dict):
            continue

        value = links.get(key)

        if isinstance(value, dict):

            return (
                value.get("@odata.id")
                or value.get("href")
            )

        if isinstance(value, str):
            return value

    # Direct property
    value = data.get(key)

    if isinstance(value, dict):

        return (
            value.get("@odata.id")
            or value.get("href")
        )

    if isinstance(value, str):
        return value

    return None


# =============================================================================
# 3. EXCEL LOADER
# =============================================================================

def load_resource_excel_fast(
    filepath,
    sheet_name,
    start_row=2,
    end_row=None,
    empty_row_stop=25,
):

    if start_row is None:
        start_row = 2

    start_row = int(start_row)

    if start_row < 2:

        raise ValueError(
            "START_ROW must be >= 2."
        )

    if end_row is not None:

        end_row = int(end_row)

        if end_row < start_row:

            raise ValueError(
                "END_ROW must be >= START_ROW."
            )

    wb = openpyxl.load_workbook(
        filepath,
        read_only=True,
        data_only=True,
    )

    try:

        if sheet_name not in wb.sheetnames:

            raise ValueError(
                f"Sheet '{sheet_name}' not found."
            )

        ws = wb[sheet_name]

        header_values = next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True,
            ),
            None,
        )

        if not header_values:
            return pd.DataFrame()

        headers = [

            (
                str(cell).strip()
                if cell is not None
                and str(cell).strip() != ""
                else f"Unnamed_{index}"
            )

            for index, cell
            in enumerate(header_values)
        ]

        data = []
        excel_rows = []

        empty_count = 0

        for excel_row, row in enumerate(
            ws.iter_rows(
                min_row=2,
                values_only=True,
            ),
            start=2,
        ):

            if excel_row < start_row:
                continue

            if (
                end_row is not None
                and excel_row > end_row
            ):
                break

            is_empty = all(
                cell is None
                or str(cell).strip() == ""
                for cell in row
            )

            if is_empty:

                empty_count += 1

                if empty_count >= empty_row_stop:
                    break

                continue

            empty_count = 0

            normalized = tuple(
                row[:len(headers)]
            )

            if len(normalized) < len(headers):

                normalized += (
                    None,
                ) * (
                    len(headers)
                    - len(normalized)
                )

            data.append(normalized)
            excel_rows.append(excel_row)

        df = pd.DataFrame(
            data,
            columns=headers,
        )

        if not df.empty:

            df.insert(
                0,
                "_excel_row",
                excel_rows,
            )

        return df

    finally:

        wb.close()


# =============================================================================
# 4. SERVER PROCESSOR
# =============================================================================

class ServerProcessor:

    def __init__(
        self,
        ip,
        row_number,
        username,
        password,
    ):

        self.ip = ip
        self.row_number = row_number

        self.base_url = (
            f"https://{ip}"
        )

        self.username = username
        self.password = password

        self.system_uri = None

        self.session = requests.Session()

        self.session.verify = False

        self.session.headers.update(
            {
                "Accept": "application/json",
                "Content-Type": "application/json",
                "OData-Version": "4.0",
            }
        )

        retries = Retry(
            total=HTTP_RETRY_TOTAL,
            connect=HTTP_RETRY_TOTAL,
            read=HTTP_RETRY_TOTAL,
            status=HTTP_RETRY_TOTAL,
            backoff_factor=HTTP_RETRY_BACKOFF_FACTOR,
            status_forcelist=list(
                HTTP_RETRY_STATUS_CODES
            ),
            allowed_methods=frozenset(
                [
                    "GET",
                    "HEAD",
                    "OPTIONS",
                ]
            ),
            raise_on_status=False,
        )

        adapter = HTTPAdapter(
            max_retries=retries,
        )

        self.session.mount(
            "https://",
            adapter,
        )


    # =========================================================================
    # HTTP
    # =========================================================================

    def _request(
        self,
        method,
        uri,
        **kwargs,
    ):

        kwargs.setdefault(
            "timeout",
            REQUEST_TIMEOUT_SECONDS,
        )

        if uri.startswith("http"):

            url = uri

        else:

            url = (
                f"{self.base_url}"
                f"{uri}"
            )

        return self.session.request(
            method,
            url,
            **kwargs,
        )


    def _get_json(
        self,
        uri,
    ):

        response = self._request(
            "GET",
            uri,
        )

        response.raise_for_status()

        return response.json()


    # =========================================================================
    # AUTH
    # =========================================================================

    def authenticate(
        self,
        retries=AUTH_RETRIES,
    ):

        for attempt in range(
            1,
            retries + 1,
        ):

            try:

                response = self._request(
                    "POST",
                    (
                        "/redfish/v1/"
                        "SessionService/"
                        "Sessions/"
                    ),
                    json={
                        "UserName":
                            self.username,
                        "Password":
                            self.password,
                    },
                )

                response.raise_for_status()

                token = (
                    response.headers.get(
                        "X-Auth-Token"
                    )
                )

                if not token:

                    raise Exception(
                        "iLO did not return X-Auth-Token."
                    )

                self.session.headers.update(
                    {
                        "X-Auth-Token":
                            token
                    }
                )

                self.discover_system()

                return

            except Exception as exc:

                if attempt >= retries:

                    raise Exception(
                        "Authentication failed "
                        f"after {retries} attempts: "
                        f"{exc}"
                    )

                log(
                    self.row_number,
                    self.ip,
                    (
                        "Authentication failed/busy. "
                        f"Retrying in {AUTH_RETRY_DELAY_SECONDS} seconds... "
                        f"(Attempt "
                        f"{attempt + 1}/{retries})"
                    ),
                )

                time.sleep(AUTH_RETRY_DELAY_SECONDS)


    # =========================================================================
    # SYSTEM DISCOVERY
    # =========================================================================

    def discover_system(self):

        root = self._get_json(
            "/redfish/v1/"
        )

        systems_uri = (
            root.get(
                "Systems",
                {},
            )
            .get(
                "@odata.id"
            )
        )

        if not systems_uri:

            raise Exception(
                "Systems resource not found."
            )

        systems = self._get_json(
            systems_uri
        )

        members = systems.get(
            "Members",
            [],
        )

        if not members:

            raise Exception(
                "No ComputerSystem found."
            )

        self.system_uri = (
            get_member_uri(
                members[0]
            )
        )

        if not self.system_uri:

            raise Exception(
                "Unable to determine ComputerSystem URI."
            )


    def get_system_info(self):

        if not self.system_uri:
            self.discover_system()

        data = self._get_json(
            self.system_uri
        )

        oem = (
            data.get(
                "Oem",
                {},
            )
            or {}
        )

        hpe = (
            oem.get(
                "Hpe",
                oem.get(
                    "Hp",
                    {},
                ),
            )
            or {}
        )

        return {
            "PowerState":
                data.get(
                    "PowerState",
                    "Unknown",
                ),

            "PostState":
                hpe.get(
                    "PostState",
                    "Unknown",
                ),

            "Data":
                data,
        }


    # =========================================================================
    # SMART STORAGE INVENTORY
    #
    # IMPORTANT:
    #
    # This is the READ/INVENTORY interface.
    #
    # Existing RAID volumes are discovered here even when
    # SmartStorageConfig is unavailable.
    # =========================================================================

    def get_smartstorage_inventory(self):

        if not self.system_uri:
            self.discover_system()

        collection_uri = (
            self.system_uri.rstrip("/")
            + "/SmartStorage/ArrayControllers/"
        )

        try:

            response = self._request(
                "GET",
                collection_uri,
            )

            if response.status_code == 404:

                return []

            response.raise_for_status()

            members = (
                response.json()
                .get(
                    "Members",
                    [],
                )
            )

        except Exception:

            return []

        controllers = []

        for member in members:

            controller_uri = (
                get_member_uri(
                    member
                )
            )

            if not controller_uri:
                continue

            try:

                controller_data = (
                    self._get_json(
                        controller_uri
                    )
                )

            except Exception:

                continue

            logical_drives = []
            physical_drives = []

            # -----------------------------------------------------------------
            # Logical Drives
            # -----------------------------------------------------------------

            logical_uri = (
                get_link_uri(
                    controller_data,
                    "LogicalDrives",
                )
            )

            if not logical_uri:

                logical_uri = (
                    controller_uri.rstrip("/")
                    + "/LogicalDrives/"
                )

            try:

                logical_collection = (
                    self._get_json(
                        logical_uri
                    )
                )

                for ld_member in (
                    logical_collection.get(
                        "Members",
                        [],
                    )
                ):

                    ld_uri = (
                        get_member_uri(
                            ld_member
                        )
                    )

                    if not ld_uri:
                        continue

                    try:

                        ld_data = (
                            self._get_json(
                                ld_uri
                            )
                        )

                        ld_data[
                            "_uri"
                        ] = ld_uri

                        logical_drives.append(
                            ld_data
                        )

                    except Exception:

                        pass

            except Exception:

                pass

            # -----------------------------------------------------------------
            # Physical Drives
            # -----------------------------------------------------------------

            drive_uri = (
                get_link_uri(
                    controller_data,
                    "DiskDrives",
                )
            )

            if not drive_uri:

                drive_uri = (
                    controller_uri.rstrip("/")
                    + "/DiskDrives/"
                )

            try:

                drive_collection = (
                    self._get_json(
                        drive_uri
                    )
                )

                for drive_member in (
                    drive_collection.get(
                        "Members",
                        [],
                    )
                ):

                    member_uri = (
                        get_member_uri(
                            drive_member
                        )
                    )

                    if not member_uri:
                        continue

                    try:

                        drive_data = (
                            self._get_json(
                                member_uri
                            )
                        )

                        drive_data[
                            "_uri"
                        ] = member_uri

                        physical_drives.append(
                            drive_data
                        )

                    except Exception:

                        pass

            except Exception:

                pass

            controllers.append(
                {
                    "uri":
                        controller_uri,

                    "data":
                        controller_data,

                    "logical_drives":
                        logical_drives,

                    "physical_drives":
                        physical_drives,
                }
            )

        return controllers


    # =========================================================================
    # DMTF REDFISH STORAGE MODEL
    #
    # HPE Gen10/Gen10+ controllers can expose the standard DMTF /Storage model
    # even when the legacy HPE SmartStorage/SmartStorageConfig OEM resources
    # are unavailable. This implementation uses DMTF Storage for both
    # inventory and RAID writes when POST capability is advertised.
    # =========================================================================

    @staticmethod
    def _volume_raid_type(volume_data):

        return str(
            volume_data.get(
                "RAIDType",
                volume_data.get(
                    "Raid",
                    volume_data.get(
                        "VolumeType",
                        "",
                    ),
                ),
            )
            or ""
        ).strip()


    @staticmethod
    def _volume_drive_uris(volume_data):

        links = (
            volume_data.get(
                "Links",
                {},
            )
            or {}
        )

        drives = (
            links.get(
                "Drives",
                [],
            )
            or []
        )

        result = []

        for drive in drives:

            uri = (
                get_member_uri(
                    drive
                )
            )

            if uri:

                result.append(
                    uri
                )

        return sorted(
            set(
                result
            ),
            key=natural_sort_key,
        )


    @staticmethod
    def _is_real_raid_volume(volume_data):

        raid_type = (
            ServerProcessor
            ._volume_raid_type(
                volume_data
            )
            .lower()
        )

        # RAIDType=None is generally an HBA/JBOD representation, not a
        # configured RAID logical volume.
        return raid_type not in (
            "",
            "none",
            "jbod",
            "nonredundant",
        )


    def _get_collection_response(
        self,
        collection_uri,
    ):

        response = self._request(
            "GET",
            collection_uri,
        )

        if response.status_code == 404:

            return (
                response,
                {},
            )

        response.raise_for_status()

        try:

            data = response.json()

        except Exception:

            data = {}

        return (
            response,
            data,
        )


    def _get_dmtf_controller_label(
        self,
        storage_uri,
        storage_data,
    ):

        labels = []

        for key in (
            "Name",
            "Id",
            "Description",
        ):

            value = (
                storage_data.get(
                    key
                )
            )

            if value:

                labels.append(
                    str(value)
                )

        storage_controllers = (
            storage_data.get(
                "StorageControllers"
            )
        )

        if isinstance(
            storage_controllers,
            list,
        ):

            for controller in storage_controllers:

                if not isinstance(
                    controller,
                    dict,
                ):

                    continue

                for key in (
                    "Name",
                    "Model",
                    "Manufacturer",
                    "FirmwareVersion",
                ):

                    value = (
                        controller.get(
                            key
                        )
                    )

                    if value:

                        labels.append(
                            str(value)
                        )

        controllers_uri = (
            get_link_uri(
                storage_data,
                "Controllers",
            )
        )

        if not controllers_uri:

            controllers_uri = (
                storage_uri.rstrip("/")
                + "/Controllers/"
            )

        try:

            controller_collection = (
                self._get_json(
                    controllers_uri
                )
            )

            for member in (
                controller_collection.get(
                    "Members",
                    [],
                )
                or []
            ):

                controller_uri = (
                    get_member_uri(
                        member
                    )
                )

                if not controller_uri:

                    continue

                try:

                    controller_data = (
                        self._get_json(
                            controller_uri
                        )
                    )

                except Exception:

                    continue

                for key in (
                    "Name",
                    "Model",
                    "Manufacturer",
                    "FirmwareVersion",
                ):

                    value = (
                        controller_data.get(
                            key
                        )
                    )

                    if value:

                        labels.append(
                            str(value)
                        )

        except Exception:

            pass

        unique = []

        for value in labels:

            if value not in unique:

                unique.append(
                    value
                )

        if unique:

            return " | ".join(
                unique
            )

        return storage_uri


    def _get_dmtf_drive_uris(
        self,
        storage_uri,
        storage_data,
    ):

        drive_uris = []

        direct_drives = (
            storage_data.get(
                "Drives",
                [],
            )
            or []
        )

        if isinstance(
            direct_drives,
            list,
        ):

            for drive in direct_drives:

                uri = (
                    get_member_uri(
                        drive
                    )
                )

                if uri:

                    drive_uris.append(
                        uri
                    )

        drives_uri = (
            get_link_uri(
                storage_data,
                "Drives",
            )
        )

        if not drives_uri:

            drives_uri = (
                storage_uri.rstrip("/")
                + "/Drives/"
            )

        try:

            drive_collection = (
                self._get_json(
                    drives_uri
                )
            )

            for member in (
                drive_collection.get(
                    "Members",
                    [],
                )
                or []
            ):

                uri = (
                    get_member_uri(
                        member
                    )
                )

                if uri:

                    drive_uris.append(
                        uri
                    )

        except Exception:

            pass

        return sorted(
            set(
                drive_uris
            ),
            key=natural_sort_key,
        )


    def _probe_dmtf_volume_write_support(
        self,
        volumes_uri,
        collection_response,
    ):

        methods = set()

        for value in re.split(
            r"[\s,]+",
            str(
                collection_response.headers.get(
                    "Allow",
                    "",
                )
            ).upper(),
        ):

            if value:

                methods.add(
                    value
                )

        # Some controller firmware returns Allow only for HEAD or OPTIONS.
        for method in (
            "HEAD",
            "OPTIONS",
        ):

            try:

                response = (
                    self._request(
                        method,
                        volumes_uri,
                    )
                )

                for value in re.split(
                    r"[\s,]+",
                    str(
                        response.headers.get(
                            "Allow",
                            "",
                        )
                    ).upper(),
                ):

                    if value:

                        methods.add(
                            value
                        )

            except Exception:

                pass

        capabilities_uri = (
            volumes_uri.rstrip("/")
            + "/Capabilities"
        )

        capabilities = None
        capabilities_status = None

        try:

            response = (
                self._request(
                    "GET",
                    capabilities_uri,
                )
            )

            capabilities_status = (
                response.status_code
            )

            if response.status_code == 200:

                capabilities = (
                    response.json()
                )

        except Exception:

            pass

        # HPE documents Volumes/Capabilities as existing only on writable
        # storage devices. Either this resource or Allow: POST is sufficient.
        writable = (
            capabilities is not None
            or "POST" in methods
        )

        return {
            "writable":
                writable,

            "methods":
                sorted(
                    methods
                ),

            "capabilities_uri":
                capabilities_uri,

            "capabilities_status":
                capabilities_status,

            "capabilities":
                capabilities,
        }


    def discover_dmtf_storage_targets(
        self,
    ):

        if not self.system_uri:

            self.discover_system()

        collection_uri = (
            self.system_uri.rstrip("/")
            + "/Storage/"
        )

        try:

            response = (
                self._request(
                    "GET",
                    collection_uri,
                )
            )

            if response.status_code == 404:

                return []

            response.raise_for_status()

            members = (
                response.json()
                .get(
                    "Members",
                    [],
                )
                or []
            )

        except Exception:

            return []

        targets = []

        for member in members:

            storage_uri = (
                get_member_uri(
                    member
                )
            )

            if not storage_uri:

                continue

            try:

                storage_data = (
                    self._get_json(
                        storage_uri
                    )
                )

            except Exception:

                continue

            volumes_uri = (
                get_link_uri(
                    storage_data,
                    "Volumes",
                )
            )

            if not volumes_uri:

                volumes_uri = (
                    storage_uri.rstrip("/")
                    + "/Volumes/"
                )

            try:

                (
                    volume_response,
                    volume_collection,
                ) = (
                    self._get_collection_response(
                        volumes_uri
                    )
                )

            except Exception:

                continue

            volumes = []

            for volume_member in (
                volume_collection.get(
                    "Members",
                    [],
                )
                or []
            ):

                volume_uri = (
                    get_member_uri(
                        volume_member
                    )
                )

                if not volume_uri:

                    continue

                try:

                    volume_data = (
                        self._get_json(
                            volume_uri
                        )
                    )

                except Exception:

                    continue

                volume_data[
                    "_uri"
                ] = volume_uri

                volume_data[
                    "_storage_uri"
                ] = storage_uri

                volume_data[
                    "_volumes_uri"
                ] = volumes_uri

                volumes.append(
                    volume_data
                )

            drive_uris = (
                self._get_dmtf_drive_uris(
                    storage_uri,
                    storage_data,
                )
            )

            write_support = (
                self._probe_dmtf_volume_write_support(
                    volumes_uri,
                    volume_response,
                )
            )

            controller_label = (
                self._get_dmtf_controller_label(
                    storage_uri,
                    storage_data,
                )
            )

            label_upper = (
                controller_label.upper()
            )

            is_boot_device = (
                "NS204" in label_upper
                or "BOOT DEVICE" in label_upper
            )

            raid_volumes = [
                volume
                for volume in volumes
                if self._is_real_raid_volume(
                    volume
                )
            ]

            targets.append(
                {
                    "storage_uri":
                        storage_uri,

                    "storage_data":
                        storage_data,

                    "volumes_uri":
                        volumes_uri,

                    "volumes":
                        volumes,

                    "raid_volumes":
                        raid_volumes,

                    "drive_uris":
                        drive_uris,

                    "controller_label":
                        controller_label,

                    "is_boot_device":
                        is_boot_device,

                    "writable":
                        write_support[
                            "writable"
                        ],

                    "allow_methods":
                        write_support[
                            "methods"
                        ],

                    "capabilities_uri":
                        write_support[
                            "capabilities_uri"
                        ],

                    "capabilities_status":
                        write_support[
                            "capabilities_status"
                        ],

                    "capabilities":
                        write_support[
                            "capabilities"
                        ],
                }
            )

        return targets


    def log_dmtf_storage_targets(
        self,
        targets,
    ):

        if not targets:

            log(
                self.row_number,
                self.ip,
                (
                    "DMTF Storage discovery: "
                    "no /Storage targets found."
                ),
            )

            return

        for target in targets:

            log(
                self.row_number,
                self.ip,
                (
                    "DMTF Storage target: "
                    f"{target['storage_uri']} | "
                    f"Controller={target['controller_label']} | "
                    f"Drives={len(target['drive_uris'])} | "
                    f"Volumes={len(target['volumes'])} | "
                    f"RAID volumes={len(target['raid_volumes'])} | "
                    f"Writable="
                    f"{'YES' if target['writable'] else 'NO'} | "
                    f"Allow="
                    f"{','.join(target['allow_methods']) or 'not advertised'} | "
                    f"CapabilitiesHTTP="
                    f"{target['capabilities_status']}"
                ),
            )


    def select_dmtf_storage_target(
        self,
        targets=None,
    ):

        if targets is None:

            targets = (
                self.discover_dmtf_storage_targets()
            )

        if not targets:

            return None

        ranked = []

        for target in targets:

            enough_drives = (
                len(
                    target[
                        "drive_uris"
                    ]
                )
                >= RAID_DATA_DRIVE_COUNT
            )

            score = (
                1
                if target[
                    "writable"
                ]
                else 0,

                1
                if enough_drives
                else 0,

                0
                if target[
                    "is_boot_device"
                ]
                else 1,

                len(
                    target[
                        "raid_volumes"
                    ]
                ),

                len(
                    target[
                        "drive_uris"
                    ]
                ),
            )

            ranked.append(
                (
                    score,
                    target,
                )
            )

        ranked.sort(
            key=lambda item:
                item[0],
            reverse=True,
        )

        return (
            ranked[
                0
            ][
                1
            ]
        )


    def get_dmtf_existing_volumes(
        self,
    ):

        volumes = []

        for target in (
            self.discover_dmtf_storage_targets()
        ):

            for volume in (
                target[
                    "raid_volumes"
                ]
            ):

                volumes.append(
                    volume
                )

        return volumes


    # =========================================================================
    # EXISTING RAID DETECTION
    # =========================================================================

    def get_existing_logical_drives(
        self,
    ):

        controllers = (
            self.get_smartstorage_inventory()
        )

        logical_drives = []

        for controller in controllers:

            logical_drives.extend(
                controller[
                    "logical_drives"
                ]
            )

        if logical_drives:

            return (
                logical_drives,
                "SmartStorage",
                controllers,
            )

        dmtf_volumes = (
            self.get_dmtf_existing_volumes()
        )

        if dmtf_volumes:

            return (
                dmtf_volumes,
                "DMTF Storage",
                controllers,
            )

        return (
            [],
            "None",
            controllers,
        )


    # =========================================================================
    # DMTF RAID WRITE
    # =========================================================================

    @staticmethod
    def _dmtf_raid_type():

        value = str(
            RAID_LEVEL
        ).strip().upper()

        if value == "RAID1":

            return "RAID1"

        return value


    def _wait_for_async_operation(
        self,
        response,
        timeout_seconds=30,
    ):

        if response.status_code != 202:

            return

        location = (
            response.headers.get(
                "Location"
            )
        )

        if not location:

            return

        deadline = (
            time.time()
            + timeout_seconds
        )

        while time.time() < deadline:

            time.sleep(
                2
            )

            try:

                task_response = (
                    self._request(
                        "GET",
                        location,
                    )
                )

                if task_response.status_code in (
                    404,
                    410,
                ):

                    return

                task_response.raise_for_status()

                task_data = (
                    task_response.json()
                )

                task_state = str(
                    task_data.get(
                        "TaskState",
                        task_data.get(
                            "JobState",
                            "",
                        ),
                    )
                ).strip()

                if task_state.lower() in (
                    "completed",
                    "completedok",
                    "success",
                    "succeeded",
                ):

                    return

                if task_state.lower() in (
                    "exception",
                    "killed",
                    "cancelled",
                    "failed",
                ):

                    raise Exception(
                        "Redfish storage task failed: "
                        f"{task_state}; "
                        f"{str(task_data)[:1000]}"
                    )

            except requests.exceptions.RequestException:

                continue

        log(
            self.row_number,
            self.ip,
            (
                "Warning: asynchronous storage task "
                "did not reach a terminal state within "
                f"{timeout_seconds}s. Continuing because "
                "iLO 5 storage changes may remain pending "
                "until reboot."
            ),
        )


    def delete_dmtf_volumes(
        self,
        target,
    ):

        volumes = list(
            target[
                "volumes"
            ]
        )

        # Use deterministic reverse ordering for volume deletion.
        volumes.sort(
            key=lambda item:
                natural_sort_key(
                    item.get(
                        "_uri",
                        "",
                    )
                ),
            reverse=True,
        )

        for volume in volumes:

            volume_uri = (
                volume.get(
                    "_uri"
                )
            )

            if not volume_uri:

                continue

            raid_type = (
                self._volume_raid_type(
                    volume
                )
                or "Unknown"
            )

            log(
                self.row_number,
                self.ip,
                (
                    "Overwrite: deleting existing DMTF "
                    f"volume {volume_uri} "
                    f"(RAIDType={raid_type})..."
                ),
            )

            response = (
                self._request(
                    "DELETE",
                    volume_uri,
                    timeout=WRITE_TIMEOUT_SECONDS,
                )
            )

            if response.status_code == 404:

                continue

            if response.status_code not in (
                200,
                202,
                204,
            ):

                raise Exception(
                    "Failed to delete existing DMTF "
                    f"volume {volume_uri}. "
                    f"HTTP {response.status_code}: "
                    f"{response.text[:1000]}"
                )

            self._wait_for_async_operation(
                response
            )


    def stage_dmtf_raid1_configuration(
        self,
        target,
        overwrite_raid,
    ):

        drive_uris = sorted(
            target[
                "drive_uris"
            ],
            key=natural_sort_key,
        )

        if (
            len(
                drive_uris
            )
            < RAID_DATA_DRIVE_COUNT
        ):

            raise Exception(
                f"Only {len(drive_uris)} "
                "physical drive(s) are visible through "
                "DMTF Storage; RAID1 requires "
                f"{RAID_DATA_DRIVE_COUNT}."
            )

        selected_drives = (
            drive_uris[
                :RAID_DATA_DRIVE_COUNT
            ]
        )

        if overwrite_raid:

            self.delete_dmtf_volumes(
                target
            )

        payload = {
            "RAIDType":
                self._dmtf_raid_type(),

            "Links": {
                "Drives": [
                    {
                        "@odata.id":
                            drive_uri
                    }
                    for drive_uri
                    in selected_drives
                ]
            },
        }

        log(
            self.row_number,
            self.ip,
            (
                "Creating RAID1 through DMTF Redfish "
                f"POST {target['volumes_uri']} "
                "using drives: "
                f"{', '.join(selected_drives)}"
            ),
        )

        response = self._request(
            "POST",
            target[
                "volumes_uri"
            ],
            json=payload,
            timeout=WRITE_TIMEOUT_SECONDS,
        )

        if response.status_code not in (
            200,
            201,
            202,
            204,
        ):

            raise Exception(
                "DMTF RAID1 create request failed. "
                f"HTTP {response.status_code}: "
                f"{response.text[:1500]}"
            )

        log(
            self.row_number,
            self.ip,
            (
                "DMTF RAID1 create request accepted "
                f"(HTTP {response.status_code})."
            ),
        )

        self._wait_for_async_operation(
            response
        )

        return selected_drives


    # =========================================================================
    # DMTF RAID1 APPLY / INITIALIZE / VERIFY
    # =========================================================================

    @staticmethod
    def _uri_tail(uri):

        return str(
            uri
            or ""
        ).rstrip("/").split("/")[-1]


    def _volume_matches_selected_drives(
        self,
        volume_data,
        selected_drive_uris,
    ):
        """
        Compare drive identities without requiring identical URI namespaces.

        HPE MR firmware can expose the same drive through Storage/Drives and
        Chassis/Drives URI forms. Comparing the final drive identifier avoids
        rejecting a correct RAID1 solely because those URI prefixes differ.
        """
        expected_ids = {
            self._uri_tail(
                uri
            )
            for uri
            in (
                selected_drive_uris
                or []
            )
            if uri
        }

        actual_uris = (
            self._volume_drive_uris(
                volume_data
            )
        )

        actual_ids = {
            self._uri_tail(
                uri
            )
            for uri
            in actual_uris
            if uri
        }

        # If firmware does not return Links/Drives on the Volume resource,
        # the selected storage target + RAIDType are still sufficient.
        if not actual_ids:

            return True

        if not expected_ids:

            return True

        return expected_ids.issubset(
            actual_ids
        )


    def wait_for_dmtf_raid1_volume(
        self,
        storage_uri,
        selected_drive_uris=None,
        timeout_seconds=None,
        context="RAID apply",
        expected_controller_name=None,
    ):
        """
        Continuously poll for RAID1 on the matching physical controller.

        HPE MR-controller DMTF Storage resource IDs can change across a reboot
        or storage reconfiguration. The pre-reboot Storage URI is therefore
        preferred but is not treated as a permanent controller identity.

        Fallback matching uses:
          1. controller Name/model;
          2. selected drive IDs;
          3. a non-boot storage target with enough drives.

        Polling starts immediately and repeats every
        DMTF_POLL_INTERVAL_SECONDS until RAID1 is visible or the maximum
        timeout expires.
        """
        if timeout_seconds is None:
            timeout_seconds = DMTF_APPLY_TIMEOUT_SECONDS

        deadline = time.time() + timeout_seconds
        attempt = 0
        storage_seen = False
        drives_seen = False
        last_post_state = None
        last_power_state = None
        announced_changed_uris = set()

        expected_uri = str(storage_uri or "").rstrip("/")
        expected_controller_name = str(
            expected_controller_name or ""
        ).strip()

        expected_drive_ids = {
            self._uri_tail(uri)
            for uri in (selected_drive_uris or [])
            if uri
        }

        log(
            self.row_number,
            self.ip,
            (
                "Starting continuous DMTF storage polling for {}. "
                "Poll interval={}s; maximum wait={}s; "
                "pre-reboot storage URI={}; controller={}."
            ).format(
                context,
                DMTF_POLL_INTERVAL_SECONDS,
                timeout_seconds,
                expected_uri or "unknown",
                expected_controller_name or "unknown",
            ),
        )

        while time.time() < deadline:
            attempt += 1
            power_state = "Unknown"
            post_state = "Unknown"

            try:
                system_info = self.get_system_info()
                power_state = str(
                    system_info.get(
                        "PowerState",
                        "Unknown",
                    )
                )
                post_state = str(
                    system_info.get(
                        "PostState",
                        "Unknown",
                    )
                )
                last_power_state = power_state
                last_post_state = post_state
            except Exception:
                pass

            try:
                targets = self.discover_dmtf_storage_targets()

                exact_target = None
                controller_matches = []
                drive_matches = []
                fallback_targets = []

                for target in targets:
                    current_uri = str(
                        target.get(
                            "storage_uri",
                            "",
                        )
                    ).rstrip("/")

                    storage_data = (
                        target.get(
                            "storage_data",
                            {},
                        )
                        or {}
                    )
                    current_name = str(
                        storage_data.get(
                            "Name",
                            "",
                        )
                        or ""
                    ).strip()

                    current_drive_ids = {
                        self._uri_tail(uri)
                        for uri in (
                            target.get(
                                "drive_uris",
                                [],
                            )
                            or []
                        )
                        if uri
                    }

                    if (
                        expected_uri
                        and current_uri == expected_uri
                    ):
                        exact_target = target

                    if (
                        expected_controller_name
                        and current_name.lower()
                        == expected_controller_name.lower()
                    ):
                        controller_matches.append(target)

                    if (
                        expected_drive_ids
                        and expected_drive_ids.issubset(
                            current_drive_ids
                        )
                    ):
                        drive_matches.append(target)

                    if (
                        not target.get(
                            "is_boot_device",
                            False,
                        )
                        and len(
                            target.get(
                                "drive_uris",
                                [],
                            )
                            or []
                        ) >= RAID_DATA_DRIVE_COUNT
                    ):
                        fallback_targets.append(target)

                candidates = []
                candidate_uris = set()

                def add_candidate(target):
                    if not target:
                        return

                    uri = str(
                        target.get(
                            "storage_uri",
                            "",
                        )
                    ).rstrip("/")

                    if uri in candidate_uris:
                        return

                    candidate_uris.add(uri)
                    candidates.append(target)

                add_candidate(exact_target)

                for target in controller_matches:
                    add_candidate(target)

                for target in drive_matches:
                    add_candidate(target)

                for target in fallback_targets:
                    add_candidate(target)

                if not candidates:
                    if attempt == 1 or attempt % 6 == 0:
                        available = [
                            "{} ({})".format(
                                target.get(
                                    "storage_uri",
                                    "unknown",
                                ),
                                (
                                    target.get(
                                        "storage_data",
                                        {},
                                    )
                                    or {}
                                ).get(
                                    "Name",
                                    "unknown",
                                ),
                            )
                            for target in targets
                        ]

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "Polling {}: no suitable DMTF "
                                "storage target is currently visible. "
                                "PowerState={}; PostState={}; "
                                "Available={}; Attempt={}."
                            ).format(
                                context,
                                power_state,
                                post_state,
                                available or "none",
                                attempt,
                            ),
                        )

                for target in candidates:
                    current_uri = str(
                        target.get(
                            "storage_uri",
                            "",
                        )
                    ).rstrip("/")

                    storage_data = (
                        target.get(
                            "storage_data",
                            {},
                        )
                        or {}
                    )
                    current_name = str(
                        storage_data.get(
                            "Name",
                            "",
                        )
                        or ""
                    ).strip()

                    if (
                        expected_uri
                        and current_uri != expected_uri
                        and current_uri
                        not in announced_changed_uris
                    ):
                        announced_changed_uris.add(
                            current_uri
                        )

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "DMTF Storage resource URI changed "
                                "after reboot/reconfiguration: "
                                "{} -> {}. Continuing by matching "
                                "controller '{}'."
                            ).format(
                                expected_uri,
                                current_uri,
                                (
                                    current_name
                                    or expected_controller_name
                                    or "unknown"
                                ),
                            ),
                        )

                    if not storage_seen:
                        storage_seen = True

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "Matching DMTF storage controller "
                                "is reachable during boot. "
                                "URI={}; Controller={}; "
                                "PowerState={}; PostState={}; "
                                "Drives={}; Volumes={}."
                            ).format(
                                current_uri,
                                current_name or "unknown",
                                power_state,
                                post_state,
                                len(
                                    target.get(
                                        "drive_uris",
                                        [],
                                    )
                                    or []
                                ),
                                len(
                                    target.get(
                                        "volumes",
                                        [],
                                    )
                                    or []
                                ),
                            ),
                        )

                    if (
                        len(
                            target.get(
                                "drive_uris",
                                [],
                            )
                            or []
                        ) >= RAID_DATA_DRIVE_COUNT
                        and not drives_seen
                    ):
                        drives_seen = True

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "Required physical drives are "
                                "visible through DMTF Storage "
                                "during the pre-OS/System Utilities "
                                "phase: {}"
                            ).format(
                                target.get(
                                    "drive_uris",
                                    [],
                                )
                            ),
                        )

                    volume_summaries = []

                    for volume in (
                        target.get(
                            "volumes",
                            [],
                        )
                        or []
                    ):
                        raid_value = (
                            self._volume_raid_type(
                                volume
                            )
                        )

                        volume_summaries.append(
                            "{}:{}".format(
                                volume.get(
                                    "_uri",
                                    "unknown",
                                ),
                                raid_value or "None",
                            )
                        )

                        if (
                            str(raid_value)
                            .strip()
                            .upper()
                            != "RAID1"
                        ):
                            continue

                        if not (
                            self._volume_matches_selected_drives(
                                volume,
                                selected_drive_uris,
                            )
                        ):
                            continue

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "RAID1 volume is visible on the "
                                "matching DMTF storage controller. "
                                "StorageURI={}; PowerState={}; "
                                "PostState={}; Volume={}"
                            ).format(
                                current_uri,
                                power_state,
                                post_state,
                                volume.get(
                                    "_uri",
                                    "unknown",
                                ),
                            ),
                        )

                        return volume

                    if attempt == 1 or attempt % 6 == 0:
                        log(
                            self.row_number,
                            self.ip,
                            (
                                "Polling {}: StorageURI={}; "
                                "Controller={}; PowerState={}; "
                                "PostState={}; Drives={}; "
                                "Volumes={}; Attempt={}."
                            ).format(
                                context,
                                current_uri,
                                current_name or "unknown",
                                power_state,
                                post_state,
                                len(
                                    target.get(
                                        "drive_uris",
                                        [],
                                    )
                                    or []
                                ),
                                (
                                    ", ".join(
                                        volume_summaries
                                    )
                                    or "none"
                                ),
                                attempt,
                            ),
                        )

            except Exception as exc:
                if attempt == 1 or attempt % 6 == 0:
                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Polling {}: temporary Redfish "
                            "storage error: {}. "
                            "PowerState={}; PostState={}; "
                            "Attempt={}."
                        ).format(
                            context,
                            exc,
                            power_state,
                            post_state,
                            attempt,
                        ),
                    )

            remaining = deadline - time.time()

            if remaining <= 0:
                break

            time.sleep(
                min(
                    DMTF_POLL_INTERVAL_SECONDS,
                    remaining,
                )
            )

        log(
            self.row_number,
            self.ip,
            (
                "Timed out waiting for RAID1 during {} "
                "after {}s. StorageSeen={}; "
                "DrivesSeen={}; LastPowerState={}; "
                "LastPostState={}."
            ).format(
                context,
                timeout_seconds,
                (
                    "YES"
                    if storage_seen
                    else "NO"
                ),
                (
                    "YES"
                    if drives_seen
                    else "NO"
                ),
                last_power_state or "Unknown",
                last_post_state or "Unknown",
            ),
        )

        return None


    def _get_volume_initialize_action(
        self,
        volume_data,
    ):

        actions = (
            volume_data.get(
                "Actions",
                {},
            )
            or {}
        )

        for action_name, action_data in actions.items():

            if (
                "volume.initialize"
                in str(
                    action_name
                ).lower()
                and isinstance(
                    action_data,
                    dict,
                )
            ):

                return (
                    action_name,
                    action_data,
                )

        return (
            None,
            None,
        )


    def wait_for_volume_initialize_action(
        self,
        volume_uri,
        timeout_seconds,
    ):
        """
        Wait until the newly-created MR volume advertises #Volume.Initialize.

        A newly exposed RDE volume can become visible before all action
        metadata is ready. Do not construct an action URI when the resource
        does not advertise the action; wait for the supported target instead.
        """
        deadline = (
            time.time()
            + timeout_seconds
        )

        attempt = 0

        while time.time() < deadline:

            attempt += 1

            try:

                volume_data = (
                    self._get_json(
                        volume_uri
                    )
                )

                volume_data[
                    "_uri"
                ] = volume_uri

                (
                    action_name,
                    action_data,
                ) = (
                    self._get_volume_initialize_action(
                        volume_data
                    )
                )

                if (
                    action_name
                    and action_data
                    and action_data.get(
                        "target"
                    )
                ):

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Volume.Initialize action is "
                            "advertised by the MR controller: "
                            f"{action_data.get('target')}"
                        ),
                    )

                    return (
                        volume_data,
                        action_name,
                        action_data,
                    )

                if (
                    attempt == 1
                    or attempt % 6 == 0
                ):

                    advertised_actions = sorted(
                        (
                            volume_data.get(
                                "Actions",
                                {},
                            )
                            or {}
                        ).keys()
                    )

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Waiting for Volume.Initialize "
                            "action metadata. AdvertisedActions="
                            f"{advertised_actions or 'none'}; "
                            f"Attempt={attempt}."
                        ),
                    )

            except Exception as exc:

                if (
                    attempt == 1
                    or attempt % 6 == 0
                ):

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Waiting for Volume.Initialize "
                            f"action metadata: {exc}; "
                            f"Attempt={attempt}."
                        ),
                    )

            remaining = (
                deadline
                - time.time()
            )

            if remaining <= 0:

                break

            time.sleep(
                min(
                    INITIALIZE_POLL_INTERVAL_SECONDS,
                    remaining,
                )
            )

        return (
            None,
            None,
            None,
        )


    @staticmethod
    def _initialization_operations(
        volume_data,
    ):

        operations = (
            volume_data.get(
                "Operations",
                [],
            )
            or []
        )

        active = []

        for operation in operations:

            if not isinstance(
                operation,
                dict,
            ):

                continue

            name = str(
                operation.get(
                    "OperationName",
                    operation.get(
                        "Operation",
                        "",
                    ),
                )
                or ""
            ).strip()

            if (
                "initializ"
                in name.lower()
            ):

                active.append(
                    operation
                )

        return active


    def wait_for_volume_initialization(
        self,
        volume_uri,
        timeout_seconds,
    ):
        """
        Poll HPE MR Volume.Operations until initialization completes.

        HPE documents that the action can return before the operation itself
        has completed. Progress is exposed through Operations and
        PercentageComplete.
        """
        deadline = (
            time.time()
            + timeout_seconds
        )

        attempt = 0
        saw_initialization = False

        while time.time() < deadline:

            attempt += 1

            volume_data = (
                self._get_json(
                    volume_uri
                )
            )

            volume_data[
                "_uri"
            ] = volume_uri

            raid_value = (
                self._volume_raid_type(
                    volume_data
                )
            )

            if (
                str(
                    raid_value
                ).strip().upper()
                != "RAID1"
            ):

                raise Exception(
                    "RAIDType changed during initialization: "
                    f"{raid_value}"
                )

            active_operations = (
                self._initialization_operations(
                    volume_data
                )
            )

            initialize_method = str(
                volume_data.get(
                    "InitializeMethod",
                    "",
                )
                or ""
            ).strip()

            status = (
                volume_data.get(
                    "Status",
                    {},
                )
                or {}
            )

            state = str(
                status.get(
                    "State",
                    "",
                )
                or ""
            ).strip()

            if active_operations:

                saw_initialization = True

                if (
                    attempt == 1
                    or attempt % 6 == 0
                ):

                    progress_parts = []

                    for operation in active_operations:

                        name = str(
                            operation.get(
                                "OperationName",
                                operation.get(
                                    "Operation",
                                    "Initializing",
                                ),
                            )
                            or "Initializing"
                        )

                        percentage = (
                            operation.get(
                                "PercentageComplete"
                            )
                        )

                        if percentage is None:

                            progress_parts.append(
                                name
                            )

                        else:

                            progress_parts.append(
                                "{}={}%".format(
                                    name,
                                    percentage,
                                )
                            )

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "RAID1 initialization in progress: "
                            f"{', '.join(progress_parts)}; "
                            f"InitializeMethod="
                            f"{initialize_method or 'unknown'}; "
                            f"State={state or 'unknown'}."
                        ),
                    )

            else:

                log(
                    self.row_number,
                    self.ip,
                    (
                        "RAID1 initialization completed. "
                        f"InitializeMethod="
                        f"{initialize_method or 'not reported'}; "
                        f"State={state or 'unknown'}; "
                        f"OperationSeen="
                        f"{'YES' if saw_initialization else 'NO (completed quickly)'}."
                    ),
                )

                return volume_data

            remaining = (
                deadline
                - time.time()
            )

            if remaining <= 0:

                break

            time.sleep(
                min(
                    INITIALIZE_POLL_INTERVAL_SECONDS,
                    remaining,
                )
            )

        raise Exception(
            "RAID1 initialization did not complete within "
            f"{timeout_seconds}s."
        )


    def initialize_dmtf_volume(
        self,
        volume_data,
        initialize_type,
    ):
        """
        Initialize an HPE MR RAID1 using the action advertised by the volume.

        HPE MR quick initialization:
            InitializeMethod = Foreground
            InitializeType   = Fast

        HPE MR full initialization:
            InitializeMethod = Foreground
            InitializeType   = Slow
        """
        volume_uri = (
            volume_data.get(
                "_uri"
            )
            or volume_data.get(
                "@odata.id"
            )
        )

        if not volume_uri:

            raise Exception(
                "Cannot initialize DMTF volume: "
                "volume URI is missing."
            )

        (
            current_volume,
            action_name,
            initialize_action,
        ) = (
            self.wait_for_volume_initialize_action(
                volume_uri,
                INITIALIZE_ACTION_DISCOVERY_TIMEOUT_SECONDS,
            )
        )

        if not initialize_action:

            raise Exception(
                "HPE MR volume does not advertise "
                "#Volume.Initialize within "
                f"{INITIALIZE_ACTION_DISCOVERY_TIMEOUT_SECONDS}s. "
                "Initialization was requested by policy, so "
                "the RAID is not considered successfully prepared."
            )

        action_uri = (
            initialize_action.get(
                "target"
            )
        )

        allowed_types = (
            initialize_action.get(
                "InitializeType@Redfish.AllowableValues",
                [],
            )
            or []
        )

        allowed_methods = (
            initialize_action.get(
                "InitializeMethod@Redfish.AllowableValues",
                [],
            )
            or []
        )

        requested_type = str(
            initialize_type
            or "Fast"
        ).strip()

        requested_method = str(
            INITIALIZE_METHOD
            or "Foreground"
        ).strip()

        chosen_type = requested_type

        if allowed_types:

            allowed_type_map = {
                str(
                    value
                ).strip().lower():
                    str(
                        value
                    ).strip()
                for value
                in allowed_types
            }

            chosen_type = (
                allowed_type_map.get(
                    requested_type.lower()
                )
            )

            if chosen_type is None:

                raise Exception(
                    "Requested initialization type "
                    f"'{requested_type}' is not supported. "
                    "Controller reports allowable values: "
                    f"{allowed_types}"
                )

        chosen_method = requested_method

        if allowed_methods:

            allowed_method_map = {
                str(
                    value
                ).strip().lower():
                    str(
                        value
                    ).strip()
                for value
                in allowed_methods
            }

            chosen_method = (
                allowed_method_map.get(
                    requested_method.lower()
                )
            )

            if chosen_method is None:

                raise Exception(
                    "Requested initialization method "
                    f"'{requested_method}' is not supported. "
                    "Controller reports allowable values: "
                    f"{allowed_methods}"
                )

        log(
            self.row_number,
            self.ip,
            (
                "Initializing RAID1 volume "
                f"{volume_uri} using HPE MR Redfish action "
                f"{action_uri}. "
                f"InitializeMethod={chosen_method}; "
                f"InitializeType={chosen_type}; "
                f"AllowableMethods="
                f"{allowed_methods or 'not advertised'}; "
                f"AllowableTypes="
                f"{allowed_types or 'not advertised'}."
            ),
        )

        payload = {
            "InitializeMethod":
                chosen_method,

            "InitializeType":
                chosen_type,
        }

        response = (
            self._request(
                "POST",
                action_uri,
                json=payload,
                timeout=WRITE_TIMEOUT_SECONDS,
            )
        )

        if response.status_code not in (
            200,
            201,
            202,
            204,
        ):

            raise Exception(
                "Volume initialization request failed. "
                f"HTTP {response.status_code}: "
                f"{response.text[:1500]}"
            )

        log(
            self.row_number,
            self.ip,
            (
                "Volume initialization request accepted "
                f"(HTTP {response.status_code}) with "
                f"InitializeMethod={chosen_method}, "
                f"InitializeType={chosen_type}."
            ),
        )

        self._wait_for_async_operation(
            response,
            timeout_seconds=(
                INITIALIZE_TIMEOUT_SECONDS
            ),
        )

        initialized_volume = (
            self.wait_for_volume_initialization(
                volume_uri,
                INITIALIZE_TIMEOUT_SECONDS,
            )
        )

        time.sleep(
            INITIALIZE_SETTLE_SECONDS
        )

        log(
            self.row_number,
            self.ip,
            (
                "RAID1 volume initialization verified. "
                "Old boot/partition metadata has been cleared."
            ),
        )

        return initialized_volume


    # =========================================================================
    # SMART STORAGE CONFIG DISCOVERY
    #
    # This is the WRITE/STAGING interface.
    # =========================================================================

    def _collect_config_links(
        self,
        obj,
        found=None,
    ):

        if found is None:
            found = set()

        if isinstance(
            obj,
            dict,
        ):

            uri = (
                obj.get(
                    "@odata.id"
                )
            )

            if isinstance(
                uri,
                str,
            ):

                normalized = (
                    uri.rstrip("/")
                )

                if (
                    "smartstorageconfig"
                    in normalized.lower()
                    and not normalized
                    .lower()
                    .endswith(
                        "/settings"
                    )
                ):

                    found.add(
                        normalized
                    )

            for value in obj.values():

                self._collect_config_links(
                    value,
                    found,
                )

        elif isinstance(
            obj,
            list,
        ):

            for value in obj:

                self._collect_config_links(
                    value,
                    found,
                )

        return found


    def discover_smartstorage_configs(self):

        if not self.system_uri:
            self.discover_system()

        candidates = set()

        # Discover advertised links from ComputerSystem.
        try:

            system_data = (
                self._get_json(
                    self.system_uri
                )
            )

            candidates.update(
                self._collect_config_links(
                    system_data
                )
            )

        except Exception:

            pass

        # Known iLO 5 URI forms.
        candidates.add(
            self.system_uri.rstrip("/")
            + "/SmartStorageConfig"
        )

        for index in range(
            2,
            17,
        ):

            candidates.add(
                self.system_uri.rstrip("/")
                + f"/SmartStorageConfig{index}"
            )

        discovered = []

        for uri in sorted(
            candidates,
            key=natural_sort_key,
        ):

            try:

                response = self._request(
                    "GET",
                    uri + "/",
                )

                if response.status_code in (
                    404,
                    405,
                ):
                    continue

                response.raise_for_status()

                data = response.json()

                canonical_uri = (
                    str(
                        data.get(
                            "@odata.id",
                            uri,
                        )
                    )
                    .rstrip("/")
                )

                discovered.append(
                    {
                        "uri":
                            canonical_uri,

                        "data":
                            data,
                    }
                )

            except Exception:

                continue

        # Deduplicate
        unique = {}

        for config in discovered:

            unique[
                config["uri"]
            ] = config

        return list(
            unique.values()
        )


    # =========================================================================
    # CONFIG DRIVE LOCATIONS
    # =========================================================================

    def get_config_drive_locations(
        self,
        config_data,
    ):

        locations = []

        for drive in (
            config_data.get(
                "PhysicalDrives",
                [],
            )
            or []
        ):

            if not isinstance(
                drive,
                dict,
            ):
                continue

            location = (
                drive.get(
                    "Location"
                )
            )

            if location:

                locations.append(
                    str(location)
                )

        return sorted(
            locations,
            key=natural_sort_key,
        )


    # =========================================================================
    # SELECT SMARTSTORAGECONFIG
    # =========================================================================

    def select_smartstorage_config(self):

        configs = (
            self.discover_smartstorage_configs()
        )

        if not configs:

            return (
                None,
                [],
                None,
            )

        ranked = []

        for config in configs:

            locations = (
                self.get_config_drive_locations(
                    config["data"]
                )
            )

            # Prefer configs containing enough physical disks.
            score = (
                1
                if len(locations)
                >= RAID_DATA_DRIVE_COUNT
                else 0,

                len(locations),
            )

            ranked.append(
                (
                    score,
                    config,
                    locations,
                )
            )

        ranked.sort(
            key=lambda item:
                item[0],
            reverse=True,
        )

        (
            _,
            selected,
            locations,
        ) = ranked[0]

        return (
            selected[
                "uri"
            ],
            locations,
            selected[
                "data"
            ],
        )


    # =========================================================================
    # BOOT TARGET
    # =========================================================================

    @staticmethod
    def _unable_to_modify_during_post(
        response,
    ):

        try:
            response_text = str(
                response.text
                or ""
            )
        except Exception:
            response_text = ""

        return (
            "UnableToModifyDuringSystemPOST"
            in response_text
        )


    def wait_for_power_state(
        self,
        desired_state,
        timeout_seconds,
        poll_seconds=5,
    ):

        desired = str(
            desired_state
        ).strip().lower()

        deadline = (
            time.time()
            + timeout_seconds
        )

        while time.time() < deadline:

            try:
                current = str(
                    self.get_system_info()
                    .get(
                        "PowerState",
                        "Unknown",
                    )
                ).strip()

                if current.lower() == desired:
                    return True

            except Exception:
                pass

            time.sleep(
                poll_seconds
            )

        return False


    def force_power_off(
        self,
        reason=None,
    ):

        try:
            current_state = str(
                self.get_system_info()
                .get(
                    "PowerState",
                    "Unknown",
                )
            )
        except Exception:
            current_state = "Unknown"

        if current_state.lower() == "off":

            log(
                self.row_number,
                self.ip,
                "Server is already powered off.",
            )

            return True

        if reason:

            log(
                self.row_number,
                self.ip,
                reason,
            )

        log(
            self.row_number,
            self.ip,
            (
                "Forcing server power OFF so iLO boot "
                "settings can be modified safely..."
            ),
        )

        self.reset_server(
            "ForceOff"
        )

        powered_off = (
            self.wait_for_power_state(
                "Off",
                POWER_OFF_TIMEOUT_SECONDS,
                poll_seconds=5,
            )
        )

        if not powered_off:

            raise Exception(
                "Server did not reach PowerState=Off "
                f"within {POWER_OFF_TIMEOUT_SECONDS}s."
            )

        log(
            self.row_number,
            self.ip,
            "Server reached PowerState=Off.",
        )

        # Give iLO a moment to release POST-only configuration locks.
        time.sleep(
            2
        )

        return True


    def _patch_system_utilities_boot_override(
        self,
        target,
    ):

        payload = {
            "Boot": {
                "BootSourceOverrideTarget":
                    target,
                "BootSourceOverrideEnabled":
                    "Once",
            }
        }

        return self._request(
            "PATCH",
            self.system_uri,
            json=payload,
        )


    def set_boot_to_system_utilities(self):

        info = (
            self.get_system_info()
        )

        system_data = (
            info[
                "Data"
            ]
        )

        boot = (
            system_data.get(
                "Boot",
                {},
            )
            or {}
        )

        allowable = (
            boot.get(
                "BootSourceOverrideTarget@Redfish.AllowableValues",
                [],
            )
            or []
        )

        if (
            not allowable
            or "BiosSetup" in allowable
        ):
            target = "BiosSetup"
        elif "Utilities" in allowable:
            target = "Utilities"
        else:
            raise Exception(
                "iLO does not advertise BiosSetup or "
                f"Utilities boot target. "
                f"Allowable targets: {allowable}"
            )

        log(
            self.row_number,
            self.ip,
            (
                "Setting one-time boot target "
                f"to System Utilities ({target})..."
            ),
        )

        response = (
            self._patch_system_utilities_boot_override(
                target
            )
        )

        if response.status_code < 400:
            return target

        if not (
            BOOT_OVERRIDE_POST_RECOVERY
            and self._unable_to_modify_during_post(
                response
            )
        ):

            raise Exception(
                "Failed to set System Utilities "
                f"boot override: HTTP "
                f"{response.status_code}: "
                f"{response.text[:500]}"
            )

        try:
            current_info = self.get_system_info()
            current_power = str(
                current_info.get(
                    "PowerState",
                    "Unknown",
                )
            )
            current_post = str(
                current_info.get(
                    "PostState",
                    "Unknown",
                )
            )
        except Exception:
            current_power = "Unknown"
            current_post = "Unknown"

        log(
            self.row_number,
            self.ip,
            (
                "iLO rejected the System Utilities boot "
                "override because the server is in POST "
                "(UnableToModifyDuringSystemPOST). "
                f"PowerState={current_power}; "
                f"PostState={current_post}."
            ),
        )

        self.force_power_off(
            reason=(
                "POST-safe recovery: powering the server "
                "off before retrying the BiosSetup boot override."
            ),
        )

        retry_deadline = (
            time.time()
            + BOOT_OVERRIDE_RETRY_TIMEOUT_SECONDS
        )

        attempt = 0
        last_response = response

        while time.time() < retry_deadline:

            attempt += 1

            log(
                self.row_number,
                self.ip,
                (
                    "Retrying System Utilities boot override "
                    f"while server is OFF (Attempt {attempt})..."
                ),
            )

            last_response = (
                self._patch_system_utilities_boot_override(
                    target
                )
            )

            if last_response.status_code < 400:

                log(
                    self.row_number,
                    self.ip,
                    (
                        "System Utilities boot override "
                        "accepted successfully while server "
                        "is powered off."
                    ),
                )

                return target

            if not self._unable_to_modify_during_post(
                last_response
            ):

                raise Exception(
                    "Failed to set System Utilities "
                    "boot override after POST-safe power-off. "
                    f"HTTP {last_response.status_code}: "
                    f"{last_response.text[:500]}"
                )

            remaining = (
                retry_deadline
                - time.time()
            )

            if remaining <= 0:
                break

            time.sleep(
                min(
                    BOOT_OVERRIDE_RETRY_INTERVAL_SECONDS,
                    remaining,
                )
            )

        raise Exception(
            "System Utilities boot override remained "
            "locked by POST after powering the server off "
            f"and retrying for "
            f"{BOOT_OVERRIDE_RETRY_TIMEOUT_SECONDS}s. "
            f"Last response: HTTP "
            f"{last_response.status_code}: "
            f"{last_response.text[:500]}"
        )


    def set_boot_to_normal(self):

        try:

            response = self._request(
                "PATCH",
                self.system_uri,
                json={
                    "Boot": {
                        "BootSourceOverrideEnabled":
                            "Disabled",
                    }
                },
            )

            # Compatibility fallback
            if response.status_code >= 400:

                self._request(
                    "PATCH",
                    self.system_uri,
                    json={
                        "Boot": {
                            "BootSourceOverrideTarget":
                                "None"
                        }
                    },
                )

        except Exception:

            pass


    # =========================================================================
    # RESET ACTION DISCOVERY
    # =========================================================================

    def get_reset_action(self):

        info = (
            self.get_system_info()
        )

        data = (
            info[
                "Data"
            ]
        )

        actions = (
            data.get(
                "Actions",
                {},
            )
            or {}
        )

        reset_action = (
            actions.get(
                "#ComputerSystem.Reset",
                {},
            )
            or {}
        )

        target = (
            reset_action.get(
                "target"
            )
            or reset_action.get(
                "Target"
            )
        )

        if not target:

            target = (
                self.system_uri.rstrip("/")
                + "/Actions/ComputerSystem.Reset"
            )

        allowable = (
            reset_action.get(
                "ResetType@Redfish.AllowableValues",
                [],
            )
            or []
        )

        return (
            target,
            allowable,
        )


    def reset_server(
        self,
        reset_type,
    ):

        target, allowable = (
            self.get_reset_action()
        )

        if (
            allowable
            and reset_type not in allowable
        ):

            raise Exception(
                f"ResetType '{reset_type}' "
                "not supported. "
                f"Allowable={allowable}"
            )

        response = self._request(
            "POST",
            target,
            json={
                "ResetType":
                    reset_type
            },
            timeout=WRITE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:

            raise Exception(
                f"{reset_type} failed: "
                f"HTTP {response.status_code}: "
                f"{response.text[:500]}"
            )


    # =========================================================================
    # VERIFY REBOOT
    # =========================================================================

    def wait_for_post_transition(
        self,
        initial_post_state,
        timeout_seconds,
    ):

        deadline = (
            time.time()
            + timeout_seconds
        )

        while time.time() < deadline:

            time.sleep(
                REBOOT_DETECTION_POLL_SECONDS
            )

            try:

                info = (
                    self.get_system_info()
                )

                post_state = str(
                    info.get(
                        "PostState",
                        "Unknown",
                    )
                )

                power_state = str(
                    info.get(
                        "PowerState",
                        "Unknown",
                    )
                )

                if (
                    power_state.lower()
                    != "on"
                ):

                    return True

                if (
                    post_state
                    not in (
                        initial_post_state,
                        "Unknown",
                        "",
                    )
                ):

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Reboot/POST transition detected: "
                            f"{initial_post_state} -> "
                            f"{post_state}"
                        ),
                    )

                    return True

                if post_state not in (
                    "FinishedPost",
                    "Unknown",
                    "",
                ):

                    return True

            except Exception:

                # Temporary Redfish interruption itself is evidence
                # that iLO/host is transitioning.
                continue

        return False


    # =========================================================================
    # FORCE POWER CYCLE
    # =========================================================================

    def force_power_cycle(self):

        log(
            self.row_number,
            self.ip,
            (
                "ForceRestart did not produce a "
                "detectable POST transition. "
                "Falling back to ForceOff -> On..."
            ),
        )

        try:
            self.force_power_off()
        except Exception as exc:
            raise Exception(
                "ForceRestart did not initiate POST "
                "and ForceOff fallback failed: "
                f"{exc}"
            )

        log(
            self.row_number,
            self.ip,
            "Server is powered off. Powering on...",
        )

        self.reset_server(
            "On"
        )

        time.sleep(
            POWER_ON_INITIAL_WAIT_SECONDS
        )


    # =========================================================================
    # REBOOT INTO SYSTEM UTILITIES
    # =========================================================================

    def reboot_to_system_utilities(self):

        boot_target = (
            self.set_boot_to_system_utilities()
        )

        info = (
            self.get_system_info()
        )

        power_state = str(
            info.get(
                "PowerState",
                "Unknown",
            )
        )

        initial_post_state = str(
            info.get(
                "PostState",
                "Unknown",
            )
        )

        if (
            power_state.lower()
            == "off"
        ):

            log(
                self.row_number,
                self.ip,
                (
                    "Server is powered off. "
                    "Powering on into System Utilities..."
                ),
            )

            self.reset_server(
                "On"
            )

            return boot_target

        log(
            self.row_number,
            self.ip,
            (
                "Rebooting server into "
                "System Utilities..."
            ),
        )

        try:

            self.reset_server(
                "ForceRestart"
            )

        except Exception as exc:

            log(
                self.row_number,
                self.ip,
                (
                    "ForceRestart request failed: "
                    f"{exc}"
                ),
            )

            self.force_power_cycle()

            return boot_target

        # ---------------------------------------------------------------------
        # Confirm that ForceRestart actually caused a reboot.
        # ---------------------------------------------------------------------

        reboot_detected = (
            self.wait_for_post_transition(
                initial_post_state,
                REBOOT_DETECTION_TIMEOUT_SECONDS,
            )
        )

        if not reboot_detected:

            self.force_power_cycle()

        return boot_target


    # =========================================================================
    # NORMAL REBOOT
    # =========================================================================

    def reboot_normal(
        self,
    ):

        self.set_boot_to_normal()

        info = (
            self.get_system_info()
        )

        power_state = str(
            info.get(
                "PowerState",
                "Unknown",
            )
        )

        initial_post_state = str(
            info.get(
                "PostState",
                "Unknown",
            )
        )

        if (
            power_state.lower()
            == "off"
        ):

            log(
                self.row_number,
                self.ip,
                "Server is OFF. Powering on...",
            )

            self.reset_server(
                "On"
            )

            return

        try:

            self.reset_server(
                "ForceRestart"
            )

        except Exception as exc:

            log(
                self.row_number,
                self.ip,
                (
                    "Normal ForceRestart failed: "
                    f"{exc}. Falling back to "
                    "ForceOff -> On."
                ),
            )

            self.force_power_cycle()

            return

        reboot_detected = (
            self.wait_for_post_transition(
                initial_post_state,
                REBOOT_DETECTION_TIMEOUT_SECONDS,
            )
        )

        if not reboot_detected:

            self.force_power_cycle()


    def wait_for_post_complete(
        self,
        timeout_seconds,
    ):

        deadline = (
            time.time()
            + timeout_seconds
        )

        saw_post_activity = False

        while time.time() < deadline:

            time.sleep(
                REBOOT_DETECTION_POLL_SECONDS
            )

            try:

                info = (
                    self.get_system_info()
                )

                power_state = str(
                    info.get(
                        "PowerState",
                        "Unknown",
                    )
                )

                post_state = str(
                    info.get(
                        "PostState",
                        "Unknown",
                    )
                )

                if post_state not in (
                    "FinishedPost",
                    "Unknown",
                    "",
                ):

                    saw_post_activity = True

                if (
                    power_state.lower()
                    == "on"
                    and post_state
                    == "FinishedPost"
                    and saw_post_activity
                ):

                    log(
                        self.row_number,
                        self.ip,
                        (
                            "Server POST completed "
                            "after RAID configuration reboot."
                        ),
                    )

                    return True

            except Exception:

                pass

        log(
            self.row_number,
            self.ip,
            (
                "Warning: POST completion was not "
                f"confirmed within {timeout_seconds}s. "
                "Continuing with RAID inventory verification."
            ),
        )

        return False


    # =========================================================================
    # WAIT FOR A WRITABLE RAID INTERFACE
    # =========================================================================

    def wait_for_raid_write_interface(
        self,
        timeout_minutes,
    ):

        max_attempts = max(
            1,
            int(
                (
                    timeout_minutes
                    * 60
                )
                / POLL_INTERVAL_SECONDS
            ),
        )

        for attempt in range(
            1,
            max_attempts + 1,
        ):

            dmtf_targets = (
                self.discover_dmtf_storage_targets()
            )

            dmtf_target = (
                self.select_dmtf_storage_target(
                    dmtf_targets
                )
            )

            if (
                dmtf_target
                and dmtf_target[
                    "writable"
                ]
                and len(
                    dmtf_target[
                        "drive_uris"
                    ]
                )
                >= RAID_DATA_DRIVE_COUNT
            ):

                return {
                    "method":
                        "DMTF",

                    "dmtf_target":
                        dmtf_target,
                }

            (
                config_uri,
                drive_locations,
                config_data,
            ) = (
                self.select_smartstorage_config()
            )

            if (
                config_uri
                and len(
                    drive_locations
                )
                >= RAID_DATA_DRIVE_COUNT
            ):

                return {
                    "method":
                        "SmartStorageConfig",

                    "config_uri":
                        config_uri,

                    "drive_locations":
                        drive_locations,

                    "config_data":
                        config_data,
                }

            if (
                attempt == 1
                or attempt % 2 == 1
            ):

                log(
                    self.row_number,
                    self.ip,
                    (
                        "Waiting for a writable RAID "
                        "interface... "
                        f"(Attempt {attempt}/{max_attempts}; "
                        f"DMTF="
                        f"{'writable' if dmtf_target and dmtf_target['writable'] else 'not writable'}; "
                        f"SmartStorageConfig="
                        f"{'available' if config_uri else 'not available'})"
                    ),
                )

            time.sleep(
                POLL_INTERVAL_SECONDS
            )

        return None


    # =========================================================================
    # WAIT FOR STORAGE
    # =========================================================================

    def wait_for_storage(
        self,
        timeout_minutes,
        require_config=False,
    ):

        max_attempts = max(
            1,
            int(
                (
                    timeout_minutes
                    * 60
                )
                / POLL_INTERVAL_SECONDS
            ),
        )

        for attempt in range(
            1,
            max_attempts + 1,
        ):

            controllers = (
                self.get_smartstorage_inventory()
            )

            config_uri = None
            drive_locations = []
            config_data = None

            if require_config:

                (
                    config_uri,
                    drive_locations,
                    config_data,
                ) = (
                    self.select_smartstorage_config()
                )

            inventory_ok = bool(
                controllers
            )

            config_ok = (
                bool(config_uri)
                if require_config
                else True
            )

            if (
                inventory_ok
                and config_ok
            ):

                return (
                    controllers,
                    config_uri,
                    drive_locations,
                    config_data,
                )

            if (
                attempt == 1
                or attempt % 2 == 1
            ):

                log(
                    self.row_number,
                    self.ip,
                    (
                        "Waiting for Smart Array "
                        "inventory/configuration... "
                        f"(Attempt "
                        f"{attempt}/{max_attempts})"
                    ),
                )

            time.sleep(
                POLL_INTERVAL_SECONDS
            )

        return (
            [],
            None,
            [],
            None,
        )


    # =========================================================================
    # STAGE RAID1
    # =========================================================================

    def stage_raid1_configuration(
        self,
        config_uri,
        drive_locations,
    ):

        if (
            len(drive_locations)
            < RAID_DATA_DRIVE_COUNT
        ):

            raise Exception(
                f"Only {len(drive_locations)} "
                "physical drive(s) are visible "
                "through SmartStorageConfig; "
                f"RAID1 requires "
                f"{RAID_DATA_DRIVE_COUNT}."
            )

        selected_drives = sorted(
            drive_locations,
            key=natural_sort_key,
        )[
            :RAID_DATA_DRIVE_COUNT
        ]

        settings_uri = (
            config_uri.rstrip("/")
            + "/Settings/"
        )

        payload = {

            "DataGuard":
                "Disabled",

            "LogicalDrives": [
                {
                    "LogicalDriveName":
                        RAID_DISPLAY_NAME,

                    "Raid":
                        RAID_LEVEL,

                    "CapacityGiB":
                        RAID_CAPACITY_GIB,

                    "DataDrives":
                        selected_drives,

                    "LegacyBootPriority":
                        "Primary",
                }
            ],
        }

        log(
            self.row_number,
            self.ip,
            (
                f"Staging {RAID_LEVEL} "
                "using physical drives: "
                f"{', '.join(selected_drives)}"
            ),
        )

        response = self._request(
            "PUT",
            settings_uri,
            json=payload,
            timeout=WRITE_TIMEOUT_SECONDS,
        )

        if response.status_code >= 400:

            raise Exception(
                "Failed to stage RAID configuration. "
                f"HTTP {response.status_code}: "
                f"{response.text[:1000]}"
            )

        return selected_drives


    # =========================================================================
    # RAID VERIFICATION
    # =========================================================================

    def poll_for_raid(
        self,
        baseline_count,
        timeout_minutes,
        expected_storage_uri=None,
        expected_drive_uris=None,
    ):

        # DMTF writes are verified with the dedicated short-poll routine. This
        # avoids the old 30-minute loop when HPE returns equivalent drive links
        # under a different URI namespace.
        if expected_storage_uri:

            volume = (
                self.wait_for_dmtf_raid1_volume(
                    expected_storage_uri,
                    expected_drive_uris,
                    timeout_seconds=(
                        DMTF_APPLY_TIMEOUT_SECONDS
                    ),
                )
            )

            if volume:

                log(
                    self.row_number,
                    self.ip,
                    (
                        "SUCCESS: RAID1 volume verified "
                        "through DMTF Storage."
                    ),
                )

                return True

            return False

        max_attempts = max(
            1,
            int(
                (
                    timeout_minutes
                    * 60
                )
                / POLL_INTERVAL_SECONDS
            ),
        )

        for attempt in range(
            1,
            max_attempts + 1,
        ):

            time.sleep(
                POLL_INTERVAL_SECONDS
            )

            try:

                (
                    logical_drives,
                    source,
                    controllers,
                ) = (
                    self.get_existing_logical_drives()
                )

                for logical_drive in logical_drives:

                    raid_value = str(
                        logical_drive.get(
                            "Raid",
                            logical_drive.get(
                                "RAIDType",
                                logical_drive.get(
                                    "VolumeType",
                                    "",
                                ),
                            ),
                        )
                    ).lower()

                    if (
                        "raid1" in raid_value
                        or "mirrored" in raid_value
                    ):

                        log(
                            self.row_number,
                            self.ip,
                            (
                                "SUCCESS: RAID1 logical "
                                f"volume detected via {source}."
                            ),
                        )

                        return True

            except Exception:

                pass

            if (
                attempt == 1
                or attempt % 2 == 1
            ):

                log(
                    self.row_number,
                    self.ip,
                    (
                        "Polling RAID inventory... "
                        f"(Attempt {attempt}/{max_attempts})"
                    ),
                )

        return False


# =============================================================================
# 5. SERVER WORKER
# =============================================================================

def process_server(
    server_data,
    overwrite_raid,
    ctrl_timeout,
    raid_timeout,
):

    ip = (
        server_data[
            "ilo_ip"
        ]
    )

    row_number = (
        server_data[
            "excel_row"
        ]
    )

    server = ServerProcessor(
        ip,
        row_number,
        USERNAME,
        PASSWORD,
    )

    booted_to_utilities = False

    try:

        log(
            row_number,
            ip,
            (
                "Starting DL380 Gen10+ "
                "RAID configuration process..."
            ),
        )

        server.authenticate()

        # =====================================================================
        # READ CURRENT RAID
        # =====================================================================

        log(
            row_number,
            ip,
            (
                "Reading existing storage/RAID "
                "inventory while server is online..."
            ),
        )

        (
            existing_ld,
            inventory_source,
            controllers,
        ) = (
            server.get_existing_logical_drives()
        )

        if existing_ld:

            log(
                row_number,
                ip,
                (
                    f"Existing RAID detected via "
                    f"{inventory_source}: "
                    f"{len(existing_ld)} logical "
                    "volume(s)."
                ),
            )

            if not overwrite_raid:

                log(
                    row_number,
                    ip,
                    (
                        "Existing RAID found. "
                        "Skipping because Overwrite=No."
                    ),
                )

                return {
                    "row":
                        row_number,

                    "ip":
                        ip,

                    "status":
                        "Skipped",

                    "reason":
                        (
                            "Existing RAID found "
                            f"via {inventory_source} "
                            f"({len(existing_ld)} volume(s))"
                        ),
                }

            log(
                row_number,
                ip,
                (
                    "Overwrite enabled; existing "
                    "RAID will be replaced."
                ),
            )

        else:

            log(
                row_number,
                ip,
                (
                    "No existing RAID logical drive "
                    "detected from online inventory."
                ),
            )

        baseline_count = len(
            existing_ld
        )

        # =====================================================================
        # DISCOVER BOTH RAID WRITE MODELS
        # =====================================================================

        dmtf_targets = (
            server.discover_dmtf_storage_targets()
        )

        server.log_dmtf_storage_targets(
            dmtf_targets
        )

        dmtf_target = (
            server.select_dmtf_storage_target(
                dmtf_targets
            )
        )

        (
            config_uri,
            drive_locations,
            config_data,
        ) = (
            server.select_smartstorage_config()
        )

        write_interface = None

        if (
            dmtf_target
            and dmtf_target[
                "writable"
            ]
            and len(
                dmtf_target[
                    "drive_uris"
                ]
            )
            >= RAID_DATA_DRIVE_COUNT
        ):

            write_interface = {
                "method":
                    "DMTF",

                "dmtf_target":
                    dmtf_target,
            }

            log(
                row_number,
                ip,
                (
                    "Using writable DMTF Redfish Storage "
                    "interface. Legacy SmartStorageConfig "
                    "is not required for this controller."
                ),
            )

        elif (
            config_uri
            and len(
                drive_locations
            )
            >= RAID_DATA_DRIVE_COUNT
        ):

            write_interface = {
                "method":
                    "SmartStorageConfig",

                "config_uri":
                    config_uri,

                "drive_locations":
                    drive_locations,

                "config_data":
                    config_data,
            }

            log(
                row_number,
                ip,
                (
                    "Using legacy HPE "
                    "SmartStorageConfig write interface."
                ),
            )

        # =====================================================================
        # FALLBACK TO SYSTEM UTILITIES ONLY IF NEITHER MODEL IS READY
        # =====================================================================

        if not write_interface:

            log(
                row_number,
                ip,
                (
                    "No writable RAID interface is currently "
                    "ready. Rebooting to System Utilities and "
                    "retrying both DMTF Storage and "
                    "SmartStorageConfig..."
                ),
            )

            server.reboot_to_system_utilities()

            booted_to_utilities = True

            write_interface = (
                server.wait_for_raid_write_interface(
                    ctrl_timeout
                )
            )

            if not write_interface:

                return {
                    "row":
                        row_number,

                    "ip":
                        ip,

                    "status":
                        "Failed",

                    "reason":
                        (
                            "Neither writable DMTF Storage "
                            "nor SmartStorageConfig became "
                            "available after System Utilities "
                            f"for {ctrl_timeout} mins"
                        ),
                }

            (
                existing_after_boot,
                inventory_after_boot,
                _,
            ) = (
                server.get_existing_logical_drives()
            )

            if existing_after_boot:

                baseline_count = len(
                    existing_after_boot
                )

                log(
                    row_number,
                    ip,
                    (
                        "Existing RAID visible after "
                        "pre-OS transition via "
                        f"{inventory_after_boot}: "
                        f"{baseline_count} volume(s)."
                    ),
                )

        # =====================================================================
        # CREATE/STAGE RAID1
        # =====================================================================

        selected_drives = []
        expected_storage_uri = None
        expected_controller_name = None
        expected_drive_uris = None

        if (
            write_interface[
                "method"
            ]
            == "DMTF"
        ):

            target = (
                write_interface[
                    "dmtf_target"
                ]
            )

            log(
                row_number,
                ip,
                (
                    "Selected DMTF storage target: "
                    f"{target['storage_uri']} | "
                    f"{target['controller_label']}"
                ),
            )

            log(
                row_number,
                ip,
                (
                    "Physical drives exposed through "
                    "DMTF Storage: "
                    f"{target['drive_uris']}"
                ),
            )

            selected_drives = (
                server.stage_dmtf_raid1_configuration(
                    target,
                    overwrite_raid,
                )
            )

            expected_storage_uri = (
                target[
                    "storage_uri"
                ]
            )

            expected_controller_name = str(
                (
                    target.get(
                        "storage_data",
                        {},
                    )
                    or {}
                ).get(
                    "Name",
                    "",
                )
                or ""
            ).strip()

            expected_drive_uris = list(
                selected_drives
            )

        else:

            config_uri = (
                write_interface[
                    "config_uri"
                ]
            )

            drive_locations = (
                write_interface[
                    "drive_locations"
                ]
            )

            log(
                row_number,
                ip,
                (
                    "SmartStorageConfig available: "
                    f"{config_uri}"
                ),
            )

            log(
                row_number,
                ip,
                (
                    "Physical drives exposed for "
                    f"configuration: {drive_locations}"
                ),
            )

            selected_drives = (
                server.stage_raid1_configuration(
                    config_uri,
                    drive_locations,
                )
            )

        # =====================================================================
        # APPLY / INITIALIZE / VERIFY
        # =====================================================================

        if (
            write_interface[
                "method"
            ]
            == "DMTF"
        ):

            # iLO 5 applies MR-controller volume creation on reboot. Boot into
            # System Utilities so the old OS is never allowed to start before
            # the recreated RAID volume has been initialized.
            log(
                row_number,
                ip,
                (
                    "RAID configuration request accepted. "
                    "Rebooting into System Utilities to apply "
                    "the new DMTF RAID configuration while "
                    "preventing the old OS from booting..."
                ),
            )

            server.reboot_to_system_utilities()

            booted_to_utilities = True

            # Do NOT wait for PostState=FinishedPost here.
            #
            # System Utilities/BiosSetup is itself a pre-OS POST destination,
            # and iLO can legitimately remain at InPostDiscoveryComplete while
            # the utility screen is already available. Start polling the DMTF
            # controller and RAID volume immediately after the reboot
            # transition instead.
            created_volume = (
                server.wait_for_dmtf_raid1_volume(
                    expected_storage_uri,
                    expected_drive_uris,
                    timeout_seconds=(
                        DMTF_APPLY_TIMEOUT_SECONDS
                    ),
                    context=(
                        "System Utilities RAID apply"
                    ),
                    expected_controller_name=(
                        expected_controller_name
                    ),
                )
            )

            if not created_volume:

                return {
                    "row":
                        row_number,

                    "ip":
                        ip,

                    "status":
                        "Failed",

                    "reason":
                        (
                            "RAID1 create request was accepted "
                            "but the new RAID1 volume did not "
                            "appear on the expected DMTF storage "
                            f"target within "
                            f"{DMTF_APPLY_TIMEOUT_SECONDS}s"
                        ),
                }

            # A brand-new RAID volume is always initialized.  When an
            # existing RAID was overwritten/recreated, initialization follows
            # the centralized policy in vars.py.
            existing_raid_was_overwritten = (
                overwrite_raid
                and baseline_count > 0
            )

            initialize_volume = (
                True
                if not existing_raid_was_overwritten
                else INITIALIZE_OVERWRITTEN_VOLUME
            )

            if initialize_volume:

                if existing_raid_was_overwritten:

                    log(
                        row_number,
                        ip,
                        (
                            "Existing RAID was overwritten. "
                            "Initialization policy is ENABLED; "
                            f"initializing recreated RAID1 using "
                            f"{INITIALIZE_TYPE} initialization."
                        ),
                    )

                else:

                    log(
                        row_number,
                        ip,
                        (
                            "New RAID1 volume detected. "
                            "New RAID volumes are always initialized "
                            f"using {INITIALIZE_TYPE} initialization."
                        ),
                    )

                server.initialize_dmtf_volume(
                    created_volume,
                    INITIALIZE_TYPE,
                )

                initialization_status = (
                    "{}-initialized"
                    .format(
                        INITIALIZE_TYPE
                    )
                )

            else:

                log(
                    row_number,
                    ip,
                    (
                        "Existing RAID was overwritten. "
                        "Initialization policy is DISABLED; "
                        "the recreated RAID1 will NOT be initialized."
                    ),
                )

                initialization_status = (
                    "not initialized by policy"
                )

            # Restore normal boot only after RAID1 exists and the configured
            # initialization policy has been applied.
            log(
                row_number,
                ip,
                (
                    "RAID1 creation is verified and the configured "
                    "initialization policy has been applied. "
                    "Restoring normal boot and rebooting the server..."
                ),
            )

            server.reboot_normal()

            booted_to_utilities = False

            final_volume = (
                server.wait_for_dmtf_raid1_volume(
                    expected_storage_uri,
                    expected_drive_uris,
                    timeout_seconds=(
                        DMTF_APPLY_TIMEOUT_SECONDS
                    ),
                    context=(
                        "final post-reboot verification"
                    ),
                    expected_controller_name=(
                        expected_controller_name
                    ),
                )
            )

            if not final_volume:

                return {
                    "row":
                        row_number,

                    "ip":
                        ip,

                    "status":
                        "Failed",

                    "reason":
                        (
                            "RAID1 was created and initialized "
                            "but could not be verified after "
                            "the final reboot"
                        ),
                }

            return {
                "row":
                    row_number,

                "ip":
                    ip,

                "status":
                    "Successful",

                "reason":
                    (
                        "RAID1 created, "
                        f"{initialization_status}, "
                        "and verified using DMTF on "
                        f"{', '.join(selected_drives)}"
                    ),
            }

        # Legacy SmartStorageConfig path.
        log(
            row_number,
            ip,
            (
                "RAID configuration staged. "
                "Rebooting server to apply "
                "SmartStorageConfig settings..."
            ),
        )

        server.reboot_normal()

        server.wait_for_post_complete(
            max(
                180,
                int(
                    ctrl_timeout
                    * 60
                ),
            )
        )

        success = (
            server.poll_for_raid(
                baseline_count,
                raid_timeout,
            )
        )

        if success:

            return {
                "row":
                    row_number,

                "ip":
                    ip,

                "status":
                    "Successful",

                "reason":
                    (
                        "RAID1 created successfully "
                        "using SmartStorageConfig on "
                        f"{', '.join(selected_drives)}"
                    ),
            }

        return {
            "row":
                row_number,

            "ip":
                ip,

            "status":
                "Failed",

            "reason":
                (
                    "SmartStorageConfig RAID1 was staged "
                    "but verification timed out after "
                    f"{raid_timeout} mins"
                ),
        }

    except Exception as exc:

        try:

            if booted_to_utilities:

                server.set_boot_to_normal()

        except Exception:

            pass

        return {
            "row":
                row_number,

            "ip":
                ip,

            "status":
                "Failed",

            "reason":
                str(exc),
        }

    finally:

        server.session.close()


# =============================================================================
# 6. MAIN
# =============================================================================

def main():

    print(
        "SCRIPT BUILD: {}".format(
            SCRIPT_BUILD
        )
    )

    excel_path = EXCEL_PATH

    if not os.path.exists(
        excel_path
    ):

        print(
            "ERROR: Could not find "
            f"Excel file at {excel_path}"
        )

        sys.exit(1)

    try:

        print(
            f"Loading data from "
            f"{excel_path} "
            f"(Sheet: '{SHEET_NAME}')..."
        )

        end_text = (
            END_ROW
            if END_ROW is not None
            else "end of data"
        )

        print(
            (
                "Excel row range: "
                f"{START_ROW} through "
                f"{end_text} (inclusive)"
            )
        )

        print(
            (
                "Optimized load enabled: "
                "within the selected range, "
                "reading stops after "
                f"{EXCEL_EMPTY_ROW_STOP} "
                "consecutive empty rows."
            )
        )

        df = (
            load_resource_excel_fast(
                excel_path,
                SHEET_NAME,
                START_ROW,
                END_ROW,
                EXCEL_EMPTY_ROW_STOP,
            )
        )

    except Exception as exc:

        print(
            f"Failed to read Excel file: {exc}"
        )

        sys.exit(1)

    if df.empty:

        print(
            "No populated rows found "
            "in selected range."
        )

        sys.exit(0)

    # -------------------------------------------------------------------------
    # Validate columns
    # -------------------------------------------------------------------------

    if (
        "equipment_type"
        not in df.columns
    ):

        print(
            "ERROR: 'equipment_type' "
            "column is missing."
        )

        sys.exit(1)

    missing_columns = [
        column
        for column
        in REQUIRED_COLUMNS
        if column not in df.columns
    ]

    if missing_columns:

        print(
            "ERROR: Missing required "
            "column(s): "
            + ", ".join(
                missing_columns
            )
        )

        sys.exit(1)

    # -------------------------------------------------------------------------
    # Equipment filter
    # -------------------------------------------------------------------------

    equipment_series = (
        df[
            "equipment_type"
        ]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    df_filtered = df[
        equipment_series.isin(
            VALID_EQUIPMENT_TYPES
        )
    ].copy()

    if df_filtered.empty:

        print(
            "No matching servers found "
            "for equipment types: "
            + ", ".join(
                VALID_EQUIPMENT_TYPES
            )
        )

        sys.exit(0)

    servers_to_process = []
    final_report = []

    for _, row in (
        df_filtered.iterrows()
    ):

        excel_row = int(
            row[
                "_excel_row"
            ]
        )

        raw_ip = (
            row.get(
                "ilo_ip"
            )
        )

        ip = (
            str(
                raw_ip
            ).strip()
            if pd.notna(
                raw_ip
            )
            else "Unknown"
        )

        missing = [

            column

            for column
            in REQUIRED_COLUMNS

            if (
                column not in row
                or pd.isna(
                    row[
                        column
                    ]
                )
                or str(
                    row[
                        column
                    ]
                ).strip()
                == ""
            )
        ]

        if missing:

            final_report.append(
                {
                    "row":
                        excel_row,

                    "ip":
                        ip,

                    "status":
                        "Skipped",

                    "reason":
                        (
                            "Missing required value(s): "
                            + ", ".join(
                                missing
                            )
                        ),
                }
            )

            continue

        servers_to_process.append(
            {
                "excel_row":
                    excel_row,

                "ilo_ip":
                    ip,
            }
        )

    total_found = len(
        servers_to_process
    )

    print(
        "\nTotal valid DL380 Gen10+ "
        "servers found for processing: "
        f"{total_found}"
    )

    if total_found == 0:

        print(
            "No valid servers left "
            "to process."
        )

        sys.exit(0)

    user_input = input(
        "\nDo you want to OVERWRITE "
        "servers with already configured "
        "RAID volumes? (yes/no): "
    ).strip().lower()

    overwrite_raid = (
        user_input == "yes"
    )

    print(
        "\n"
        + "=" * 90
    )

    print(
        "STARTING PARALLEL "
        "DL380 GEN10+ RAID CONFIGURATION"
    )

    print(
        "Equipment types : "
        + ", ".join(
            VALID_EQUIPMENT_TYPES
        )
    )

    print(
        f"RAID level      : {RAID_LEVEL}"
    )

    print(
        f"Data drives     : "
        f"{RAID_DATA_DRIVE_COUNT}"
    )

    print(
        (
            "Overwrite RAID  : "
            + (
                "YES"
                if overwrite_raid
                else "NO"
            )
        )
    )

    print(
        f"Max workers     : {MAX_WORKERS}"
    )

    print(
        "=" * 90
    )

    safe_workers = min(
        MAX_WORKERS,
        total_found,
    )

    # =========================================================================
    # PARALLEL EXECUTION
    # =========================================================================

    with concurrent.futures.ThreadPoolExecutor(
        max_workers=safe_workers
    ) as executor:

        futures = {

            executor.submit(
                process_server,
                server,
                overwrite_raid,
                CONTROLLER_TIMEOUT_MINUTES,
                RAID_TIMEOUT_MINUTES,
            ):
                (server, time.time())

            for server
            in servers_to_process
        }

        for future in (
            concurrent.futures
            .as_completed(
                futures
            )
        ):

            (
                server_data,
                submitted_at,
            ) = futures[
                future
            ]

            try:

                result = (
                    future.result()
                )

            except Exception as exc:

                result = {
                    "row":
                        server_data[
                            "excel_row"
                        ],

                    "ip":
                        server_data[
                            "ilo_ip"
                        ],

                    "status":
                        "Failed",

                    "reason":
                        (
                            "Unhandled worker exception: "
                            f"{exc}"
                        ),
                }

            result[
                "time_seconds"
            ] = round(
                time.time()
                - submitted_at,
                2,
            )

            final_report.append(
                result
            )

            log(
                result[
                    "row"
                ],
                result[
                    "ip"
                ],
                (
                    "FINISHED - Status: "
                    f"{result['status']}"
                ),
            )

    # =========================================================================
    # REPORT
    # =========================================================================

    final_report = sorted(
        final_report,
        key=lambda item:
            item[
                "row"
            ],
    )

    print(
        "\n\n"
        + "=" * 110
    )

    print(
        "FINAL EXECUTION REPORT: "
        "HPE DL380 Gen10+"
    )

    print(
        "=" * 110
    )

    print(
        f"{'ROW':<7} | "
        f"{'iLO IP ADDRESS':<16} | "
        f"{'STATUS':<12} | "
        f"REASON / DETAILS"
    )

    print(
        "-" * 110
    )

    for server_result in (
        final_report
    ):

        print(
            f"{server_result['row']:<7} | "
            f"{server_result['ip']:<16} | "
            f"{server_result['status']:<12} | "
            f"{server_result['reason']}"
        )

    print(
        "=" * 110
        + "\n"
    )

    # =========================================================================
    # CSV
    # =========================================================================

    if final_report:

        log_dir = LOG_DIR

        os.makedirs(
            log_dir,
            exist_ok=True,
        )

        timestamp = (
            time.strftime(
                "%Y%m%d-%H%M%S"
            )
        )

        range_text = (
            f"rows_{START_ROW}-"
            f"{END_ROW if END_ROW is not None else 'end'}"
        )

        report_filename = (
            f"{REPORT_PREFIX}_"
            f"{range_text}_"
            f"{timestamp}.csv"
        )

        report_path = (
            os.path.join(
                log_dir,
                report_filename,
            )
        )

        report_df = (
            pd.DataFrame(
                final_report
            )
        )

        report_df = report_df[
            [
                "row",
                "ip",
                "status",
                "time_seconds",
                "reason",
            ]
        ]

        report_df.to_csv(
            report_path,
            index=False,
        )

        print(
            "Detailed CSV saved to: "
            f"{report_path}\n"
        )


# =============================================================================
# ENTRY POINT
# =============================================================================

    summary_rows = [
        make_summary_row(
            row=item.get("row", "-"),
            item_type="RAID Rackmount",
            name="DL380 Gen10+",
            target=item.get("ip", ""),
            status=item.get("status", "Unknown"),
            time_seconds=item.get("time_seconds", 0.0),
            details=item.get("reason", ""),
        )
        for item in final_report
    ]

    print_summary_report(
        summary_rows,
        title="FINAL RAID RACK-MOUNT SUMMARY",
    )

    write_summary_csv(
        summary_rows,
        vars.SCRIPT_ARTIFACT_PREFIXES[
            "configure_raid_RM"
        ],
    )

    return (
        1
        if any(
            item.get("status") == "Failed"
            for item in final_report
        )
        else 0
    )


if __name__ == "__main__":

    sys.exit(
        run_logged_main(
            main,
            log_prefix=vars.SCRIPT_ARTIFACT_PREFIXES[
                "configure_raid_RM"
            ],
            title="RACK-MOUNT RAID CONFIGURATION",
        )
    )